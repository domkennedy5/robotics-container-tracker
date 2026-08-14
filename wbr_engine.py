"""
wbr_engine.py — WBR metric calculations for Robotics Destination Dray

Methodology (finalized Aug 2026 — QBR spec):
- Population:  GVT Terminal Gateout date in week (Mon–Sun ISO), Movement Key = 1.
               De-duplicated: one row per container.
- AV→OA:      AV = GVT Available SSL/Dray Carrier (pre-resolved by GVT, no tier logic needed).
               OA = GVT Terminal Gateout.
               SLA clock starts NEXT business day after AV date.
               Elapsed = business days (Mon–Fri, US federal holidays excluded). SLA ≤ 3.
               Missing AV → exclude container from denominator entirely.
- OA→Del:     OA = GVT Terminal Gateout. SLA clock starts next business day after OA.
               4-tier delivery source:
                 Tier 1: GVT Arrived at Site
                 Tier 2: OBLT EVENT_526 / status LY
                 Tier 3: OBLT EVENT_511 / status RD
                 Tier 4: GVT Enter Facility
               Containers with no delivery date in any tier = in-transit; excluded from denom.
               Excl A320-RBTCS (non-operational). Elapsed = business days. SLA ≤ 3.
- EL→CL:      EL 3-tier: OBLT EVENT_511/RD → OBLT EVENT_517/ON → GVT Container Empty.
               CL 4-tier: OBLT EVENT_563/Y5 → OBLT EVENT_518/ED → OBLT EVENT_519/ER
                           → GVT Return to Port.
               No next-business-day shift (clock starts on EL date).
               Containers with EL but no CL = in-progress; excluded from denom.
               Stored in empty_term_pct DB/display key for backward compatibility.
- E2E:         OBLT VD (Vessel Departure) → GVT Enter Facility; completed only.
               Calendar days. Exclude negative-day deltas (data error).
- OTP:         GVT Enter Facility (date-only) ≤ GVT Current Dray SLA (date-only).
               Delivered containers only. GVT is sole source.

Data source roles (finalized):
  GVT    — PRIMARY: population (Terminal Gateout, Movement Key), AV (pre-resolved),
            OA, OA→Del Tier 1 + 4, EL Tier 3, CL Tier 4, E2E endpoint, OTP
  OBLT   — OA→Del Tiers 2+3, EL Tiers 1+2, CL Tiers 1-3, E2E start (VD)
  IL/ISS — Supplementary context only; NOT used in any metric calculation

Backward compatibility:
  When GVT file does not contain terminal_gateout data (pre-Aug 2026 uploads), the engine
  automatically falls back to the legacy OBLT-primary logic so historical re-runs still work.
"""
from __future__ import annotations

import io
import math
from datetime import date, datetime, timedelta
from typing import Any

import numpy as np
import openpyxl
import pandas as pd

AV_OA_SLA_DAYS  = 3
OA_DEL_SLA_DAYS = 3
EMPTY_TERM_SLA  = 3
EXCL_FACILITY   = 'A320-RBTCS'


def week_bounds(week_num: int, year: int) -> tuple:
    """Return (week_start, week_end) as Mon–Sun per ISO week (finalized QBR spec)."""
    iso_monday = date.fromisocalendar(year, week_num, 1)
    iso_sunday = iso_monday + timedelta(days=6)
    return iso_monday, iso_sunday


# ── Business-day utilities (finalized QBR methodology) ───────────────────────

def us_federal_holidays(year: int) -> frozenset:
    """All 11 US federal holidays for year as a frozenset of date objects.
    Weekend shifts applied: Saturday → Friday observed, Sunday → Monday observed.
    """
    import calendar as _cal

    def _nth(y, m, wd, n):          # nth (1-based) weekday (0=Mon) in month
        d = date(y, m, 1)
        d += timedelta(days=(wd - d.weekday()) % 7)
        return d + timedelta(weeks=n - 1)

    def _last(y, m, wd):            # last occurrence of weekday in month
        last = date(y, m, _cal.monthrange(y, m)[1])
        return last - timedelta(days=(last.weekday() - wd) % 7)

    def _obs(d):                    # observed date: Sat→Fri, Sun→Mon
        if d.weekday() == 5: return d - timedelta(days=1)
        if d.weekday() == 6: return d + timedelta(days=1)
        return d

    return frozenset({
        _obs(date(year,  1,  1)),   # New Year's Day
        _nth(year,  1, 0, 3),       # MLK Day            3rd Mon Jan
        _nth(year,  2, 0, 3),       # Presidents' Day    3rd Mon Feb
        _last(year, 5, 0),          # Memorial Day       last Mon May
        _obs(date(year,  6, 19)),   # Juneteenth
        _obs(date(year,  7,  4)),   # Independence Day
        _nth(year,  9, 0, 1),       # Labor Day          1st Mon Sep
        _nth(year, 10, 0, 2),       # Columbus Day       2nd Mon Oct
        _obs(date(year, 11, 11)),   # Veterans Day
        _nth(year, 11, 3, 4),       # Thanksgiving       4th Thu Nov
        _obs(date(year, 12, 25)),   # Christmas Day
    })


def _hols(start: date, end: date) -> list:
    """Sorted holiday date-strings for np.busday_count covering [start, end]."""
    all_h: set = set()
    for y in range(start.year, end.year + 1):
        all_h.update(us_federal_holidays(y))
    return [str(d) for d in sorted(all_h) if start <= d <= end]


def next_business_day(d: date) -> date:
    """First Mon–Fri non-holiday day strictly after d."""
    hols: set = set()
    for y in {d.year, (d + timedelta(days=7)).year}:
        hols.update(us_federal_holidays(y))
    nxt = d + timedelta(days=1)
    while nxt.weekday() >= 5 or nxt in hols:
        nxt += timedelta(days=1)
    return nxt


def business_days_elapsed(start: date, end: date) -> int:
    """Business days in [start, end] inclusive. Returns 0 for negative delta.

    AV→OA and OA→Del: pass start = next_business_day(event_date).
    EL→CL: pass start = el_date directly (no next-day shift per spec).
    """
    if pd.isna(start) or pd.isna(end):
        return 0
    if isinstance(start, pd.Timestamp): start = start.date()
    if isinstance(end,   pd.Timestamp): end   = end.date()
    if end < start:
        return 0
    h = _hols(start, end)
    # np.busday_count([a,b)) exclusive of b — add one day for inclusive end
    return max(0, int(np.busday_count(str(start), str(end + timedelta(days=1)), holidays=h)))


def guess_week(ready_dates, report_date=None) -> tuple:
    """Return (week_num, year) of the WBR week.

    GVT files span 6-12 prior weeks so a naive most-common picks the wrong
    week.  When report_date is provided we narrow to a 7-day backward window
    (the WBR Sun-Sat week sits immediately before report_date Monday).
    """
    dates = pd.to_datetime(ready_dates, errors='coerce').dropna()
    if dates.empty:
        return None, date.today().year
    if report_date is not None:
        rpt_ts = pd.Timestamp(report_date)
        window = dates[
            (dates >= rpt_ts - pd.Timedelta(days=7)) &
            (dates <= rpt_ts + pd.Timedelta(days=1))
        ]
        if not window.empty:
            dates = window
    iso_weeks = dates.apply(lambda d: (d + timedelta(days=1)).isocalendar()[:2])
    most_common = pd.Series(iso_weeks).value_counts().idxmax()
    return int(most_common[1]), int(most_common[0])


def guess_week_from_oblt(oblt: pd.DataFrame, report_date=None) -> tuple:
    """Return (week_num, year) from OBLT EVENT_511 (RD/gate-out) dates.

    Preferred over GVT-based detection when OBLT is the population source.
    For Robotics live unloads, EVENT_511 (RD) is the population ingestion trigger.
    Falls back to None if no RD events are present.
    """
    rd_dates = oblt[oblt['status'] == 'RD']['date'].dropna()
    if rd_dates.empty:
        return None, date.today().year
    return guess_week(rd_dates, report_date)


def _to_date(v):
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    if isinstance(v, str):
        try: return datetime.fromisoformat(v).date()
        except: return None
    return None


def _to_ts(v):
    if v is None: return pd.NaT
    if isinstance(v, (datetime, date)): return pd.Timestamp(v)
    if isinstance(v, str):
        try: return pd.Timestamp(v)
        except: return pd.NaT
    return pd.NaT


def load_gvt(file_bytes: bytes) -> pd.DataFrame:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    hdrs = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    h = {v: i for i, v in enumerate(hdrs) if v}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        ctr = row[h.get('Container', -1)]
        if not ctr: continue
        rows.append({
            'container':          str(ctr).strip(),
            'ready_date':         _to_date(row[h.get('Ready Date/Time', -1)]),
            'port':               row[h.get('Discharged Port', -1)],
            'facility':           row[h.get('Facility', -1)],
            'market':             row[h.get('Market', -1)],
            'scac':               row[h.get('Dray SCAC(FL)', -1)],
            'enter_facility':     _to_ts(row[h.get('Enter Facility', -1)]),
            'container_empty':    _to_ts(row[h.get('Container Empty', -1)]),
            'return_to_port':     _to_ts(row[h.get('Return to Port', -1)]),
            'status':             row[h.get('Container Status', -1)],
            'terminal_avail':     _to_ts(row[h.get('Terminal Avail', -1)]),
            'current_dray_sla':   _to_ts(row[h.get('Current Dray SLA', -1)]),
            'category':           row[h.get('Category', -1)],
            # QBR methodology columns (present in all weekly GVT files)
            'terminal_gateout':   _to_ts(row[h.get('Terminal Gateout', -1)]),
            'available_ssl_dray': _to_ts(row[h.get('Available SSL/Dray Carrier', -1)]),
            'movement_key':       row[h.get('Container Gateout Movement Key', -1)],
            'arrived_at_site':    _to_ts(row[h.get('Arrived at Site', -1)]),
            'port_region':        row[h.get('Port Region', -1)],
        })
    wb.close()
    return pd.DataFrame(rows)


def load_oblt(file_bytes: bytes) -> pd.DataFrame:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb['ObltData'] if 'ObltData' in wb.sheetnames else wb.active
    hdrs = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    h = {v: i for i, v in enumerate(hdrs) if v}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        ctr = row[h.get('tracking_id', -1)]
        status = row[h.get('status', -1)]
        if not ctr or status not in ('AV','OA','RD','VD','LY','ON','Y5','ED','ER'): continue
        rows.append({
            'container': str(ctr).strip(),
            'status':    status,
            'event':     str(row[h.get('status_event', -1)] or '').strip(),
            'date':      _to_ts(row[h.get('status_date', -1)]),
        })
    wb.close()
    df = pd.DataFrame(rows)
    df['date'] = pd.to_datetime(df['date'], errors='coerce')
    if 'event' not in df.columns:
        df['event'] = ''
    return df


def load_inbound_loads(file_bytes: bytes) -> pd.DataFrame:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb['Inbound Loads'] if 'Inbound Loads' in wb.sheetnames else list(wb)[0]
    rows_raw = list(ws.iter_rows(values_only=True))
    if not rows_raw: return pd.DataFrame()
    hdrs = rows_raw[0]
    h = {v: i for i, v in enumerate(hdrs) if v}
    rows = []
    for row in rows_raw[1:]:
        ctr = row[h.get('Container Id', -1)]
        if not ctr: continue
        rows.append({
            'container':      str(ctr).strip(),
            'po_promised':    _to_date(row[h.get('PO Promised Date', -1)]),
            'actual_arrival': _to_date(row[h.get('Actual Arrival At Final Destination', -1)]),
        })
    wb.close()
    return pd.DataFrame(rows)


def _latest(oblt: pd.DataFrame, status: str):
    sub = oblt[oblt['status'] == status].sort_values('date')
    return sub.groupby('container')['date'].last()


def _latest_by_event(oblt: pd.DataFrame, event: str, status: str | None = None) -> pd.Series:
    """Latest date per container for a specific OBLT event code (e.g. 'EVENT_511').
    Optionally filter by status value as well. Returns empty Series if event column absent.
    """
    if 'event' not in oblt.columns or oblt.empty:
        return pd.Series(dtype='datetime64[ns]')
    mask = (oblt['event'] == event) & oblt['date'].notna()
    if status:
        mask &= (oblt['status'] == status)
    sub = oblt[mask].sort_values('date')
    return sub.groupby('container')['date'].last()


def compute_metrics(gvt, oblt, inbound, week_start, week_end, report_date) -> dict:
    """Route to QBR methodology (v2) when GVT contains terminal_gateout data,
    or fall back to legacy OBLT-primary logic (v1) for older file uploads.
    Signature and output dict shape are unchanged — app.py needs no updates.
    """
    has_tgo = ('terminal_gateout' in gvt.columns and gvt['terminal_gateout'].notna().any())
    if has_tgo:
        return _compute_metrics_v2(gvt, oblt, week_start, week_end)
    return _compute_metrics_v1(gvt, oblt, week_start, week_end, report_date)


def _compute_metrics_v1(gvt, oblt, week_start, week_end, report_date) -> dict:
    """Legacy OBLT-primary methodology (pre-Aug 2026 file format).
    Kept for backward compatibility when terminal_gateout column is absent.
    """
    rd_wk = oblt[
        (oblt['status'] == 'RD') &
        oblt['date'].notna() &
        (oblt['date'].dt.date >= week_start) &
        (oblt['date'].dt.date <= week_end)
    ].sort_values('date')
    pop_rd   = rd_wk.groupby('container').last().reset_index()
    pop_ctrs = set(pop_rd['container'])
    n = len(pop_ctrs)
    if n == 0:
        return _empty_metrics()

    rd_pop_ev = pop_rd.set_index('container')['date']

    gvt_pop = gvt[gvt['container'].isin(pop_ctrs)].copy()
    if not gvt_pop.empty:
        gvt_pop = (gvt_pop.sort_values('enter_facility', na_position='last')
                          .groupby('container').last().reset_index())
    gvt_facility = (gvt_pop.set_index('container')['facility']
                    if not gvt_pop.empty else pd.Series(dtype=str))

    av_ev = _latest(oblt, 'AV')
    rd_ev = _latest(oblt, 'RD')
    vd_ev = _latest(oblt, 'VD')
    rpt   = pd.Timestamp(report_date)

    if not gvt_pop.empty and 'terminal_avail' in gvt_pop.columns:
        ta = gvt_pop.set_index('container')['terminal_avail'].dropna()
        missing_av = ta[~ta.index.isin(av_ev.index)]
        if not missing_av.empty:
            av_ev = pd.concat([av_ev, missing_av])

    ao_days = []; ao_met = 0; ao_den = 0
    for c in pop_ctrs:
        hav = c in av_ev.index; hoa = c in rd_pop_ev.index
        if hoa and hav:
            d = (rd_pop_ev[c] - av_ev[c]).total_seconds() / 86400
            if d > 0: ao_days.append(d)
            ao_den += 1
            if d <= AV_OA_SLA_DAYS: ao_met += 1
        elif hoa and not hav:
            ao_den += 1; ao_met += 1
        elif hav and not hoa:
            if (rpt - av_ev[c]).total_seconds() / 86400 > AV_OA_SLA_DAYS:
                ao_den += 1

    pop_scored_ctrs = {c for c in pop_ctrs if gvt_facility.get(c, '') != EXCL_FACILITY}
    od_days = []; od_met = 0; od_den = 0
    for c in pop_scored_ctrs:
        oa = rd_pop_ev[c]
        d  = (rd_ev[c] - oa).total_seconds() / 86400 if c in rd_ev.index else 0.0
        if d >= 0: od_days.append(d)
        od_den += 1
        if d <= OA_DEL_SLA_DAYS: od_met += 1

    et_days = [0.0] * n; et_met = n; et_den = n

    gvt_ef = (gvt_pop.set_index('container')['enter_facility']
              if not gvt_pop.empty and 'enter_facility' in gvt_pop.columns
              else pd.Series(dtype='datetime64[ns]'))
    e2e = []
    for c in pop_ctrs:
        if c not in vd_ev.index: continue
        ef = gvt_ef.get(c)
        if pd.isna(ef): continue
        d = (ef - vd_ev[c]).total_seconds() / 86400
        if d > 0: e2e.append(d)

    otp_met = 0; otp_den = 0
    if (not gvt_pop.empty and 'current_dray_sla' in gvt_pop.columns
            and gvt_pop['current_dray_sla'].notna().any()):
        scored = gvt_pop[
            gvt_pop['enter_facility'].notna() & gvt_pop['current_dray_sla'].notna()
        ].copy()
        if not scored.empty:
            scored['del_date'] = scored['enter_facility'].dt.normalize()
            scored['sla_date'] = scored['current_dray_sla'].dt.normalize()
            otp_met = int((scored['del_date'] <= scored['sla_date']).sum())
            otp_den = len(scored)

    def _pct(m, d): return round(m/d*100) if d else None
    def _avg(lst):  return round(float(np.mean(lst)), 1) if lst else None
    def _p90(lst):  return math.ceil(float(np.percentile(lst, 90))) if lst else None

    return {
        'containers':     n,
        'av_oa_avg':      _avg(ao_days),
        'av_oa_sla_pct':  _pct(ao_met, ao_den),
        'av_oa_p90':      _p90(ao_days),
        'oa_del_avg':     _avg(od_days),
        'oa_del_sla_pct': _pct(od_met, od_den),
        'oa_del_p90':     _p90(od_days),
        'empty_term_pct': _pct(et_met, et_den) if et_den else None,
        'e2e_avg':        round(float(np.mean(e2e))) if e2e else None,
        'otp_pct':        _pct(otp_met, otp_den),
        'week_start':     week_start.isoformat(),
        'week_end':       week_end.isoformat(),
    }


def _compute_metrics_v2(gvt, oblt, week_start, week_end) -> dict:
    """QBR-finalized methodology (Aug 2026+).

    Population: GVT Terminal Gateout in week (Mon-Sun), Movement Key = 1.
    AV→OA: business days, GVT pre-resolved AV, next-bday clock start.
    OA→Del: 4-tier delivery, business days, next-bday clock start.
    EL→CL: 3+4-tier, business days, same-day clock start.
    E2E + OTP: unchanged from v1.
    """
    def _pct(m, d): return round(m/d*100) if d else None
    def _avg(lst):  return round(float(np.mean(lst)), 1) if lst else None
    def _p90(lst):  return math.ceil(float(np.percentile(lst, 90))) if lst else None
    def _ts_date(v):
        if pd.isna(v): return None
        return v.date() if isinstance(v, pd.Timestamp) else v

    # ── Population ────────────────────────────────────────────────────────────
    pop_mask = (
        gvt['terminal_gateout'].notna() &
        (gvt['terminal_gateout'].dt.date >= week_start) &
        (gvt['terminal_gateout'].dt.date <= week_end)
    )
    if 'movement_key' in gvt.columns:
        pop_mask &= (gvt['movement_key'] == 1)
    pop_gvt = gvt[pop_mask].copy()
    # De-dup: one row per container (keep latest terminal_gateout)
    pop_gvt = pop_gvt.sort_values('terminal_gateout').groupby('container').last().reset_index()
    pop_ctrs = set(pop_gvt['container'])
    n = len(pop_ctrs)
    if n == 0:
        return _empty_metrics()

    gvt_idx = pop_gvt.set_index('container')

    # ── OBLT event lookups (full file, not week-restricted) ───────────────────
    oblt_526_ly = _latest_by_event(oblt, 'EVENT_526', 'LY')  # OA→Del Tier 2
    oblt_511_rd = _latest_by_event(oblt, 'EVENT_511', 'RD')  # OA→Del Tier 3 + EL Tier 1
    oblt_517_on = _latest_by_event(oblt, 'EVENT_517', 'ON')  # EL Tier 2
    oblt_563_y5 = _latest_by_event(oblt, 'EVENT_563', 'Y5')  # CL Tier 1
    oblt_518_ed = _latest_by_event(oblt, 'EVENT_518', 'ED')  # CL Tier 2
    oblt_519_er = _latest_by_event(oblt, 'EVENT_519', 'ER')  # CL Tier 3
    vd_ev       = _latest(oblt, 'VD')                         # E2E start

    # ── AV→OA ────────────────────────────────────────────────────────────────
    # AV = GVT available_ssl_dray (pre-resolved). OA = GVT terminal_gateout.
    # Clock starts next business day after AV. Missing AV → skip (not in denom).
    ao_days = []; ao_met = 0; ao_den = 0
    for c in pop_ctrs:
        if c not in gvt_idx.index: continue
        row = gvt_idx.loc[c]
        av_d = _ts_date(row.get('available_ssl_dray'))
        oa_d = _ts_date(row.get('terminal_gateout'))
        if av_d is None or oa_d is None:
            continue  # can't compute without both dates
        clock = next_business_day(av_d)
        elapsed = business_days_elapsed(clock, oa_d)
        ao_days.append(elapsed)
        ao_den += 1
        if elapsed <= AV_OA_SLA_DAYS: ao_met += 1

    # ── OA→Del (4-tier, excl A320-RBTCS) ─────────────────────────────────────
    # OA = GVT terminal_gateout. Clock starts next business day after OA.
    # Containers with no delivery date in any tier are in-transit → not in denom.
    od_days = []; od_met = 0; od_den = 0
    for c in pop_ctrs:
        if c not in gvt_idx.index: continue
        row = gvt_idx.loc[c]
        if str(row.get('facility', '')).strip() == EXCL_FACILITY:
            continue
        oa_d = _ts_date(row.get('terminal_gateout'))
        if oa_d is None: continue
        clock = next_business_day(oa_d)

        # Tier cascade: first non-null delivery date wins
        del_d = None
        if del_d is None: del_d = _ts_date(row.get('arrived_at_site'))
        if del_d is None and c in oblt_526_ly.index: del_d = _ts_date(oblt_526_ly[c])
        if del_d is None and c in oblt_511_rd.index: del_d = _ts_date(oblt_511_rd[c])
        if del_d is None: del_d = _ts_date(row.get('enter_facility'))

        if del_d is None:
            continue  # in-transit: no delivery date yet → exclude from denom

        elapsed = business_days_elapsed(clock, del_d)
        od_days.append(elapsed)
        od_den += 1
        if elapsed <= OA_DEL_SLA_DAYS: od_met += 1

    # ── EL→CL (stored as empty_term_pct for DB/display compat) ───────────────
    # EL: EVENT_511/RD → EVENT_517/ON → GVT container_empty.
    # CL: EVENT_563/Y5 → EVENT_518/ED → EVENT_519/ER → GVT return_to_port.
    # No next-bday shift. Containers with EL but no CL are in-progress → not in denom.
    el_days = []; el_met = 0; el_den = 0
    for c in pop_ctrs:
        row = gvt_idx.loc[c] if c in gvt_idx.index else None

        # EL tier cascade
        el_d = None
        if c in oblt_511_rd.index: el_d = _ts_date(oblt_511_rd[c])
        if el_d is None and c in oblt_517_on.index: el_d = _ts_date(oblt_517_on[c])
        if el_d is None and row is not None: el_d = _ts_date(row.get('container_empty'))
        if el_d is None: continue  # no EL event → skip

        # CL tier cascade
        cl_d = None
        if c in oblt_563_y5.index: cl_d = _ts_date(oblt_563_y5[c])
        if cl_d is None and c in oblt_518_ed.index: cl_d = _ts_date(oblt_518_ed[c])
        if cl_d is None and c in oblt_519_er.index: cl_d = _ts_date(oblt_519_er[c])
        if cl_d is None and row is not None: cl_d = _ts_date(row.get('return_to_port'))
        if cl_d is None: continue  # not yet returned → exclude from denom

        elapsed = business_days_elapsed(el_d, cl_d)  # same-day start, no next-bday shift
        el_days.append(elapsed)
        el_den += 1
        if elapsed <= EMPTY_TERM_SLA: el_met += 1

    # ── E2E: OBLT VD → GVT Enter Facility (calendar days, unchanged) ─────────
    gvt_ef = (gvt_idx['enter_facility']
              if 'enter_facility' in gvt_idx.columns
              else pd.Series(dtype='datetime64[ns]'))
    e2e = []
    for c in pop_ctrs:
        if c not in vd_ev.index: continue
        ef = gvt_ef.get(c)
        if pd.isna(ef): continue
        d = (ef - vd_ev[c]).total_seconds() / 86400
        if d > 0: e2e.append(d)

    # ── OTP: GVT Enter Facility ≤ GVT Current Dray SLA (unchanged) ───────────
    otp_met = 0; otp_den = 0
    if ('current_dray_sla' in gvt_idx.columns and gvt_idx['current_dray_sla'].notna().any()):
        scored = pop_gvt[
            pop_gvt['enter_facility'].notna() & pop_gvt['current_dray_sla'].notna()
        ].copy()
        if not scored.empty:
            scored['del_date'] = scored['enter_facility'].dt.normalize()
            scored['sla_date'] = scored['current_dray_sla'].dt.normalize()
            otp_met = int((scored['del_date'] <= scored['sla_date']).sum())
            otp_den = len(scored)

    return {
        'containers':     n,
        'av_oa_avg':      _avg(ao_days),
        'av_oa_sla_pct':  _pct(ao_met, ao_den),
        'av_oa_p90':      _p90(ao_days),
        'oa_del_avg':     _avg(od_days),
        'oa_del_sla_pct': _pct(od_met, od_den),
        'oa_del_p90':     _p90(od_days),
        'empty_term_pct': _pct(el_met, el_den) if el_den else None,
        'e2e_avg':        round(float(np.mean(e2e))) if e2e else None,
        'otp_pct':        _pct(otp_met, otp_den),
        'week_start':     week_start.isoformat(),
        'week_end':       week_end.isoformat(),
    }



def compute_carrier_scorecard(gvt, oblt, week_start, week_end, report_date) -> list:
    """Per-carrier AV→OA and OA→Del breakdown for the WBR carrier scorecard.
    Routes to QBR methodology (v2) when terminal_gateout is present, else legacy (v1).
    Returns list of dicts sorted by volume descending — shape unchanged.
    """
    has_tgo = ('terminal_gateout' in gvt.columns and gvt['terminal_gateout'].notna().any())
    if has_tgo:
        return _scorecard_v2(gvt, oblt, week_start, week_end)
    return _scorecard_v1(gvt, oblt, week_start, week_end, report_date)


def _scorecard_v1(gvt, oblt, week_start, week_end, report_date) -> list:
    """Legacy OBLT-primary scorecard (pre-Aug 2026 file format)."""
    rd_wk = oblt[
        (oblt['status'] == 'RD') &
        oblt['date'].notna() &
        (oblt['date'].dt.date >= week_start) &
        (oblt['date'].dt.date <= week_end)
    ].sort_values('date')
    pop_rd   = rd_wk.groupby('container').last().reset_index()
    pop_ctrs = set(pop_rd['container'])
    if not pop_ctrs:
        return []

    rd_pop_ev = pop_rd.set_index('container')['date']

    gvt_pop = gvt[gvt['container'].isin(pop_ctrs)].copy()
    if not gvt_pop.empty:
        gvt_pop = (gvt_pop.sort_values('enter_facility', na_position='last')
                          .groupby('container').last().reset_index())
    gvt_scac     = (gvt_pop.set_index('container')['scac']
                    if not gvt_pop.empty and 'scac' in gvt_pop.columns
                    else pd.Series(dtype=str))
    gvt_facility = (gvt_pop.set_index('container')['facility']
                    if not gvt_pop.empty else pd.Series(dtype=str))

    av_ev = _latest(oblt, 'AV')
    rd_ev = _latest(oblt, 'RD')
    rpt   = pd.Timestamp(report_date)

    carrier_map: dict = {}
    for c in pop_ctrs:
        scac = gvt_scac.get(c)
        label = str(scac).strip() if scac and not pd.isna(scac) and str(scac).strip() else 'Unknown'
        carrier_map.setdefault(label, set()).add(c)

    scorecard = []
    for carrier_label, ctrs in carrier_map.items():
        ao_days = []; ao_met = 0; ao_den = 0
        for c in ctrs:
            hav = c in av_ev.index; hoa = c in rd_pop_ev.index
            if hoa and hav:
                d = (rd_pop_ev[c] - av_ev[c]).total_seconds() / 86400
                if d > 0: ao_days.append(d)
                ao_den += 1
                if d <= AV_OA_SLA_DAYS: ao_met += 1
            elif hoa and not hav:
                ao_den += 1; ao_met += 1
            elif hav and not hoa:
                if (rpt - av_ev[c]).total_seconds() / 86400 > AV_OA_SLA_DAYS:
                    ao_den += 1

        scored_ctrs = {c for c in ctrs if gvt_facility.get(c, '') != EXCL_FACILITY}
        od_days = []; od_met = 0; od_den = 0
        for c in scored_ctrs:
            oa = rd_pop_ev[c]
            d  = (rd_ev[c] - oa).total_seconds() / 86400 if c in rd_ev.index else 0.0
            if d >= 0: od_days.append(d)
            od_den += 1
            if d <= OA_DEL_SLA_DAYS: od_met += 1

        scorecard.append({
            'carrier':        carrier_label,
            'volume':         len(ctrs),
            'av_oa_sla_pct':  round(ao_met/ao_den*100) if ao_den else None,
            'av_oa_misses':   (ao_den - ao_met) if ao_den else 0,
            'av_oa_avg':      round(float(np.mean(ao_days)), 1) if ao_days else None,
            'av_oa_p90':      round(float(np.percentile(ao_days, 90))) if ao_days else None,
            'oa_del_sla_pct': round(od_met/od_den*100) if od_den else None,
            'oa_del_misses':  (od_den - od_met) if od_den else 0,
            'oa_del_avg':     round(float(np.mean(od_days)), 1) if od_days else None,
            'oa_del_p90':     round(float(np.percentile(od_days, 90))) if od_days else None,
        })

    return sorted(scorecard, key=lambda x: -x['volume'])


def _scorecard_v2(gvt, oblt, week_start, week_end) -> list:
    """QBR-finalized per-carrier scorecard (Aug 2026+).
    Population: GVT Terminal Gateout in week, Movement Key = 1.
    AV→OA: business days, pre-resolved GVT AV, next-bday clock start.
    OA→Del: 4-tier delivery, business days, next-bday clock start.
    """
    def _ts_date(v):
        if pd.isna(v): return None
        return v.date() if isinstance(v, pd.Timestamp) else v

    pop_mask = (
        gvt['terminal_gateout'].notna() &
        (gvt['terminal_gateout'].dt.date >= week_start) &
        (gvt['terminal_gateout'].dt.date <= week_end)
    )
    if 'movement_key' in gvt.columns:
        pop_mask &= (gvt['movement_key'] == 1)
    pop_gvt = gvt[pop_mask].copy()
    pop_gvt = pop_gvt.sort_values('terminal_gateout').groupby('container').last().reset_index()
    pop_ctrs = set(pop_gvt['container'])
    if not pop_ctrs:
        return []

    gvt_idx = pop_gvt.set_index('container')

    # OBLT delivery lookups
    oblt_526_ly = _latest_by_event(oblt, 'EVENT_526', 'LY')
    oblt_511_rd = _latest_by_event(oblt, 'EVENT_511', 'RD')

    # Per-carrier grouping from GVT scac
    carrier_map: dict = {}
    for c in pop_ctrs:
        row = gvt_idx.loc[c] if c in gvt_idx.index else None
        scac = row.get('scac') if row is not None else None
        label = str(scac).strip() if scac and not pd.isna(scac) and str(scac).strip() else 'Unknown'
        carrier_map.setdefault(label, set()).add(c)

    scorecard = []
    for carrier_label, ctrs in carrier_map.items():
        ao_days = []; ao_met = 0; ao_den = 0
        for c in ctrs:
            if c not in gvt_idx.index: continue
            row = gvt_idx.loc[c]
            av_d = _ts_date(row.get('available_ssl_dray'))
            oa_d = _ts_date(row.get('terminal_gateout'))
            if av_d is None or oa_d is None: continue
            clock = next_business_day(av_d)
            elapsed = business_days_elapsed(clock, oa_d)
            ao_days.append(elapsed)
            ao_den += 1
            if elapsed <= AV_OA_SLA_DAYS: ao_met += 1

        scored_ctrs = {c for c in ctrs
                       if str(gvt_idx.loc[c].get('facility','') if c in gvt_idx.index else '').strip() != EXCL_FACILITY}
        od_days = []; od_met = 0; od_den = 0
        for c in scored_ctrs:
            if c not in gvt_idx.index: continue
            row = gvt_idx.loc[c]
            oa_d = _ts_date(row.get('terminal_gateout'))
            if oa_d is None: continue
            clock = next_business_day(oa_d)
            del_d = None
            if del_d is None: del_d = _ts_date(row.get('arrived_at_site'))
            if del_d is None and c in oblt_526_ly.index: del_d = _ts_date(oblt_526_ly[c])
            if del_d is None and c in oblt_511_rd.index: del_d = _ts_date(oblt_511_rd[c])
            if del_d is None: del_d = _ts_date(row.get('enter_facility'))
            if del_d is None: continue
            elapsed = business_days_elapsed(clock, del_d)
            od_days.append(elapsed)
            od_den += 1
            if elapsed <= OA_DEL_SLA_DAYS: od_met += 1

        scorecard.append({
            'carrier':        carrier_label,
            'volume':         len(ctrs),
            'av_oa_sla_pct':  round(ao_met/ao_den*100) if ao_den else None,
            'av_oa_misses':   (ao_den - ao_met) if ao_den else 0,
            'av_oa_avg':      round(float(np.mean(ao_days)), 1) if ao_days else None,
            'av_oa_p90':      round(float(np.percentile(ao_days, 90))) if ao_days else None,
            'oa_del_sla_pct': round(od_met/od_den*100) if od_den else None,
            'oa_del_misses':  (od_den - od_met) if od_den else 0,
            'oa_del_avg':     round(float(np.mean(od_days)), 1) if od_days else None,
            'oa_del_p90':     round(float(np.percentile(od_days, 90))) if od_days else None,
        })

    return sorted(scorecard, key=lambda x: -x['volume'])

def _empty_metrics() -> dict:
    return {k: None for k in [
        'containers','av_oa_avg','av_oa_sla_pct','av_oa_p90',
        'oa_del_avg','oa_del_sla_pct','oa_del_p90',
        'empty_term_pct','e2e_avg','otp_pct','week_start','week_end'
    ]}


def compute_totals(weeks_data: list) -> dict:
    valid = [w for w in weeks_data if w and w.get('containers')]
    if not valid: return _empty_metrics()
    total_vol = sum(w['containers'] for w in valid)

    def wavg(k):
        vals = [(w[k], w['containers']) for w in valid if w.get(k) is not None]
        return round(sum(v*c for v,c in vals)/sum(c for _,c in vals), 1) if vals else None

    def _raw(k):
        vals = [(w[k], w['containers']) for w in valid if w.get(k) is not None]
        return sum(v*c for v,c in vals)/sum(c for _,c in vals) if vals else None

    def wavgi(k):
        v = _raw(k); return int(v + 0.5) if v is not None else None

    return {
        'containers':     total_vol,
        'av_oa_avg':      wavg('av_oa_avg'),
        'av_oa_sla_pct':  wavgi('av_oa_sla_pct'),
        'av_oa_p90':      wavgi('av_oa_p90'),
        'oa_del_avg':     wavg('oa_del_avg'),
        'oa_del_sla_pct': wavgi('oa_del_sla_pct'),
        'oa_del_p90':     wavgi('oa_del_p90'),
        'empty_term_pct': wavgi('empty_term_pct'),
        'e2e_avg':        wavgi('e2e_avg'),
        'otp_pct':        wavgi('otp_pct'),
        'week_start':     None, 'week_end': None,
    }


WBR_SCHEMA = '''
CREATE TABLE IF NOT EXISTS wbr_results (
    id             INTEGER PRIMARY KEY AUTOINCREMENT,
    year           INTEGER NOT NULL,
    week_num       INTEGER NOT NULL,
    week_label     TEXT,
    week_start     TEXT,
    week_end       TEXT,
    containers     INTEGER,
    av_oa_avg      REAL,
    av_oa_sla_pct  INTEGER,
    av_oa_p90      INTEGER,
    oa_del_avg     REAL,
    oa_del_sla_pct INTEGER,
    oa_del_p90     INTEGER,
    empty_term_pct INTEGER,
    e2e_avg        INTEGER,
    otp_pct        INTEGER,
    generated_at   TEXT,
    UNIQUE(year, week_num)
);
'''


def save_week_to_db(conn, year, week_num, m, generated_at):
    conn.execute('''
        INSERT OR REPLACE INTO wbr_results
        (year,week_num,week_label,week_start,week_end,containers,
         av_oa_avg,av_oa_sla_pct,av_oa_p90,
         oa_del_avg,oa_del_sla_pct,oa_del_p90,
         empty_term_pct,e2e_avg,otp_pct,generated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    ''', (year, week_num, f'W{week_num}', m.get('week_start'), m.get('week_end'), m.get('containers'),
          m.get('av_oa_avg'), m.get('av_oa_sla_pct'), m.get('av_oa_p90'),
          m.get('oa_del_avg'), m.get('oa_del_sla_pct'), m.get('oa_del_p90'),
          m.get('empty_term_pct'), m.get('e2e_avg'), m.get('otp_pct'), generated_at))
    conn.commit()


def load_weeks_from_db(conn, year, week_nums):
    placeholders = ','.join('?' * len(week_nums))
    # Exclude seed/placeholder rows — only load data from real WBR runs
    rows = conn.execute(
        f'SELECT * FROM wbr_results WHERE year=? AND week_num IN ({placeholders})'
        f' AND (generated_at IS NULL OR generated_at != ?)',
        [year] + list(week_nums) + ['seed']
    ).fetchall()
    result = {}
    for row in rows:
        d = dict(row)
        result[d['week_num']] = {
            'containers':     d['containers'],
            'av_oa_avg':      d['av_oa_avg'],
            'av_oa_sla_pct':  d['av_oa_sla_pct'],
            'av_oa_p90':      d['av_oa_p90'],
            'oa_del_avg':     d['oa_del_avg'],
            'oa_del_sla_pct': d['oa_del_sla_pct'],
            'oa_del_p90':     d['oa_del_p90'],
            'empty_term_pct': d['empty_term_pct'],
            'e2e_avg':        d['e2e_avg'],
            'otp_pct':        d['otp_pct'],
            'week_start':     d.get('week_start'),
            'week_end':       d.get('week_end'),
        }
    return result




def load_market_context(conn, days_back: int = 14) -> dict:
    """
    Load the latest port intel, freight rate, and macro signal data from DB.
    Returns a dict ready for bridge injection.
    Called at WBR generate time — data populated daily by port-intel-scraper Lambda.

    Returns:
        {
          'freight':  str  — Drewry WCI headline (latest)
          'macro':    str  — GSCPI reading (latest)
          'ports': {
              'USSAV': [headlines...],
              'USORF': [headlines...],
              'USLAX': [headlines...],
              'USBOS': [headlines...],
              'ALL':   [ocean freight headlines...],
          },
          'has_data': bool
        }
    """
    from datetime import datetime, timedelta, timezone
    cutoff = (datetime.now(timezone.utc) - timedelta(days=days_back)).strftime("%Y-%m-%dT%H:%M:%SZ")

    result = {"freight": None, "macro": None, "ports": {}, "has_data": False}

    try:
        # Freight rate — latest Drewry WCI
        row = conn.execute(
            "SELECT headline, summary FROM port_intel "
            "WHERE source='Drewry WCI' ORDER BY scraped_at DESC LIMIT 1"
        ).fetchone()
        if row:
            result["freight"] = row[0]
            result["has_data"] = True

        # Macro signal — latest GSCPI
        row = conn.execute(
            "SELECT value, period FROM macro_signals "
            "WHERE source='NY Fed' AND metric='GSCPI' ORDER BY scraped_at DESC LIMIT 1"
        ).fetchone()
        if row:
            v, period = row
            if v > 1.0:
                level = "significantly elevated"
            elif v > 0.5:
                level = "elevated"
            elif v < -1.0:
                level = "significantly below normal"
            elif v < -0.5:
                level = "below normal"
            else:
                level = "near normal"
            result["macro"] = f"GSCPI {v:+.2f}σ ({period}) — supply chain pressure {level}"
            result["has_data"] = True

        # Port headlines — top 3 per port from past days_back days
        for port in ("USSAV", "USORF", "USLAX", "USBOS", "ALL"):
            rows = conn.execute(
                "SELECT headline, source, published_date FROM port_intel "
                "WHERE port_code=? AND source != 'Drewry WCI' AND scraped_at >= ? "
                "ORDER BY scraped_at DESC LIMIT 3",
                (port, cutoff)
            ).fetchall()
            if rows:
                result["ports"][port] = [
                    {"headline": r[0], "source": r[1], "date": r[2]} for r in rows
                ]
                result["has_data"] = True

    except Exception as e:
        # Table may not exist yet (Lambda not yet deployed) — fail silently
        print(f"load_market_context: {e}")

    return result


def format_market_context_block(ctx: dict) -> str:
    """
    Format market context dict into a plain-text bridge block.
    Returns empty string if no data available.
    """
    if not ctx.get("has_data"):
        return ""

    lines = ["[Market Context]"]

    if ctx.get("freight"):
        lines.append(f"Freight Rates: {ctx['freight']}")

    if ctx.get("macro"):
        lines.append(f"Macro: {ctx['macro']}")

    port_labels = {
        "USSAV": "Savannah (USSAV)",
        "USORF": "Norfolk/Virginia (USORF)",
        "USLAX": "Los Angeles (USLAX)",
        "USBOS": "Boston (USBOS)",
        "ALL":   "Ocean / Industry",
    }

    port_section = []
    for port, label in port_labels.items():
        headlines = ctx["ports"].get(port, [])
        if headlines:
            # Use the most recent headline only for bridge brevity
            h = headlines[0]
            port_section.append(f"  {label}: {h['headline']}")

    if port_section:
        lines.append("Port Intelligence:")
        lines.extend(port_section)

    return "\n".join(lines)


def parse_wbr_pdf(pdf_bytes: bytes):
    """Parse a previously generated WBR PDF and extract the weekly table data.

    Coordinate-independent: scans all text on the page, locates week-column headers
    by pattern matching (e.g. "W25", "W 25", "2026 W 25"), then reads cell values
    by proximity. Works regardless of which app version generated the PDF.

    Returns:
        week_labels : list[str]  – e.g. ["W25","W26","W27","W28","W29"]
        week_data   : dict[int, dict]  – keyed by week_num int, values are metrics dicts
    """
    import fitz
    import re

    ROW_KEYS = [
        "containers",  "av_oa_avg",  "av_oa_sla_pct",  "av_oa_p90",
        "oa_del_avg",  "oa_del_sla_pct",  "oa_del_p90",
        "empty_term_pct", "e2e_avg", "otp_pct",
    ]
    INT_KEYS = {"containers", "av_oa_sla_pct", "av_oa_p90",
                "oa_del_sla_pct", "oa_del_p90",
                "empty_term_pct", "e2e_avg", "otp_pct"}

    doc  = fitz.open(stream=pdf_bytes, filetype="pdf")
    page = doc[0]
    PAGE_H = page.rect.height

    # Extract ALL words with bounding boxes: (x0,y0,x1,y1,text,block,line,word)
    all_words = page.get_text("words")
    doc.close()

    if not all_words:
        return [], {}

    words_list = [(w[0], w[1], w[2], w[3], w[4]) for w in all_words]  # x0,y0,x1,y1,text

    # ── Step 1: Find week-column headers by scanning for "W XX" / "WXX" tokens ──
    col_headers = {}   # week_num → x_center
    col_header_ys = {}  # week_num → y0 of the "W" token — tracked in same scan as col_headers
    i = 0
    while i < len(words_list):
        x0, y0, x1, y1, txt = words_list[i]
        # Match "W25" as single token
        m = re.fullmatch(r"W(\d{1,2})", txt.strip(), re.IGNORECASE)
        if m:
            wn = int(m.group(1))
            if 1 <= wn <= 53:
                col_headers[wn] = (x0 + x1) / 2
                col_header_ys[wn] = y0
            i += 1
            continue
        # Match "W" followed by a digit token on the same y-band (e.g. "2026 W 25")
        if txt.strip().upper() == "W" and i + 1 < len(words_list):
            nx0, ny0, nx1, ny1, ntxt = words_list[i + 1]
            if abs(ny0 - y0) < 8:
                try:
                    wn = int(ntxt.strip())
                    if 1 <= wn <= 53:
                        col_headers[wn] = (x0 + nx1) / 2
                        col_header_ys[wn] = y0
                    i += 2
                    continue
                except ValueError:
                    pass
        i += 1

    if not col_headers:
        return [], {}

    # ── Step 2: Determine header row y, then derive all row y-positions ──
    # Use y0 from the SAME scan that built col_headers (last assignment = main table).
    # This avoids the fallback PAGE_H*0.57 which was 4.6pt off, causing every row
    # window to capture the row above instead of the correct one.
    hdr_y = float(np.median(list(col_header_ys.values()))) if col_header_ys else PAGE_H * 0.57

    ROW_H = 16.0  # fixed row height in both app and Quick AI slides
    # row_ys[ri] = y-center of data row ri (ri=0 → Containers, ri=9 → OTP)
    # In fitz coords (y increases downward), rows are below the header
    row_ys = [hdr_y + ROW_H * (ri + 1) for ri in range(len(ROW_KEYS))]

    # ── Step 3: Extract cell values by proximity ──────────────────────────
    # Sort col_headers by x position
    sorted_cols = sorted(col_headers.items(), key=lambda kv: kv[1])  # (wn, x_center)
    week_nums_sorted = [kv[0] for kv in sorted_cols]
    col_xs = [kv[1] for kv in sorted_cols]

    # Column width estimate
    if len(col_xs) > 1:
        col_w = min(abs(col_xs[j+1] - col_xs[j]) for j in range(len(col_xs)-1))
    else:
        col_w = 80.0

    # Row height estimate
    if len(row_ys) > 1:
        row_h = np.median([abs(row_ys[j+1] - row_ys[j]) for j in range(len(row_ys)-1)])
    else:
        row_h = ROW_H_EST

    # Skip "Total" column (typically the rightmost; x > last data col + col_w*0.6)
    # It's already excluded since we only captured "W XX" headers (not "Total")

    week_data = {}
    for ri, key in enumerate(ROW_KEYS):
        ry = row_ys[ri]
        for ci, wn in enumerate(week_nums_sorted):
            cx = col_xs[ci]
            # Clip rect centered on (cx, ry) with ±col_w/2 x, ±row_h/2 y
            half_w = col_w * 0.48
            half_h = row_h * 0.55
            cell_words = [
                (x0, y0, x1, y1, txt)
                for x0, y0, x1, y1, txt in words_list
                if abs((x0+x1)/2 - cx) < half_w and abs((y0+y1)/2 - ry) < half_h
            ]
            if not cell_words:
                continue
            # Join all text in cell, strip %, dashes
            raw = " ".join(w[4] for w in sorted(cell_words, key=lambda w: w[0]))
            raw = raw.replace("%", "").replace("\u2013", "").replace("\u2014", "").replace("–", "").replace("—", "").strip()
            if not raw:
                continue
            try:
                val = float(raw)
                if wn not in week_data:
                    week_data[wn] = {}
                week_data[wn][key] = int(round(val)) if key in INT_KEYS else round(val, 1)
            except ValueError:
                pass

    week_labels = [f"W{wn}" for wn in week_nums_sorted]
    return week_labels, week_data
