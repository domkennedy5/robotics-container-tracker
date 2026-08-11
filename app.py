from __future__ import annotations
import streamlit as st
import pandas as pd
import sqlite3
import openpyxl
import io
import os
from datetime import datetime, date, timedelta
from zoneinfo import ZoneInfo

_EASTERN = ZoneInfo("America/New_York")
_PHX     = ZoneInfo("America/Phoenix")   # MST year-round (no DST)
_CENTRAL = ZoneInfo("America/Chicago")


def _ct_to_mst(hour: int, minute: int = 0) -> str:
    """Return formatted MST equivalent of a CT clock time (uses today's offset)."""
    _t = datetime.now(_CENTRAL).replace(hour=hour, minute=minute, second=0, microsecond=0)
    return _t.astimezone(_PHX).strftime('%I:%M %p').lstrip('0')
from inbound_forecast import render_inbound_forecast_tab

from streamlit_js_eval import streamlit_js_eval
from utils import (
    normalize_container, containers_match,
    parse_container_list, parse_carrier_file,
    parse_carrier_template, TEMPLATE_SHEET_MAP,
)
import data_sync

# ── page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Robotics Container Tracker",
    page_icon=None,
    layout="wide",
    initial_sidebar_state="expanded",
)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH  = os.path.join(BASE_DIR, "tracker.db")
VENDOR_PORTAL_URL     = "https://robotics-container-tracker-7uf88f7ez9tga3k44phfjm.streamlit.app/vendor_upload"
CARRIER_PLAN_BASE_URL = "https://robotics-container-tracker-7uf88f7ez9tga3k44phfjm.streamlit.app/carrier_plan"

# Friendly display labels for carrier codes
CARRIER_DISPLAY = {
    "HDDR": "HDDR (Maersk/HUDD)",
}

def _apply_carrier_display(df: "pd.DataFrame") -> "pd.DataFrame":
    """Map raw carrier codes to display names in-place (copy)."""
    if "carrier_name" in df.columns:
        df = df.copy()
        df["carrier_name"] = df["carrier_name"].map(
            lambda x: CARRIER_DISPLAY.get(x, x)
        )
    return df

# ── AWS secrets (graceful fallback if not configured) ─────────────────────────
def _aws_cfg():
    try:
        return (
            st.secrets["aws"]["AWS_ACCESS_KEY_ID"],
            st.secrets["aws"]["AWS_SECRET_ACCESS_KEY"],
            st.secrets["aws"]["AWS_REGION"],
            st.secrets["aws"]["S3_BUCKET"],
        )
    except Exception:
        return None, None, None, None

AWS_KEY, AWS_SECRET, AWS_REGION, S3_BUCKET = _aws_cfg()
S3_ENABLED = all([AWS_KEY, AWS_SECRET, AWS_REGION, S3_BUCKET])

# ── password gate ──────────────────────────────────────────────────────────────
def _check_password():
    if st.session_state.get("authenticated"):
        return True

    try:
        correct = st.secrets["password"]
    except Exception:
        correct = "robotics2026"

    st.title("Robotics Container Tracker")
    pwd = st.text_input("Password", type="password")
    if st.button("Sign in"):
        if pwd == correct:
            st.session_state.authenticated = True
            st.rerun()
        else:
            st.error("Incorrect password.")
    return False

if not _check_password():
    st.stop()

# ── S3 startup sync ────────────────────────────────────────────────────────────
if S3_ENABLED and not st.session_state.get("s3_startup_done"):
    data_sync.pull_db_from_s3(AWS_KEY, AWS_SECRET, AWS_REGION, S3_BUCKET)
    # Load last generated WBR PDF from S3 so dashboard is always visible on fresh load
    if not st.session_state.get("wbr_last_pdf"):
        _pdf_s3, _meta_s3 = data_sync.pull_wbr_pdf_from_s3(AWS_KEY, AWS_SECRET, AWS_REGION, S3_BUCKET)
        if _pdf_s3:
            st.session_state["wbr_last_pdf"]  = _pdf_s3
            st.session_state["wbr_last_week"] = _meta_s3.get("week")
            st.session_state["wbr_last_rd"]   = _meta_s3.get("date", "")
    st.session_state.s3_startup_done = True

# ── database ───────────────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS carrier_submissions (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            submitted_at     TEXT    NOT NULL,
            carrier_name     TEXT    NOT NULL,
            container_id     TEXT    NOT NULL,
            sheet_type       TEXT,
            port             TEXT,
            terminal         TEXT,
            fc_building      TEXT,
            flexi_id         TEXT,
            outgate_date     TEXT,
            delivery_date    TEXT,
            status           TEXT,
            within_sla       TEXT,
            sla_notes        TEXT,
            empty_return_due TEXT,
            appointment_date TEXT,
            accessorial_type TEXT,
            notes            TEXT,
            source_file      TEXT,
            source           TEXT DEFAULT 'web'
        );
        CREATE TABLE IF NOT EXISTS lookup_log (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            searched_at     TEXT    NOT NULL,
            requester       TEXT,
            container_ids   TEXT    NOT NULL,
            found_count     INTEGER,
            not_found_count INTEGER
        );
        CREATE TABLE IF NOT EXISTS alerts_log (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            alert_type   TEXT NOT NULL,
            container_id TEXT NOT NULL,
            sent_at      TEXT NOT NULL,
            recipient    TEXT
        );
        CREATE TABLE IF NOT EXISTS carrier_rates (
            id                  INTEGER PRIMARY KEY AUTOINCREMENT,
            carrier_name        TEXT NOT NULL UNIQUE,
            rate_per_container  REAL DEFAULT 0.0,
            notes               TEXT
        );
        CREATE TABLE IF NOT EXISTS rate_lanes (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            scac         TEXT NOT NULL,
            carrier_name TEXT,
            port         TEXT NOT NULL,
            destination  TEXT NOT NULL,
            dest_address TEXT,
            base_rate    REAL NOT NULL,
            lane_type    TEXT,
            lane_id      TEXT,
            rate_source  TEXT,
            imported_at  TEXT
        );
        CREATE TABLE IF NOT EXISTS route_lanes (
            id                INTEGER PRIMARY KEY AUTOINCREMENT,
            origin_port       TEXT,
            dest_port         TEXT,
            transit_notes     TEXT,
            node_code         TEXT NOT NULL,
            node_address      TEXT,
            fc_code           TEXT,
            node_type         TEXT,
            program_id        TEXT,
            warehouse_partner TEXT,
            dray_bid_done     TEXT,
            cost_cards_done   TEXT,
            imported_at       TEXT
        );
        CREATE TABLE IF NOT EXISTS shipment_status (
            id                  INTEGER PRIMARY KEY AUTOINCREMENT,
            container_no        TEXT,
            current_status      TEXT,
            fc                  TEXT,
            vendor_name         TEXT,
            po_no               TEXT,
            edd                 TEXT,
            vessel_name         TEXT,
            voyage_no           TEXT,
            carrier_code        TEXT,
            discharge_city      TEXT,
            discharge_eta       TEXT,
            final_dest_eta      TEXT,
            container_available TEXT,
            container_out       TEXT,
            empty_return        TEXT,
            delivered           TEXT,
            customs_cleared     TEXT,
            container_size      TEXT,
            load_port_country   TEXT,
            load_port_city      TEXT,
            source_file         TEXT,
            imported_at         TEXT
        );
        CREATE TABLE IF NOT EXISTS delivery_plan (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            week_start   TEXT NOT NULL,
            appt_date    TEXT,
            appt_time    TEXT,
            slot_num     INTEGER,
            container_id TEXT NOT NULL,
            carrier      TEXT NOT NULL,
            product_type TEXT DEFAULT 'SHELF, HDTP',
            qty          INTEGER,
            notes        TEXT,
            status       TEXT DEFAULT 'SCHEDULED',
            created_at   TEXT,
            updated_at   TEXT
        );
                CREATE TABLE IF NOT EXISTS inbound_loads (
            id                       INTEGER PRIMARY KEY AUTOINCREMENT,
            container_id             TEXT,
            supplier                 TEXT,
            po_number                TEXT,
            location_code            TEXT,
            destination_fc           TEXT,
            origin_port              TEXT,
            dest_port                TEXT,
            vessel_name              TEXT,
            eta_discharge            TEXT,
            actual_arrival_discharge TEXT,
            gateout_terminal         TEXT,
            last_free_detention      TEXT,
            last_free_demurrage      TEXT,
            est_appointment          TEXT,
            eta_final                TEXT,
            actual_arrival_final     TEXT,
            empty_return_time        TEXT,
            source_file              TEXT,
            imported_at              TEXT
        );
        CREATE TABLE IF NOT EXISTS dbr_receipts (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            carrier       TEXT NOT NULL,
            week_start    TEXT NOT NULL,
            received_date TEXT NOT NULL,
            received_via  TEXT DEFAULT 'manual',
            file_name     TEXT,
            notes         TEXT,
            logged_at     TEXT NOT NULL,
            UNIQUE(carrier, week_start, received_date)
        );
        CREATE TABLE IF NOT EXISTS wbr_results (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            year           INTEGER NOT NULL,
            week_num       INTEGER NOT NULL,
            week_label     TEXT,
            week_start     TEXT,
            week_end       TEXT,
            containers     INTEGER,
            av_oa_avg      REAL,
            av_oa_sla_pct  INTEGER,
            av_oa_p90      INTEGER,
            oa_del_avg     REAL,
            oa_del_sla_pct INTEGER,
            oa_del_p90     INTEGER,
            empty_term_pct INTEGER,
            e2e_avg        INTEGER,
            otp_pct        INTEGER,
            generated_at   TEXT,
            UNIQUE(year, week_num)
        );
        CREATE TABLE IF NOT EXISTS wbr_carrier_results (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            year            INTEGER NOT NULL,
            week_num        INTEGER NOT NULL,
            carrier         TEXT NOT NULL,
            volume          INTEGER,
            av_oa_sla_pct   INTEGER,
            av_oa_avg       REAL,
            av_oa_misses    INTEGER,
            oa_del_sla_pct  INTEGER,
            oa_del_avg      REAL,
            oa_del_misses   INTEGER,
            generated_at    TEXT,
            UNIQUE(year, week_num, carrier)
        );
        CREATE TABLE IF NOT EXISTS wbr_context_notes (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            year       INTEGER NOT NULL,
            week_num   INTEGER NOT NULL,
            notes      TEXT,
            updated_at TEXT,
            UNIQUE(year, week_num)
        );
    """)
    conn.commit()
    conn.close()

init_db()

def migrate_db():
    """Apply schema migrations for existing DBs — safe to run on every startup."""
    conn = get_db()

    # ── carrier_submissions column additions ─────────────────────────────────
    cs_cols = [
        ("sheet_type",       "TEXT"),
        ("port",             "TEXT"),
        ("fc_building",      "TEXT"),
        ("flexi_id",         "TEXT"),
        ("outgate_date",     "TEXT"),
        ("delivery_date",    "TEXT"),
        ("within_sla",       "TEXT"),
        ("sla_notes",        "TEXT"),
        ("empty_return_due", "TEXT"),
        ("appointment_date", "TEXT"),
        ("accessorial_type", "TEXT"),
    ]
    for col, col_type in cs_cols:
        try:
            conn.execute(f"ALTER TABLE carrier_submissions ADD COLUMN {col} {col_type}")
        except sqlite3.OperationalError:
            pass  # column already exists

    # ── wbr_results: add week_end column if missing ──────────────────────────
    try:
        conn.execute("ALTER TABLE wbr_results ADD COLUMN week_end TEXT")
    except (sqlite3.OperationalError, sqlite3.DatabaseError):
        pass  # table doesn't exist yet (created on first WBR generate) or column already there

    # ── rate_lanes: dedup exact (scac, port, destination) duplicates, then add UNIQUE index ──
    try:
        conn.execute("""
            DELETE FROM rate_lanes WHERE id NOT IN (
                SELECT MIN(id) FROM rate_lanes GROUP BY scac, port, destination
            )
        """)
        conn.execute("""
            CREATE UNIQUE INDEX IF NOT EXISTS idx_rate_lanes_lane
            ON rate_lanes(scac, port, destination)
        """)
    except sqlite3.OperationalError:
        pass

    # ── delivery_plan: add UNIQUE index to prevent double-booking same container same week ──
    try:
        conn.execute("""
            CREATE UNIQUE INDEX IF NOT EXISTS idx_delivery_plan_uniq
            ON delivery_plan(week_start, container_id)
        """)
    except sqlite3.OperationalError:
        pass  # index already exists or conflict — don't fail startup

    # ── carrier_submissions: dedup exact matches, then add UNIQUE index ────────
    # Keeps the first (lowest id) row for each (carrier_name, container_id, sheet_type, submitted_at)
    try:
        conn.execute("""
            DELETE FROM carrier_submissions WHERE id NOT IN (
                SELECT MIN(id) FROM carrier_submissions
                GROUP BY carrier_name, container_id, sheet_type, submitted_at
            )
        """)
        conn.execute("""
            CREATE UNIQUE INDEX IF NOT EXISTS idx_cs_dedup
            ON carrier_submissions(carrier_name, container_id, sheet_type, submitted_at)
        """)
    except sqlite3.OperationalError:
        pass

    # ── SRF forecast tables ──────────────────────────────────────────────────────
    conn.execute('''CREATE TABLE IF NOT EXISTS srf_submissions (
        id               INTEGER PRIMARY KEY AUTOINCREMENT,
        submission_week  INTEGER NOT NULL,
        submission_year  INTEGER NOT NULL,
        submitted_at     TEXT,
        submitted_by     TEXT,
        UNIQUE(submission_week, submission_year)
    )''')
    conn.execute('''CREATE TABLE IF NOT EXISTS srf_forecast (
        id                INTEGER PRIMARY KEY AUTOINCREMENT,
        submission_week   INTEGER NOT NULL,
        submission_year   INTEGER NOT NULL,
        forecast_week     INTEGER NOT NULL,
        forecast_year     INTEGER NOT NULL,
        forecasted_volume INTEGER,
        required_capacity INTEGER,
        notice_required   INTEGER DEFAULT 0,
        notice_sent       INTEGER DEFAULT 0,
        notice_sent_at    TEXT,
        UNIQUE(submission_week, submission_year, forecast_week, forecast_year)
    )''')
    conn.commit()

    # ── port_intel / freight_rates / macro_signals: create if missing ────────────
    for _tbl_sql in [
        """CREATE TABLE IF NOT EXISTS port_intel (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            port_code TEXT, source TEXT, headline TEXT, summary TEXT,
            url TEXT, published_date TEXT, scraped_at TEXT)""",
        """CREATE TABLE IF NOT EXISTS freight_rates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source TEXT, route TEXT, rate_usd REAL, wow_change_pct REAL,
            report_date TEXT, scraped_at TEXT,
            UNIQUE(source, route, report_date))""",
        """CREATE TABLE IF NOT EXISTS macro_signals (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source TEXT, metric TEXT, value REAL, period TEXT, scraped_at TEXT,
            UNIQUE(source, metric, period))""",
    ]:
        try:
            conn.execute(_tbl_sql)
        except Exception:
            pass
    conn.commit()

    # ── wbr_context_notes: create if missing (new in July 2026) ────────────────
    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS wbr_context_notes (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                year       INTEGER NOT NULL,
                week_num   INTEGER NOT NULL,
                notes      TEXT,
                updated_at TEXT,
                UNIQUE(year, week_num)
            )
        """)
    except sqlite3.OperationalError:
        pass

    # ── inbound_containers + carrier_contacts ───────────────────────────────────
    for _new_tbl in [
        """CREATE TABLE IF NOT EXISTS inbound_containers (
            id                INTEGER PRIMARY KEY AUTOINCREMENT,
            container_id      TEXT    NOT NULL UNIQUE,
            carrier           TEXT,
            port_region       TEXT,
            yard_dest         TEXT,
            yard_sched        TEXT,
            yard_actual       TEXT,
            fc_dest           TEXT,
            fc_sched          TEXT,
            fc_actual         TEXT,
            empty_returned    TEXT,
            status            TEXT    DEFAULT 'pending',
            notes             TEXT,
            source_email_date TEXT,
            live_drop         TEXT,
            date_received     TEXT,
            last_updated      TEXT,
            source_file       TEXT
        )""",
        """CREATE TABLE IF NOT EXISTS carrier_contacts (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            scac             TEXT    NOT NULL UNIQUE,
            display_name     TEXT,
            contact_name     TEXT,
            email            TEXT,
            reminder_enabled INTEGER DEFAULT 0,
            created_at       TEXT    DEFAULT (datetime('now')),
            updated_at       TEXT    DEFAULT (datetime('now'))
        )""",
    ]:
        try:
            conn.execute(_new_tbl)
        except Exception:
            pass
    # ── carrier_contacts: portal_password column (Option A per-carrier auth) ──
    try:
        conn.execute("ALTER TABLE carrier_contacts ADD COLUMN portal_password TEXT")
    except Exception:
        pass  # column already exists

    # ── carrier_contacts: seed the 5 known carriers (INSERT OR IGNORE — never overwrites) ──
    _seed_carriers = [
        ("ATMI", "Cargomatic",      "Tyler Domingues", None),
        ("ARVY", "Road One",        "Tyler Spangler",  None),
        ("HDDR", "Hudd Distribution","Sandji Ruffin",  None),
        ("RKNE", "Rocking E",       "Mark Brennan",    None),
        ("TGHE", "TGH Equipment",   None,              None),
    ]
    for _sc, _dn, _cn, _em in _seed_carriers:
        try:
            conn.execute(
                "INSERT OR IGNORE INTO carrier_contacts (scac, display_name, contact_name, email) VALUES (?,?,?,?)",
                (_sc, _dn, _cn, _em)
            )
        except Exception:
            pass
    conn.commit()

    # ── wbr_results: auto-seed historical W19–W28 if not present ──────────────
    # Uses INSERT OR IGNORE so existing rows (user-generated) are never overwritten.
    _hist = [
        (2026, 19, "W19", "2026-05-03", "2026-05-09",  69,  8.2, 61, 29, 5.0, 48, 13, 100, 45, None),
        (2026, 20, "W20", "2026-05-10", "2026-05-16",  92,  8.2, 49, 22, 3.5, 61, 11, 100, 42,   58),
        (2026, 21, "W21", "2026-05-17", "2026-05-23",  12,  7.8, 17, 11, 2.2, 73,  7, 100, 48,   85),
        (2026, 22, "W22", "2026-05-24", "2026-05-30", 134,  6.2, 60, 15, 4.1, 50,  8, 100, 44,   58),
        (2026, 23, "W23", "2026-05-31", "2026-06-06",  94,  4.1, 56,  8, 2.9, 52,  5, 100, 53,   85),
        (2026, 24, "W24", "2026-06-07", "2026-06-13",  89,  2.0, 68,  3, 0.5,100,  1, 100, 44,   97),
        (2026, 25, "W25", "2026-06-14", "2026-06-20", 101,  1.1,100,  1, 2.9, 71, 10, 100, 45,   45),
        (2026, 26, "W26", "2026-06-21", "2026-06-27",  83,  2.9, 70,  4, 7.0, 31, 14, 100, 41,   66),
        (2026, 27, "W27", "2026-06-28", "2026-07-04",  60,  1.7,100,  3, 4.8, 28,  9, 100, 36,   48),
        (2026, 28, "W28", "2026-07-05", "2026-07-11",  32,  0.8,100,  2, 3.1, 38,  5, 100, 55,   46),
        (2026, 29, "W29", "2026-07-13", "2026-07-19",  55,  2.0, 71,  4, 1.5, 62,  4, 100, 58,  100),
    ]
    for row in _hist:
        try:
            conn.execute(
                "INSERT OR IGNORE INTO wbr_results "
                "(year,week_num,week_label,week_start,week_end,containers,"
                "av_oa_avg,av_oa_sla_pct,av_oa_p90,"
                "oa_del_avg,oa_del_sla_pct,oa_del_p90,"
                "empty_term_pct,e2e_avg,otp_pct,generated_at) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,'seed')",
                row
            )
        except Exception:
            pass  # table may not exist yet on very first run; migrate_db re-runs after init_db

    conn.commit()
    conn.close()
    # Push seeded data to S3 so every session gets it (no-op if S3 not configured)
    if S3_ENABLED:
        try:
            data_sync.push_db_to_s3(AWS_KEY, AWS_SECRET, AWS_REGION, S3_BUCKET)
        except Exception:
            pass  # non-fatal

migrate_db()

def db_write(sql: str, params: tuple):
    """Write to DB then push to S3."""
    conn = get_db()
    conn.execute(sql, params)
    conn.commit()
    conn.close()
    if S3_ENABLED:
        data_sync.push_db_to_s3(AWS_KEY, AWS_SECRET, AWS_REGION, S3_BUCKET)

# ── DBR config & loader ────────────────────────────────────────────────────────
SHEET_CFG = {
    "Delivery Appointments": {
        "container_col": "Container #",
        "display_cols": ["Location", "Container #", "Status", "Terminal ",
                         "FC/Building", "Del Appointment Date/Time", "LFD", "Outgate Date"],
        "status_label": "Delivery Appointments",
    },
    "Empty Returns": {
        "container_col": "Container #",
        "display_cols": ["Container #", "Terminal", "Empty Return Due Date",
                         "Appointment Date", "Status", "If Outside Window - Reason"],
        "status_label": "Empty Returns",
    },
    "On Vessel": {
        "container_col": "Container #",
        "display_cols": ["Location", "Container #", "Status", "Terminal ",
                         "LFD", "Appointment Date/Time", "Flexi ID"],
        "status_label": "On Vessel",
    },
    "CANCELED": {
        "container_col": "Container #",
        "display_cols": ["Location", "Container #", "Status", "Terminal ",
                         "FC/Building", "Del Appointment Date/Time", "LFD"],
        "status_label": "Canceled",
    },
    "Demurrage": {
        "container_col": "Container #",
        "display_cols": ["Container #", "Terminal ", "Type of Hold", "Days in Hold", "Bridge"],
        "status_label": "Demurrage",
    },
    "Accessorials": {
        "container_col": "Container #",
        "display_cols": ["Container #", "Accessorial", "Date(s) Incurred",
                         "Reason for Charge", "Notes"],
        "status_label": "Accessorials",
    },
}

@st.cache_data(show_spinner="Parsing DBR…", ttl=3600)
def load_dbr(file_bytes: bytes) -> dict:
    # pd.read_excel handles merged cells, varying row widths, and None headers
    sheets = {}
    for sheet_name, cfg in SHEET_CFG.items():
        try:
            df = pd.read_excel(
                io.BytesIO(file_bytes),
                sheet_name=sheet_name,
                header=0,
                dtype=str,
                engine="openpyxl",
            )
        except Exception:
            continue
        df.columns = [str(c).strip() if pd.notna(c) else f"_col_{i}"
                      for i, c in enumerate(df.columns)]
        cont_col = cfg["container_col"]
        if cont_col in df.columns:
            df = df[df[cont_col].notna() & (df[cont_col].str.strip() != "")].copy()
            df["_norm"] = df[cont_col].apply(lambda x: normalize_container(str(x)))
        sheets[sheet_name] = df
    return sheets


# ── Report parsers (Import Shipment Status + Inbound Loads) ───────────────────
_SS_COLS = {
    "ContainerNo": "container_no", "CurrentStatus": "current_status",
    "FC": "fc", "VendorName": "vendor_name", "PONo": "po_no",
    "EDD": "edd", "VesselName": "vessel_name", "VoyageNo": "voyage_no",
    "CarrierCode": "carrier_code", "DischargeCity": "discharge_city",
    "DischargePortETA": "discharge_eta", "FinalDestETA": "final_dest_eta",
    "EDIContainerAvailable": "container_available", "EDIContainerOut": "container_out",
    "EDIEmptyReturn": "empty_return", "EDIDelivered": "delivered",
    "EDIClearedCustoms": "customs_cleared", "ContainerSize": "container_size",
    "LoadPortCountry": "load_port_country", "LoadPortCity": "load_port_city",
}
_IL_COLS = {
    "Container Id": "container_id", "Supplier": "supplier",
    "PO Number": "po_number", "Location Code": "location_code",
    "Destination FC": "destination_fc", "International Origin": "origin_port",
    "International Destination": "dest_port", "Vessel Name": "vessel_name",
    "ETA Arrival At Discharge Port": "eta_discharge",
    "Actual Arrival At Discharge Port": "actual_arrival_discharge",
    "Gateout Terminal Time": "gateout_terminal",
    "Last Free Detention Time": "last_free_detention",
    "Last Free Demurrage Time": "last_free_demurrage",
    "Estimated Appointment Date": "est_appointment",
    "ETA To Final Destination": "eta_final",
    "Actual Arrival At Final Destination": "actual_arrival_final",
    "Empty Container Return Time": "empty_return_time",
}

def parse_shipment_status_file(file_bytes: bytes) -> pd.DataFrame:
    df = pd.read_excel(io.BytesIO(file_bytes), dtype=str, engine="openpyxl")
    available = {k: v for k, v in _SS_COLS.items() if k in df.columns}
    out = df[list(available.keys())].rename(columns=available).copy()
    out = out[out["container_no"].notna() & (out["container_no"].str.strip() != "")]
    out = out.drop_duplicates(subset=["container_no"], keep="last")
    return out.reset_index(drop=True)

def parse_inbound_loads_file(file_bytes: bytes) -> pd.DataFrame:
    df = pd.read_excel(io.BytesIO(file_bytes), dtype=str, engine="openpyxl")
    available = {k: v for k, v in _IL_COLS.items() if k in df.columns}
    out = df[list(available.keys())].rename(columns=available).copy()
    out = out[out["container_id"].notna() & (out["container_id"].str.strip() != "")]
    out = out.drop_duplicates(subset=["container_id"], keep="last")
    return out.reset_index(drop=True)

def upsert_report(df: pd.DataFrame, table: str, key_col: str, source_file: str):
    """Delete existing rows for containers in this upload, then insert fresh."""
    now = datetime.now(_EASTERN).isoformat()
    df = df.copy()
    df["source_file"] = source_file
    df["imported_at"]  = now
    conn = get_db()
    # remove stale rows for these container IDs
    ids = df[key_col].dropna().unique().tolist()
    if ids:
        placeholders = ",".join("?" * len(ids))
        conn.execute(f"DELETE FROM {table} WHERE {key_col} IN ({placeholders})", ids)
    df.to_sql(table, conn, if_exists="append", index=False)
    conn.commit()
    conn.close()
    if S3_ENABLED:
        data_sync.push_db_to_s3(AWS_KEY, AWS_SECRET, AWS_REGION, S3_BUCKET)
    return len(df)


def lookup_containers(dbr_sheets: dict, query_ids: list) -> tuple:
    norm_map = {normalize_container(cid): cid for cid in query_ids}
    results, matched = [], set()

    for sheet_name, cfg in SHEET_CFG.items():
        if sheet_name not in dbr_sheets:
            continue
        df = dbr_sheets[sheet_name].copy()
        if "_norm" not in df.columns:
            continue

        def find_match(dbr_norm):
            for qn in norm_map:
                if containers_match(qn, dbr_norm):
                    return qn
            return None

        df["_mq"] = df["_norm"].apply(find_match)
        hits = df[df["_mq"].notna()].copy()
        if hits.empty:
            continue

        keep = [c for c in cfg["display_cols"] if c in hits.columns]
        hits = hits[keep + ["_mq"]].copy()
        hits.rename(columns={cfg["container_col"]: "Container #"}, inplace=True, errors="ignore")
        hits.insert(0, "Sheet", cfg["status_label"])
        hits.insert(1, "Searched As", hits["_mq"].map(lambda n: norm_map.get(n, n)))
        hits.drop(columns=["_mq"], inplace=True)
        hits.columns = [str(c).strip() for c in hits.columns]
        results.append(hits)
        for qn in df[df["_mq"].notna()]["_mq"].tolist():
            matched.add(qn)

    not_found = [orig for norm, orig in norm_map.items() if norm not in matched]
    combined = pd.concat(results, ignore_index=True) if results else pd.DataFrame()
    return combined, not_found

# ── sidebar ────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("DBR File")

    # try to auto-load from S3 on first run
    if S3_ENABLED and not st.session_state.get("dbr_auto_loaded"):
        with st.spinner("Checking S3 for latest DBR…"):
            dbr_bytes_s3 = data_sync.pull_dbr_from_s3(AWS_KEY, AWS_SECRET, AWS_REGION, S3_BUCKET)
            if dbr_bytes_s3:
                st.session_state.dbr_bytes = dbr_bytes_s3
                last_updated = data_sync.get_dbr_last_updated(AWS_KEY, AWS_SECRET, AWS_REGION, S3_BUCKET)
                st.session_state.dbr_last_updated = last_updated
        st.session_state.dbr_auto_loaded = True

    dbr_file = st.file_uploader(
        "Upload new DBR Excel", type=["xlsx"],
        help="Upload a new weekly DBR. It will be saved to S3 automatically."
    )

    if dbr_file:
        raw = dbr_file.read()
        st.session_state.dbr_bytes = raw
        st.session_state.dbr_last_updated = datetime.utcnow().isoformat()
        if S3_ENABLED:
            with st.spinner("Saving to S3…"):
                ok = data_sync.push_dbr_to_s3(raw, AWS_KEY, AWS_SECRET,
                                               AWS_REGION, S3_BUCKET, dbr_file.name)
            st.success("Saved to S3" if ok else "Saved locally (S3 unavailable)")
        else:
            st.success(f"Loaded: {dbr_file.name}")

    if st.session_state.get("dbr_bytes"):
        dbr_sheets = load_dbr(st.session_state.dbr_bytes)
        ts = st.session_state.get("dbr_last_updated", "")
        if ts:
            try:
                friendly = datetime.fromisoformat(ts).strftime("%b %d, %Y %H:%M UTC")
            except Exception:
                friendly = ts
            st.caption(f"DBR last updated: {friendly}")
        st.caption(f"Delivery Appts: {len(dbr_sheets.get('Delivery Appointments', pd.DataFrame()))} rows")
        st.caption(f"Empty Returns: {len(dbr_sheets.get('Empty Returns', pd.DataFrame()))} rows")
    else:
        dbr_sheets = None
        st.info("No DBR loaded yet." if S3_ENABLED else "Upload the DBR to enable lookups.")

    st.divider()
    st.markdown("**Additional Reports**")
    st.caption("Upload emailed reports to enrich tracking data.")

    # Import Shipment Status
    ss_file = st.file_uploader(
        "Import Shipment Status", type=["xlsx"],
        help="Weekly shipment status report — adds EDI dates, vessel, customs status per container",
        key="ss_upload"
    )
    if ss_file:
        with st.spinner("Parsing…"):
            ss_bytes = ss_file.read()
            ss_df = parse_shipment_status_file(ss_bytes)
            n_ss = upsert_report(ss_df, "shipment_status", "container_no", ss_file.name)
            st.session_state.ss_loaded = True
            st.session_state.ss_source = ss_file.name
        st.success(f"{n_ss:,} containers loaded")

    if st.session_state.get("ss_loaded"):
        st.caption(f"Shipment Status: {st.session_state.get('ss_source','')}")

    # Inbound Loads Report
    il_file = st.file_uploader(
        "Inbound Loads Report", type=["xlsx"],
        help="AR Inbound Loads — adds detention/demurrage free-time deadlines per container",
        key="il_upload"
    )
    if il_file:
        with st.spinner("Parsing…"):
            il_bytes = il_file.read()
            il_df = parse_inbound_loads_file(il_bytes)
            n_il = upsert_report(il_df, "inbound_loads", "container_id", il_file.name)
            st.session_state.il_loaded = True
            st.session_state.il_source = il_file.name
        st.success(f"{n_il:,} containers loaded")

    if st.session_state.get("il_loaded"):
        st.caption(f"Inbound Loads: {st.session_state.get('il_source','')}")

    st.divider()

    if S3_ENABLED:
        if st.button("🔄 Refresh data from S3", use_container_width=True, help="Re-download the latest database from S3"):
            st.session_state.pop("s3_startup_done", None)
            st.rerun()

    # ── Timezone-aware timestamp ───────────────────────────────────────────────
    _US_ZONES = [
        ("America/New_York",    "ET"),
        ("America/Chicago",     "CT"),
        ("America/Denver",      "MT"),
        ("America/Los_Angeles", "PT"),
    ]
    _US_IANA = [tz for tz, _ in _US_ZONES]

    # Detect browser timezone via JS (returns None on first render, resolves on next)
    _tz_raw = streamlit_js_eval(
        js_expressions="Intl.DateTimeFormat().resolvedOptions().timeZone",
        key="user_tz_detect"
    )
    if _tz_raw and _tz_raw != st.session_state.get("_user_tz"):
        st.session_state["_user_tz"] = _tz_raw
    _user_tz = st.session_state.get("_user_tz", "America/New_York")

    # Determine primary US zone (or flag as non-US)
    _primary_tz, _primary_lbl = next(
        ((tz, lbl) for tz, lbl in _US_ZONES if tz == _user_tz),
        ("America/New_York", "ET")   # fallback if non-US or unresolved
    )
    _non_us = _user_tz not in _US_IANA

    _now_utc = datetime.now(ZoneInfo("UTC"))
    _pdt = _now_utc.astimezone(ZoneInfo(_primary_tz))

    # Format: "Thursday, July 24, 2026  10:30 PM ET"
    _fmt_date = f"{_pdt.strftime('%A, %B')} {_pdt.day}, {_pdt.year}"
    _fmt_time = _pdt.strftime("%I:%M %p").lstrip("0")
    _primary_line = f"{_fmt_date}  {_fmt_time} **{_primary_lbl}**"

    # Reference zones (the 3 that aren't primary)
    _ref_parts = []
    for tz, lbl in _US_ZONES:
        if tz == _primary_tz:
            continue
        _rdt = _now_utc.astimezone(ZoneInfo(tz))
        _ref_parts.append(f"{_rdt.strftime('%I:%M %p').lstrip('0')} {lbl}")
    _ref_line = "  ·  ".join(_ref_parts)

    # Non-US local time note
    _local_note = ""
    if _non_us:
        _ldt = _now_utc.astimezone(ZoneInfo(_user_tz))
        _local_note = f"\nLocal: {_ldt.strftime('%I:%M %p').lstrip('0')} ({_user_tz})"

    # Data last updated — most recent submission in DB
    try:
        _ts_conn = get_db()
        _last_sub = _ts_conn.execute("SELECT MAX(submitted_at) FROM carrier_submissions").fetchone()[0]
        _ts_conn.close()
        if _last_sub:
            _ldt_raw = datetime.fromisoformat(str(_last_sub))
            if _ldt_raw.tzinfo is None:
                _ldt_raw = _ldt_raw.replace(tzinfo=ZoneInfo("UTC"))
            _ldt_local = _ldt_raw.astimezone(ZoneInfo(_primary_tz))
            _data_updated = (
                f"{_ldt_local.strftime('%b')} {_ldt_local.day}, "
                f"{_ldt_local.strftime('%Y')}  "
                f"{_ldt_local.strftime('%I:%M %p').lstrip('0')} {_primary_lbl}"
            )
        else:
            _data_updated = "—"
    except Exception:
        _data_updated = "—"

    st.markdown(
        f"{_primary_line}  \n"
        f"<span style=\'font-size:0.78em;color:#888\'>{_ref_line}{_local_note}</span>",
        unsafe_allow_html=True
    )
    st.caption(f"Data last updated: {_data_updated}")
    st.caption("Container Tracker v1.2" + (" · S3 sync active" if S3_ENABLED else ""))

# ── Global styles ─────────────────────────────────────────────────────────────
st.markdown("""
<style>
/* Tighten main content top padding */
.main .block-container { padding-top: 1.25rem; padding-bottom: 2rem; }

/* Active tab — bold only, no color override */
[data-testid="stTabs"] button[role="tab"][aria-selected="true"] { font-weight: 700; }

/* Section header spacing */
h3 { margin-top: 1.2rem !important; margin-bottom: 0.15rem !important; }
h4 { margin-top: 0.9rem !important; margin-bottom: 0.15rem !important; }

/* Divider spacing */
hr { margin: 0.85rem 0 !important; }
</style>
""", unsafe_allow_html=True)

# ── tabs ───────────────────────────────────────────────────────────────────────
st.title("Robotics Container Tracker")
tab9, tab7, tab6, tab1, tabC, tabDBR, tab10 = st.tabs(["WBR", "Planning", "Insights", "Container Lookup", "Carriers", "DBR Dashboard", "Inbound Forecast"])


# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — Container Lookup
# ══════════════════════════════════════════════════════════════════════════════
with tab1:
    st.subheader("Container Lookup")
    st.caption("**Purpose:** Find the current status of any container across all DBR sheets. Paste a list of IDs and get instant results — no manual spreadsheet searching required.")
    st.caption("**How to use:** Upload the DBR in the sidebar, paste container IDs (one per line) into the box, then click Search.")

    col_a, col_b = st.columns([2, 1])
    with col_a:
        lookup_input = st.text_area(
            "Container IDs", height=260, label_visibility="collapsed",
            placeholder="TCNU389902-4\nONEU624470-0\nGCXU542076\n...",
        )
    with col_b:
        requester = st.text_input("Your name (optional)", placeholder="e.g. Dominique")
        st.caption("Searches across all DBR sheets: Delivery, Empty Returns, On Vessel, Canceled, Demurrage, and Accessorials.")
        st.caption("Container IDs work with or without the check digit (e.g. TCNU389902-4 and TCNU389902 both match).")
        search_btn = st.button("Search", type="primary", use_container_width=True)

    if search_btn:
        if not dbr_sheets:
            st.error("No DBR loaded — upload one in the sidebar.")
        elif not lookup_input.strip():
            st.warning("Paste at least one container ID.")
        else:
            query_ids = parse_container_list(lookup_input)
            results_df, not_found = lookup_containers(dbr_sheets, query_ids)

            db_write(
                "INSERT INTO lookup_log (searched_at, requester, container_ids, found_count, not_found_count) VALUES (?,?,?,?,?)",
                (datetime.now(_EASTERN).isoformat(), requester or "anonymous",
                 "\n".join(query_ids), len(query_ids) - len(not_found), len(not_found))
            )

            # ── Status + Detail helpers ────────────────────────────────────────────────
            def _smart_status(sheet, row):
                sv = str(row.get("Status", "") or "").strip().lower()
                if sheet == "Delivery Appointments":
                    if "delivered" in sv: return "Delivered"
                    if "in transit" in sv: return "In Transit"
                    return "Scheduled for Delivery"
                if sheet == "Empty Returns":
                    if "terminat" in sv or "returned" in sv:
                        return "Delivered / Empty Returned"
                    return "Delivered / Pending Return"
                if sheet == "On Vessel":    return "On Vessel"
                if sheet == "Canceled":     return "Canceled"
                if sheet == "Demurrage":    return "Demurrage Hold"
                if sheet == "Accessorials": return "Accessorial Charge"
                return sheet

            def _key_detail(sheet, row):
                def g(col): return str(row.get(col) or "").strip()
                if sheet == "Delivery Appointments":
                    appt = g("Del Appointment Date/Time") or g("Outgate Date")
                    return " · ".join(p for p in [g("FC/Building"), appt] if p)
                if sheet == "Empty Returns":
                    due = g("Empty Return Due Date")
                    return " · ".join(p for p in [g("Terminal"), ("Due " + due) if due else ""] if p)
                if sheet == "On Vessel":
                    lfd = g("LFD")
                    return " · ".join(p for p in [g("Location"), ("LFD " + lfd) if lfd else ""] if p)
                if sheet == "Canceled":
                    return " · ".join(p for p in [g("FC/Building"), g("Del Appointment Date/Time")] if p)
                if sheet == "Demurrage":
                    days = g("Days in Hold")
                    return " · ".join(p for p in [g("Type of Hold"), (days + " days") if days else ""] if p)
                if sheet == "Accessorials":
                    return " · ".join(p for p in [g("Accessorial"), g("Date(s) Incurred")] if p)
                return ""

            # ── Build summary table ───────────────────────────────────────────────
            summary_rows = []
            if not results_df.empty:
                for _, row in results_df.iterrows():
                    sheet = row.get("Sheet", "")
                    container = row.get("Container #") or row.get("Searched As", "")
                    summary_rows.append({
                        "Container": str(container).strip(),
                        "Status":    _smart_status(sheet, row),
                        "Detail":    _key_detail(sheet, row),
                    })
            summary_df = (pd.DataFrame(summary_rows) if summary_rows
                          else pd.DataFrame(columns=["Container", "Status", "Detail"]))

            # ── Status bucket metrics ──────────────────────────────────────────────────
            BUCKET_ORDER = [
                "Scheduled for Delivery", "In Transit", "On Vessel",
                "Delivered", "Delivered / Pending Return", "Delivered / Empty Returned",
                "Demurrage Hold", "Accessorial Charge", "Canceled",
            ]
            counts = {b: 0 for b in BUCKET_ORDER}
            for s in (summary_df["Status"] if not summary_df.empty else []):
                if s in counts: counts[s] += 1
            active = [(k, v) for k, v in counts.items() if v > 0]
            active.append(("Not Found", len(not_found)))
            metric_cols = st.columns(len(active))
            for col, (label, val) in zip(metric_cols, active):
                col.metric(label, val)

            st.divider()

            # ── Consolidated summary table ────────────────────────────────────────────
            if not summary_df.empty:
                st.dataframe(
                    summary_df, use_container_width=True, hide_index=True,
                    column_config={
                        "Container": st.column_config.TextColumn("Container", width="small"),
                        "Status":    st.column_config.TextColumn("Status",    width="medium"),
                        "Detail":    st.column_config.TextColumn("Detail",    width="large"),
                    },
                )
            else:
                st.warning("None of the queried containers found in the DBR.")

            if not_found:
                st.markdown(f"**Not found in this DBR ({len(not_found)}):**")
                st.dataframe(pd.DataFrame({"Container ID": not_found}),
                             use_container_width=True, hide_index=True)

            if not summary_df.empty or not_found:
                buf = io.BytesIO()
                with pd.ExcelWriter(buf, engine="openpyxl") as w:
                    if not summary_df.empty:
                        summary_df.to_excel(w, index=False, sheet_name="Summary")
                        results_df.to_excel(w, index=False, sheet_name="Full Details")
                    if not_found:
                        pd.DataFrame({"Not Found": not_found}).to_excel(
                            w, index=False, sheet_name="Not Found")
                buf.seek(0)
                st.download_button(
                    "⬇️ Download results (.xlsx)", data=buf,
                    file_name=f"container_lookup_{date.today()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            if not results_df.empty:
                with st.expander("🔍 Full details (all fields)"):
                    st.dataframe(results_df, use_container_width=True, hide_index=True)




# ══════════════════════════════════════════════════════════════════════════════
# TAB C — Carriers (Submission + Data)
# ─────────────────────────────────────────────────────────────────────────────

with tabC:
    _car_submit, _car_data = st.tabs(["Submit", "Data"])

    with _car_submit:
        st.subheader("Carrier Submission Portal")
        st.caption("**Purpose:** The single intake point for all carrier status updates. Carriers upload their weekly template here so you have one place to review submissions instead of managing files across email.")
        st.caption("**How to use:** Upload a completed AGL Carrier Template and click Confirm & Submit, or share the vendor portal link with carriers to submit directly.")

        _vp_col, _tmpl_col = st.columns([3, 1])
        with _vp_col:
            st.info(
                f"**Share this link with your carriers:**  \n"
                f"[{VENDOR_PORTAL_URL}]({VENDOR_PORTAL_URL})  \n"
                "Carriers submit directly through this portal — no manual upload needed on your end."
            )
        with _tmpl_col:
            tmpl_path = os.path.join(BASE_DIR, "agl_carrier_template.xlsx")
            if os.path.exists(tmpl_path):
                with open(tmpl_path, "rb") as f:
                    st.download_button(
                        "⬇️ Download Carrier Template",
                        data=f.read(),
                        file_name="AGL_Carrier_Template.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )

        st.divider()

        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # DBR Receipt Tracker
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        st.markdown("### DBR Receipt Tracker")
        st.caption("Track whether each carrier has submitted their weekly DBR. Use the week selector to navigate; missing submissions are flagged automatically.")

        _today_rd  = datetime.now(_EASTERN).date()
        _this_mon  = _today_rd - timedelta(days=_today_rd.weekday())
        _wk_input  = st.date_input(
            "Week (select any day)", value=_this_mon, key="receipt_wk_sel",
            label_visibility="collapsed",
        )
        _wk_start = _wk_input - timedelta(days=_wk_input.weekday())
        _wk_end   = _wk_start + timedelta(days=4)
        st.caption(f"**{_wk_start.strftime('%b %d').lstrip('0')} – {_wk_end.strftime('%b %d, %Y').lstrip('0')}**")

        TRACKED_CARRIERS = ["ATMI", "ARVY", "HDDR", "RKNE", "TGHE"]

        _conn_rd = get_db()
        _rec_df  = pd.read_sql(
            "SELECT carrier, received_date FROM dbr_receipts WHERE week_start = ?",
            _conn_rd, params=(_wk_start.isoformat(),)
        )
        _car_cc_df = pd.read_sql("SELECT scac, contact_name, email FROM carrier_contacts", _conn_rd)
        _conn_rd.close()
        _car_cc_map = {r["scac"]: r for _, r in _car_cc_df.iterrows()} if not _car_cc_df.empty else {}

        _rec_map = {}
        for _, _r in _rec_df.iterrows():
            _rec_map.setdefault(_r["carrier"], set()).add(_r["received_date"])

        _days     = [_wk_start + timedelta(days=i) for i in range(5)]
        _day_cols = [f"{d.strftime('%a')} {d.month}/{d.day}" for d in _days]

        _grid_rows, _missing_carriers = [], []
        for _carrier in TRACKED_CARRIERS:
            _row = {"Carrier": _carrier}
            _got = False
            for _d, _dc in zip(_days, _day_cols):
                if _d.isoformat() in _rec_map.get(_carrier, set()):
                    _row[_dc] = "✅"
                    _got = True
                else:
                    _row[_dc] = ""
            if _got:
                _row["Status"] = "✅ Received"
            elif _wk_start <= _today_rd:
                _row["Status"] = "⚠️ Missing"
                _missing_carriers.append(_carrier)
            else:
                _row["Status"] = "—"
            _grid_rows.append(_row)

        _grid_df = pd.DataFrame(_grid_rows)
        st.dataframe(
            _grid_df, use_container_width=True, hide_index=True,
            column_config={
                "Carrier": st.column_config.TextColumn("Carrier", width="small"),
                "Status":  st.column_config.TextColumn("Status",  width="medium"),
                **{dc: st.column_config.TextColumn(dc, width="small") for dc in _day_cols},
            },
        )

        if _missing_carriers and _wk_start <= _today_rd:
            st.warning(f"**{len(_missing_carriers)} carrier(s) missing this week:** {', '.join(_missing_carriers)}")
            for _mc in _missing_carriers:
                _cc_info  = _car_cc_map.get(_mc, {})
                _em_val   = _cc_info.get("email") or ""
                _cn_val   = _cc_info.get("contact_name") or (_mc + " Team")
                with st.expander(f"Send reminder — {_mc}"):
                    _dl_c = _today_rd.strftime("%B %d, %Y").lstrip("0")
                    _body_c = (
                        f"Hi {_cn_val},\n\n"
                        f"We have not received your container status DBR for the week of "
                        f"{_wk_start.strftime('%B %d, %Y').lstrip('0')}.\n"
                        f"Could you please submit your update at your earliest convenience?\n\n"
                        f"Submit via the AGL carrier portal:\n{VENDOR_PORTAL_URL}\n\n"
                        f"Thank you,\nAGL Robotics Logistics"
                    )
                    _draft_c  = st.text_area("Email draft", _body_c, height=180, key="c_draft_" + _mc)
                    _to_c     = st.text_input("To", value=_em_val, placeholder="carrier@example.com", key="c_to_" + _mc)
                    _cs1, _cs2 = st.columns([1, 3])
                    _do_send_c = _cs1.button("Send", type="primary", disabled=not bool(_to_c.strip()), key="c_send_" + _mc)
                    if not _to_c.strip():
                        _cs2.caption("Add carrier email in DBR Dashboard > By Carrier to enable.")
                    if _do_send_c and _to_c.strip():
                        try:
                            import boto3 as _b3c
                            _ses_c = _b3c.client("ses",
                                aws_access_key_id=AWS_KEY,
                                aws_secret_access_key=AWS_SECRET,
                                region_name=(AWS_REGION or "us-east-1"))
                            _ses_c.send_email(
                                Source="agl-robotics-no-reply@amazon.com",
                                Destination={"ToAddresses": [_to_c.strip()]},
                                Message={
                                    "Subject": {"Data": "[AGL] DBR Reminder — Week of " + _wk_start.strftime("%b %d")},
                                    "Body":    {"Text": {"Data": _draft_c}},
                                },
                            )
                            st.success("Reminder sent to " + _to_c.strip())
                        except Exception as _se_c:
                            st.error("Send failed: " + str(_se_c))

        st.divider()

        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # Bulk DBR File Upload
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        st.markdown("### Submit DBR Files")
        st.caption("Upload one or more files in one shot — mix carriers, mix weeks. Carrier and file date are auto-detected from the filename; override either if needed.")

        def _detect_carrier_from_name(fname: str) -> str:
            fn = fname.upper()
            for sc in ["ATMI", "ARVY", "RKNE", "TGHE"]:
                if sc in fn: return sc
            if "HUDD" in fn or "HDDR" in fn: return "HDDR"
            return ""

        def _detect_date_from_name(fname: str) -> date:
            import re as _re
            for pat in [r"(\d{1,2})[.\-](\d{1,2})[.\-](\d{2,4})", r"(\d{2})(\d{2})(\d{2,4})"]:
                m = _re.search(pat, fname)
                if m:
                    try:
                        mo, dy, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
                        if yr < 100: yr += 2000
                        if 1 <= mo <= 12 and 1 <= dy <= 31:
                            return date(yr, mo, dy)
                    except Exception:
                        pass
            return datetime.now(_EASTERN).date()

        _CARRIER_OPTS = ["", "ATMI", "ARVY", "HDDR", "RKNE", "TGHE"]

        # Key counter — incrementing resets the file uploader widget
        if "bulk_upload_key" not in st.session_state:
            st.session_state.bulk_upload_key = 0

        # ── Acknowledge-and-clear banner ──────────────────────────────────────
        if st.session_state.get("bulk_upload_result"):
            _bres = st.session_state.bulk_upload_result
            _bfiles_lbl = ", ".join(_bres.get("files", []))
            st.success(
                "\u2705  **Submission successful.**\n\n"
                "**" + str(_bres["logged"]) + "** container records logged across "
                "**" + str(_bres["carriers"]) + "** carrier(s)  \n"
                "Files: " + _bfiles_lbl
            )
            st.info("Your data has been captured and stored. Click below to clear and submit another batch.")
            if st.button("\u2714\ufe0f  Acknowledge & Clear", type="primary", key="bulk_ack_btn"):
                del st.session_state["bulk_upload_result"]
                st.session_state.bulk_upload_key += 1
                st.rerun()
        else:
            bulk_files = st.file_uploader(
                "Drop DBR files here",
                type=["xlsx", "csv"],
                accept_multiple_files=True,
                key="bulk_dbr_upload_" + str(st.session_state.bulk_upload_key),
                label_visibility="collapsed",
            )

            if bulk_files:
                st.caption(str(len(bulk_files)) + " file(s) selected — confirm carrier for each:")

                _file_carrier_map = {}
                _hdr1, _hdr2, _hdr3 = st.columns([4, 1, 2])
                _hdr2.caption("Carrier")
                _hdr3.caption("File date — which day is this data from?")
                for _fi, _uf in enumerate(bulk_files):
                    _det      = _detect_carrier_from_name(_uf.name)
                    _det_date = _detect_date_from_name(_uf.name)
                    _hd1, _hd2, _hd3 = st.columns([4, 1, 2])
                    with _hd1:
                        st.markdown("&nbsp;&nbsp;\U0001f4c4 `" + _uf.name + "`")
                    with _hd2:
                        _chosen = st.selectbox(
                            "Carrier", _CARRIER_OPTS,
                            index=_CARRIER_OPTS.index(_det) if _det in _CARRIER_OPTS else 0,
                            key="bulk_c_" + str(_fi), label_visibility="collapsed",
                        )
                    with _hd3:
                        _file_date = st.date_input(
                            "File date", value=_det_date,
                            key="bulk_d_" + str(_fi), label_visibility="collapsed",
                        )
                    _file_carrier_map[_fi] = (_uf, _chosen, _file_date)

                _unassigned = [_uf.name for _, (_uf, _c, _fd) in _file_carrier_map.items() if not _c]
                if _unassigned:
                    st.warning("Assign carrier for: " + ", ".join(_unassigned))
                else:
                    _all_parsed = {}
                    _parse_errs = []
                    for _fi, (_uf, _carrier, _file_date) in _file_carrier_map.items():
                        _ck = "bulk_" + _uf.name + "_" + str(_uf.size)
                        if _ck not in st.session_state:
                            st.session_state[_ck] = _uf.read()
                        _raw = st.session_state[_ck]
                        try:
                            _p = parse_carrier_template(_raw, _uf.name)
                            if _p:
                                if _carrier not in _all_parsed:
                                    _all_parsed[_carrier] = {}
                                for _stype, _rows in _p.items():
                                    _all_parsed[_carrier].setdefault(_stype, [])
                                    for _r in _rows:
                                        _r["_src"] = _uf.name
                                    _all_parsed[_carrier][_stype].extend(_rows)
                            else:
                                _parse_errs.append(_uf.name + ": no container data found")
                        except Exception as _e:
                            _parse_errs.append(_uf.name + ": " + str(_e))

                    for _err in _parse_errs:
                        st.warning(_err)

                    if _all_parsed:
                        _tot = sum(len(_r) for _cd in _all_parsed.values() for _r in _cd.values())
                        st.success("Ready: **" + str(_tot) + " containers** across **" + str(len(_all_parsed)) + " carrier(s)**")

                        for _carrier, _sheets in _all_parsed.items():
                            _ctot = sum(len(_r) for _r in _sheets.values())
                            with st.expander("\U0001f50e Preview \u2014 " + _carrier + " (" + str(_ctot) + " containers)"):
                                _slabels = {"delivery": "Delivery", "delivery_ilm1": "ILM1",
                                            "delivery_ric6": "RIC6", "empty_return": "Empty Returns",
                                            "demurrage": "Demurrage", "accessorial": "Accessorials", "ody": "ODY"}
                                _ptabs = st.tabs([_slabels.get(_k, _k) for _k in _sheets])
                                for _ptab, (_st, _rows) in zip(_ptabs, _sheets.items()):
                                    with _ptab:
                                        _df_p = pd.DataFrame(_rows).drop(columns=["_raw_container","_src"], errors="ignore")
                                        st.dataframe(_df_p, use_container_width=True, hide_index=True)

                        if st.button("\u2705 Confirm & Submit All", type="primary", key="bulk_submit_btn"):
                            _logged_at = datetime.now(_EASTERN).isoformat()
                            _conn = get_db()
                            _logged = 0
                            _DEL_T = {"delivery","delivery_ilm1","delivery_ric6","delivery_dbm6"}
                            _FC_MP = {"delivery_ilm1":"ILM1","delivery_ric6":"RIC6","delivery_dbm6":"DBM6"}
                            _carrier_file_dates = {}
                            for _fi, (_uf, _carrier, _fdate) in _file_carrier_map.items():
                                if _carrier:
                                    _carrier_file_dates.setdefault(_carrier, _fdate)
                            # ── Step 1: Log dbr_receipts for ALL assigned files ─────────────
                            # Use today (upload date) as received_date — not the file's data date.
                            # This ensures receipts always land in the current week regardless of
                            # what date is in the filename. One receipt per carrier per week.
                            _today_recv = datetime.now(_EASTERN).date()
                            _wk_recv    = (_today_recv - timedelta(days=_today_recv.weekday())).isoformat()
                            _rcpt_logged = set()
                            for _fi2, (_uf2, _c2, _fd2) in _file_carrier_map.items():
                                if _c2 and _c2 not in _rcpt_logged:
                                    _conn.execute(
                                        "INSERT OR IGNORE INTO dbr_receipts"
                                        " (carrier, week_start, received_date, received_via, file_name, logged_at)"
                                        " VALUES (?,?,?,?,?,?)",
                                        (_c2, _wk_recv, _today_recv.isoformat(), "web", _uf2.name, _logged_at),
                                    )
                                    _rcpt_logged.add(_c2)

                            # ── Step 2: Log container data for successfully parsed files ────
                            for _carrier, _sheets in _all_parsed.items():
                                _fdate = _carrier_file_dates.get(_carrier, datetime.now(_EASTERN).date())
                                for _stype, _rows in _sheets.items():
                                    for _row in _rows:
                                        # carrier_submissions log
                                        _conn.execute(
                                            "INSERT OR IGNORE INTO carrier_submissions"
                                            " (submitted_at, carrier_name, container_id, sheet_type,"
                                            " port, terminal, fc_building, flexi_id, outgate_date,"
                                            " delivery_date, status, within_sla, sla_notes,"
                                            " empty_return_due, appointment_date, accessorial_type,"
                                            " notes, source_file, source)"
                                            " VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                                            (_fdate.isoformat(), _carrier, _row["container_id"], _stype,
                                             _row.get("port"), _row.get("terminal"), _row.get("fc_building"),
                                             _row.get("flexi_id"), _row.get("outgate_date"),
                                             _row.get("delivery_date"), _row.get("status"),
                                             _row.get("within_sla"), _row.get("sla_notes"),
                                             _row.get("empty_return_due"), _row.get("appointment_date"),
                                             _row.get("accessorial_type"), _row.get("notes"),
                                             _row.get("_src"), "web"),
                                        )
                                        _logged += 1
                                        # inbound_containers upsert (feeds DBR Dashboard)
                                        if _stype in _DEL_T:
                                            _hd = bool(_row.get("delivery_date"))
                                            _ho = bool(_row.get("outgate_date"))
                                            if _hd:   _ic_st, _fa = "delivered", str(_row["delivery_date"])
                                            elif _ho: _ic_st, _fa = "at_yard", None
                                            else:     _ic_st, _fa = "pending", None
                                            _fd = _FC_MP.get(_stype) or _row.get("fc_building")
                                            _conn.execute(
                                                "INSERT INTO inbound_containers"
                                                " (container_id,carrier,port_region,fc_dest,"
                                                "  yard_actual,fc_actual,status,notes,last_updated,source_file)"
                                                " VALUES (?,?,?,?,?,?,?,?,?,?)"
                                                " ON CONFLICT(container_id) DO UPDATE SET"
                                                "  carrier=COALESCE(excluded.carrier,carrier),"
                                                "  port_region=COALESCE(excluded.port_region,port_region),"
                                                "  fc_dest=COALESCE(excluded.fc_dest,fc_dest),"
                                                "  yard_actual=COALESCE(excluded.yard_actual,yard_actual),"
                                                "  fc_actual=COALESCE(excluded.fc_actual,fc_actual),"
                                                "  status=excluded.status,"
                                                "  last_updated=excluded.last_updated,"
                                                "  source_file=excluded.source_file",
                                                (_row["container_id"], _carrier, _row.get("port"), _fd,
                                                 _row.get("outgate_date"), _fa,
                                                 _ic_st, _row.get("notes"), _logged_at, _row.get("_src")),
                                            )
                                        elif _stype == "empty_return":
                                            _ret = str(_row.get("status") or "").lower()
                                            if any(kw in _ret for kw in ("return","complete","done","at port")):
                                                _rd = _row.get("appointment_date") or _fdate.isoformat()
                                                _conn.execute(
                                                    "UPDATE inbound_containers SET"
                                                    " empty_returned=COALESCE(empty_returned,?),"
                                                    " status='empty_returned', last_updated=?"
                                                    " WHERE container_id=?",
                                                    (_rd, _logged_at, _row["container_id"]),
                                                )
                            _conn.commit()
                            _conn.close()
                            if S3_ENABLED:
                                data_sync.push_db_to_s3(AWS_KEY, AWS_SECRET, AWS_REGION, S3_BUCKET)
                            st.session_state["bulk_upload_result"] = {
                                "logged": _logged,
                                "carriers": len(_all_parsed),
                                "files": [_file_carrier_map[i][0].name for i in _file_carrier_map],
                            }
                            st.rerun()

        st.divider()

        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # Manual entry (fallback)
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        with st.expander("⌨️ Enter containers manually (fallback — use file upload when possible)"):
            st.caption("Use this only if a carrier cannot provide a file. Paste container IDs directly.")
            with st.form("carrier_form_manual", clear_on_submit=True):
                _fm1, _fm2 = st.columns(2)
                with _fm1:
                    _m_carrier  = st.selectbox("Carrier *", ["", "ATMI", "ARVY", "HDDR", "RKNE", "TGHE"], key="man_carrier")
                    _m_terminal = st.text_input("Terminal", placeholder="EWR, BOS, CHI…", key="man_term")
                with _fm2:
                    _m_status = st.selectbox("Status", [
                        "Delivery Scheduled", "Delivered pending empty return",
                        "Empty Return Scheduled", "TERMINATED", "On Vessel",
                        "Customs Hold", "Other",
                    ], key="man_status")
                    _m_notes = st.text_input("Notes", placeholder="Delays, exceptions…", key="man_notes")
                _m_ids_text = st.text_area("Container IDs (one per line) *", height=140, key="man_ids")
                _m_submit   = st.form_submit_button("Submit", type="primary")
            if _m_submit:
                if not _m_carrier:
                    st.error("Select a carrier.")
                elif not _m_ids_text.strip():
                    st.error("Paste at least one container ID.")
                else:
                    _mids = parse_container_list(_m_ids_text)
                    _mn   = datetime.now(_EASTERN).isoformat()
                    _mc   = get_db()
                    for _cid in _mids:
                        _mc.execute(
                            "INSERT INTO carrier_submissions (submitted_at, carrier_name, container_id, terminal, status, notes, source) VALUES (?,?,?,?,?,?,?)",
                            (_mn, _m_carrier, normalize_container(_cid), _m_terminal.strip() or None, _m_status, _m_notes.strip() or None, "manual")
                        )
                    _mwk = (datetime.now(_EASTERN).date() - timedelta(days=datetime.now(_EASTERN).date().weekday())).isoformat()
                    _mc.execute(
                        "INSERT OR IGNORE INTO dbr_receipts (carrier, week_start, received_date, received_via, logged_at) VALUES (?,?,?,?,?)",
                        (_m_carrier, _mwk, datetime.now(_EASTERN).date().isoformat(), "manual", _mn)
                    )
                    _mc.commit(); _mc.close()
                    if S3_ENABLED:
                        data_sync.push_db_to_s3(AWS_KEY, AWS_SECRET, AWS_REGION, S3_BUCKET)
                    st.success(f"Logged {len(_mids)} containers from **{_m_carrier}**")

        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # Submission Log
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        st.markdown("### Submission Log")

        _sl_conn = get_db()
        _log_df  = pd.read_sql(
            """SELECT submitted_at, carrier_name, container_id, sheet_type,
                      terminal, status, notes, source_file, source
               FROM (
                   SELECT *, ROW_NUMBER() OVER (
                       PARTITION BY carrier_name, container_id, sheet_type
                       ORDER BY submitted_at DESC
                   ) AS rn
                   FROM carrier_submissions
               ) WHERE rn = 1
               ORDER BY submitted_at DESC""",
            _sl_conn
        )
        _sl_conn.close()
        _log_df = _apply_carrier_display(_log_df)

        if _log_df.empty:
            st.info("No submissions yet.")
        else:
            _sc1, _sc2, _sc3 = st.columns(3)
            with _sc1:
                _sf_carrier = st.multiselect("Carrier", sorted(_log_df["carrier_name"].unique()), key="sl_carrier")
            with _sc2:
                _sf_status  = st.multiselect("Status",  sorted(_log_df["status"].dropna().unique()), key="sl_status")
            with _sc3:
                _sf_source  = st.multiselect("Source",  sorted(_log_df["source"].dropna().unique()), key="sl_source")

            _filt = _log_df.copy()
            if _sf_carrier: _filt = _filt[_filt["carrier_name"].isin(_sf_carrier)]
            if _sf_status:  _filt = _filt[_filt["status"].isin(_sf_status)]
            if _sf_source:  _filt = _filt[_filt["source"].isin(_sf_source)]

            st.dataframe(_filt, use_container_width=True, hide_index=True)

            _sl_buf = io.BytesIO()
            _filt.to_excel(_sl_buf, index=False)
            _sl_buf.seek(0)
            st.download_button(
                "⬇️ Export log (.xlsx)", data=_sl_buf,
                file_name=f"carrier_log_{date.today()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )




    with _car_data:
        st.subheader("Carrier Data")
        st.caption("**Purpose:** A full record of every container status submission from carriers. Use this to verify what a carrier reported, check SLA performance, or audit delivery and return activity.")
        st.caption("**How to use:** Filter by carrier or sheet type at the top, then drill into sub-tabs for Delivery, Empty Returns, Demurrage, and Accessorials.")

        conn = get_db()
        all_df = pd.read_sql(
            """SELECT submitted_at, carrier_name, container_id, sheet_type,
                      port, terminal, fc_building, delivery_date, status,
                      within_sla, empty_return_due, appointment_date,
                      accessorial_type, notes, source
               FROM (
                   SELECT *, ROW_NUMBER() OVER (
                       PARTITION BY carrier_name, container_id, sheet_type
                       ORDER BY submitted_at DESC
                   ) AS rn
                   FROM carrier_submissions
               ) WHERE rn = 1
               ORDER BY submitted_at DESC""",
            conn
        )
        conn.close()
        all_df = _apply_carrier_display(all_df)

        if all_df.empty:
            st.info("No carrier submissions yet.")
        else:
            st.caption("📌 Showing **latest status per container** — duplicate weekly file entries are collapsed. Full history is retained in the database.")
            # top filters
            f1, f2, f3 = st.columns(3)
            with f1:
                f_carrier = st.multiselect("Carrier", sorted(all_df["carrier_name"].unique()), key="cd_carrier")
            with f2:
                f_sheet = st.multiselect("Sheet type", sorted(all_df["sheet_type"].dropna().unique()), key="cd_sheet")
            with f3:
                f_sla = st.multiselect("Within SLA?", sorted(all_df["within_sla"].dropna().unique()), key="cd_sla")

            filt = all_df.copy()
            if f_carrier: filt = filt[filt["carrier_name"].isin(f_carrier)]
            if f_sheet:   filt = filt[filt["sheet_type"].isin(f_sheet)]
            if f_sla:     filt = filt[filt["within_sla"].isin(f_sla)]

            # metrics
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Total containers", len(filt))
            m2.metric("Carriers", filt["carrier_name"].nunique())
            sla_counts = filt["within_sla"].value_counts()
            m3.metric("Within SLA", int(sla_counts.get("YES", 0)))
            m4.metric("Outside SLA", int(sla_counts.get("NO", 0)))

            st.divider()

            # tabbed by sheet type
            sheet_types = sorted(filt["sheet_type"].dropna().unique())
            if sheet_types:
                sheet_labels = {
                    "delivery":       "Delivery",
                    "delivery_ilm1":  "ILM1",
                    "delivery_ric6":  "RIC6",
                    "empty_return":   "Empty Returns",
                    "ody":            "ODY/Storage",
                    "demurrage":      "Demurrage",
                    "accessorial":    "Accessorials",
                }
                stabs = st.tabs([sheet_labels.get(s, s) for s in sheet_types])
                for stab, stype in zip(stabs, sheet_types):
                    with stab:
                        sdf = filt[filt["sheet_type"] == stype].copy()
                        # show only non-empty columns
                        nonempty = [c for c in sdf.columns if sdf[c].notna().any() and sdf[c].astype(str).str.strip().ne("").any()]
                        st.dataframe(sdf[nonempty], use_container_width=True, hide_index=True)

            st.divider()
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as w:
                for stype in filt["sheet_type"].dropna().unique():
                    sdf = filt[filt["sheet_type"] == stype]
                    sdf.to_excel(w, index=False, sheet_name=stype[:31])
            buf.seek(0)
            st.download_button("⬇️ Export all carrier data (.xlsx)", data=buf,
                file_name=f"carrier_data_{date.today()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        st.divider()
        with st.expander("Backfill Historical Container Data", expanded=False):
            st.caption(
                "Import a historical DBR Tracker Excel (master format) to populate the container lifecycle view. "
                "Existing records are updated non-destructively — only blank fields are filled in."
            )
            if "cd_backfill_key" not in st.session_state:
                st.session_state.cd_backfill_key = 0
            if st.session_state.get("cd_backfill_result"):
                _bfr = st.session_state.cd_backfill_result
                st.success(
                    "\u2705  Import complete — "
                    "**" + str(_bfr["new"]) + "** new  \u00b7  **" + str(_bfr["updated"]) + "** updated  \u00b7  "
                    + ", ".join(_bfr["files"])
                )
                if st.button("\u2714\ufe0f  Clear", key="cd_bf_ack"):
                    del st.session_state["cd_backfill_result"]
                    st.session_state.cd_backfill_key += 1
                    st.rerun()
            else:
                _bf_files = st.file_uploader(
                    "Upload DBR Tracker Excel file(s)",
                    type=["xlsx"],
                    accept_multiple_files=True,
                    key="cd_bf_up_" + str(st.session_state.cd_backfill_key),
                )
                if _bf_files:
                    _bf_rows, _bf_errs, _bf_names = [], [], []
                    for _bf in _bf_files:
                        _bfb = _bf.read()
                        with st.spinner("Parsing " + _bf.name + "..."):
                            _bfr2, _bfe2 = _parse_dbr_tracker(_bfb)
                        _bf_rows.extend(_bfr2); _bf_errs.extend(_bfe2); _bf_names.append(_bf.name)
                    if _bf_errs:
                        with st.expander(str(len(_bf_errs)) + " parse warnings"):
                            for _e in _bf_errs: st.caption(_e)
                    if _bf_rows:
                        st.info("Ready to import **" + str(len(_bf_rows)) + "** containers from **" + str(len(_bf_files)) + "** file(s).")
                        if st.button("Import All", type="primary", key="cd_bf_go"):
                            _bfc  = get_db()
                            _bfin = _bfup = 0
                            _bfts = datetime.now(_PHX).isoformat()
                            for _r in _bf_rows:
                                _cid = _r.get("container_id")
                                if not _cid: continue
                                _ex = _bfc.execute("SELECT id FROM inbound_containers WHERE container_id=?", (_cid,)).fetchone()
                                if _ex:
                                    _bfc.execute(
                                        "UPDATE inbound_containers SET carrier=COALESCE(carrier,?),"
                                        "fc_dest=COALESCE(fc_dest,?),fc_sched=COALESCE(fc_sched,?),"
                                        "status=?,notes=COALESCE(notes,?),last_updated=? WHERE container_id=?",
                                        (_r.get("carrier"),_r.get("site_code"),_r.get("appt_date"),
                                         _r.get("status","pending"),_r.get("notes"),_bfts,_cid)
                                    ); _bfup += 1
                                else:
                                    _bfc.execute(
                                        "INSERT OR IGNORE INTO inbound_containers"
                                        "(container_id,carrier,fc_dest,fc_sched,status,notes,last_updated,source_file)"
                                        " VALUES (?,?,?,?,?,?,?,?)",
                                        (_cid,_r.get("carrier"),_r.get("site_code"),_r.get("appt_date"),
                                         _r.get("status","pending"),_r.get("notes"),_bfts,"; ".join(_bf_names))
                                    ); _bfin += 1
                            _bfc.commit(); _bfc.close()
                            if S3_ENABLED:
                                data_sync.push_db_to_s3(AWS_KEY, AWS_SECRET, AWS_REGION, S3_BUCKET)
                            st.session_state["cd_backfill_result"] = {"new":_bfin,"updated":_bfup,"files":_bf_names}
                            st.rerun()
                    else:
                        st.warning("No containers found. Ensure the file has a Delivery Plan sheet.")



# TAB 6 — Insights (Overview + Empty Returns + Lane Costs)
# ─────────────────────────────────────────────────────────────────────────────

with tab6:
    _ins_overview, _ins_empty, _ins_lanes = st.tabs([
        "Overview", "Empty Returns", "Lane Costs"
    ])

    with _ins_overview:
        st.subheader("Weekly Insights")
        st.caption("**Purpose:** Your weekly operations briefing — auto-generated from all loaded data sources. Start here to see what needs attention before you open anything else.")
        st.caption("**How to use:** Upload the DBR and supplemental reports in the sidebar — sections activate automatically. Start with Detention & Demurrage Risk.")

        today_ts = pd.Timestamp(date.today())

        def _risk_flag(d):
            if pd.isna(d): return "—"
            if d < 0:   return f"{abs(int(d))}d overdue"
            if d == 0:  return "Due today"
            if d <= 3:  return f"{int(d)}d"
            return f"{int(d)}d ok"

        # ── load supplemental data ────────────────────────────────────────────────
        conn = get_db()
        ss_df = pd.read_sql("SELECT * FROM shipment_status  ORDER BY imported_at DESC", conn)
        il_df = pd.read_sql("SELECT * FROM inbound_loads    ORDER BY imported_at DESC", conn)
        sub_df = pd.read_sql(
            "SELECT submitted_at, carrier_name, container_id, sheet_type, "
            "terminal, fc_building, status, within_sla, source FROM carrier_submissions "
            "WHERE carrier_name IS NOT NULL AND carrier_name != '' ORDER BY submitted_at DESC",
            conn,
        )
        conn.close()

        # normalise date columns
        for col in ["last_free_detention", "last_free_demurrage", "eta_discharge",
                    "gateout_terminal", "eta_final", "actual_arrival_final", "empty_return_time"]:
            if col in il_df.columns:
                il_df[col] = pd.to_datetime(il_df[col], errors="coerce")

        for col in ["discharge_eta", "final_dest_eta", "container_available",
                    "container_out", "empty_return", "delivered"]:
            if col in ss_df.columns:
                ss_df[col] = pd.to_datetime(ss_df[col], errors="coerce")

        if not sub_df.empty:
            sub_df["submitted_at"] = pd.to_datetime(sub_df["submitted_at"], errors="coerce")

        has_il = not il_df.empty
        has_ss = not ss_df.empty

        # ══════════════════════════════════════════════════════════════════════
        # SECTION 1 — DBR Snapshot
        # ══════════════════════════════════════════════════════════════════════
        st.markdown("### DBR Snapshot")
        if not dbr_sheets:
            st.info("Load a DBR in the sidebar to activate this section.")
        else:
            _del = dbr_sheets.get("Delivery Appointments", pd.DataFrame())
            _er  = dbr_sheets.get("Empty Returns",         pd.DataFrame())
            _vsl = dbr_sheets.get("On Vessel",             pd.DataFrame())
            _dem = dbr_sheets.get("Demurrage",             pd.DataFrame())

            overdue_ct = due_soon_ct = 0
            if not _er.empty and "Empty Return Due Date" in _er.columns:
                er_c = _er.copy()
                er_c["_erd"]  = pd.to_datetime(er_c["Empty Return Due Date"], errors="coerce")
                er_c["_days"] = (er_c["_erd"] - today_ts).dt.days
                er_act = er_c if "Status" not in er_c.columns else \
                         er_c[er_c["Status"].fillna("").str.upper().ne("TERMINATED")]
                overdue_ct  = int((er_act["_days"] < 0).sum())
                due_soon_ct = int(((er_act["_days"] >= 0) & (er_act["_days"] <= 3)).sum())

            c1,c2,c3,c4,c5,c6 = st.columns(6)
            c1.metric("Delivery Appts",    len(_del))
            c2.metric("Empty Returns",     len(_er))
            c3.metric("On Vessel",         len(_vsl))
            c4.metric("Demurrage",         len(_dem))
            c5.metric("ER Overdue",     overdue_ct)
            c6.metric("ER Due ≤3 days", due_soon_ct)

            # LFD risk
            if not _del.empty and "LFD" in _del.columns:
                lfd = _del.copy()
                lfd["_lfd"]  = pd.to_datetime(lfd["LFD"], errors="coerce")
                lfd["Days to LFD"] = (lfd["_lfd"] - today_ts).dt.days
                risk = lfd[lfd["Days to LFD"].notna() & (lfd["Days to LFD"] <= 5)].copy()
                if not risk.empty:
                    risk["Alert"] = risk["Days to LFD"].apply(_risk_flag)
                    show = [c for c in ["Alert","Container #","Terminal ","FC/Building","Status","LFD","Days to LFD"] if c in risk.columns]
                    st.divider()
                    st.markdown(f"#### LFD Risk — {len(risk)} containers within 5 days")
                    st.dataframe(risk[show].sort_values("Days to LFD"), use_container_width=True, hide_index=True)

            # Status + terminal breakdown
            if not _del.empty:
                s_col, t_col_name = st.columns(2)
                if "Status" in _del.columns:
                    with s_col:
                        st.markdown("#### Delivery Status")
                        vc = _del["Status"].fillna("Unknown").value_counts().reset_index()
                        vc.columns = ["Status","Count"]
                        st.dataframe(vc, use_container_width=True, hide_index=True)
                term_col = "Terminal " if "Terminal " in _del.columns else ("Terminal" if "Terminal" in _del.columns else None)
                if term_col:
                    with t_col_name:
                        st.markdown("#### By Terminal")
                        tc = _del[term_col].fillna("Unknown").value_counts().head(10).reset_index()
                        tc.columns = ["Terminal","Count"]
                        st.bar_chart(tc.set_index("Terminal")["Count"])

        # ══════════════════════════════════════════════════════════════════════
        # SECTION 2 — Detention & Demurrage Risk (from Inbound Loads)
        # ══════════════════════════════════════════════════════════════════════
        st.divider()
        st.markdown("### Detention & Demurrage Risk")

        if not has_il:
            st.info("Upload the **Inbound Loads Report** in the sidebar to activate detention/demurrage tracking.")
        else:
            # detention = clock starts when container available at terminal;
            #             container not yet gated out; last_free_detention approaching
            det = il_df.copy()
            det["days_to_det"] = (det["last_free_detention"] - today_ts).dt.days
            det["days_to_dem"] = (det["last_free_demurrage"] - today_ts).dt.days

            # active detention risk: has free-time deadline, not yet gated out
            det_risk = det[
                det["last_free_detention"].notna() &
                (det["gateout_terminal"].isna() | (det["gateout_terminal"] > today_ts)) &
                (det["days_to_det"] <= 7)
            ].copy()

            dem_risk = det[
                det["last_free_demurrage"].notna() &
                (det["actual_arrival_discharge"].isna() | det["gateout_terminal"].isna()) &
                (det["days_to_dem"] <= 7)
            ].copy()

            d1, d2, d3, d4 = st.columns(4)
            det_overdue = int((det["days_to_det"].dropna() < 0).sum())
            dem_overdue = int((det["days_to_dem"].dropna() < 0).sum())
            d1.metric("Detention Overdue",  det_overdue)
            d2.metric("Detention ≤7 days",  max(0, len(det_risk) - det_overdue))
            d3.metric("Demurrage Overdue",  dem_overdue)
            d4.metric("Demurrage ≤7 days",  max(0, len(dem_risk) - dem_overdue))

            if not det_risk.empty or not dem_risk.empty:
                drisk_tab, ddem_tab = st.tabs(["Detention Risk", "Demurrage Risk"])

                with drisk_tab:
                    if det_risk.empty:
                        st.success("No detention risk in next 7 days.")
                    else:
                        det_risk["Alert"] = det_risk["days_to_det"].apply(_risk_flag)
                        show_d = [c for c in ["Alert","container_id","supplier","destination_fc",
                                              "dest_port","last_free_detention","days_to_det",
                                              "eta_discharge","gateout_terminal"]
                                  if c in det_risk.columns]
                        st.dataframe(
                            det_risk[show_d].sort_values("days_to_det"),
                            use_container_width=True, hide_index=True
                        )

                with ddem_tab:
                    if dem_risk.empty:
                        st.success("No demurrage risk in next 7 days.")
                    else:
                        dem_risk["Alert"] = dem_risk["days_to_dem"].apply(_risk_flag)
                        show_m = [c for c in ["Alert","container_id","supplier","destination_fc",
                                              "dest_port","last_free_demurrage","days_to_dem",
                                              "eta_discharge","actual_arrival_discharge"]
                                  if c in dem_risk.columns]
                        st.dataframe(
                            dem_risk[show_m].sort_values("days_to_dem"),
                            use_container_width=True, hide_index=True
                        )

        # ══════════════════════════════════════════════════════════════════════
        # SECTION 3 — In-Transit Pipeline (from Shipment Status)
        # ══════════════════════════════════════════════════════════════════════
        st.divider()
        st.markdown("### In-Transit Pipeline")

        if not has_ss:
            st.info("Upload the **Import Shipment Status** report in the sidebar to activate pipeline tracking.")
        else:
            # active = not yet delivered
            active_ss = ss_df[ss_df["delivered"].isna()].copy() if "delivered" in ss_df.columns else ss_df.copy()

            p1,p2,p3,p4 = st.columns(4)
            p1.metric("Total Active Containers",  len(active_ss))
            p2.metric("Unique FCs",     active_ss["fc"].nunique() if "fc" in active_ss.columns else 0)
            p3.metric("Unique Vessels", active_ss["vessel_name"].nunique() if "vessel_name" in active_ss.columns else 0)
            p4.metric("Unique Carriers",active_ss["carrier_code"].nunique() if "carrier_code" in active_ss.columns else 0)

            col_pip1, col_pip2 = st.columns(2)

            with col_pip1:
                st.markdown("#### By Status")
                if "current_status" in active_ss.columns:
                    sc = active_ss["current_status"].fillna("Unknown").value_counts().reset_index()
                    sc.columns = ["Status","Containers"]
                    st.dataframe(sc, use_container_width=True, hide_index=True)

            with col_pip2:
                st.markdown("#### By Destination FC")
                if "fc" in active_ss.columns:
                    fc_c = active_ss["fc"].fillna("Unknown").value_counts().head(12).reset_index()
                    fc_c.columns = ["FC","Containers"]
                    st.bar_chart(fc_c.set_index("FC")["Containers"])

            # ETA table: upcoming arrivals in next 30 days
            if "discharge_eta" in active_ss.columns:
                soon = active_ss[
                    active_ss["discharge_eta"].notna() &
                    (active_ss["discharge_eta"] >= today_ts) &
                    (active_ss["discharge_eta"] <= today_ts + pd.Timedelta(days=30))
                ].copy()
                if not soon.empty:
                    soon["Days to ETA"] = (soon["discharge_eta"] - today_ts).dt.days
                    st.divider()
                    st.markdown(f"#### Arriving at Port — Next 30 Days ({len(soon)} containers)")
                    show_s = [c for c in ["container_no","vendor_name","fc","vessel_name",
                                          "discharge_city","discharge_eta","Days to ETA","current_status"]
                              if c in soon.columns]
                    st.dataframe(
                        soon[show_s].sort_values("Days to ETA"),
                        use_container_width=True, hide_index=True
                    )

        # ══════════════════════════════════════════════════════════════════════
        # SECTION 4 — Carrier Intelligence
        # ══════════════════════════════════════════════════════════════════════
        st.divider()
        st.markdown("### Carrier Intelligence")

        if sub_df.empty:
            st.info("No carrier submissions yet — insights will appear as data comes in.")
        else:
            s1,s2,s3,s4 = st.columns(4)
            s1.metric("Total Submissions", len(sub_df))
            s2.metric("Unique Carriers",   sub_df["carrier_name"].nunique())
            s3.metric("Unique Containers", sub_df["container_id"].nunique())
            last_s = sub_df["submitted_at"].max()
            s4.metric("Last Submission", last_s.strftime("%b %d %H:%M") if pd.notna(last_s) else "—")

            ci1, ci2 = st.columns(2)
            with ci1:
                st.markdown("#### Submission Volume")
                vol = sub_df.groupby("carrier_name").size().sort_values(ascending=False).reset_index()
                vol.columns = ["Carrier","Submissions"]
                st.bar_chart(vol.set_index("Carrier")["Submissions"])

            with ci2:
                if "within_sla" in sub_df.columns and sub_df["within_sla"].notna().any():
                    st.markdown("#### SLA Rate by Carrier")
                    sla = sub_df.dropna(subset=["within_sla"]).copy()
                    sla["within_sla"] = sla["within_sla"].str.upper().str.strip()
                    sp = sla.groupby(["carrier_name","within_sla"]).size().unstack(fill_value=0).reset_index()
                    for c in ["YES","NO"]:
                        if c not in sp.columns: sp[c] = 0
                    sp["Total"] = sp["YES"] + sp["NO"]
                    sp["SLA %"] = (sp["YES"] / sp["Total"].replace(0,1) * 100).round(1)
                    sp = sp.rename(columns={"carrier_name":"Carrier","YES":"Within SLA","NO":"Outside SLA"})
                    sp = sp[["Carrier","Within SLA","Outside SLA","Total","SLA %"]].sort_values("SLA %",ascending=False)
                    try:
                        st.dataframe(
                            sp.style.format({"SLA %":"{:.1f}%"})
                              .background_gradient(subset=["SLA %"],cmap="RdYlGn",vmin=0,vmax=100),
                            use_container_width=True, hide_index=True
                        )
                    except Exception:
                        st.dataframe(sp, use_container_width=True, hide_index=True)

            # freshness
            fresh = (sub_df.groupby("carrier_name")["submitted_at"].max()
                     .reset_index().rename(columns={"carrier_name":"Carrier","submitted_at":"Last Submitted"})
                     .sort_values("Last Submitted",ascending=False))
            # Strip tz-awareness before arithmetic — submitted_at may be naive or tz-aware
            _ls_naive = fresh["Last Submitted"].apply(
                lambda x: x.tz_localize(None) if pd.notna(x) and getattr(x, "tzinfo", None) is not None else x
            )
            fresh["Days Since"] = (today_ts - _ls_naive.dt.normalize()).dt.days
            fresh["Alert"] = fresh["Days Since"].apply(
                lambda d: "Stale (7+ days)" if d>=7 else ("4-6 days" if d>=4 else "Recent")
            )
            fresh["Last Submitted"] = fresh["Last Submitted"].dt.strftime("%b %d %H:%M")
            st.markdown("#### Submission Freshness")
            st.dataframe(fresh, use_container_width=True, hide_index=True)

        # ══════════════════════════════════════════════════════════════════════
        # SECTION 5 — Action Items
        # ══════════════════════════════════════════════════════════════════════
        st.divider()
        st.markdown("### Action Items")
        st.caption("Priority flags auto-generated from all loaded data sources.")

        actions = []

        # From DBR
        if dbr_sheets:
            if overdue_ct > 0:
                actions.append(("Critical", f"{overdue_ct} container(s) with overdue empty returns — DBR",
                                "Go to Empty Returns tab for details."))
            if due_soon_ct > 0:
                actions.append(("Urgent", f"{due_soon_ct} container(s) with empty return due ≤3 days — DBR",
                                "Schedule appointments immediately."))
            if not _dem.empty:
                actions.append(("Critical", f"{len(_dem)} container(s) currently in demurrage — DBR",
                                "Costs accruing daily. Expedite gateout."))
            if dbr_sheets and not _del.empty and "LFD" in _del.columns:
                _lfd_chk = _del.copy()
                _lfd_chk["_lfd"]  = pd.to_datetime(_lfd_chk["LFD"], errors="coerce")
                _lfd_chk["_days"] = (_lfd_chk["_lfd"] - today_ts).dt.days
                lfd_r = _lfd_chk[_lfd_chk["_days"].notna() & (_lfd_chk["_days"] <= 2)]
                if not lfd_r.empty:
                    actions.append(("Critical", f"{len(lfd_r)} container(s) with LFD within 2 days — DBR",
                                    "Expedite or notify carrier."))

        # From Inbound Loads
        if has_il:
            if det_overdue > 0:
                actions.append(("Critical", f"{det_overdue} container(s) past detention free time — Inbound Loads",
                                "Detention fees accruing. Contact dray carrier immediately."))
            if dem_overdue > 0:
                actions.append(("Critical", f"{dem_overdue} container(s) past demurrage free time — Inbound Loads",
                                "Demurrage fees accruing. Contact ocean carrier / terminal."))
            det_warn = len(det_risk) - det_overdue
            if det_warn > 0:
                actions.append(("Urgent", f"{det_warn} container(s) with detention deadline within 7 days",
                                "Schedule dray pickup before free time expires."))

        # From Carrier Submissions
        if not sub_df.empty and "fresh" in dir():
            stale = fresh[fresh["Days Since"] >= 7] if "Days Since" in fresh.columns else pd.DataFrame()
            if not stale.empty:
                carriers_stale = ", ".join(stale["Carrier"].tolist())
                actions.append(("Attention", f"{len(stale)} carrier(s) with no submission in 7+ days",
                                f"Follow up: {carriers_stale}"))

        if not actions:
            st.success("No open action items across all loaded data sources.")
        else:
            priority_order = {"Critical": 0, "Urgent": 1, "Attention": 2}
            actions.sort(key=lambda x: priority_order.get(x[0], 9))
            for priority, title, detail in actions:
                with st.container(border=True):
                    pr_col, txt_col = st.columns([1, 5])
                    pr_col.markdown(f"**{priority}**")
                    txt_col.markdown(f"**{title}**  \n{detail}")




    with _ins_empty:
        st.subheader("Empty Returns")
        st.caption("**Purpose:** Shows which containers still owe a terminal return after FC delivery, ranked by urgency. Overdue returns can trigger fees — this view tells you what needs carrier follow-up today.")
        st.caption("**How to use:** Upload the DBR in the sidebar, then check Overdue and Due Soon first — those need immediate carrier follow-up.")

        if not dbr_sheets:
            st.info("Upload the DBR file in the sidebar to see empty returns data.")
        else:
            er_raw = dbr_sheets.get("Empty Returns", pd.DataFrame()).copy()
            if er_raw.empty:
                st.warning("No Empty Returns sheet found in this DBR.")
            else:
                er_raw.columns = [str(c).strip() for c in er_raw.columns]
                er_raw["Empty Return Due Date"] = pd.to_datetime(
                    er_raw["Empty Return Due Date"], errors="coerce")
                today = pd.Timestamp(date.today())

                # Terminated = resolved. Only open containers are tracked for risk.
                terminated = er_raw[
                    er_raw["Status"].fillna("").str.upper() == "TERMINATED"
                ] if "Status" in er_raw.columns else pd.DataFrame()

                active = er_raw[
                    er_raw["Status"].fillna("").str.upper().ne("TERMINATED") &
                    er_raw["Empty Return Due Date"].notna()
                ].copy()
                active["Days Until Due"] = (active["Empty Return Due Date"] - today).dt.days
                active["Alert"] = active["Days Until Due"].apply(
                    lambda d: "OVERDUE" if d < 0 else ("Due soon" if d <= 3 else "OK")
                )

                overdue  = active[active["Days Until Due"] < 0]
                due_soon = active[(active["Days Until Due"] >= 0) & (active["Days Until Due"] <= 3)]
                on_track = active[active["Days Until Due"] > 3]

                m1, m2, m3, m4, m5 = st.columns(5)
                m1.metric("Open",         len(active))
                m2.metric("Overdue",      len(overdue))
                m3.metric("Due ≤ 3 days", len(due_soon))
                m4.metric("On track",     len(on_track))
                m5.metric("Terminated",   len(terminated), help="Return confirmed or obligation closed — no action needed")
                st.divider()

                view = st.radio(
                    "Show",
                    ["Overdue", "Due Soon (≤3 days)", "All Open"],
                    horizontal=True,
                )
                display = (
                    overdue  if view == "Overdue" else
                    due_soon if view == "Due Soon (≤3 days)" else
                    active
                )

                show_cols = [c for c in [
                    "Alert", "Container #", "Terminal",
                    "Empty Return Due Date", "Appointment Date",
                    "Status", "Days Until Due", "If Outside Window - Reason",
                ] if c in display.columns]

                if display.empty:
                    st.success("No containers in this category.")
                else:
                    st.dataframe(
                        display[show_cols].sort_values("Days Until Due", ascending=True)
                        if "Days Until Due" in display.columns else display[show_cols],
                        use_container_width=True, hide_index=True,
                    )
                    buf = io.BytesIO()
                    display.to_excel(buf, index=False)
                    buf.seek(0)
                    st.download_button(
                        "Export (.xlsx)", data=buf,
                        file_name=f"empty_returns_{date.today()}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

                # Terminated — collapsed by default, clearly labelled as resolved
                if not terminated.empty:
                    with st.expander(f"Terminated / Resolved ({len(terminated)} containers)", expanded=False):
                        st.caption("These containers have had their return obligation closed. No action needed.")
                        term_cols = [c for c in ["Container #", "Terminal", "Empty Return Due Date",
                                                  "Status", "If Outside Window - Reason"]
                                     if c in terminated.columns]
                        st.dataframe(terminated[term_cols] if term_cols else terminated,
                                     use_container_width=True, hide_index=True)


    with _ins_lanes:
        st.subheader("Lane Cost Simulator")
        st.caption("**Purpose:** Compare drayage rates across carriers by lane and model the total cost of a delivery scenario. Use this when allocating containers across carriers or evaluating cost trade-offs.")
        st.caption("**How to use:** Check the Rate Matrix for cheapest carrier per lane, then use the Scenario Builder to model cost by entering lane and container count.")

        # ── Load rate + route data ────────────────────────────────────────────────
        conn = get_db()
        rates_df  = pd.read_sql("SELECT * FROM rate_lanes  ORDER BY port, destination, base_rate", conn)
        routes_df = pd.read_sql(
            "SELECT DISTINCT node_code, fc_code, node_address, dest_port, node_type, warehouse_partner, dray_bid_done "
            "FROM route_lanes ORDER BY node_code", conn
        )
        conn.close()

        if rates_df.empty:
            st.info(
                "No rate data loaded yet. Rate cards are imported by the AGL ops team — "
                "contact the app admin to refresh."
            )

        if not rates_df.empty:
            # ── Build node lookups ────────────────────────────────────────────────────
            # rate_lanes.destination matches either node_code (A100) or fc_code (ARW0)
            fc_to_node   = routes_df.dropna(subset=["fc_code"]).set_index("fc_code")["node_code"].to_dict()
            node_to_fc   = routes_df.dropna(subset=["fc_code"]).set_index("node_code")["fc_code"].to_dict()
            node_to_addr = routes_df.set_index("node_code")["node_address"].to_dict()
            node_to_type = routes_df.set_index("node_code")["node_type"].to_dict()

            # normalize destination to node_code where possible
            rates_df["node_code"] = rates_df["destination"].apply(
                lambda d: d if str(d) in node_to_fc else fc_to_node.get(str(d), str(d))
            )

            # ── Global filters ────────────────────────────────────────────────────────
            gf1, gf2, gf3 = st.columns(3)
            all_ports   = sorted(rates_df["port"].dropna().unique())
            all_sources = sorted(rates_df["rate_source"].dropna().unique())
            all_types   = sorted(rates_df["lane_type"].dropna().unique())

            with gf1:
                g_port = st.selectbox("Port of Arrival", ["All"] + all_ports, key="g_port")
            with gf2:
                g_sources = st.multiselect("Rate Source", all_sources,
                    default=["robotics_2026"] if "robotics_2026" in all_sources else all_sources,
                    key="g_sources")
            with gf3:
                g_types = st.multiselect("Lane Type", all_types, default=all_types, key="g_types")

            rfilt = rates_df.copy()
            if g_port != "All":
                rfilt = rfilt[rfilt["port"] == g_port]
            if g_sources:
                rfilt = rfilt[rfilt["rate_source"].isin(g_sources)]
            if g_types:
                rfilt = rfilt[rfilt["lane_type"].isin(g_types)]

            if rfilt.empty:
                st.warning("No rates for the selected filters.")

            # ── Rate Matrix ────────────────────────────────────────────────────────────
            with st.expander("Rate Matrix — all carrier options by lane", expanded=True):
                st.caption("Green = cheapest carrier per lane  ·  Red = most expensive  ·  — = no rate on this lane")

                pivot = rfilt.pivot_table(
                    index=["port", "node_code"],
                    columns="carrier_name",
                    values="base_rate",
                    aggfunc="min",
                ).reset_index()
                pivot.columns.name = None
                carrier_cols = [c for c in pivot.columns if c not in ["port", "node_code"]]

                pivot["FC"] = pivot["node_code"].map(node_to_fc)
                pivot["Type"] = pivot["node_code"].map(node_to_type)
                pivot["Facility"] = pivot["node_code"].map(
                    lambda n: (node_to_addr.get(n) or "")[:55]
                )

                if carrier_cols:
                    pivot["Cheapest"] = pivot[carrier_cols].idxmin(axis=1)
                    pivot["Low $"]  = pivot[carrier_cols].min(axis=1)
                    pivot["Spread"] = pivot[carrier_cols].max(axis=1) - pivot["Low $"]

                col_order = (["port", "node_code", "FC", "Type", "Facility"]
                             + carrier_cols
                             + ["Cheapest", "Low $", "Spread"])
                col_order = [c for c in col_order if c in pivot.columns]
                disp = pivot[col_order].rename(columns={"port": "Port", "node_code": "Node"}).sort_values(["Port", "Node"])

                fmt = {c: "${:,.0f}" for c in carrier_cols}
                fmt.update({"Low $": "${:,.0f}", "Spread": "${:,.0f}"})

                try:
                    styled = disp.style.format(fmt, na_rep="—")
                    if len(carrier_cols) > 1:
                        styled = (styled
                            .highlight_min(subset=carrier_cols, axis=1, color="#c8f7c5")
                            .highlight_max(subset=carrier_cols, axis=1, color="#fad7d7"))
                    st.dataframe(styled, use_container_width=True, hide_index=True)
                except Exception:
                    st.dataframe(disp, use_container_width=True, hide_index=True)

            st.divider()

            # ── Scenario Builder ───────────────────────────────────────────────────────
            st.markdown("### Scenario Builder")
            st.caption("Select port → node → carrier, enter container count, build your cost projection.")

            sb_ports = sorted(rfilt["port"].dropna().unique())
            sb1, sb2, sb3, sb4 = st.columns([1.2, 1.5, 2, 1])

            with sb1:
                sb_port = st.selectbox("Port", sb_ports, key="sb_port")

            sb_nodes = sorted(rfilt[rfilt["port"] == sb_port]["node_code"].dropna().unique())
            with sb2:
                sb_node = st.selectbox(
                    "Node", sb_nodes, key="sb_node",
                    format_func=lambda n: f"{n}  ({node_to_fc.get(n, '?')})",
                )

            carriers_here = (
                rfilt[(rfilt["port"] == sb_port) & (rfilt["node_code"] == sb_node)]
                [["carrier_name", "base_rate", "lane_type"]]
                .drop_duplicates()
                .sort_values("base_rate")
            )
            c_opts = [
                f"{r.carrier_name}  |  ${r.base_rate:,.0f}  [{r.lane_type}]"
                for _, r in carriers_here.iterrows()
            ]

            with sb3:
                sb_carrier_sel = st.selectbox("Carrier & Rate", c_opts if c_opts else ["No rates on this lane"], key="sb_carrier")
            with sb4:
                sb_count = st.number_input("Containers", min_value=1, value=10, step=1, key="sb_count")

            if st.button("Add Lane to Scenario", type="primary"):
                if c_opts and sb_carrier_sel != "No rates on this lane":
                    carrier_nm = sb_carrier_sel.split("  |  $")[0].strip()
                    rate_val   = float(sb_carrier_sel.split("$")[1].split("  ")[0].replace(",", ""))
                    fc   = node_to_fc.get(sb_node, "")
                    addr = (node_to_addr.get(sb_node) or "")[:60]
                    if "scenario_v2" not in st.session_state:
                        st.session_state.scenario_v2 = []
                    st.session_state.scenario_v2.append({
                        "Port":       sb_port,
                        "Node":       sb_node,
                        "FC":         fc,
                        "Facility":   addr,
                        "Carrier":    carrier_nm,
                        "Containers": sb_count,
                        "Rate ($)":   rate_val,
                        "Cost ($)":   round(sb_count * rate_val, 2),
                    })
                    st.rerun()

            # ── Portfolio ─────────────────────────────────────────────────────────────
            items = st.session_state.get("scenario_v2", [])

            if not items:
                st.info("Add lanes above to build your cost scenario.")
            else:
                sc_df = pd.DataFrame(items)
                total_ctr  = int(sc_df["Containers"].sum())
                total_cost = float(sc_df["Cost ($)"].sum())
                avg_rate   = total_cost / total_ctr if total_ctr else 0

                pm1, pm2, pm3 = st.columns(3)
                pm1.metric("Total Containers",    f"{total_ctr:,}")
                pm2.metric("Total Projected Cost", f"${total_cost:,.0f}")
                pm3.metric("Avg Cost / Container", f"${avg_rate:,.0f}")

                edited_sc = st.data_editor(
                    sc_df,
                    column_config={
                        "Port":       st.column_config.TextColumn("Port",     disabled=True),
                        "Node":       st.column_config.TextColumn("Node",     disabled=True),
                        "FC":         st.column_config.TextColumn("FC",       disabled=True),
                        "Facility":   st.column_config.TextColumn("Facility", disabled=True),
                        "Carrier":    st.column_config.TextColumn("Carrier",  disabled=True),
                        "Containers": st.column_config.NumberColumn("Containers", min_value=0),
                        "Rate ($)":   st.column_config.NumberColumn("Rate ($)",   format="$%d", disabled=True),
                        "Cost ($)":   st.column_config.NumberColumn("Cost ($)",   format="$%d", disabled=True),
                    },
                    use_container_width=True, hide_index=True, num_rows="dynamic", key="sc_editor",
                )

                # recalc costs if containers edited
                if not edited_sc.equals(sc_df):
                    edited_sc["Cost ($)"] = edited_sc["Containers"] * edited_sc["Rate ($)"]
                    st.session_state.scenario_v2 = edited_sc.to_dict("records")
                    sc_df = edited_sc.copy()
                    total_cost = float(sc_df["Cost ($)"].sum())
                    total_ctr  = int(sc_df["Containers"].sum())

                dl_c, cl_c = st.columns([3, 1])
                with dl_c:
                    buf = io.BytesIO()
                    sc_df.to_excel(buf, index=False)
                    buf.seek(0)
                    st.download_button("⬇️ Export scenario (.xlsx)", buf,
                        file_name=f"dray_scenario_{date.today()}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                with cl_c:
                    if st.button("Clear All", type="secondary", use_container_width=True):
                        st.session_state.scenario_v2 = []
                        st.rerun()

                st.divider()

                # ── By carrier + by port breakdown ────────────────────────────────────
                lc, rc = st.columns(2)

                with lc:
                    st.markdown("#### Cost by Carrier")
                    by_c = sc_df.groupby("Carrier").agg(
                        Containers=("Containers", "sum"),
                        Cost=("Cost ($)", "sum"),
                    ).reset_index()
                    by_c["% Total"] = (by_c["Cost"] / total_cost * 100).round(1) if total_cost else 0
                    by_c.columns = ["Carrier", "Containers", "Cost ($)", "% Total"]
                    st.dataframe(
                        by_c.style.format({"Cost ($)": "${:,.0f}", "% Total": "{:.1f}%"}),
                        use_container_width=True, hide_index=True,
                    )
                    st.bar_chart(by_c.set_index("Carrier")["Cost ($)"])

                with rc:
                    st.markdown("#### Cost by Port")
                    by_p = sc_df.groupby("Port").agg(
                        Containers=("Containers", "sum"),
                        Cost=("Cost ($)", "sum"),
                    ).reset_index()
                    by_p["% Total"] = (by_p["Cost"] / total_cost * 100).round(1) if total_cost else 0
                    by_p.columns = ["Port", "Containers", "Cost ($)", "% Total"]
                    st.dataframe(
                        by_p.style.format({"Cost ($)": "${:,.0f}", "% Total": "{:.1f}%"}),
                        use_container_width=True, hide_index=True,
                    )
                    st.bar_chart(by_p.set_index("Port")["Cost ($)"])

                # ── Savings vs cheapest ───────────────────────────────────────────────
                st.divider()
                st.markdown("#### Savings vs Cheapest Available")

                sav_rows = []
                for _, row in sc_df.iterrows():
                    lane_r = rfilt[(rfilt["port"] == row["Port"]) & (rfilt["node_code"] == row["Node"])]
                    if lane_r.empty:
                        continue
                    min_idx      = lane_r["base_rate"].idxmin()
                    min_rate     = float(lane_r.loc[min_idx, "base_rate"])
                    min_carrier  = str(lane_r.loc[min_idx, "carrier_name"])
                    current_cost = float(row["Cost ($)"])
                    cheapest     = row["Containers"] * min_rate
                    sav_rows.append({
                        "Lane":          f"{row['Port']} → {row['Node']} ({row['FC']})",
                        "Selected":      row["Carrier"],
                        "Rate ($)":      row["Rate ($)"],
                        "Cheapest":      min_carrier,
                        "Cheapest Rate": min_rate,
                        "Containers":    row["Containers"],
                        "Current Cost":  current_cost,
                        "Min Cost":      cheapest,
                        "Savings":       current_cost - cheapest,
                    })

                if sav_rows:
                    sav_df = pd.DataFrame(sav_rows)
                    total_sav = float(sav_df["Savings"].sum())

                    if total_sav > 1:
                        st.info(f"Switching to cheapest available on all lanes would save **${total_sav:,.0f}**")
                    else:
                        st.success("You're already using the cheapest available carrier on all lanes.")

                    st.dataframe(
                        sav_df.style.format({
                            "Rate ($)":      "${:,.0f}",
                            "Cheapest Rate": "${:,.0f}",
                            "Current Cost":  "${:,.0f}",
                            "Min Cost":      "${:,.0f}",
                            "Savings":       "${:,.0f}",
                        }),
                        use_container_width=True, hide_index=True,
                    )




# TAB 7 — Planning / Scheduler (v2)
# ══════════════════════════════════════════════════════════════════════════════

_PLAN_SLOTS = [
    ("7:30 AM", 1), ("8:30 AM", 2), ("9:30 AM", 3),
    ("10:30 AM", 4), ("11:30 AM", 5),
    ("1:00 PM", 6), ("2:00 PM", 7), ("3:00 PM", 8), ("3:30 PM", 9), ("4:00 PM", 10),
]
_PLAN_PRODUCTS = ["SHELF, HDTP", "STRAP, HDTP", "SHELF, CLDTP", "HDPT 2.0", "BASE ASSEMBLY, HDTP", "OTHER"]
_SCAC_COLOR = {
    "ATMI": "#E8A020", "ARVY": "#4285F4",
    "HDDR": "#34A853", "RKNE": "#9C27B0", "TGHE": "#E91E63",
}
_PLAN_STATUSES = ["SCHEDULED", "DELIVERED", "REJECTED", "RESCHEDULED", "HOLD", "PENDING"]
_STATUS_BADGE = {
    "SCHEDULED": "Scheduled", "DELIVERED": "Delivered", "REJECTED": "Rejected",
    "RESCHEDULED": "Rescheduled", "HOLD": "Hold", "PENDING": "Pending",
}

def _init_plan_config():
    conn = get_db()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS plan_sites (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            site_code    TEXT NOT NULL UNIQUE,
            fc_name      TEXT,
            port         TEXT,
            region       TEXT,
            capacity_day INTEGER DEFAULT 9,
            active       INTEGER DEFAULT 1,
            notes        TEXT
        );
        CREATE TABLE IF NOT EXISTS plan_carriers (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            scac         TEXT NOT NULL UNIQUE,
            carrier_name TEXT NOT NULL,
            ops_contact  TEXT,
            ops_email    TEXT,
            active       INTEGER DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS plan_site_carrier (
            id                 INTEGER PRIMARY KEY AUTOINCREMENT,
            site_code          TEXT NOT NULL,
            scac               TEXT NOT NULL,
            site_contact       TEXT,
            site_contact_email TEXT,
            priority_time      TEXT,
            allocation_pct     REAL DEFAULT 0,
            active             INTEGER DEFAULT 1,
            notes              TEXT,
            UNIQUE(site_code, scac)
        );
        CREATE TABLE IF NOT EXISTS plan_carrier_tokens (
            scac        TEXT NOT NULL UNIQUE,
            token       TEXT NOT NULL UNIQUE,
            created_at  TEXT
        );
    """)
    conn.commit()
    try:
        conn.execute("ALTER TABLE delivery_plan ADD COLUMN site_code TEXT")
        conn.commit()
    except sqlite3.OperationalError:
        pass
    if conn.execute("SELECT COUNT(*) FROM plan_sites").fetchone()[0] == 0:
        conn.executemany(
            "INSERT OR IGNORE INTO plan_sites (site_code,fc_name,port,region,capacity_day,active,notes) VALUES (?,?,?,?,?,?,?)",
            [
                ("RIC6","Richmond FC",       "ORF","East",      9,1,""),
                ("ILM1","Wilmington NC",     "ORF","East",      9,1,"Knightrider (NCKR) for overflow"),
                ("IAG1","Niagara NY",        "ORF","East",      4,1,"Via XPH1 transload — 53ft trailer"),
                ("DBM6","DBM6",              "ORF","East",      6,1,""),
                ("ATL2","ATL2",              "SAV","Southeast", 6,1,""),
                ("DRO1","DRO1",              "ORF","East",      4,1,""),
                ("LAX", "LAX / West Coast",  "LAX","West",      6,1,""),
                ("OAK", "OAK / Northern CA", "OAK","West",      6,1,""),
                ("RSW9","Fort Myers FL",     "MIA","Southeast", 4,0,"NEW — starts 11/2/2026"),
            ]
        )
    if conn.execute("SELECT COUNT(*) FROM plan_carriers").fetchone()[0] == 0:
        conn.executemany(
            "INSERT OR IGNORE INTO plan_carriers (scac,carrier_name,ops_contact,ops_email) VALUES (?,?,?,?)",
            [
                ("ATMI","Cargomatic",       "Tyler Domingues",    "tdomingues@cargomatic.com"),
                ("ARVY","Arrive Logistics", "Tyler Spangler",     ""),
                ("HDDR","Maersk/HUDD",      "Emily Christenbury", ""),
                ("RKNE","Road One",         "Mark Brennan",       ""),
                ("TGHE","Tighe Logistics",  "Jose Alvarado",      "jalvarado@tighe-co.com"),
            ]
        )
    if conn.execute("SELECT COUNT(*) FROM plan_site_carrier").fetchone()[0] == 0:
        conn.executemany(
            "INSERT OR IGNORE INTO plan_site_carrier (site_code,scac,site_contact,site_contact_email,priority_time,allocation_pct,active,notes) VALUES (?,?,?,?,?,?,?,?)",
            [
                ("RIC6","ATMI","Tyler Domingues",    "tdomingues@cargomatic.com","8:30 AM",60.0,1,""),
                ("RIC6","HDDR","Sandji Ruffin",      "",                         "7:30 AM",20.0,1,"Before lunch"),
                ("RIC6","ARVY","Tyler Spangler",     "",                         "8:30 AM",20.0,1,""),
                ("ILM1","HDDR","Jerry Nesbit",       "",                         "7:30 AM",100.0,1,""),
                ("ILM1","ATMI","Tyler Domingues",    "tdomingues@cargomatic.com","8:30 AM",0.0, 1,"As needed"),
                ("IAG1","ATMI","Tyler Domingues",    "tdomingues@cargomatic.com","8:30 AM",100.0,1,"Via XPH1 transload"),
                ("DBM6","ATMI","Tyler Domingues",    "tdomingues@cargomatic.com","8:30 AM",100.0,1,""),
                ("DRO1","ATMI","Tyler Domingues",    "tdomingues@cargomatic.com","8:30 AM",100.0,1,""),
                ("ATL2","ATMI","Tyler Domingues",    "tdomingues@cargomatic.com","8:30 AM",50.0,1,""),
                ("ATL2","RKNE","Michelle Fileccia",  "",                         "8:30 AM",50.0,1,""),
                ("LAX", "ATMI","Anel Nunez",         "",                         "8:30 AM",50.0,1,""),
                ("LAX", "HDDR","Ailua Osoimalo",     "",                         "7:30 AM",50.0,1,"Desirae Swain backup"),
                ("OAK", "ATMI","Tyler Domingues",    "tdomingues@cargomatic.com","8:30 AM",20.0,1,""),
                ("OAK", "HDDR","Emily Christenbury", "",                         "7:30 AM",40.0,1,""),
                ("OAK", "ARVY","Tyler Spangler",     "",                         "8:30 AM",40.0,1,""),
                ("RSW9","RKNE","Chuck Henderson",    "",                         "8:30 AM",100.0,0,"NEW 11/2/26"),
            ]
        )
    # week config table — stores active receiving days per week
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS plan_week_config (
            week_start  TEXT NOT NULL UNIQUE,
            active_days TEXT NOT NULL DEFAULT 'Mon,Tue,Wed,Thu,Fri',
            updated_at  TEXT
        );
    """)
    conn.commit()
    conn.close()

_init_plan_config()

def _migrate_plan_config():
    """One-time migrations to fill gaps in seeded config data."""
    conn = get_db()
    # Carrier emails
    conn.executemany(
        "UPDATE plan_carriers SET ops_email=? WHERE scac=? AND (ops_email IS NULL OR ops_email='')",
        [
            ("tspangler@arrivelogistics.com", "ARVY"),
        ]
    )
    # Site-carrier contact emails (HDDR contacts per site)
    conn.executemany(
        "UPDATE plan_site_carrier SET site_contact_email=? WHERE site_code=? AND scac=? AND (site_contact_email IS NULL OR site_contact_email='')",
        [
            ("sandji.ruffin@maersk.com",  "RIC6", "HDDR"),
            ("jerry.nesbit@maersk.com",   "ILM1", "HDDR"),
            ("desirae.swain@maersk.com",  "LAX",  "HDDR"),
        ]
    )
    # Add SJC8 site if missing
    conn.execute(
        """INSERT OR IGNORE INTO plan_sites (site_code,fc_name,port,region,capacity_day,active,notes)
           VALUES ('SJC8','San Jose CA','OAK','West',6,1,'West coast; via OAK port')"""
    )
    # Add SJC8 carrier mapping (ATMI)
    conn.execute(
        """INSERT OR IGNORE INTO plan_site_carrier
           (site_code,scac,site_contact,site_contact_email,priority_time,allocation_pct,active,notes)
           VALUES ('SJC8','ATMI','Tyler Domingues','tdomingues@cargomatic.com','8:30 AM',100.0,1,'')"""
    )
    # New site constraint columns
    for _col_ddl in [
        "ALTER TABLE plan_sites ADD COLUMN capacity_sat INTEGER DEFAULT 5",
        "ALTER TABLE plan_sites ADD COLUMN last_appt_time TEXT DEFAULT '4:00 PM'",
        "ALTER TABLE plan_sites ADD COLUMN last_appt_sat TEXT DEFAULT '11:30 AM'",
        "ALTER TABLE plan_sites ADD COLUMN capacity_week INTEGER DEFAULT 0",
    ]:
        try:
            conn.execute(_col_ddl)
        except sqlite3.OperationalError:
            pass
    # RIC6 hard constraints from SOP
    conn.execute("""UPDATE plan_sites SET capacity_day=10, capacity_sat=5,
        last_appt_time='4:00 PM', last_appt_sat='11:30 AM', capacity_week=115
        WHERE site_code='RIC6'""")
    # delivery_plan confirmation columns
    for _col in [
        "ALTER TABLE delivery_plan ADD COLUMN carrier_confirmed INTEGER DEFAULT 0",
        "ALTER TABLE delivery_plan ADD COLUMN confirmed_at TEXT",
        "ALTER TABLE delivery_plan ADD COLUMN carrier_note TEXT",
    ]:
        try:
            conn.execute(_col)
        except sqlite3.OperationalError:
            pass
    conn.commit()
    conn.close()

_migrate_plan_config()

# ── Planning: slot tables, HUDD rule, smart scheduler ────────────────────────
_WEEKDAY_SLOTS  = [
    ("7:30 AM",1),("8:30 AM",2),("9:30 AM",3),("10:30 AM",4),("11:30 AM",5),
    ("1:00 PM",6),("2:00 PM",7),("3:00 PM",8),("3:30 PM",9),("4:00 PM",10),
]
_SATURDAY_SLOTS = [("7:30 AM",1),("8:30 AM",2),("9:30 AM",3),("10:30 AM",4),("11:30 AM",5)]
_HUDD_SCACS     = {"HUDD","HDDR"}


def _get_site_constraints(site_code: str) -> dict:
    """Return site constraint dict from plan_sites with safe defaults."""
    conn = get_db()
    try:
        row = conn.execute(
            "SELECT capacity_day, capacity_sat, last_appt_time, last_appt_sat, capacity_week "
            "FROM plan_sites WHERE site_code=?", (site_code,)
        ).fetchone()
    except Exception:
        row = None
    conn.close()
    if row:
        return {
            "capacity_day":   int(row[0] or 10),
            "capacity_sat":   int(row[1] or 5),
            "last_appt_time": row[2] or "4:00 PM",
            "last_appt_sat":  row[3] or "11:30 AM",
            "capacity_week":  int(row[4] or 0),
        }
    return {"capacity_day":10,"capacity_sat":5,"last_appt_time":"4:00 PM",
            "last_appt_sat":"11:30 AM","capacity_week":0}


def _get_reschedulable_pool(site_code: str, current_week_iso: str) -> pd.DataFrame:
    """Pull containers needing rescheduling from prior plan weeks.

    Returns two buckets tagged in pool_reason:
    - 'Undelivered': status SCHEDULED/HOLD with appt_date before today (never delivered)
    - 'Refused/Rejected': status REJECTED (delivery was turned away)
    """
    today_iso = date.today().isoformat()
    conn = get_db()
    try:
        df = pd.read_sql_query(
            """
            SELECT *, 'Undelivered' AS pool_reason FROM delivery_plan
            WHERE site_code=? AND week_start < ?
              AND status IN ('SCHEDULED','HOLD')
              AND appt_date IS NOT NULL AND appt_date < ?
            UNION ALL
            SELECT *, 'Refused/Rejected' AS pool_reason FROM delivery_plan
            WHERE site_code=? AND status='REJECTED'
            ORDER BY appt_date ASC
            """,
            conn, params=(site_code, current_week_iso, today_iso, site_code)
        )
    except Exception:
        df = pd.DataFrame()
    conn.close()
    if df.empty:
        return df
    # Deduplicate: keep REJECTED over Undelivered for the same container
    df = df.sort_values("pool_reason").drop_duplicates("container_id", keep="last")
    return df.reset_index(drop=True)


def _auto_schedule_pool(containers: list, week_start, work_days: list, site_code: str) -> tuple:
    """Smart slot filler.

    containers: list of dicts {cid, scac, product, site, notes, status, ...}
    Assumes already FIFO-sorted (refused first, then carried-over, then new GVT).

    Rules:
    - HUDD/HDDR: 7:30 AM first slot, one per weekday
    - Other carriers: fill remaining slots round-robin (level-load) across all active days
    - Saturday: 5 slots max (7:30-11:30 AM only)
    - Overflow = containers that could not be placed

    Returns (scheduled: list[dict], overflow: list[dict])
    """
    slots_by_day = {}
    for day in work_days:
        is_sat = day.weekday() == 5
        slots_by_day[day] = list(_SATURDAY_SLOTS if is_sat else _WEEKDAY_SLOTS)

    scheduled, overflow = [], []
    hudd_q  = [c for c in containers if c.get("scac","") in _HUDD_SCACS]
    other_q = [c for c in containers if c.get("scac","") not in _HUDD_SCACS]

    # Assign HUDD: one per weekday at 7:30 AM (Saturday excluded)
    for day in [d for d in work_days if d.weekday() != 5]:
        if not hudd_q:
            break
        c = hudd_q.pop(0)
        slots = slots_by_day[day]
        idx = next((i for i,(t,_) in enumerate(slots) if t == "7:30 AM"), None)
        if idx is not None:
            slots.pop(idx)
            scheduled.append({**c, "appt_date": day, "appt_time": "7:30 AM",
                               "slot_num": 1, "status": "SCHEDULED"})
        else:
            overflow.append(c)
    overflow.extend(hudd_q)  # HUDD with no remaining weekday

    # Level-load others: round-robin across all active days
    days_list = list(work_days)
    d_idx = 0
    for c in other_q:
        placed = False
        for attempt in range(len(days_list)):
            day = days_list[(d_idx + attempt) % len(days_list)]
            if slots_by_day[day]:
                t, sn = slots_by_day[day].pop(0)
                scheduled.append({**c, "appt_date": day, "appt_time": t,
                                   "slot_num": sn, "status": "SCHEDULED"})
                d_idx = (d_idx + attempt + 1) % len(days_list)
                placed = True
                break
        if not placed:
            overflow.append(c)

    return scheduled, overflow


def _get_or_create_carrier_token(scac: str) -> str:
    """Return existing token for SCAC or generate a new UUID token."""
    import uuid
    conn = get_db()
    row = conn.execute("SELECT token FROM plan_carrier_tokens WHERE scac=?", (scac,)).fetchone()
    if row:
        conn.close()
        return row["token"]
    token = str(uuid.uuid4()).replace("-", "")
    conn.execute(
        "INSERT INTO plan_carrier_tokens (scac,token,created_at) VALUES (?,?,?)",
        (scac, token, datetime.now().isoformat())
    )
    conn.commit()
    conn.close()
    return token


def _regenerate_carrier_token(scac: str) -> str:
    """Force-generate a new token for SCAC (invalidates old link)."""
    import uuid
    conn = get_db()
    token = str(uuid.uuid4()).replace("-", "")
    conn.execute(
        "INSERT OR REPLACE INTO plan_carrier_tokens (scac,token,created_at) VALUES (?,?,?)",
        (scac, token, datetime.now().isoformat())
    )
    conn.commit()
    conn.close()
    return token


_ALL_DAYS   = ["Sun","Mon","Tue","Wed","Thu","Fri","Sat"]
_DEFAULT_DAYS = ["Mon","Tue","Wed","Thu","Fri"]

def _get_week_config(week_iso: str) -> list[str]:
    conn = get_db()
    row = conn.execute("SELECT active_days FROM plan_week_config WHERE week_start=?", (week_iso,)).fetchone()
    conn.close()
    if row:
        return [d.strip() for d in row[0].split(",") if d.strip()]
    return list(_DEFAULT_DAYS)

def _save_week_config(week_iso: str, active_days: list[str]):
    conn = get_db()
    conn.execute(
        """INSERT INTO plan_week_config (week_start, active_days, updated_at)
           VALUES (?,?,?)
           ON CONFLICT(week_start) DO UPDATE SET active_days=excluded.active_days, updated_at=excluded.updated_at""",
        (week_iso, ",".join(active_days), datetime.utcnow().isoformat())
    )
    conn.commit()
    conn.close()



# ── plan DB helpers ───────────────────────────────────────────────────────────

def _plan_week_start(d: date) -> date:
    return d - timedelta(days=d.weekday())

def _get_plan(week_iso: str) -> pd.DataFrame:
    conn = get_db()
    df = pd.read_sql_query(
        "SELECT * FROM delivery_plan WHERE week_start=? ORDER BY appt_date, slot_num, id",
        conn, params=(week_iso,)
    )
    conn.close()
    return df

def _get_sites_df() -> pd.DataFrame:
    conn = get_db()
    df = pd.read_sql_query("SELECT * FROM plan_sites ORDER BY site_code", conn)
    conn.close()
    return df

def _get_carriers_df() -> pd.DataFrame:
    conn = get_db()
    df = pd.read_sql_query("SELECT * FROM plan_carriers ORDER BY scac", conn)
    conn.close()
    return df

def _get_scm_df() -> pd.DataFrame:
    conn = get_db()
    df = pd.read_sql_query(
        """SELECT m.*, s.fc_name, c.carrier_name
           FROM plan_site_carrier m
           LEFT JOIN plan_sites s ON m.site_code=s.site_code
           LEFT JOIN plan_carriers c ON m.scac=c.scac
           ORDER BY m.site_code, m.scac""", conn)
    conn.close()
    return df

def _add_plan_entry(week_start, appt_date, appt_time, slot_num,
                    container_id, scac, site_code, product_type, qty, notes, status):
    db_write(
        """INSERT INTO delivery_plan
           (week_start,appt_date,appt_time,slot_num,container_id,
            carrier,site_code,product_type,qty,notes,status,created_at,updated_at)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (week_start.isoformat(),
         appt_date.isoformat() if appt_date else None,
         appt_time, slot_num, container_id.upper().strip(),
         scac, site_code, product_type,
         int(qty) if qty else None,
         notes.strip() if notes else None, status,
         datetime.utcnow().isoformat(), datetime.utcnow().isoformat())
    )

def _update_plan_entry(row_id: int, **kwargs):
    conn = get_db()
    kwargs["updated_at"] = datetime.utcnow().isoformat()
    set_clause = ", ".join(f"{k}=?" for k in kwargs)
    conn.execute(f"UPDATE delivery_plan SET {set_clause} WHERE id=?",
                 list(kwargs.values()) + [row_id])
    conn.commit(); conn.close()
    if S3_ENABLED:
        data_sync.push_db_to_s3(AWS_KEY, AWS_SECRET, AWS_REGION, S3_BUCKET)

def _delete_plan_entry(row_id: int):
    conn = get_db()
    conn.execute("DELETE FROM delivery_plan WHERE id=?", (row_id,))
    conn.commit(); conn.close()
    if S3_ENABLED:
        data_sync.push_db_to_s3(AWS_KEY, AWS_SECRET, AWS_REGION, S3_BUCKET)

def _config_write(sql: str, params: tuple):
    conn = get_db()
    conn.execute(sql, params)
    conn.commit(); conn.close()


# ── DBR Tracker historical importer ──────────────────────────────────────────
def _parse_dbr_tracker(file_bytes: bytes) -> tuple:
    """Parse ToteASERs Robotics DBR Tracker.xlsx -> list of delivery_plan dicts.
    Returns (rows: list[dict], errors: list[str]).
    """
    import io as _bio
    wb = openpyxl.load_workbook(_bio.BytesIO(file_bytes), data_only=True)

    if "Delivery Plan" not in wb.sheetnames:
        return [], ["File missing 'Delivery Plan' sheet — is this the DBR Tracker?"]

    # Build container -> (product_type, qty) from Consolidated sheet
    prod_lookup: dict = {}
    if "Consolidated" in wb.sheetnames:
        for i, row in enumerate(wb["Consolidated"].iter_rows(values_only=True)):
            if i == 0:
                continue
            cid = row[2]
            if cid:
                prod_lookup[str(cid).strip()] = (
                    str(row[8]).strip() if row[8] else None,
                    int(row[9]) if row[9] else None,
                )

    def _scac(carrier_str):
        # "ARVY (Arrive)" -> "ARVY"; normalize HUDD → HDDR (Maersk)
        if not carrier_str:
            return None
        code = str(carrier_str).strip().split()[0].upper()
        return "HDDR" if code == "HUDD" else code

    STATUS_MAP = {
        "Delivered": "delivered",
        "At Yard":   "at_yard",
        "Pending":   "planned",
    }

    rows, errors = [], []
    for i, row in enumerate(wb["Delivery Plan"].iter_rows(values_only=True)):
        if i <= 1:
            continue  # row 0 = grouping label, row 1 = headers
        container = row[3]
        if not container:
            continue
        container = str(container).strip()
        fc_dest  = row[7]
        fc_sched = row[8]
        if not fc_dest or not fc_sched:
            errors.append(f"Row {i+1}: missing FC Destination or FC Sched Del for {container} — skipped")
            continue
        # Resolve appt_date
        if hasattr(fc_sched, "date"):
            appt_date = fc_sched.date()
        else:
            try:
                appt_date = datetime.strptime(str(fc_sched), "%Y-%m-%d").date()
            except Exception:
                errors.append(f"Row {i+1}: unparseable date '{fc_sched}' for {container} — skipped")
                continue
        # Notes: merge Notes field + Live/Drop
        note_parts = []
        if row[12]:
            note_parts.append(str(row[12]).strip())
        if row[14]:
            note_parts.append(f"[{str(row[14]).strip()}]")
        prod_type, qty = prod_lookup.get(container, (None, None))
        rows.append({
            "week_start":   _plan_week_start(appt_date).isoformat(),
            "appt_date":    appt_date.isoformat(),
            "appt_time":    None,
            "slot_num":     None,
            "container_id": container,
            "carrier":      _scac(row[1]),
            "site_code":    str(fc_dest).strip().upper(),
            "product_type": prod_type,
            "qty":          qty,
            "notes":        " | ".join(note_parts) if note_parts else None,
            "status":       STATUS_MAP.get(str(row[11]).strip() if row[11] else "", "planned"),
        })
    return rows, errors

# ── SharePoint DBR fetch ─────────────────────────────────────────────────────

@st.cache_data(ttl=3600, show_spinner="Fetching DBR Tracker from SharePoint...")
def _fetch_dbr_from_sharepoint(tenant_id: str, client_id: str, client_secret: str) -> bytes | None:
    """Download the master DBR Tracker from SharePoint via MSAL client credentials.
    Cached for 1 hour to avoid hammering SharePoint on every rerun.
    Returns raw xlsx bytes, or None on failure.
    """
    try:
        import msal
        import requests as _rq
        app = msal.ConfidentialClientApplication(
            client_id,
            authority=f"https://login.microsoftonline.com/{tenant_id}",
            client_credential=client_secret,
        )
        result = app.acquire_token_for_client(
            scopes=["https://amazon.sharepoint.com/.default"]
        )
        token = result.get("access_token")
        if not token:
            return None
        # Download via SharePoint REST API using the server-relative path
        file_path = "/sites/AGLRobotics/Shared%20Documents/ToteASERs%20Robotics%20DBR%20Tracker.xlsx"
        url = (
            "https://amazon.sharepoint.com/sites/AGLRobotics"
            f"/_api/web/GetFileByServerRelativePath(decodedurl='{file_path}')/$value"
        )
        resp = _rq.get(
            url,
            headers={"Authorization": f"Bearer {token}", "Accept": "application/octet-stream"},
            timeout=30,
        )
        resp.raise_for_status()
        return resp.content
    except Exception as _sp_exc:
        return str(_sp_exc)  # return error string so caller can distinguish


def _sp_secrets() -> tuple[str | None, str | None, str | None]:
    """Return (tenant_id, client_id, client_secret) from Streamlit secrets, or Nones."""
    try:
        cfg = st.secrets["sharepoint"]
        return (
            cfg.get("TENANT_ID") or cfg.get("tenant_id"),
            cfg.get("CLIENT_ID")  or cfg.get("client_id"),
            cfg.get("CLIENT_SECRET") or cfg.get("client_secret"),
        )
    except Exception:
        return None, None, None


# ── request parser ────────────────────────────────────────────────────────────
import re as _re

def _parse_site_request(text: str) -> list[dict]:
    """
    Parse a stakeholder request message into a list of material line items.
    Returns: [{"material": str, "qty": int, "unit": str, "level_load": bool, "include_sat": bool, "note": str}]
    """
    results = []
    lines = text.strip().splitlines()
    mat_pattern = _re.compile(
        r"([A-Z][A-Z ,./&-]+?)[\s:]+(\d+)\s*(TLs?|containers?)?\s*(.*)$",
        _re.IGNORECASE
    )
    for line in lines:
        line = line.strip().lstrip("•-*").strip()
        m = mat_pattern.match(line)
        if not m:
            continue
        material = m.group(1).strip().upper().rstrip(",").rstrip()
        qty_str  = m.group(2)
        unit     = (m.group(3) or "containers").lower()
        rest     = (m.group(4) or "").lower()
        # normalise material names
        material = material.replace("  ", " ")
        if "shelf" in material.lower() and "hdtp" in material.lower():
            material = "SHELF, HDTP"
        elif "strap" in material.lower() and "hdtp" in material.lower():
            material = "STRAP, HDTP"
        elif "base" in material.lower():
            material = "BASE ASSEMBLY, HDTP"
        elif "column" in material.lower():
            material = "COLUMN"
        level_load  = "level load" in rest
        include_sat = "saturday" in rest or "sat" in rest
        results.append({
            "material":    material,
            "qty":         int(qty_str),
            "unit":        "TL" if "tl" in unit else "container",
            "level_load":  level_load,
            "include_sat": include_sat,
            "note":        rest.strip("() "),
        })
    return results

def _generate_level_load_schedule(
    containers: list[str],
    site_code: str,
    scac: str,
    product_type: str,
    week_start: date,
    include_sat: bool,
    priority_time: str,
) -> list[dict]:
    """
    Distribute container IDs evenly across Mon–Sat (or Mon–Fri).
    Returns list of plan entry dicts ready to insert.
    """
    n_days   = 6 if include_sat else 5
    work_days = [week_start + timedelta(days=i) for i in range(n_days)]
    slot_time = priority_time if priority_time else "8:30 AM"
    slot_num  = dict(_PLAN_SLOTS).get(slot_time, 2)

    entries = []
    for idx, cid in enumerate(containers):
        day = work_days[idx % n_days]
        entries.append({
            "week_start":   week_start,
            "appt_date":    day,
            "appt_time":    slot_time,
            "slot_num":     slot_num,
            "container_id": cid.upper().strip(),
            "scac":         scac,
            "site_code":    site_code,
            "product_type": product_type,
            "qty":          1,
            "notes":        "",
            "status":       "SCHEDULED",
        })
    return entries


# ── Excel export helpers ──────────────────────────────────────────────────────

def _xl_setup():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    thin     = Side(style="thin")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    al_ctr   = Alignment(horizontal="center", vertical="center")
    al_mid   = Alignment(vertical="center")
    return hdr_fill, hdr_font, border, al_ctr, al_mid

def _xl_carrier_fills():
    from openpyxl.styles import PatternFill
    return {
        "ATMI": PatternFill("solid", fgColor="FFF2CC"),
        "ARVY": PatternFill("solid", fgColor="DEEAF1"),
        "HDDR": PatternFill("solid", fgColor="E2EFDA"),
        "RKNE": PatternFill("solid", fgColor="F3E5F5"),
        "TGHE": PatternFill("solid", fgColor="FCE4EC"),
    }

def _xl_header(ws, cols):
    from openpyxl.utils import get_column_letter
    hdr_fill, hdr_font, border, al_ctr, _ = _xl_setup()
    ws.append(cols)
    for cell in ws[1]:
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.border = border; cell.alignment = al_ctr
    ws.row_dimensions[1].height = 18

def _xl_col_w(ws, widths):
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _to_bytes(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def _export_compiled_excel(plan_df: pd.DataFrame, week_days: list, week_start: date) -> bytes:
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    fills = _xl_carrier_fills()
    _, _, border, _, al_mid = _xl_setup()
    COLS = ["#","Day","Site","Appt Time","Container #","Carrier","Product Type","Qty","Notes","Status"]
    W    = [4,12,8,10,16,8,14,8,28,12]

    # summary sheet first
    ws0 = wb.create_sheet("All Sites Summary")
    _xl_header(ws0, COLS)
    active = plan_df[plan_df["status"] != "PENDING"].sort_values(["appt_date","site_code","slot_num"])
    for n, r in enumerate(active.itertuples(), 1):
        try: day_s = date.fromisoformat(str(r.appt_date)).strftime("%A")
        except: day_s = str(r.appt_date or "TBD")
        ws0.append([n, day_s, r.site_code or "", r.appt_time,
                    r.container_id, r.carrier, r.product_type or "",
                    r.qty or "", r.notes or "", r.status])
        fill = fills.get(r.carrier)
        for cell in ws0[ws0.max_row]:
            cell.border = border; cell.alignment = al_mid
            if fill: cell.fill = fill
    _xl_col_w(ws0, W)

    # day sheets
    day_abbr = ["Mon","Tue","Wed","Thu","Fri","Sat"]
    for dd, abbr in zip(week_days, day_abbr):
        lbl = f"{abbr} {dd.month}-{dd.day}"
        ws = wb.create_sheet(lbl)
        _xl_header(ws, COLS)
        day_df = plan_df[(plan_df["appt_date"]==dd.isoformat()) & (plan_df["status"]!="PENDING")].sort_values(["site_code","slot_num"])
        for n, r in enumerate(day_df.itertuples(), 1):
            ws.append([n, dd.strftime("%A"), r.site_code or "", r.appt_time,
                       r.container_id, r.carrier, r.product_type or "",
                       r.qty or "", r.notes or "", r.status])
            fill = fills.get(r.carrier)
            for cell in ws[ws.max_row]:
                cell.border = border; cell.alignment = al_mid
                if fill: cell.fill = fill
        _xl_col_w(ws, W)

    return _to_bytes(wb)

def _export_site_excel(plan_df: pd.DataFrame, site_code: str, week_start: date) -> bytes:
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    fills = _xl_carrier_fills()
    _, _, border, _, al_mid = _xl_setup()
    from openpyxl.styles import Font as XlFont
    COLS = ["#","Day","Appt Time","Container #","Carrier","Product Type","Qty","Notes","Status"]
    W    = [4,12,10,16,8,14,8,28,12]
    site_df = plan_df[plan_df["site_code"]==site_code]

    ws = wb.create_sheet(f"{site_code} — All Carriers")
    _xl_header(ws, COLS)
    for n, r in enumerate(site_df[site_df["status"]!="PENDING"].sort_values(["appt_date","slot_num"]).itertuples(), 1):
        try: day_s = date.fromisoformat(str(r.appt_date)).strftime("%A")
        except: day_s = "TBD"
        ws.append([n, day_s, r.appt_time, r.container_id, r.carrier,
                   r.product_type or "", r.qty or "", r.notes or "", r.status])
        fill = fills.get(r.carrier)
        for cell in ws[ws.max_row]:
            cell.border = border; cell.alignment = al_mid
            if fill: cell.fill = fill
    _xl_col_w(ws, W)

    for scac in sorted(site_df["carrier"].dropna().unique()):
        ws2 = wb.create_sheet(f"{site_code} — {scac}")
        _xl_header(ws2, COLS)
        cdf = site_df[site_df["carrier"]==scac]
        for n, r in enumerate(cdf[cdf["status"]!="PENDING"].sort_values(["appt_date","slot_num"]).itertuples(), 1):
            try: day_s = date.fromisoformat(str(r.appt_date)).strftime("%A")
            except: day_s = "TBD"
            ws2.append([n, day_s, r.appt_time, r.container_id, r.carrier,
                        r.product_type or "", r.qty or "", r.notes or "", r.status])
            fill = fills.get(r.carrier)
            for cell in ws2[ws2.max_row]:
                cell.border = border; cell.alignment = al_mid
                if fill: cell.fill = fill
        _xl_col_w(ws2, W)

    return _to_bytes(wb)

def _export_carrier_excel(plan_df: pd.DataFrame, scac: str, carrier_name: str, week_start: date) -> bytes:
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    fills = _xl_carrier_fills(); fill = fills.get(scac)
    _, _, border, _, al_mid = _xl_setup()
    from openpyxl.styles import Font as XlFont
    COLS = ["#","Day","Site","Appt Time","Container #","Product Type","Qty","Notes"]
    W    = [4,12,8,10,16,14,8,28]
    c_df = plan_df[plan_df["carrier"]==scac]

    ws = wb.create_sheet(f"{carrier_name} — All Sites")
    _xl_header(ws, COLS)
    for n, r in enumerate(c_df[c_df["status"]!="PENDING"].sort_values(["appt_date","site_code","slot_num"]).itertuples(), 1):
        try: day_s = date.fromisoformat(str(r.appt_date)).strftime("%A")
        except: day_s = "TBD"
        ws.append([n, day_s, r.site_code or "", r.appt_time,
                   r.container_id, r.product_type or "", r.qty or "", r.notes or ""])
        for cell in ws[ws.max_row]:
            cell.border = border; cell.alignment = al_mid
            if fill: cell.fill = fill
    pend = c_df[c_df["status"]=="PENDING"]
    if not pend.empty:
        ws.append([])
        ws.append(["PENDING — confirm availability:"] + [""]*7)
        for cell in ws[ws.max_row]:
            cell.font = XlFont(bold=True, italic=True, color="7F7F7F")
        for r in pend.itertuples():
            ws.append([None,"TBD",r.site_code or "","TBD",r.container_id,r.product_type or "","",r.notes or ""])
            for cell in ws[ws.max_row]: cell.border = border
    _xl_col_w(ws, W)

    for site in sorted(c_df["site_code"].dropna().unique()):
        ws2 = wb.create_sheet(site)
        _xl_header(ws2, COLS)
        for n, r in enumerate(c_df[(c_df["site_code"]==site)&(c_df["status"]!="PENDING")].sort_values(["appt_date","slot_num"]).itertuples(), 1):
            try: day_s = date.fromisoformat(str(r.appt_date)).strftime("%A")
            except: day_s = "TBD"
            ws2.append([n, day_s, site, r.appt_time, r.container_id,
                        r.product_type or "", r.qty or "", r.notes or ""])
            for cell in ws2[ws2.max_row]:
                cell.border = border; cell.alignment = al_mid
                if fill: cell.fill = fill
        _xl_col_w(ws2, W)

    return _to_bytes(wb)


# ── shared UI helpers ─────────────────────────────────────────────────────────

def _plan_row_table(df: pd.DataFrame, cols: list) -> pd.DataFrame:
    rows = []
    for r in df.itertuples():
        try:
            day_s = date.fromisoformat(str(r.appt_date)).strftime("%a %m/%d")
        except Exception:
            day_s = "TBD"
        all_cols = {
            "Day":         day_s if getattr(r,"status","") != "PENDING" else "TBD",
            "Time":        r.appt_time if getattr(r,"status","") != "PENDING" else "TBD",
            "Site":        getattr(r,"site_code","") or "",
            "Container #": r.container_id,
            "Carrier":     r.carrier,
            "Product":     r.product_type or "",
            "Qty":         r.qty or "",
            "Notes":       r.notes or "",
            "Status":      _STATUS_BADGE.get(r.status, r.status),
            "Confirmed":   "✅" if getattr(r, "carrier_confirmed", 0) else "—",
        }
        rows.append({c: all_cols[c] for c in cols if c in all_cols})
    return pd.DataFrame(rows) if rows else pd.DataFrame(columns=cols)

def _status_editor(df: pd.DataFrame, key_prefix: str):
    if df.empty:
        return
    _sites_e   = _get_sites_df()
    _carriers_e = _get_carriers_df()
    _site_opts  = _sites_e["site_code"].tolist() if not _sites_e.empty else []
    _scac_opts  = _carriers_e["scac"].tolist() if not _carriers_e.empty else list(_SCAC_COLOR.keys())

    ids    = df["id"].tolist()
    labels = [
        f"{r.container_id}  |  {getattr(r,'site_code','') or '?'}  |  {r.carrier}  |  "
        f"{r.appt_time}  ({_STATUS_BADGE.get(r.status, r.status)})"
        for r in df.itertuples()
    ]
    with st.expander("Edit Entry", expanded=False):
        sel    = st.selectbox("Select container", labels, key=f"{key_prefix}_sel")
        row_id = ids[labels.index(sel)]
        sel_row = df[df["id"] == row_id].iloc[0]

        ef1, ef2, ef3 = st.columns(3)
        with ef1:
            st.caption("**Schedule**")
            # date
            try:
                _cur_date = date.fromisoformat(str(sel_row["appt_date"])) if pd.notna(sel_row.get("appt_date")) else date.today()
            except Exception:
                _cur_date = date.today()
            new_date = st.date_input("Delivery date", value=_cur_date, key=f"{key_prefix}_date")
            # time
            _t_opts = [t for t, _ in _PLAN_SLOTS]
            _cur_time = str(sel_row.get("appt_time", "8:30 AM") or "8:30 AM")
            _t_idx = _t_opts.index(_cur_time) if _cur_time in _t_opts else 1
            new_time = st.selectbox("Appointment time", _t_opts, index=_t_idx, key=f"{key_prefix}_time")
            new_slot = dict(_PLAN_SLOTS)[new_time]

        with ef2:
            st.caption("**Assignment**")
            _cur_site = str(sel_row.get("site_code") or "")
            _s_idx = _site_opts.index(_cur_site) if _cur_site in _site_opts else 0
            new_site = st.selectbox("Site", _site_opts or ["RIC6"], index=_s_idx, key=f"{key_prefix}_site")
            _cur_scac = str(sel_row.get("carrier") or "")
            _c_idx = _scac_opts.index(_cur_scac) if _cur_scac in _scac_opts else 0
            new_scac = st.selectbox("Carrier", _scac_opts, index=_c_idx, key=f"{key_prefix}_carrier")
            _prod_opts = _PLAN_PRODUCTS
            _cur_prod = str(sel_row.get("product_type") or _prod_opts[0])
            _p_idx = _prod_opts.index(_cur_prod) if _cur_prod in _prod_opts else 0
            new_prod = st.selectbox("Product type", _prod_opts, index=_p_idx, key=f"{key_prefix}_prod")

        with ef3:
            st.caption("**Status & Notes**")
            _cur_stat = str(sel_row.get("status") or "SCHEDULED")
            _st_idx = _PLAN_STATUSES.index(_cur_stat) if _cur_stat in _PLAN_STATUSES else 0
            new_stat  = st.selectbox("Status", _PLAN_STATUSES, index=_st_idx, key=f"{key_prefix}_stat")
            _cur_qty  = int(sel_row.get("qty") or 1)
            new_qty   = st.number_input("Qty", value=_cur_qty, step=10, min_value=0, key=f"{key_prefix}_qty")
            new_notes = st.text_input("Notes", value=str(sel_row.get("notes") or ""), key=f"{key_prefix}_notes")
            del_flag  = st.checkbox("Delete this entry", key=f"{key_prefix}_del")

        if st.button("Save changes", key=f"{key_prefix}_apply", type="primary"):
            if del_flag:
                _delete_plan_entry(row_id)
                st.success("Entry deleted.")
            else:
                _update_plan_entry(row_id,
                    appt_date=new_date.isoformat(),
                    appt_time=new_time,
                    slot_num=new_slot,
                    site_code=new_site,
                    carrier=new_scac,
                    product_type=new_prod,
                    status=new_stat,
                    qty=int(new_qty),
                    notes=new_notes.strip(),
                    week_start=_plan_week_start(new_date).isoformat(),
                )
                st.success(f"Saved — {sel_row['container_id']} updated.")
            st.rerun()

def _week_grid(plan_df: pd.DataFrame, week_days: list, constraints: dict | None = None):
    """Week grid with dynamic daily cap indicators.

    Saturday uses constraints["capacity_sat"] (default 5);
    weekdays use constraints["capacity_day"] (default 10).
    Color: green < 70%, orange >= 70%, red = full.
    """
    _c = constraints or {}
    cap_wd  = _c.get("capacity_day", 10)
    cap_sat = _c.get("capacity_sat", 5)
    day_abbrs = ["Mon","Tue","Wed","Thu","Fri","Sat"]
    gcols = st.columns(6)
    for ci, (gc, gd) in enumerate(zip(gcols, week_days)):
        day_iso = gd.isoformat()
        ent = plan_df[(plan_df["appt_date"]==day_iso)&(plan_df["status"]!="PENDING")].sort_values("slot_num")
        cap = cap_sat if gd.weekday() == 5 else cap_wd
        with gc:
            dc = len(ent)
            st.markdown(f"**{day_abbrs[ci]} {gd.month}/{gd.day}**")
            cap_color = "red" if dc >= cap else ("#E8A020" if dc >= cap * 0.7 else "#34A853")
            cap_label = "FULL" if dc >= cap else f"{dc}/{cap}"
            st.markdown(f"<small style='color:{cap_color};font-weight:600'>{cap_label}</small>", unsafe_allow_html=True)
            if ent.empty:
                st.markdown("<small style='color:#555'>—</small>", unsafe_allow_html=True)
            else:
                for _, er in ent.iterrows():
                    clr  = _SCAC_COLOR.get(er["carrier"], "#888")
                    emj  = _STATUS_BADGE.get(er["status"], "")
                    site = er.get("site_code") or ""
                    st.markdown(
                        f"""<div style="background:{clr}22;border-left:3px solid {clr};
                            padding:4px 6px;margin:3px 0;border-radius:3px;font-size:11px;line-height:1.5">
                            <b>{er["appt_time"]}</b> {emj}<br/>
                            <code style="font-size:10px">{er["container_id"]}</code><br/>
                            <span style="color:{clr};font-weight:600">{er["carrier"]}</span>
                            {"<span style='color:#666'> · " + site + "</span>" if site else ""}
                        </div>""",
                        unsafe_allow_html=True
                    )


# ══════════════════════════════════════════════════════════════════════════════
# TAB 7 BODY
# ══════════════════════════════════════════════════════════════════════════════
# ═══════════════════════════════════════════════════════════════════════════════
# SRF — Short Range Forecast helpers
# ═══════════════════════════════════════════════════════════════════════════════
import math as _math

def _srf_required_capacity(volume: int) -> int:
    """LSP must maintain capacity for at least 20% more than forecasted volume."""
    return _math.ceil(volume * 1.20)

def _srf_save_submission(conn, submission_week: int, year: int, forecasts: dict, user: str = "kennewdo"):
    """
    Save a new SRF submission.
    forecasts = {forecast_week: volume, ...}  (6 entries)
    Returns the submission id.
    """
    now = datetime.now(_EASTERN).isoformat()
    conn.execute(
        "INSERT OR REPLACE INTO srf_submissions (submission_week, submission_year, submitted_at, submitted_by) "
        "VALUES (?,?,?,?)", (submission_week, year, now, user)
    )
    sub_id = conn.execute(
        "SELECT id FROM srf_submissions WHERE submission_week=? AND submission_year=?",
        (submission_week, year)
    ).fetchone()[0]
    for fw, vol in forecasts.items():
        if vol is None: continue
        cap = _srf_required_capacity(vol)
        conn.execute(
            "INSERT OR REPLACE INTO srf_forecast "
            "(submission_week, submission_year, forecast_week, forecast_year, "
            " forecasted_volume, required_capacity) VALUES (?,?,?,?,?,?)",
            (submission_week, year, fw, year, vol, cap)
        )
    conn.commit()
    return sub_id

def _srf_get_latest(conn, year: int) -> dict:
    """Return the most recent SRF submission rows as {forecast_week: row_dict}."""
    sub = conn.execute(
        "SELECT submission_week FROM srf_submissions WHERE submission_year=? "
        "ORDER BY submission_week DESC LIMIT 1", (year,)
    ).fetchone()
    if not sub:
        return {}
    sw = sub[0]
    rows = conn.execute(
        "SELECT forecast_week, forecasted_volume, required_capacity, "
        "notice_required, notice_sent, notice_sent_at, submission_week "
        "FROM srf_forecast WHERE submission_week=? AND submission_year=? "
        "ORDER BY forecast_week", (sw, year)
    ).fetchall()
    return {r[0]: {"vol": r[1], "cap": r[2], "notice_req": r[3],
                   "notice_sent": r[4], "notice_sent_at": r[5],
                   "submission_week": r[6]} for r in rows}

def _srf_get_prior(conn, year: int, before_week: int) -> dict:
    """Return the SRF from the submission immediately before before_week."""
    sub = conn.execute(
        "SELECT submission_week FROM srf_submissions "
        "WHERE submission_year=? AND submission_week < ? "
        "ORDER BY submission_week DESC LIMIT 1", (year, before_week)
    ).fetchone()
    if not sub:
        return {}
    sw = sub[0]
    rows = conn.execute(
        "SELECT forecast_week, forecasted_volume, required_capacity "
        "FROM srf_forecast WHERE submission_week=? AND submission_year=?",
        (sw, year)
    ).fetchall()
    return {r[0]: {"vol": r[1], "cap": r[2]} for r in rows}

def _srf_mark_notice_sent(conn, submission_week: int, year: int, forecast_week: int):
    now = datetime.now(_EASTERN).isoformat()
    conn.execute(
        "UPDATE srf_forecast SET notice_sent=1, notice_sent_at=? "
        "WHERE submission_week=? AND submission_year=? AND forecast_week=?",
        (now, submission_week, year, forecast_week)
    )
    conn.commit()

def _srf_compute_notice_flags(latest: dict, prior: dict, current_week: int) -> dict:
    """
    For each week in latest, determine whether a capacity increase notice is needed.
    Returns enhanced dict with notice fields populated.
    """
    result = {}
    for fw, row in latest.items():
        r = dict(row)
        prior_vol = prior.get(fw, {}).get("vol")
        weeks_out = fw - current_week
        vol_increase = (prior_vol is not None and row["vol"] > prior_vol)
        if vol_increase:
            delta = row["vol"] - prior_vol
            delta_pct = round(delta / prior_vol * 100) if prior_vol > 0 else None
            r["prior_vol"] = prior_vol
            r["delta"] = delta
            r["delta_pct"] = delta_pct
            if weeks_out <= 1:
                r["notice_status"] = "URGENT"   # already within / past window
            elif weeks_out == 2:
                r["notice_status"] = "DUE_NOW"  # exactly on the 2-week line
            else:
                r["notice_status"] = "SCHEDULED"  # time remains, due by W{fw-2}
                r["notice_due_week"] = fw - 2
        else:
            r["prior_vol"] = prior_vol
            r["delta"] = 0
            r["delta_pct"] = 0
            r["notice_status"] = "OK" if not row["notice_sent"] else "SENT"
        result[fw] = r
    return result


def _srf_auto_forecast(conn, current_week: int, year: int, n_weeks: int = 6, history: int = 8) -> dict:
    """
    Auto-generate 6-week SRF volume forecast from historical data.

    Data priority:
      1. wbr_results.containers  -- weekly totals from WBR Bridge (most reliable)
      2. delivery_plan COUNT by week_start -- fallback when wbr_results is sparse

    Method: weighted moving average (recent weeks get higher weight) + linear trend
    detection. Trend applied only if slope > 3%/week; otherwise flat WMA baseline.

    Returns dict:
      forecast       {forecast_week: projected_volume}
      method         wbr_results | delivery_plan | default
      wma_baseline   rounded WMA value
      trend_pct      % weekly slope (informational)
      history        [(week_num, volume), ...] chronological
      vols           [volume, ...] chronological
    """
    # 1. Pull historical weekly volume
    rows = conn.execute(
        "SELECT week_num, containers FROM wbr_results "
        "WHERE year=? AND week_num < ? AND containers IS NOT NULL AND containers > 0 "
        "ORDER BY week_num DESC LIMIT ?",
        (year, current_week, history)
    ).fetchall()
    method = "wbr_results"

    # Fallback: aggregate delivery_plan by ISO week
    if len(rows) < 3:
        dp_rows = conn.execute(
            "SELECT week_start, COUNT(*) as cnt "
            "FROM delivery_plan "
            "WHERE week_start IS NOT NULL AND week_start != '' "
            "GROUP BY week_start ORDER BY week_start DESC LIMIT ?",
            (history,)
        ).fetchall()
        rows = []
        for _rdp in dp_rows:
            if _rdp[0]:
                try:
                    _iso = date.fromisoformat(str(_rdp[0])).isocalendar()[1]
                    rows.append((_iso, int(_rdp[1])))
                except Exception:
                    pass
        method = "delivery_plan"

    # Default -- not enough history anywhere
    if len(rows) < 2:
        baseline = 65
        fcast = {(current_week + i) % 53 or 53: baseline for i in range(1, n_weeks + 1)}
        return {"forecast": fcast, "method": "default",
                "wma_baseline": baseline, "trend_pct": 0.0,
                "history": [], "vols": []}

    # 2. Chronological order (oldest first)
    hist = list(reversed(rows))
    vols = [float(r[1]) for r in hist]
    n = len(vols)

    # 3. Weighted moving average: weight = rank (newest = highest)
    weights  = list(range(1, n + 1))
    wma      = sum(w * v for w, v in zip(weights, vols)) / sum(weights)

    # 4. Linear trend via least-squares slope
    x_vals    = list(range(n))
    x_mean    = sum(x_vals) / n
    v_mean    = sum(vols) / n
    slope_num = sum((xi - x_mean) * (vi - v_mean) for xi, vi in zip(x_vals, vols))
    slope_den = sum((xi - x_mean) ** 2 for xi in x_vals)
    slope     = slope_num / slope_den if slope_den > 0 else 0.0
    trend_pct = round(slope / v_mean * 100, 1) if v_mean > 0 else 0.0

    # 5. Project 6 weeks: apply trend only if > 3%/wk; else flat WMA
    apply_trend = abs(trend_pct) > 3.0
    last_vol    = vols[-1]
    forecast = {}
    for i in range(1, n_weeks + 1):
        fw = (current_week + i) % 53 or 53
        projected = (last_vol + slope * i) if apply_trend else wma
        forecast[fw] = max(1, round(projected))

    return {
        "forecast":     forecast,
        "method":       method,
        "wma_baseline": round(wma),
        "trend_pct":    trend_pct,
        "history":      hist,
        "vols":         vols,
    }


with tab7:
    st.subheader("Delivery Plan Scheduler")
    st.caption("**Purpose:** Build, manage, and distribute the weekly container delivery plan across all sites and carriers.")

    # ══════════════════════════════════════════════════════════════════════════
    # ── SOP header callout ────────────────────────────────────────────────────
    with st.expander("📋 Weekly Planning SOP — Quick Reference", expanded=False):
        _sop_c1, _sop_c2 = st.columns(2)
        with _sop_c1:
            st.markdown("""
**Owner:** Dominique Kennedy (kennewdo)  
**Frequency:** Weekly — Every **Friday**  
**Plan Deadline:** 3:00 PM ET · 2:00 PM CT · 12:00 PM PT  
**Carrier DBR Deadline:** 3:00 PM CT Thursday (primary input)  
**Sites:** RIC6, ILM1, DBM6/SAV, SJC8, XPH1-IAG1, LAX/West Coast  
**SharePoint:** AGLRobotics → Carrier DBRs (subfolders: ATMI, ARVY, HUDD)
""")
        with _sop_c2:
            st.markdown("""
**Carriers & Ops Contacts:**

| SCAC | Carrier | Contact | Email |
|---|---|---|---|
| ATMI | Cargomatic | Tyler Domingues | tdomingues@cargomatic.com |
| ARVY | Arrive Logistics | Tyler Spangler | tspangler@arrivelogistics.com |
| HDDR – RIC6 | Maersk | Sandji Ruffin | sandji.ruffin@maersk.com |
| HDDR – ILM1 | Maersk | Jerry Nesbit | jerry.nesbit@maersk.com |
| HDDR – LAX | Maersk | Desirae Swain / Ailua Osoimalo | desirae.swain@maersk.com |
""")
        st.markdown("**DBR Email Subjects to expect Thursday EOD:**")
        st.code("""ATMI : [EXTERNAL] Amazon Robotics - DBR [DATE]  (Tyler Domingues)
ARVY : [EXTERNAL] Robotics DBR [DATE]  (Tyler Spangler)
HUDD RIC6 : [EXTERNAL] RIC6 Delivery Plan Update (HUDD)  (Sandji Ruffin)
HUDD ILM1 : [EXTERNAL] ILM1 Delivery Plan Update / HUDD  (Jerry Nesbit)
HUDD LAX  : RE: DBR Bridges Report - HUDD - [DATE]  (Desirae/Ailua)""", language="")
        st.markdown("**File naming:** `[CARRIER] DBR M.DD.YY.xlsx`  e.g. `ATMI DBR 7.17.26.xlsx`")

    # ── global week selector ──────────────────────────────────────────────────
    _wc1, _wc2, _wc3, _wc4 = st.columns([3,1,1,1])
    with _wc1:
        _today = date.today()
        _wk_in = st.date_input("Week (any day)", value=_plan_week_start(_today), key="plan_wk_v2")
        _ws    = _plan_week_start(_wk_in)
        _wdays = [_ws + timedelta(days=i) for i in range(6)]
        st.caption(f"**{_ws.strftime('%b %d')} – {(_ws+timedelta(days=5)).strftime('%b %d, %Y')}**")

    _pdf = _get_plan(_ws.isoformat())
    _pdf_act = _pdf[_pdf["status"] != "PENDING"]

    with _wc2: st.metric("Total loads", len(_pdf_act))
    with _wc3: st.metric("Sites", _pdf_act["site_code"].nunique() if not _pdf_act.empty else 0)
    with _wc4: st.metric("Pending", len(_pdf[_pdf["status"]=="PENDING"]))

    # ── week summary banner ──────────────────────────────────────────────────
    _iso_wk_cur = _ws.isocalendar()[1]
    if not _pdf_act.empty:
        _bsc = _pdf_act.groupby("site_code")["id"].count()
        _bsc_str = "  ·  ".join(f"{s}: **{c}**" for s, c in sorted(_bsc.items()))
        st.info(
            f"**W{_iso_wk_cur} ({_ws.strftime('%b %d')})** — "
            f"{len(_pdf_act)} loads  ·  {_bsc_str}",
            icon="📋"
        )
    else:
        st.info(
            f"**W{_iso_wk_cur} ({_ws.strftime('%b %d')})** — No loads scheduled yet.",
            icon="📋"
        )



    # ── receiving day selector (persisted per week) ───────────────────────────
    _saved_days   = _get_week_config(_ws.isoformat())
    _day_abbr_map = {"Sun": 6, "Mon": 0, "Tue": 1, "Wed": 2, "Thu": 3, "Fri": 4, "Sat": 5}

    st.markdown("**Receiving days this week:**")
    _day_cols_row = st.columns(9)
    _new_active = []
    for _di, _dname in enumerate(_ALL_DAYS):
        with _day_cols_row[_di]:
            _checked = st.checkbox(_dname, value=(_dname in _saved_days), key=f"daytoggle_{_dname}_{_ws.isoformat()}")
            if _checked:
                _new_active.append(_dname)

    # detect changes + save
    if set(_new_active) != set(_saved_days):
        # check for containers on newly-deactivated days
        _deactivated = [d for d in _saved_days if d not in _new_active]
        _conflicts = []
        if _deactivated and not _pdf.empty:
            for _dday in _deactivated:
                _didx   = _day_abbr_map.get(_dday)
                _day_dates = [_ws + timedelta(days=(_didx - _ws.weekday()) % 7)]
                for _dd2 in _day_dates:
                    _hits = _pdf[(_pdf["appt_date"]==_dd2.isoformat())&(_pdf["status"].isin(["SCHEDULED","HOLD"]))]
                    if not _hits.empty:
                        _conflicts.append((_dday, _dd2.strftime("%b %d"), len(_hits)))
        _save_week_config(_ws.isoformat(), _new_active)
        if _conflicts:
            for _cday, _cdate, _ccnt in _conflicts:
                st.warning(
                    f"**{_ccnt} container(s) still assigned to {_cday} {_cdate}** — "
                    f"go to Plan Builder → Edit to reassign, or use **Redistribute** below."
                )
        else:
            st.rerun()

    # redistribute prompt (shown when deactivated-day containers exist)
    _active_day_dates = []
    for _adn in _new_active:
        _adidx = _day_abbr_map.get(_adn, 0)
        _addate = _ws + timedelta(days=(_adidx - _ws.weekday()) % 7)
        if _addate >= _ws and _addate <= _ws + timedelta(days=6):
            _active_day_dates.append(_addate)

    _orphan_df = pd.DataFrame()
    if not _pdf.empty and _active_day_dates:
        _active_isos = {d.isoformat() for d in _active_day_dates}
        _orphan_df = _pdf[
            (~_pdf["appt_date"].isin(_active_isos)) &
            (_pdf["appt_date"].notna()) &
            (_pdf["status"].isin(["SCHEDULED","HOLD"]))
        ]

    if not _orphan_df.empty:
        st.error(f"**{len(_orphan_df)} container(s) assigned to inactive days.** Redistribute or manually reassign.")
        with st.expander("Redistribute to active days", expanded=False):
            st.dataframe(
                _plan_row_table(_orphan_df, ["Day","Site","Container #","Carrier","Product","Status"]),
                use_container_width=True, hide_index=True
            )
            if st.button("Redistribute evenly across active days", key="redistribute_btn", type="primary"):
                if _active_day_dates:
                    sorted_active = sorted(_active_day_dates)
                    for idx, row in enumerate(_orphan_df.itertuples()):
                        new_day = sorted_active[idx % len(sorted_active)]
                        _scm3 = _get_scm_df()
                        scm_r = _scm3[(_scm3["site_code"]==getattr(row,"site_code",""))&(_scm3["scac"]==row.carrier)]
                        def_t = scm_r["priority_time"].iloc[0] if not scm_r.empty and pd.notna(scm_r["priority_time"].iloc[0]) else "8:30 AM"
                        sn    = dict(_PLAN_SLOTS).get(def_t, 2)
                        _update_plan_entry(row.id,
                            appt_date=new_day.isoformat(),
                            appt_time=def_t,
                            slot_num=sn,
                            week_start=_ws.isoformat()
                        )
                    st.success(f"Redistributed {len(_orphan_df)} container(s) across {len(sorted_active)} active day(s).")
                    st.rerun()
                else:
                    st.error("No active days selected — select at least one day first.")

    st.divider()

    _tp1, _tp2, _tp3, _tp4, _tp5, _tp6, _tp7, _tp8, _tp9 = st.tabs([
        "Plan Builder", "All Sites", "By Site",
        "Carrier View", "WoW / History", "SRF", "Config", "Import History", "SOP Guide",
    ])

    # ══════════════════════════════════════════════════════════════════════════
    # PLAN BUILDER
    # ══════════════════════════════════════════════════════════════════════════
    with _tp1:
        _sites_df   = _get_sites_df()
        _cdf        = _get_carriers_df()
        _scm        = _get_scm_df()
        _act_sites  = _sites_df[_sites_df["active"]==1]["site_code"].tolist() if not _sites_df.empty else []
        _act_scacs  = _cdf[_cdf["active"]==1]["scac"].tolist() if not _cdf.empty else list(_SCAC_COLOR.keys())

        # site selector + constraint bar
        _pb_site = st.selectbox(
            "Site", _act_sites or ["RIC6"], key="pb_site_select",
            help="All steps below apply to this site's constraints."
        )
        _site_cfg = _get_site_constraints(_pb_site)
        _cap_d  = _site_cfg.get("capacity_day", 10)
        _cap_s  = _site_cfg.get("capacity_sat", 5)
        _cap_w  = _site_cfg.get("capacity_week", 0)
        _lat    = _site_cfg.get("last_appt_time", "4:00 PM")
        _las    = _site_cfg.get("last_appt_sat", "11:30 AM")
        _cap_w_str = f"{_cap_w}/wk  \u00b7  " if _cap_w else ""
        st.caption(
            f"**{_pb_site} constraints** \u2014 "
            f"{_cap_d}/day  \u00b7  {_cap_s}/Sat  \u00b7  {_cap_w_str}"
            f"Last appt {_lat} (weekday) / {_las} (Sat)  \u00b7  Saturday morning-only"
        )

        # ── SRF gap indicator ────────────────────────────────────────────────
        _srf_conn_pb = get_db()
        _srf_wk_pb   = _ws.isocalendar()[1]
        _srf_yr_pb   = _ws.isocalendar()[0]
        _srf_target_pb = None
        try:
            _srf_row_pb = _srf_conn_pb.execute(
                "SELECT forecasts_json FROM srf_submissions "
                "WHERE year=? ORDER BY submission_week DESC LIMIT 1",
                (_srf_yr_pb,)
            ).fetchone()
            if _srf_row_pb:
                import json as _json_pb
                _srf_fc_pb = _json_pb.loads(_srf_row_pb[0])
                _srf_target_pb = (
                    _srf_fc_pb.get(str(_srf_wk_pb))
                    or _srf_fc_pb.get(_srf_wk_pb)
                )
        except Exception:
            pass
        finally:
            _srf_conn_pb.close()
        if _srf_target_pb:
            _planned_pb = len(_pdf_act)
            _gap_pb     = _srf_target_pb - _planned_pb
            _gap_icon   = "✅" if _gap_pb <= 0 else "⚠️"
            _gap_txt    = "On target" if _gap_pb <= 0 else f"{_gap_pb} below target"
            st.caption(
                f"{_gap_icon} **SRF W{_srf_wk_pb}:** {_srf_target_pb} forecasted  ·  "
                f"{_planned_pb} planned  ·  {_gap_txt}"
            )

        # Step 1: Parse stakeholder request
        with st.expander("Step 1 \u2014 Paste Stakeholder Request", expanded=True):
            st.caption("Paste the message from Miguel or site ops. The tool will extract material needs and generate a plan structure.")
            _req_text = st.text_area(
                "Stakeholder request",
                height=160, key="req_text",
                placeholder='SHELF, HDTP: 52 containers (level loaded including saturday)',
            )
            if st.button("Parse Request", key="req_parse", type="primary") and _req_text.strip():
                parsed = _parse_site_request(_req_text)
                if parsed:
                    st.session_state["parsed_items"] = parsed
                    st.session_state["parsed_site"]  = _pb_site
                    st.success(f"Parsed {len(parsed)} material line(s).")
                else:
                    st.warning("No recognizable material lines found.")
            if st.session_state.get("parsed_items"):
                items = st.session_state["parsed_items"]
                parsed_site = st.session_state.get("parsed_site", _pb_site)
                st.markdown("**Parsed demand:**")
                for i, item in enumerate(items):
                    col_a, col_b, col_c, col_d = st.columns([3,1,2,2])
                    with col_a: st.markdown(f"**{item['material']}**")
                    with col_b: st.markdown(f"{item['qty']} {item['unit']}(s)")
                    with col_c:
                        ll = "Level load" if item["level_load"] else "Manual"
                        sat = " + Sat" if item["include_sat"] else ""
                        st.caption(f"{ll}{sat}")
                    with col_d:
                        _in_plan = len(_pdf[
                            (_pdf["site_code"]==parsed_site) &
                            (_pdf["product_type"].str.upper()==item["material"].upper())
                        ]) if not _pdf.empty else 0
                        st.caption(f"In plan: {_in_plan} / {item['qty']}")

        # Step 2: Container Pool
        with st.expander("Step 2 \u2014 Container Pool", expanded=False):
            _pool_tab_gvt, _pool_tab_co, _pool_tab_ref = st.tabs(
                ["\U0001f7e2 At Yard \u2014 GVT", "\U0001f504 Carried Over", "\U0001f6ab Refused / Rejected"]
            )

            with _pool_tab_gvt:
                from wbr_engine import load_gvt as _load_gvt_plan
                st.caption("Upload your GVT export to see which containers are at yard vs. contingent, sorted FIFO by Ready Date.")
                _gvt_plan_file = st.file_uploader("GVT Export (.xlsx)", type=["xlsx"], key="gvt_plan_upload")
                if _gvt_plan_file:
                    try:
                        _gvt_raw = _load_gvt_plan(_gvt_plan_file.read())
                        if _gvt_raw.empty:
                            st.warning("No containers found in GVT file.")
                        else:
                            _AT_YARD = {"in yard full", "available", "at yard", "in yard"}
                            _gvt_raw["_plan_status"] = _gvt_raw["status"].apply(
                                lambda s: "SCHEDULED" if str(s or "").strip().lower() in _AT_YARD else "PENDING"
                            )
                            _known_sites_gvt = _act_sites or ["RIC6","ILM1","IAG1","DBM6","ATL2","DRO1","LAX","OAK","SJC8","RSW9"]
                            def _map_fac(f):
                                fu = str(f or "").upper()
                                for sc in _known_sites_gvt:
                                    if sc in fu: return sc
                                return ""
                            _gvt_raw["_site_mapped"] = _gvt_raw["facility"].apply(_map_fac)
                            _gvt_sf_opts = ["All sites"] + _known_sites_gvt
                            _gvt_sf = st.selectbox(
                                "Filter by site", _gvt_sf_opts,
                                index=(_gvt_sf_opts.index(_pb_site) if _pb_site in _gvt_sf_opts else 0),
                                key="gvt_site_filter_v3",
                            )
                            _gvt_view = _gvt_raw.copy()
                            if _gvt_sf != "All sites":
                                _gvt_view = _gvt_view[_gvt_view["_site_mapped"] == _gvt_sf]
                            _gvt_view = _gvt_view.sort_values("ready_date", na_position="last").reset_index(drop=True)
                            _gvt_parsed_items = st.session_state.get("parsed_items", [])
                            _gvt_def_prod = _gvt_parsed_items[0]["material"] if _gvt_parsed_items else _PLAN_PRODUCTS[0]
                            _n_at_yard_gvt   = int((_gvt_view["_plan_status"] == "SCHEDULED").sum())
                            _n_not_ready_gvt = int((_gvt_view["_plan_status"] == "PENDING").sum())
                            st.caption(
                                f"**{len(_gvt_view)} containers** \u2014 "
                                f"\U0001f7e2 {_n_at_yard_gvt} at yard  "
                                f"\U0001f534 {_n_not_ready_gvt} not ready  \u00b7 FIFO by Ready Date"
                            )
                            _gvt_editor_df = pd.DataFrame({
                                "Add":          _gvt_view["_plan_status"].apply(lambda s: s == "SCHEDULED"),
                                "Container":    _gvt_view["container"].astype(str),
                                "GVT Status":   _gvt_view["status"].fillna("").astype(str),
                                "Ready Date":   _gvt_view["ready_date"].apply(
                                                    lambda d: d.strftime("%m/%d/%y") if pd.notna(d) else ""),
                                "Site":         _gvt_view["_site_mapped"].apply(lambda s: s if s else "?"),
                                "Carrier":      _gvt_view["scac"].fillna("").astype(str),
                                "Product Type": [_gvt_def_prod] * len(_gvt_view),
                                "Plan Status":  _gvt_view["_plan_status"],
                            })
                            _gvt_scac_opts      = (_act_scacs or list(_SCAC_COLOR.keys())) + [""]
                            _gvt_site_edit_opts = _known_sites_gvt + ["?"]
                            _gvt_edited = st.data_editor(
                                _gvt_editor_df,
                                column_config={
                                    "Add":          st.column_config.CheckboxColumn("Add", default=False),
                                    "Container":    st.column_config.TextColumn("Container", disabled=True),
                                    "GVT Status":   st.column_config.TextColumn("GVT Status", disabled=True),
                                    "Ready Date":   st.column_config.TextColumn("Ready Date \u2191", disabled=True),
                                    "Site":         st.column_config.SelectboxColumn("Site", options=_gvt_site_edit_opts),
                                    "Carrier":      st.column_config.SelectboxColumn("Carrier", options=_gvt_scac_opts),
                                    "Product Type": st.column_config.SelectboxColumn("Product Type", options=_PLAN_PRODUCTS),
                                    "Plan Status":  st.column_config.TextColumn("Plan Status", disabled=True),
                                },
                                use_container_width=True, hide_index=True,
                                key="gvt_editor_v3",
                                height=min(420, 38 + len(_gvt_view) * 35),
                            )
                            _gvt_sel = _gvt_edited[_gvt_edited["Add"] == True]
                            st.caption(f"**{len(_gvt_sel)} selected** \u2014 will be included in Step 3 auto-schedule.")
                            if st.button("Stage for Auto-Schedule", key="gvt_stage_btn", type="secondary"):
                                _staged = []
                                for _, _gs in _gvt_sel.iterrows():
                                    _gcid = str(_gs["Container"]).strip()
                                    _gsite = str(_gs["Site"]).strip()
                                    if not _gcid or _gsite == "?": continue
                                    _staged.append({
                                        "container": _gcid, "scac": str(_gs["Carrier"]).strip(),
                                        "site_code": _gsite, "product_type": str(_gs["Product Type"]),
                                        "plan_status": str(_gs["Plan Status"]), "source": "gvt", "priority": 3,
                                    })
                                st.session_state["pb_staged_gvt"] = _staged
                                st.success(f"Staged {len(_staged)} container(s) for auto-schedule.")
                    except Exception as _gvt_err:
                        st.error(f"Error reading GVT file: {_gvt_err}")

            with _pool_tab_co:
                st.caption("Containers from prior weeks that were never delivered. Automatically pulled from the plan DB.")
                _co_pool = _get_reschedulable_pool(_pb_site, _ws.isoformat())
                _co_carried = _co_pool[_co_pool["pool_source"] == "carried_over"] if not _co_pool.empty else pd.DataFrame()
                if _co_carried.empty:
                    st.info("No carried-over containers for this site.")
                else:
                    _co_editor_df = pd.DataFrame({
                        "Include":      [True] * len(_co_carried),
                        "Container":    _co_carried["container_id"].astype(str),
                        "Carrier":      _co_carried["scac"].fillna("").astype(str),
                        "Product Type": _co_carried["product_type"].fillna("").astype(str),
                        "Prior Week":   _co_carried["week_start"].astype(str),
                        "Prior Status": _co_carried["status"].astype(str),
                    })
                    _co_edited = st.data_editor(
                        _co_editor_df,
                        column_config={
                            "Include":      st.column_config.CheckboxColumn("Include", default=True),
                            "Container":    st.column_config.TextColumn("Container", disabled=True),
                            "Carrier":      st.column_config.TextColumn("Carrier", disabled=True),
                            "Product Type": st.column_config.TextColumn("Product Type", disabled=True),
                            "Prior Week":   st.column_config.TextColumn("Prior Week", disabled=True),
                            "Prior Status": st.column_config.TextColumn("Prior Status", disabled=True),
                        },
                        use_container_width=True, hide_index=True, key="co_editor",
                        height=min(360, 38 + len(_co_carried) * 35),
                    )
                    _co_sel = _co_edited[_co_edited["Include"] == True]
                    st.caption(f"**{len(_co_sel)} carried-over** container(s) will be included in auto-schedule (priority 2).")
                    if st.button("Stage Carried-Over", key="co_stage_btn", type="secondary"):
                        _co_staged = []
                        for _, _cr in _co_sel.iterrows():
                            _co_staged.append({
                                "container": str(_cr["Container"]).strip(),
                                "scac": str(_cr["Carrier"]).strip(),
                                "site_code": _pb_site,
                                "product_type": str(_cr["Product Type"]).strip(),
                                "plan_status": "SCHEDULED", "source": "carried_over", "priority": 2,
                            })
                        st.session_state["pb_staged_co"] = _co_staged
                        st.success(f"Staged {len(_co_staged)} carried-over container(s).")

            with _pool_tab_ref:
                st.caption("Containers refused at the gate or rejected. Highest rescheduling priority (longest dwell risk).")
                _ref_pool = _co_pool[_co_pool["pool_source"] == "refused"] if not _co_pool.empty else pd.DataFrame()
                if _ref_pool.empty:
                    st.info("No refused/rejected containers for this site.")
                else:
                    _ref_editor_df = pd.DataFrame({
                        "Include":      [True] * len(_ref_pool),
                        "Container":    _ref_pool["container_id"].astype(str),
                        "Carrier":      _ref_pool["scac"].fillna("").astype(str),
                        "Product Type": _ref_pool["product_type"].fillna("").astype(str),
                        "Prior Week":   _ref_pool["week_start"].astype(str),
                        "Prior Status": _ref_pool["status"].astype(str),
                    })
                    _ref_edited = st.data_editor(
                        _ref_editor_df,
                        column_config={
                            "Include":      st.column_config.CheckboxColumn("Include", default=True),
                            "Container":    st.column_config.TextColumn("Container", disabled=True),
                            "Carrier":      st.column_config.TextColumn("Carrier", disabled=True),
                            "Product Type": st.column_config.TextColumn("Product Type", disabled=True),
                            "Prior Week":   st.column_config.TextColumn("Prior Week", disabled=True),
                            "Prior Status": st.column_config.TextColumn("Prior Status", disabled=True),
                        },
                        use_container_width=True, hide_index=True, key="ref_editor",
                        height=min(360, 38 + len(_ref_pool) * 35),
                    )
                    _ref_sel = _ref_edited[_ref_edited["Include"] == True]
                    st.caption(f"**{len(_ref_sel)} refused/rejected** container(s) will be included in auto-schedule (priority 1).")
                    if st.button("Stage Refused/Rejected", key="ref_stage_btn", type="secondary"):
                        _ref_staged = []
                        for _, _rr in _ref_sel.iterrows():
                            _ref_staged.append({
                                "container": str(_rr["Container"]).strip(),
                                "scac": str(_rr["Carrier"]).strip(),
                                "site_code": _pb_site,
                                "product_type": str(_rr["Product Type"]).strip(),
                                "plan_status": "SCHEDULED", "source": "refused", "priority": 1,
                            })
                        st.session_state["pb_staged_ref"] = _ref_staged
                        st.success(f"Staged {len(_ref_staged)} refused/rejected container(s).")

        # Step 3: Auto-Schedule
        with st.expander("Step 3 \u2014 Auto-Schedule", expanded=False):
            st.caption(
                "Combines staged containers (refused first \u2192 carried-over \u2192 GVT at-yard), "
                "applies site constraints, and fills the week. HUDD/HDDR pinned to 7:30 AM. "
                "Overflow held below for push-to-next-week."
            )
            _all_staged = (
                st.session_state.get("pb_staged_ref", []) +
                st.session_state.get("pb_staged_co",  []) +
                st.session_state.get("pb_staged_gvt", [])
            )
            st.metric("Staged containers", len(_all_staged),
                      help="Refused/rejected + carried-over + GVT at-yard. Stage items in Step 2 first.")
            if _all_staged:
                _as_c1, _as_c2, _as_c3 = st.columns(3)
                with _as_c1: _as_incsat = st.checkbox("Include Saturday", value=True, key="as_incsat")
                with _as_c2: _as_respect_cap = st.checkbox("Respect daily/weekly caps", value=True, key="as_respect_cap")
                with _as_c3: _as_dry = st.checkbox("Preview only (don't write)", value=False, key="as_dry")
                if st.button("Run Auto-Schedule", type="primary", key="as_run"):
                    _as_ll_cfg  = _get_week_config(_ws.isoformat())
                    _as_day_map = {"Sun":6,"Mon":0,"Tue":1,"Wed":2,"Thu":3,"Fri":4,"Sat":5}
                    _as_wdays   = sorted([
                        _ws + timedelta(days=(_as_day_map[d]-_ws.weekday())%7)
                        for d in _as_ll_cfg if d in _as_day_map
                    ])
                    if not _as_wdays:
                        _as_wdays = [_ws + timedelta(days=i) for i in range(5)]
                    if not _as_incsat:
                        _as_wdays = [d for d in _as_wdays if d.weekday() != 5]
                    _scheduled_rows, _overflow = _auto_schedule_pool(_all_staged, _ws, _as_wdays, _pb_site)
                    if not _as_dry:
                        _as_added = 0
                        for _row in _scheduled_rows:
                            _add_plan_entry(
                                _ws, _row["date"], _row["appt_time"], _row["slot_num"],
                                _row["container"], _row["scac"], _row["site_code"],
                                _row["product_type"], 1,
                                f"Auto-scheduled ({_row['source']})", "SCHEDULED"
                            )
                            _as_added += 1
                        if _overflow:
                            st.session_state["pb_overflow"] = _overflow
                        for _k in ("pb_staged_ref","pb_staged_co","pb_staged_gvt"):
                            st.session_state.pop(_k, None)
                        st.success(
                            f"Scheduled {_as_added} container(s). "
                            f"{len(_overflow)} overflow \u2014 see queue below."
                        )
                        st.rerun()
                    else:
                        st.markdown(f"**Preview: {len(_scheduled_rows)} scheduled, {len(_overflow)} overflow**")
                        if _scheduled_rows:
                            _prev_df = pd.DataFrame(_scheduled_rows)[["date","appt_time","container","scac","product_type","source"]]
                            _prev_df.columns = ["Date","Time","Container","Carrier","Product","Source"]
                            st.dataframe(_prev_df, use_container_width=True, hide_index=True)
                        if _overflow:
                            st.warning(f"{len(_overflow)} container(s) would overflow:")
                            _ov_df2 = pd.DataFrame(_overflow)[["container","scac","product_type","source"]]
                            _ov_df2.columns = ["Container","Carrier","Product","Source"]
                            st.dataframe(_ov_df2, use_container_width=True, hide_index=True)

        # Step 4: Manual Add / Override
        with st.expander("Step 4 \u2014 Manual Add / Override", expanded=False):
            _add_tab_s, _add_tab_b = st.tabs(["Single", "Bulk Paste"])
            with _add_tab_s:
                _parsed_items = st.session_state.get("parsed_items", [])
                _mat_opts     = _PLAN_PRODUCTS
                _default_mat  = _parsed_items[0]["material"] if _parsed_items and _parsed_items[0]["material"] in _mat_opts else _mat_opts[0]
                sa1, sa2, sa3, sa4 = st.columns(4)
                with sa1:
                    _s_site   = st.selectbox("Site", _act_sites or ["RIC6"], index=(_act_sites.index(_pb_site) if _pb_site in _act_sites else 0), key="s_site")
                    _mapped   = _scm[_scm["site_code"]==_s_site]["scac"].tolist() if not _scm.empty else _act_scacs
                    _s_scac   = st.selectbox("Carrier", _mapped or _act_scacs, key="s_scac")
                    _s_status = st.selectbox("Status", ["SCHEDULED","PENDING","HOLD"], key="s_status")
                with sa2:
                    _s_cid     = st.text_input("Container #", placeholder="TCNU3773041", key="s_cid")
                    _s_product = st.selectbox("Product Type", _mat_opts, index=_mat_opts.index(_default_mat) if _default_mat in _mat_opts else 0, key="s_product")
                    _s_qty     = st.number_input("Qty (units)", value=1, step=1, min_value=0, key="s_qty")
                with sa3:
                    _s_date = st.date_input("Delivery Date", value=_today, key="s_date")
                    _scm_r  = _scm[(_scm["site_code"]==_s_site)&(_scm["scac"]==_s_scac)]
                    _def_t  = _scm_r["priority_time"].iloc[0] if not _scm_r.empty and pd.notna(_scm_r["priority_time"].iloc[0]) else "8:30 AM"
                    _t_opts = [t for t,_ in _PLAN_SLOTS]
                    _s_time = st.selectbox("Appt Time", _t_opts, index=_t_opts.index(_def_t) if _def_t in _t_opts else 1, key="s_time")
                    _s_slot = dict(_PLAN_SLOTS)[_s_time]
                with sa4:
                    _s_notes = st.text_area("Notes", height=120, key="s_notes", placeholder="Optional")
                if st.button("Add", type="primary", key="s_add"):
                    if _s_cid.strip():
                        _ad = _s_date if _s_status != "PENDING" else None
                        _at = _s_time if _s_status != "PENDING" else "TBD"
                        _an = _s_slot if _s_status != "PENDING" else 99
                        _add_plan_entry(_plan_week_start(_s_date),_ad,_at,_an,
                                        _s_cid,_s_scac,_s_site,_s_product,int(_s_qty),_s_notes,_s_status)
                        st.success(f"{_s_cid.upper().strip()} ({_s_scac} \u2192 {_s_site})")
                        st.rerun()
                    else:
                        st.warning("Container # required.")
            with _add_tab_b:
                st.caption(
                    "**Format (one container per line):** `CONTAINER_ID  CARRIER  SITE`  \n"
                    "Carrier and Site are optional if you set defaults below.  \n"
                    "Example: `TCNU3773041  HDDR  RIC6`"
                )
                _bulk_text = st.text_area("Paste containers", height=200, key="bulk_text",
                                          placeholder="TCNU3773041  HDDR  RIC6\nMRKU4103422  ATMI  ILM1")
                bb1, bb2, bb3, bb4, bb5 = st.columns(5)
                with bb1: _b_site    = st.selectbox("Default Site",    _act_sites or ["RIC6"], key="b_site")
                with bb2: _b_scac    = st.selectbox("Default Carrier", _act_scacs, key="b_scac")
                with bb3: _b_product = st.selectbox("Default Product", _mat_opts, key="b_product")
                with bb4: _b_date    = st.date_input("Delivery Date",  value=_today, key="b_date")
                with bb5: _b_status  = st.selectbox("Status",          ["SCHEDULED","PENDING"], key="b_status")
                _b_ll     = st.checkbox("Level load across the week", value=True, key="b_ll")
                _b_incsat = st.checkbox("Include Saturday",            value=True, key="b_incsat")
                if st.button("Parse & Add All", type="primary", key="b_add"):
                    if _bulk_text.strip():
                        raw_cids, added, skipped = [], 0, []
                        for ln in _bulk_text.strip().splitlines():
                            parts = ln.strip().split()
                            if not parts: continue
                            cid = parts[0].upper()
                            if len(cid) < 4: skipped.append(ln); continue
                            b_sc = parts[1].upper() if len(parts) > 1 and parts[1].upper() in _act_scacs else _b_scac
                            b_si = parts[2].upper() if len(parts) > 2 else _b_site
                            raw_cids.append((cid, b_sc, b_si))
                        if _b_ll and raw_cids:
                            _ll_active = _get_week_config(_plan_week_start(_b_date).isoformat())
                            _ll_day_map = {"Sun":6,"Mon":0,"Tue":1,"Wed":2,"Thu":3,"Fri":4,"Sat":5}
                            _ll_ws = _plan_week_start(_b_date)
                            work_days = sorted([
                                _ll_ws + timedelta(days=(_ll_day_map[d]-_ll_ws.weekday())%7)
                                for d in _ll_active if d in _ll_day_map
                            ])
                            if not work_days:
                                work_days = [_plan_week_start(_b_date) + timedelta(days=i) for i in range(5)]
                            n_days = len(work_days)
                            for idx, (cid, b_sc, b_si) in enumerate(raw_cids):
                                day = work_days[idx % n_days]
                                _scm_r2 = _scm[(_scm["site_code"]==b_si)&(_scm["scac"]==b_sc)]
                                _def_t2 = _scm_r2["priority_time"].iloc[0] if not _scm_r2.empty and pd.notna(_scm_r2["priority_time"].iloc[0]) else "8:30 AM"
                                _sn2 = dict(_PLAN_SLOTS).get(_def_t2, 2)
                                _add_plan_entry(_plan_week_start(day), day, _def_t2, _sn2,
                                                cid, b_sc, b_si, _b_product, 1, "", _b_status)
                                added += 1
                        else:
                            for cid, b_sc, b_si in raw_cids:
                                _scm_r2 = _scm[(_scm["site_code"]==b_si)&(_scm["scac"]==b_sc)]
                                _def_t2 = _scm_r2["priority_time"].iloc[0] if not _scm_r2.empty and pd.notna(_scm_r2["priority_time"].iloc[0]) else "8:30 AM"
                                _sn2 = dict(_PLAN_SLOTS).get(_def_t2, 2)
                                _ad2 = _b_date if _b_status != "PENDING" else None
                                _at2 = _def_t2 if _b_status != "PENDING" else "TBD"
                                _add_plan_entry(_plan_week_start(_b_date), _ad2, _at2,
                                                _sn2 if _b_status != "PENDING" else 99,
                                                cid, b_sc, b_si, _b_product, 1, "", _b_status)
                                added += 1
                        msg = f"Added {added} container(s)."
                        if skipped: msg += f" Skipped {len(skipped)} line(s)."
                        st.success(msg)
                        st.rerun()
                    else:
                        st.warning("Nothing to parse.")

        # Week grid
        _pdf = _get_plan(_ws.isoformat())
        if _pdf.empty:
            st.info("No entries yet. Use the steps above to build the plan.")
        else:
            st.divider()
            _week_grid(_pdf[_pdf["status"]!="PENDING"], _wdays, constraints=_site_cfg)
            st.divider()
            _status_editor(_pdf, "pb")

        # Overflow queue
        _overflow_pool = st.session_state.get("pb_overflow", [])
        if _overflow_pool:
            with st.expander(f"\u26a0\ufe0f Overflow Queue \u2014 {len(_overflow_pool)} container(s)", expanded=True):
                st.caption("These containers exceeded this week's capacity. Push to next week or dismiss.")
                _ov_df = pd.DataFrame(_overflow_pool)
                st.dataframe(
                    _ov_df[["container","scac","product_type","source"]].rename(
                        columns={"container":"Container","scac":"Carrier","product_type":"Product","source":"Source"}
                    ),
                    use_container_width=True, hide_index=True
                )
                _ov_c1, _ov_c2 = st.columns(2)
                with _ov_c1:
                    if st.button("Push to Next Week", type="primary", key="ov_push"):
                        _next_ws = _ws + timedelta(weeks=1)
                        _ov_pushed = 0
                        for _ov in _overflow_pool:
                            _add_plan_entry(
                                _next_ws, None, "TBD", 99,
                                _ov["container"], _ov.get("scac",""),
                                _ov.get("site_code", _pb_site),
                                _ov.get("product_type",""),
                                1, "Overflow from prior week", "PENDING"
                            )
                            _ov_pushed += 1
                        st.session_state.pop("pb_overflow", None)
                        st.success(f"Pushed {_ov_pushed} container(s) to week of {_next_ws}.")
                        st.rerun()
                with _ov_c2:
                    if st.button("Dismiss Overflow", type="secondary", key="ov_dismiss"):
                        st.session_state.pop("pb_overflow", None)
                        st.rerun()


    # ══════════════════════════════════════════════════════════════════════════
    # ALL SITES (COMPILED)
    # ══════════════════════════════════════════════════════════════════════════
    with _tp2:
        _pdf = _get_plan(_ws.isoformat())
        if _pdf.empty:
            st.info("No plan entries for this week.")
        else:
            _act = _pdf[_pdf["status"]!="PENDING"]
            # per-site metrics
            _sc_cnt = _act.groupby("site_code")["id"].count() if not _act.empty else pd.Series(dtype=int)
            _mc = st.columns(max(len(_sc_cnt), 1))
            for mi, (sc, cnt) in enumerate(sorted(_sc_cnt.items())):
                with _mc[mi % len(_mc)]: st.metric(sc, int(cnt))
            st.divider()
            COLS = ["Day","Time","Site","Container #","Carrier","Product","Qty","Notes","Status","Confirmed"]
            st.dataframe(_plan_row_table(_act.sort_values(["appt_date","site_code","slot_num"]), COLS),
                         use_container_width=True, hide_index=True)
            _pend = _pdf[_pdf["status"]=="PENDING"]
            if not _pend.empty:
                st.caption(f"**{len(_pend)} pending (unscheduled)**")
                st.dataframe(_plan_row_table(_pend, ["Site","Container #","Carrier","Product","Qty","Notes"]),
                             use_container_width=True, hide_index=True)
            xl = _export_compiled_excel(_pdf, _wdays, _ws)
            st.download_button("Export — All Sites", data=xl,
                file_name=f"Delivery Plan All Sites {_ws.strftime('%m.%d.%y')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary")

    # ══════════════════════════════════════════════════════════════════════════
    # BY SITE (INTERNAL)
    # ══════════════════════════════════════════════════════════════════════════
    with _tp3:
        _pdf = _get_plan(_ws.isoformat())
        _avail_sites = sorted(_pdf["site_code"].dropna().unique().tolist()) if not _pdf.empty else []
        if not _avail_sites:
            st.info("No entries for this week.")
        else:
            _sel_site = st.selectbox("Site", _avail_sites, key="site_sel")
            _sdf  = _pdf[_pdf["site_code"]==_sel_site]
            _sact = _sdf[_sdf["status"]!="PENDING"]
            # carrier breakdown badges
            _sc_c = _sact.groupby("carrier")["id"].count() if not _sact.empty else pd.Series(dtype=int)
            _bcs  = st.columns(max(len(_sc_c), 1))
            for bi, (sc, cnt) in enumerate(sorted(_sc_c.items())):
                with _bcs[bi]:
                    clr = _SCAC_COLOR.get(sc, "#888")
                    st.markdown(
                        f"<div style='border-left:4px solid {clr};padding:6px 10px'>"
                        f"<b style='color:{clr}'>{sc}</b><br/><span style='font-size:22px'>{cnt}</span></div>",
                        unsafe_allow_html=True)
            st.divider()
            COLS = ["Day","Time","Container #","Carrier","Product","Qty","Notes","Status","Confirmed"]
            st.dataframe(_plan_row_table(_sact.sort_values(["appt_date","slot_num"]), COLS),
                         use_container_width=True, hide_index=True)
            _sp = _sdf[_sdf["status"]=="PENDING"]
            if not _sp.empty:
                st.caption(f"**{len(_sp)} pending**")
                st.dataframe(_plan_row_table(_sp, ["Container #","Carrier","Product","Qty","Notes"]),
                             use_container_width=True, hide_index=True)
            # ── Daily notification ────────────────────────────────────────────────
            with st.expander("Daily Notification — Copy for Slack", expanded=False):
                st.caption("Generates tomorrow's delivery list for this site, ready to paste into Slack.")
                _notif_date = st.date_input(
                    "Delivery date to notify",
                    value=_today + timedelta(days=1),
                    key=f"notif_date_{_sel_site}"
                )
                _notif_df = _pdf[
                    (_pdf["site_code"] == _sel_site) &
                    (_pdf["appt_date"] == _notif_date.isoformat()) &
                    (_pdf["status"].isin(["SCHEDULED", "HOLD"]))
                ].sort_values("slot_num")

                if _notif_df.empty:
                    st.info(f"No scheduled deliveries for {_sel_site} on {_notif_date.strftime('%b %d')}.")
                else:
                    _day_label = _notif_date.strftime("%A, %B %d")
                    lines = [f"*{_sel_site} — Scheduled Deliveries for {_day_label}*", ""]
                    for r in _notif_df.itertuples():
                        _cn = _get_carriers_df()
                        _cname_row = _cn[_cn["scac"] == r.carrier]
                        _cname_n   = _cname_row["carrier_name"].iloc[0] if not _cname_row.empty else r.carrier
                        lines.append(f"{r.appt_time}  |  {r.container_id}  |  {_cname_n}  |  {r.product_type or ''}")
                    lines.append("")
                    lines.append(f"Total: {len(_notif_df)} container(s)")
                    notif_text = "\n".join(lines)
                    st.text_area("Copy this", value=notif_text, height=200, key=f"notif_txt_{_sel_site}")

            _status_editor(_sdf, f"site_{_sel_site}")
            if not _sdf.empty:
                xl = _export_site_excel(_pdf, _sel_site, _ws)
                st.download_button(f"Export {_sel_site} to Excel", data=xl,
                    file_name=f"{_sel_site} Delivery Plan {_ws.strftime('%m.%d.%y')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            # ── Mid-Week Adjustment ────────────────────────────────────────────────
            with st.expander("🔄 Mid-Week Adjustment", expanded=False):
                st.caption(
                    "Use when: site stops receiving, container rejected, carrier requests change, "
                    "or contingent container clears customs. Updates generate draft carrier messages."
                )
                _mw_scenario = st.selectbox(
                    "Scenario", [
                        "Site stopped receiving — hold and reschedule",
                        "Container rejected at site — reschedule to next available day",
                        "Container cleared customs — slot into plan",
                        "Carrier requested change",
                    ],
                    key=f"mw_scenario_{_sel_site}"
                )
                _mw_cid = st.text_input("Container # (optional — blank = all affected today)", key=f"mw_cid_{_sel_site}")
                _mw_note = st.text_area("Notes / reason", height=68, key=f"mw_note_{_sel_site}",
                                        placeholder="e.g. RIC6 stopped receiving after lunch — backlog at dock")
                _mw_new_date = st.date_input("Reschedule to", value=_today + timedelta(days=1), key=f"mw_date_{_sel_site}")

                if st.button("Generate Hold Message", key=f"mw_gen_{_sel_site}"):
                    _affected = _sdf[
                        (_sdf["appt_date"] == _today.isoformat()) &
                        (_sdf["status"].isin(["SCHEDULED","HOLD"]))
                    ]
                    if _mw_cid.strip():
                        _affected = _affected[_affected["container_id"].str.upper() == _mw_cid.strip().upper()]
                    if _affected.empty:
                        st.info("No scheduled containers found for today matching that criteria.")
                    else:
                        _carriers_hit = _affected["carrier"].unique().tolist()
                        _msg_lines = [f"Hi team,\n\n**{_mw_scenario}** — {_sel_site}\n"]
                        _msg_lines.append(f"**Affected containers ({len(_affected)}):**")
                        for _r in _affected.itertuples():
                            _msg_lines.append(f"  • {_r.container_id}  {_r.carrier}  {_r.appt_time}")
                        _msg_lines.append(f"\n**Action:** Please hold delivery and reschedule to {_mw_new_date.strftime('%A, %m/%d')}.")
                        if _mw_note:
                            _msg_lines.append(f"\n**Context:** {_mw_note}")
                        _msg_lines.append("\nThank you,\nDominique Kennedy — AGL Robotics Dray")
                        st.text_area("Draft carrier message (copy/paste to Slack or email)", 
                                     value="\n".join(_msg_lines), height=220,
                                     key=f"mw_msg_{_sel_site}")
                        st.caption(f"Carriers affected: {', '.join(_carriers_hit)}")

    # ══════════════════════════════════════════════════════════════════════════
    # CARRIER VIEW (EXTERNAL / SHAREABLE)
    # ══════════════════════════════════════════════════════════════════════════
    with _tp4:
        _pdf  = _get_plan(_ws.isoformat())
        _cdf2 = _get_carriers_df()

        # ── Carrier link panel ────────────────────────────────────────────────
        st.markdown("### Carrier Links")
        st.caption(
            "Each carrier has a unique, token-gated link to their delivery schedule. "
            "Share the link after verifying the data in the admin preview below. "
            "Use **Regenerate Token** to invalidate an old link."
        )
        # Build table of all known carriers + their tokens
        _all_carriers_df = _cdf2 if not _cdf2.empty else pd.DataFrame(
            columns=["scac","carrier_name","ops_contact","ops_email"]
        )
        _link_rows = []
        for _, _cr in _all_carriers_df.iterrows():
            _cr_scac  = str(_cr["scac"])
            _cr_name  = str(_cr.get("carrier_name",""))
            _cr_email = str(_cr.get("ops_email",""))
            _tok = _get_or_create_carrier_token(_cr_scac)
            _link = f"{CARRIER_PLAN_BASE_URL}?token={_tok}"
            _link_rows.append({
                "SCAC":    _cr_scac,
                "Carrier": _cr_name,
                "Contact": _cr_email,
                "Link":    _link,
                "Token":   _tok[:12] + "...",
            })
        if _link_rows:
            _ldf = pd.DataFrame(_link_rows)
            # One info box per carrier — copyable link
            for _, _lr in _ldf.iterrows():
                _lclr = _SCAC_COLOR.get(_lr["SCAC"], "#888")
                _lc1, _lc2 = st.columns([5, 1])
                with _lc1:
                    st.markdown(
                        f"<div style='background:{_lclr}12;border-left:3px solid {_lclr};"
                        f"padding:8px 12px;border-radius:4px;margin-bottom:4px'>"
                        f"<b>{_lr['Carrier']} ({_lr['SCAC']})</b>"
                        f"{'  <small>· ' + _lr['Contact'] + '</small>' if _lr['Contact'] and _lr['Contact'] != 'nan' else ''}  "
                        f"<br/><code style='font-size:11px'>{_lr['Link']}</code></div>",
                        unsafe_allow_html=True
                    )
                with _lc2:
                    if st.button("Regenerate", key=f"regen_{_lr['SCAC']}",
                                 help="Invalidates the old link — carrier must use new URL"):
                        _regenerate_carrier_token(_lr["SCAC"])
                        st.success(f"{_lr['SCAC']} token regenerated.")
                        st.rerun()

        st.divider()

        # ── Admin preview (view as carrier) ───────────────────────────────────
        st.markdown("### Admin Preview — View as Carrier")
        st.caption(
            "Select a carrier to see exactly what they will see at their link. "
            "Verify data before sending. Confirmation status shows once carriers have responded."
        )
        _avail_scacs = sorted(_pdf["carrier"].dropna().unique().tolist()) if not _pdf.empty else []
        if not _avail_scacs:
            st.info("No plan entries for this week.")
        else:
            _sel_scac = st.selectbox("Carrier", _avail_scacs, key="carr_sel")
            _cn_r = _cdf2[_cdf2["scac"]==_sel_scac]
            _cname    = _cn_r["carrier_name"].iloc[0] if not _cn_r.empty else _sel_scac
            _ccontact = _cn_r["ops_contact"].iloc[0] if not _cn_r.empty else ""
            _cemail   = _cn_r["ops_email"].iloc[0] if not _cn_r.empty else ""
            clr_cv    = _SCAC_COLOR.get(_sel_scac, "#888")

            st.markdown(
                f"<div style='background:{clr_cv}15;border-left:4px solid {clr_cv};"
                f"padding:10px 14px;border-radius:4px;margin-bottom:8px'>"
                f"<b style='font-size:16px'>{_cname} ({_sel_scac})</b>"
                f"{'<br/><small>Ops: ' + _ccontact + '</small>' if _ccontact else ''}"
                f"{'<br/><small>' + _cemail + '</small>' if _cemail and _cemail != 'nan' else ''}"
                f"</div>", unsafe_allow_html=True)

            _cdf_w = _pdf[_pdf["carrier"]==_sel_scac]
            _cact  = _cdf_w[_cdf_w["status"]!="PENDING"]

            # site breakdown metrics
            _csite_c = _cact.groupby("site_code")["id"].count() if not _cact.empty else pd.Series(dtype=int)
            cmc = st.columns(max(len(_csite_c), 1))
            for cmi, (cs, cc) in enumerate(sorted(_csite_c.items())):
                with cmc[cmi]: st.metric(cs, int(cc))

            # confirmation summary
            if "carrier_confirmed" in _cact.columns:
                _n_conf = int(_cact["carrier_confirmed"].fillna(0).astype(int).sum())
                _n_total = len(_cact)
                if _n_conf > 0:
                    st.success(f"✅ {_n_conf} / {_n_total} containers confirmed by carrier")
                elif _n_total > 0:
                    st.warning(f"⏳ Awaiting carrier confirmation — 0 / {_n_total} confirmed")

            st.divider()
            st.caption("**Carrier view** — no internal fields. Confirmation column shows once carrier responds.")
            ext_rows = []
            for r in _cact.sort_values(["appt_date","site_code","slot_num"]).itertuples():
                try: day_s = date.fromisoformat(str(r.appt_date)).strftime("%a %m/%d")
                except: day_s = "TBD"
                _conf = getattr(r, "carrier_confirmed", 0) or 0
                ext_rows.append({
                    "Day":          day_s,
                    "Site":         r.site_code or "",
                    "Appt Time":    r.appt_time,
                    "Container #":  r.container_id,
                    "Product Type": r.product_type or "",
                    "Qty":          r.qty or "",
                    "Notes":        r.notes or "",
                    "Confirmed":    "✅" if int(_conf) else "⬜",
                })
            if ext_rows:
                st.dataframe(pd.DataFrame(ext_rows), use_container_width=True, hide_index=True)
            _cpend = _cdf_w[_cdf_w["status"]=="PENDING"]
            if not _cpend.empty:
                st.caption(f"**{len(_cpend)} pending / contingent**")
                pext = [{"Site": r.site_code or "", "Container #": r.container_id,
                          "Product Type": r.product_type or "", "Notes": r.notes or ""}
                         for r in _cpend.itertuples()]
                st.dataframe(pd.DataFrame(pext), use_container_width=True, hide_index=True)

            if not _cdf_w.empty:
                # Pre-send validation checklist
                with st.expander("✅ Pre-Send Validation Checklist — verify before exporting", expanded=False):
                    st.caption("Check off each item before sending the plan to this carrier.")
                    _chk_key = f"chk_{_sel_scac}_{_ws.isoformat()}"
                    _checks = [
                        ("all_at_yard",   "All containers verified as at-yard with no actual FC delivery date"),
                        ("product_split", "Product type (Shelf/Strap) matches site's requested split"),
                        ("fifo_order",    "FIFO order respected — longest-dwelling containers scheduled first"),
                        ("no_lunch",      "No appointments during lunch break (12:00–1:00 PM)"),
                        ("last_appt",     "Last appointment at or before site-specific cutoff"),
                        ("level_load",    "Level-loaded across days (+/- 1–2 container difference max)"),
                        ("hddr_am",       "HDDR/HUDD containers in early morning slots (before lunch)"),
                        ("carrier_tab",   "Carrier tab is clean and ready to share"),
                        ("contingent",    "Contingent containers clearly marked as PENDING"),
                    ]
                    _all_checked = True
                    for _ck, _clabel in _checks:
                        _checked = st.checkbox(_clabel, key=f"{_chk_key}_{_ck}")
                        if not _checked:
                            _all_checked = False
                    if _all_checked:
                        st.success("All checks passed — plan is ready to send.")
                    else:
                        st.warning("Complete all checks before exporting.")

                _dl_c1, _dl_c2 = st.columns([3, 2])
                with _dl_c1:
                    xl = _export_carrier_excel(_pdf, _sel_scac, _cname, _ws)
                    st.download_button(f"Export {_cname} Plan to Excel", data=xl,
                        file_name=f"{_sel_scac} Delivery Plan {_ws.strftime('%m.%d.%y')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary")
                with _dl_c2:
                    # Show their carrier link for quick copy
                    _tok2 = _get_or_create_carrier_token(_sel_scac)
                    _clink = f"{CARRIER_PLAN_BASE_URL}?token={_tok2}"
                    st.info(f"🔗 [Carrier portal link]({_clink})")

    # ══════════════════════════════════════════════════════════════════════════
    # WoW / HISTORY
    # ══════════════════════════════════════════════════════════════════════════
    with _tp5:
        _wow_t, _hist_t = st.tabs(["Week-over-Week", "History Search"])

        with _wow_t:
            _pdf      = _get_plan(_ws.isoformat())
            _prior    = _get_plan((_ws - timedelta(weeks=1)).isoformat())
            cur_ids   = set(_pdf["container_id"].str.upper()) if not _pdf.empty else set()
            prior_ids = set(_prior["container_id"].str.upper()) if not _prior.empty else set()
            new_ids     = cur_ids - prior_ids
            rolled_ids  = cur_ids & prior_ids
            dropped_ids = prior_ids - cur_ids
            w1, w2, w3 = st.columns(3)
            with w1: st.metric("New this week",         len(new_ids))
            with w2: st.metric("Rolled from prior week", len(rolled_ids))
            with w3: st.metric("Dropped / delivered",    len(dropped_ids))
            st.divider()
            if new_ids:
                st.markdown("**New containers:**")
                st.dataframe(_plan_row_table(_pdf[_pdf["container_id"].str.upper().isin(new_ids)],
                    ["Day","Site","Container #","Carrier","Product","Status"]),
                    use_container_width=True, hide_index=True)
            if rolled_ids:
                st.markdown("**Carried over from prior week:**")
                roll_rows = []
                for r in _pdf[_pdf["container_id"].str.upper().isin(rolled_ids)].itertuples():
                    pr = _prior[_prior["container_id"].str.upper()==r.container_id.upper()]
                    pr_stat = pr["status"].iloc[0] if not pr.empty else "?"
                    try: day_s = date.fromisoformat(str(r.appt_date)).strftime("%a %m/%d")
                    except: day_s = "TBD"
                    roll_rows.append({
                        "Container #": r.container_id, "Site": getattr(r,"site_code","") or "",
                        "Carrier": r.carrier,
                        "This Week": f"{day_s} {_STATUS_BADGE.get(r.status,'')}",
                        "Prior Week Status": _STATUS_BADGE.get(pr_stat, pr_stat),
                    })
                st.dataframe(pd.DataFrame(roll_rows), use_container_width=True, hide_index=True)
            if dropped_ids:
                prior_lbl = (_ws - timedelta(weeks=1)).strftime("%b %d")
                st.markdown(f"**Not on this week (were on {prior_lbl}):**")
                st.dataframe(_plan_row_table(_prior[_prior["container_id"].str.upper().isin(dropped_ids)],
                    ["Day","Site","Container #","Carrier","Status"]),
                    use_container_width=True, hide_index=True)

        with _hist_t:
            st.caption("Search all stored plans.")
            hc1, hc2, hc3 = st.columns([2,1,1])
            with hc1: _hq  = st.text_input("Container # or keyword", key="hq")
            with hc2: _hsc = st.selectbox("Carrier", ["All"]+list(_SCAC_COLOR.keys()), key="hsc")
            with hc3: _hsi = st.text_input("Site", placeholder="RIC6", key="hsi")
            _hdr = st.date_input("Date range", value=[date(2026,1,1), _today], key="hdr")
            if st.button("Search", key="hsearch"):
                conn = get_db()
                q = "SELECT * FROM delivery_plan WHERE 1=1"
                p = []
                if _hq.strip():
                    q += " AND (container_id LIKE ? OR notes LIKE ?)"; p += [f"%{_hq.strip().upper()}%", f"%{_hq.strip()}%"]
                if _hsc != "All":
                    q += " AND carrier=?"; p.append(_hsc)
                if _hsi.strip():
                    q += " AND site_code LIKE ?"; p.append(f"%{_hsi.strip().upper()}%")
                if len(_hdr) == 2:
                    q += " AND (appt_date >= ? AND appt_date <= ?)"; p += [str(_hdr[0]), str(_hdr[1])]
                q += " ORDER BY appt_date DESC LIMIT 300"
                hdf = pd.read_sql_query(q, conn, params=tuple(p)); conn.close()
                if hdf.empty:
                    st.info("No results.")
                else:
                    st.caption(f"**{len(hdf)} result(s)**")
                    st.dataframe(_plan_row_table(hdf, ["Day","Site","Container #","Carrier","Product","Status","Notes"]),
                                 use_container_width=True, hide_index=True)

    # ══════════════════════════════════════════════════════════════════════════
    # CONFIG
    # ══════════════════════════════════════════════════════════════════════════

    # ═══════════════════════════════════════════════════════════════════════
    # SRF — SHORT RANGE FORECAST
    # ═══════════════════════════════════════════════════════════════════════
    with _tp6:
        # Contract: LSP must maintain capacity for ≥120% of weekly forecasted volume.
        # Amazon must notify LSP ≥2 weeks in advance of capacity increases.
        st.caption(
            f"▶ SRF anchored at **W{_ws.isocalendar()[1]}** — "
            f"week of {_ws.strftime('%b %d, %Y')}. "
            "Use the week selector above to change."
        )
        _srf_conn = get_db()
        _iso_wk    = _ws.isocalendar()[1]
        _iso_yr    = _ws.isocalendar()[0]

        st.markdown("**6-Week SRF — Auto-Generated Volume Forecast**")
        st.caption(
            "Projected volumes are calculated from historical weekly data using a weighted "
            "moving average (WMA). Recent weeks are weighted higher. Trend is applied when "
            "slope exceeds 3%/week. Override individual weeks below if needed."
        )

        # ── Auto-forecast ─────────────────────────────────────────────────────
        _forecast_data = _srf_auto_forecast(_srf_conn, _iso_wk, _iso_yr)
        _fc            = _forecast_data["forecast"]          # {week: projected_vol}
        _fc_method     = _forecast_data["method"]
        _fc_wma        = _forecast_data["wma_baseline"]
        _fc_trend      = _forecast_data["trend_pct"]
        _fc_vols       = _forecast_data.get("vols", [])
        _fc_hist       = _forecast_data.get("history", [])

        # Trend label
        if abs(_fc_trend) < 2:
            _trend_label = "Stable ➡️"
        elif _fc_trend > 0:
            _trend_label = f"Growing 📈 +{_fc_trend}%/wk"
        else:
            _trend_label = f"Declining 📉 {_fc_trend}%/wk"

        _met1, _met2, _met3 = st.columns(3)
        with _met1:
            st.metric("WMA Baseline", f"{_fc_wma} containers",
                      help="Weighted moving average of recent weekly volume — higher weight to newer weeks")
        with _met2:
            st.metric("Volume Trend", _trend_label,
                      help="Linear regression slope over history window")
        with _met3:
            _hist_src = f"{len(_fc_vols)} wks · {_fc_method}"
            st.metric("Data Source", _hist_src,
                      help="wbr_results preferred; falls back to delivery_plan aggregates")

        # Forecast table (read-only)
        _srf_weeks = [(_iso_wk + i) % 53 or 53 for i in range(1, 7)]
        import pandas as _pd_fc
        _fc_rows = []
        for _fw in _srf_weeks:
            _proj = _fc.get(_fw, _fc_wma)
            _cap  = _srf_required_capacity(_proj)
            _fc_rows.append({
                "Week":                   f"W{_fw}",
                "Projected Volume":       _proj,
                "Req. Capacity (×1.20)":  _cap,
                "Trend":                  _trend_label,
            })
        st.dataframe(_pd_fc.DataFrame(_fc_rows), use_container_width=True, hide_index=True)

        # History sparkline (if data available)
        if _fc_vols:
            _hist_df = _pd_fc.DataFrame({
                "Week": [f"W{r[0]}" for r in _fc_hist],
                "Volume": _fc_vols,
            })
            st.caption(f"📊 Historical volume used for forecast ({_fc_method})")
            st.line_chart(_hist_df.set_index("Week")["Volume"], height=120, use_container_width=True)

        # Override expander — optional manual adjustments
        with st.expander("✏️ Override forecast (for known demand spikes or planned events)", expanded=False):
            st.caption("Set a week to any value > 0 to override the auto-projected volume. Leave at 0 to keep auto.")
            _override_cols = st.columns(6)
            _overrides = {}
            for _ci, _fw in enumerate(_srf_weeks):
                with _override_cols[_ci]:
                    _overrides[_fw] = st.number_input(
                        f"W{_fw}", min_value=0, max_value=999,
                        value=0, step=1, key=f"srf_override_{_fw}"
                    )

        # Merge: override wins if non-zero; otherwise use auto forecast
        _final_forecast = {
            _fw: (_overrides[_fw] if _overrides.get(_fw, 0) > 0 else _fc.get(_fw, _fc_wma))
            for _fw in _srf_weeks
        }

        if st.button("💾 Save SRF Submission", type="primary", key="srf_save"):
            _srf_save_submission(_srf_conn, _iso_wk, _iso_yr, _final_forecast)
            _override_count = sum(1 for fw in _srf_weeks if _overrides.get(fw, 0) > 0)
            _note = f" ({_override_count} week(s) manually overridden)" if _override_count else ""
            st.success(
                f"✅ SRF W{_iso_wk} saved — {sum(_final_forecast.values())} total containers "
                f"across W{_srf_weeks[0]}–W{_srf_weeks[-1]}{_note}"
            )
            st.rerun()

        # ── Display current SRF with capacity calcs + notice flags ──────────
        _srf_current = _srf_get_latest(_srf_conn, _iso_yr)
        if _srf_current:
            _sub_wk = next(iter(_srf_current.values()))["submission_week"]
            _srf_prior_data = _srf_get_prior(_srf_conn, _iso_yr, _sub_wk + 1)
            _flagged = _srf_compute_notice_flags(_srf_current, _srf_prior_data, _iso_wk)

            # Notice alerts — shown above the table
            _urgent = [(fw, r) for fw, r in _flagged.items()
                       if r["notice_status"] in ("URGENT", "DUE_NOW") and not r["notice_sent"]]
            _scheduled = [(fw, r) for fw, r in _flagged.items()
                          if r["notice_status"] == "SCHEDULED" and not r["notice_sent"]]

            for _fw, _r in _urgent:
                _window_label = "⚠️ WITHIN NOTICE WINDOW" if _r["notice_status"] == "DUE_NOW" else "🚨 PAST NOTICE DEADLINE"
                st.error(
                    f"**W{_fw} Capacity Increase — {_window_label}**  \n"
                    f"Forecast increased {_r['prior_vol']}→{_r['vol']} containers "
                    f"(+{_r['delta']}, +{_r['delta_pct']}%). "
                    f"Required capacity: **{_r['cap']} moves**. "
                    f"Notify LSP immediately."
                )
                if st.button(f"✉️ Mark W{_fw} Notice Sent", key=f"srf_notice_{_fw}"):
                    _srf_mark_notice_sent(_srf_conn, _sub_wk, _iso_yr, _fw)
                    st.rerun()

            for _fw, _r in _scheduled:
                _due_wk = _r.get("notice_due_week", _fw - 2)
                st.warning(
                    f"**W{_fw} Capacity Increase Detected** — Notice due by W{_due_wk}.  \n"
                    f"Forecast: {_r['prior_vol']}→{_r['vol']} (+{_r['delta_pct']}%). "
                    f"Required capacity: {_r['cap']} moves."
                )
                if st.button(f"✉️ Mark W{_fw} Notice Sent", key=f"srf_notice_sched_{_fw}"):
                    _srf_mark_notice_sent(_srf_conn, _sub_wk, _iso_yr, _fw)
                    st.rerun()

            # SRF summary table
            _tbl_rows = []
            for _fw in sorted(_flagged.keys()):
                _r = _flagged[_fw]
                _prior_disp = str(_r["prior_vol"]) if _r["prior_vol"] is not None else "—"
                _delta_disp = (f"+{_r['delta']} (+{_r['delta_pct']}%)" if _r["delta"] > 0
                               else ("—" if _r["delta"] == 0 else str(_r["delta"])))
                _status_map = {
                    "OK":        "✅ No change",
                    "SENT":      f"✅ Notice sent {_r.get('notice_sent_at','')[:10]}",
                    "SCHEDULED": f"📅 Notice due W{_r.get('notice_due_week', _fw-2)}",
                    "DUE_NOW":   "⚠️ Notice due NOW",
                    "URGENT":    "🚨 Past deadline",
                }
                _tbl_rows.append({
                    "Week":              f"W{_fw}",
                    "Forecast (vol)":    _r["vol"],
                    "Prior Forecast":    _prior_disp,
                    "WoW Change":        _delta_disp,
                    "Req. Capacity (+20%)": _r["cap"],
                    "Notice Status":     _status_map.get(_r["notice_status"], _r["notice_status"]),
                })

            import pandas as _pd_srf
            _srf_df = _pd_srf.DataFrame(_tbl_rows)
            st.dataframe(_srf_df, use_container_width=True, hide_index=True)
            st.caption(f"SRF submitted for W{_sub_wk}. Req. capacity = ceil(forecast × 1.20). "
                       "Amazon must notify LSP ≥2 weeks in advance of capacity increases.")

            # Draft notice language
            _noticeable = [(_fw, _r) for _fw, _r in _flagged.items()
                           if _r["delta"] > 0 and not _r["notice_sent"]]
            if _noticeable:
                with st.expander("📝 Draft LSP Capacity Notice", expanded=False):
                    _notice_lines = [
                        f"Subject: Amazon Robotics — Capacity Increase Notice (W{', W'.join(str(fw) for fw,_ in _noticeable)})",
                        "",
                        "Team,",
                        "",
                        "Per our contracted SRF obligations, Amazon is providing advance notice of "
                        "anticipated volume increases requiring additional carrier capacity:",
                        "",
                    ]
                    for _fw, _r in _noticeable:
                        _notice_lines.append(
                            f"  • Week W{_fw}: Forecasted volume increased to {_r['vol']} containers "
                            f"(+{_r['delta']} from prior forecast of {_r['prior_vol']}). "
                            f"Required carrier capacity: {_r['cap']} moves (120% of forecast)."
                        )
                    _notice_lines += [
                        "",
                        "Please confirm capacity availability and any constraints by end of week.",
                        "",
                        "Thank you,",
                        "Amazon Global Logistics — Robotics Dray",
                    ]
                    st.text_area("Draft notice (edit before sending)",
                                 value="\n".join(_notice_lines), height=220,
                                 key="srf_notice_draft")
        else:
            st.info("No SRF saved yet for this cycle. Review the auto-generated forecast above and click **Save SRF Submission** to lock it in.")

        _srf_conn.close()

    with _tp7:
        _cfg1, _cfg2, _cfg3 = st.tabs(["Sites", "Carriers", "Site–Carrier Map"])

        with _cfg1:
            _sdf3 = _get_sites_df()
            st.dataframe(_sdf3.drop(columns=["id"], errors="ignore"), use_container_width=True, hide_index=True)
            st.markdown("**Site Receiving Constraints Reference**")
            _site_constraints = [
                {"Site":"RIC6","Receiving Window":"7:30 AM – 4:30 PM","Lunch (no recv)":"12:00–1:00 PM","Last Arrival":"3:30 PM","Max Loads/Day":"11–12","No-Recv Day":"Monday","Carrier Pref":"HDDR before lunch (7:30–11:30 AM)"},
                {"Site":"ILM1","Receiving Window":"TBD","Lunch (no recv)":"TBD","Last Arrival":"TBD","Max Loads/Day":"TBD","No-Recv Day":"TBD","Carrier Pref":"TBD"},
                {"Site":"DBM6","Receiving Window":"TBD","Lunch (no recv)":"TBD","Last Arrival":"TBD","Max Loads/Day":"TBD","No-Recv Day":"TBD","Carrier Pref":"TBD"},
                {"Site":"SJC8","Receiving Window":"TBD","Lunch (no recv)":"TBD","Last Arrival":"TBD","Max Loads/Day":"TBD","No-Recv Day":"TBD","Carrier Pref":"TBD"},
                {"Site":"LAX", "Receiving Window":"TBD","Lunch (no recv)":"TBD","Last Arrival":"TBD","Max Loads/Day":"TBD","No-Recv Day":"TBD","Carrier Pref":"TBD"},
                {"Site":"IAG1","Receiving Window":"TBD","Lunch (no recv)":"TBD","Last Arrival":"TBD","Max Loads/Day":"TBD","No-Recv Day":"TBD","Carrier Pref":"Via XPH1 transload — 53ft trailer"},
            ]
            st.dataframe(pd.DataFrame(_site_constraints), use_container_width=True, hide_index=True)
            st.caption("Update constraints above via Config edits as each site's parameters are confirmed.")
            with st.expander("Add / Update Site"):
                cx1, cx2, cx3 = st.columns(3)
                with cx1:
                    _csc = st.text_input("Site Code", placeholder="RIC6", key="cfg_sc")
                    _cfn = st.text_input("FC Name",   placeholder="Richmond FC", key="cfg_fn")
                    _cpo = st.text_input("Port",      placeholder="ORF", key="cfg_po")
                with cx2:
                    _crg = st.text_input("Region", placeholder="East", key="cfg_rg")
                    _ccp = st.number_input("Capacity/Day", value=9, min_value=1, key="cfg_cp")
                    _cac = st.checkbox("Active", value=True, key="cfg_ac")
                with cx3:
                    _cno = st.text_area("Notes", height=100, key="cfg_no")
                if st.button("Save Site", key="cfg_ss"):
                    if _csc.strip():
                        _config_write(
                            """INSERT INTO plan_sites (site_code,fc_name,port,region,capacity_day,active,notes)
                               VALUES (?,?,?,?,?,?,?)
                               ON CONFLICT(site_code) DO UPDATE SET fc_name=excluded.fc_name,port=excluded.port,
                               region=excluded.region,capacity_day=excluded.capacity_day,
                               active=excluded.active,notes=excluded.notes""",
                            (_csc.strip().upper(),_cfn.strip(),_cpo.strip().upper(),
                             _crg.strip(),int(_ccp),int(_cac),_cno.strip()))
                        st.success(f"Saved {_csc.upper()}."); st.rerun()
                    else: st.warning("Site code required.")

        with _cfg2:
            _cdf3 = _get_carriers_df()
            st.dataframe(_cdf3.drop(columns=["id"], errors="ignore"), use_container_width=True, hide_index=True)
            with st.expander("Add / Update Carrier"):
                cy1, cy2 = st.columns(2)
                with cy1:
                    _cc_scac = st.text_input("SCAC", key="cfg_cscac")
                    _cc_name = st.text_input("Carrier Name", key="cfg_cname")
                with cy2:
                    _cc_ctc  = st.text_input("Ops Contact", key="cfg_cctc")
                    _cc_eml  = st.text_input("Ops Email",   key="cfg_ceml")
                _cc_act = st.checkbox("Active", value=True, key="cfg_cact")
                if st.button("Save Carrier", key="cfg_sc2"):
                    if _cc_scac.strip():
                        _config_write(
                            """INSERT INTO plan_carriers (scac,carrier_name,ops_contact,ops_email,active)
                               VALUES (?,?,?,?,?)
                               ON CONFLICT(scac) DO UPDATE SET carrier_name=excluded.carrier_name,
                               ops_contact=excluded.ops_contact,ops_email=excluded.ops_email,active=excluded.active""",
                            (_cc_scac.strip().upper(),_cc_name.strip(),_cc_ctc.strip(),_cc_eml.strip(),int(_cc_act)))
                        st.success(f"Saved {_cc_scac.upper()}."); st.rerun()
                    else: st.warning("SCAC required.")

        with _cfg3:
            _scm2 = _get_scm_df()
            if not _scm2.empty:
                dc = ["site_code","scac","carrier_name","site_contact","site_contact_email",
                      "priority_time","allocation_pct","active","notes"]
                st.dataframe(_scm2[[c for c in dc if c in _scm2.columns]], use_container_width=True, hide_index=True)
            with st.expander("Add / Update Mapping"):
                _sdf4 = _get_sites_df(); _cdf4 = _get_carriers_df()
                cm1, cm2, cm3 = st.columns(3)
                with cm1:
                    _cm_site  = st.selectbox("Site",    _sdf4["site_code"].tolist() if not _sdf4.empty else [], key="cm_s")
                    _cm_scac  = st.selectbox("Carrier", _cdf4["scac"].tolist() if not _cdf4.empty else [], key="cm_c")
                    _cm_alloc = st.number_input("Allocation %", value=0.0, step=5.0, key="cm_a")
                with cm2:
                    _cm_ctc  = st.text_input("Site Contact Name",  key="cm_ctc")
                    _cm_eml  = st.text_input("Site Contact Email", key="cm_eml")
                    _cm_pt   = st.selectbox("Priority Time Slot", [t for t,_ in _PLAN_SLOTS], key="cm_pt")
                with cm3:
                    _cm_act  = st.checkbox("Active", value=True, key="cm_act")
                    _cm_notes= st.text_area("Notes", height=80, key="cm_no")
                if st.button("Save Mapping", key="cm_save"):
                    _config_write(
                        """INSERT INTO plan_site_carrier
                           (site_code,scac,site_contact,site_contact_email,priority_time,allocation_pct,active,notes)
                           VALUES (?,?,?,?,?,?,?,?)
                           ON CONFLICT(site_code,scac) DO UPDATE SET site_contact=excluded.site_contact,
                           site_contact_email=excluded.site_contact_email,priority_time=excluded.priority_time,
                           allocation_pct=excluded.allocation_pct,active=excluded.active,notes=excluded.notes""",
                        (_cm_site,_cm_scac,_cm_ctc.strip(),_cm_eml.strip(),
                         _cm_pt,float(_cm_alloc),int(_cm_act),_cm_notes.strip()))
                    st.success(f"Saved {_cm_site} ↔ {_cm_scac}."); st.rerun()

    # ══════════════════════════════════════════════════════════════════════════
    # IMPORT HISTORY
    # ══════════════════════════════════════════════════════════════════════════
    with _tp8:
        st.subheader("Import Historical Delivery Data")
        st.caption(
            "Upload the **ToteASERs Robotics DBR Tracker.xlsx** (downloaded from SharePoint) "
            "to load container history. Run weekly before planning to keep WoW comparisons current. "
            "Containers already in the database are skipped automatically."
        )

        _up = st.file_uploader(
            "ToteASERs Robotics DBR Tracker.xlsx",
            type=["xlsx"],
            key="imp_dbr_upload",
            help="Download the file from SharePoint, then upload it here.",
        )
        _imp_bytes = _up.read() if _up else None

        if _imp_bytes:
            _imp_rows, _imp_errors = _parse_dbr_tracker(_imp_bytes)

            if not _imp_rows:
                st.error("No rows parsed — confirm this is the DBR Tracker file.")
                for _e in _imp_errors[:5]:
                    st.warning(_e)
            else:
                # Deduplication check
                _con2 = get_db()
                _existing_ids = set(
                    r[0] for r in _con2.execute(
                        "SELECT DISTINCT container_id FROM delivery_plan"
                    ).fetchall()
                )
                _con2.close()
                _new_rows  = [r for r in _imp_rows if r["container_id"] not in _existing_ids]
                _dup_count = len(_imp_rows) - len(_new_rows)

                _im1, _im2, _im3, _im4 = st.columns(4)
                _im1.metric("In file",          len(_imp_rows))
                _im2.metric("New (to import)",   len(_new_rows))
                _im3.metric("Already in DB",     _dup_count)
                _im4.metric("Parse warnings",    len(_imp_errors))

                if _imp_errors:
                    with st.expander(f"Parse warnings ({len(_imp_errors)})"):
                        for _e in _imp_errors[:30]:
                            st.caption(_e)

                if _new_rows:
                    _prev_df = pd.DataFrame(_new_rows)

                    _ps1, _ps2, _ps3 = st.columns(3)
                    with _ps1:
                        st.caption("By Site")
                        st.dataframe(
                            _prev_df.groupby("site_code").size().reset_index(name="containers"),
                            hide_index=True, use_container_width=True,
                        )
                    with _ps2:
                        st.caption("By Carrier")
                        st.dataframe(
                            _prev_df.groupby("carrier").size().reset_index(name="containers"),
                            hide_index=True, use_container_width=True,
                        )
                    with _ps3:
                        st.caption("By Status")
                        st.dataframe(
                            _prev_df.groupby("status").size().reset_index(name="containers"),
                            hide_index=True, use_container_width=True,
                        )

                    _dr = sorted(_prev_df["appt_date"].dropna().tolist())
                    if _dr:
                        st.caption(f"Date range: {_dr[0]} through {_dr[-1]}")

                    with st.expander("Preview first 25 rows"):
                        st.dataframe(
                            _prev_df[[
                                "container_id","carrier","site_code","appt_date",
                                "product_type","qty","status","notes",
                            ]].head(25),
                            hide_index=True, use_container_width=True,
                        )

                    if st.button(
                        f"Import {len(_new_rows)} containers", type="primary", key="imp_go"
                    ):
                        _now = datetime.now(_EASTERN).isoformat()
                        _icon = get_db()
                        _icon.executemany(
                            """INSERT INTO delivery_plan
                               (week_start,appt_date,appt_time,slot_num,container_id,carrier,
                                site_code,product_type,qty,notes,status,created_at,updated_at)
                               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                            [(r["week_start"],r["appt_date"],r["appt_time"],r["slot_num"],
                              r["container_id"],r["carrier"],r["site_code"],r["product_type"],
                              r["qty"],r["notes"],r["status"],_now,_now)
                             for r in _new_rows],
                        )
                        _icon.commit(); _icon.close()
                        if S3_ENABLED:
                            data_sync.push_db_to_s3(AWS_KEY, AWS_SECRET, AWS_REGION, S3_BUCKET)
                        st.success(
                            f"Imported {len(_new_rows)} containers. "
                            "WoW / History tab now has full historical data."
                        )
                        st.rerun()
                else:
                    st.info("All containers in this file are already in the database — nothing new to import.")

    # ══════════════════════════════════════════════════════════════════════════
    # SOP GUIDE
    # ══════════════════════════════════════════════════════════════════════════
    with _tp9:
        st.markdown("### Weekly Delivery Planning SOP")
        st.caption("Amazon Robotics Dray Program — AGL · Owner: Dominique Kennedy (kennewdo) · Every Friday · Deadline: 3 PM ET / 2 PM CT")

        _sop_step = st.radio(
            "Jump to step:",
            ["Step 1 — Collect DBRs", "Step 2 — Pull Data", "Step 3 — Identify Available",
             "Step 4 — Classify Product", "Step 5 — FIFO Priority", "Step 6 — Build Plan",
             "Step 7 — Validate", "Step 8 — Send to Carriers", "Step 9 — Mid-Week Adjustments",
             "GVT Status Glossary", "File Naming"],
            horizontal=True, key="sop_step_sel"
        )

        st.divider()

        if _sop_step == "Step 1 — Collect DBRs":
            st.markdown("#### Step 1: Collect Carrier DBRs (Thursday EOD)")
            st.info("Ensure all carrier DBRs are received by **3:00 PM CT Thursday**. These are the primary inputs for building next week's plan.")
            st.markdown("""
| Carrier | Contact | Expected Email Subject |
|---|---|---|
| ATMI | Tyler Domingues | `[EXTERNAL] Amazon Robotics - DBR [DATE]` |
| ARVY | Tyler Spangler | `[EXTERNAL] Robotics DBR [DATE]` |
| HUDD RIC6 | Sandji Ruffin | `[EXTERNAL] RIC6 Delivery Plan Update (HUDD)` |
| HUDD ILM1 | Jerry Nesbit | `[EXTERNAL] ILM1 Delivery Plan Update / HUDD` |
| HUDD LAX | Desirae Swain / Ailua Osoimalo | `RE: DBR Bridges Report - HUDD - [DATE]` |
""")
            st.markdown("**File naming convention:** `[CARRIER] DBR M.DD.YY.xlsx`  e.g. `ATMI DBR 7.17.26.xlsx`")
            st.markdown("**SharePoint location:** AGLRobotics → Carrier DBRs → subfolders: ATMI / ARVY / HUDD")
            st.caption("The automated DBR-saving schedule runs at 12:00 PM, 2:00 PM, and 3:30 PM CT on weekdays to save carrier DBRs to SharePoint.")

        elif _sop_step == "Step 2 — Pull Data":
            st.markdown("#### Step 2: Pull Supporting Data (Friday Morning)")
            st.markdown("Pull the following three reports to cross-reference container status and product classification:")
            with st.expander("📥 GVT Report (GM DCM Reports — Inbound Container Milestone)", expanded=True):
                st.markdown("""
- **Filters:** Customer = AMZ · Equip Category = AMAZON Robotics
- **Ocean ETA range:** 6–12 weeks back, 6–12 weeks forward
- **Key columns:** Container, Status, Dray SCAC, Discharged Port, Facility
- **Use to identify:** containers at yard (In Yard Full), incoming (On Water / Not Ready), dispatched
- **Save as:** `GVT Data WK##.xlsx`
""")
            with st.expander("📥 Inbound Loads Report (Amazon Robotics Inbound Loads)", expanded=True):
                st.markdown("""
- **Provides:** Container → PO Line Item Description (product type classification)
- **Critical for:** Identifying SHELF vs STRAP containers
- **Key columns:** Container Id, PO Line Item Description, Quantity, Location Code
""")
            with st.expander("📥 CDS Report (Import Shipment Status)", expanded=True):
                st.markdown("""
- **Provides:** Most accurate port arrival times
- **Use to identify:** containers arriving at port this/next week that could become available
""")

        elif _sop_step == "Step 3 — Identify Available":
            st.markdown("#### Step 3: Identify Available Containers Per Site")
            st.markdown("For each site, determine what containers are **at yard and ready to deliver** using the carrier DBR.")
            st.markdown("""
**Verification method (per carrier DBR):**

Go to the site-specific tab (e.g., "RIC6 Rockville" on ATMI's DBR) and check each container:

| Check | YES | NO |
|---|---|---|
| Has Actual Delivery to FC date? | ✅ Already delivered — **skip** | ➡️ Available for scheduling |
| Marked as rerouted to another site? | ✅ Exclude from this site | ➡️ Include |
| Has Actual Delivery to Yard date? | ✅ Confirmed at yard — schedule | ⚠️ May be in transit — cross-ref GVT |
""")
            st.info("Cross-reference GVT for containers showing **Not Ready** (at port / customs hold) — these are contingent. Add as PENDING if likely to clear before delivery week.")

        elif _sop_step == "Step 4 — Classify Product":
            st.markdown("#### Step 4: Classify Product Type")
            st.markdown("Using the **Inbound Loads Report**, classify each available container:")
            st.markdown("""
1. Match on **Container Id**
2. Read **PO Line Item Description:**
   - Contains `SHELF` → **SHELF, HDTP**
   - Contains `STRAP` → **STRAP, HDTP**
3. Note the **Quantity** (units) per container — include in the delivery plan
""")
            st.caption("In the Plan Builder, set the Product Type dropdown before adding containers. Bulk paste supports inline override.")

        elif _sop_step == "Step 5 — FIFO Priority":
            st.markdown("#### Step 5: Prioritize by FIFO (Dwell Time)")
            st.markdown("""
Sort all available containers by **dwell time at yard** — longest dwelling = highest priority:

1. Use **Actual Delivery to Yard** date from carrier DBR
2. Calculate: **Days Dwelling = Plan Date − Yard Arrival Date**
3. Longest-dwelling containers get scheduled **first**
4. Containers **refused/rejected** in prior weeks get priority — flagged in DBR with notes like `REJECTED BY [SITE] [DATE]`
""")
            st.info("The WoW / History tab shows containers rolled from the prior week — these are your FIFO candidates.")

        elif _sop_step == "Step 6 — Build Plan":
            st.markdown("#### Step 6: Build the Delivery Plan (Friday, in this app)")
            st.markdown("""
Use **Plan Builder** → Step 2 — Add Container IDs:
- **Single:** one container at a time with full field control
- **Bulk Paste:** `CONTAINER_ID  CARRIER  SITE` one per line — level-load checkbox distributes evenly

**Level loading logic:**
- App distributes containers across active receiving days for the week
- Disable Monday (RIC6 no-recv) in the **Receiving Days** toggles before bulk-pasting
- Saturday can be included via checkbox if site requests it
""")
            st.markdown("**RIC6 Constraints:**")
            st.markdown("""
| Parameter | Value |
|---|---|
| Receiving window | 7:30 AM – 4:30 PM |
| Lunch break | 12:00 PM – 1:00 PM (no deliveries) |
| Last arrival time | 3:30 PM (must be docked by 4:30) |
| Max loads/day | 11–12 |
| Stagger | ~1 per hour |
| No receiving day | **Monday** |
| HUDD preference | Deliver before lunch (7:30 AM – 11:30 AM) |
""")
            st.caption("All time slots in the app skip the 12–1 PM window automatically. Monday can be toggled off per week.")

        elif _sop_step == "Step 7 — Validate":
            st.markdown("#### Step 7: Review & Validate Before Sending")
            st.info("The **Pre-Send Validation Checklist** is available in the Carrier View tab before each carrier export.")
            st.markdown("""
☐ All containers verified as at-yard with no actual FC delivery date  
☐ Product type (Shelf/Strap) matches site's requested split  
☐ FIFO order respected (longest dwelling first)  
☐ No appointments during lunch break  
☐ Last appointment at or before 3:30 PM (or site-specific cutoff)  
☐ Level-loaded across days (+/− 1–2 container difference max)  
☐ HDDR/HUDD containers in early morning slots (before lunch)  
☐ Carrier tabs are clean and ready to copy/paste  
☐ Contingent containers clearly marked as PENDING with status note  
""")

        elif _sop_step == "Step 8 — Send to Carriers":
            st.markdown("#### Step 8: Send to Carriers")
            st.markdown("""
Copy/paste each carrier's schedule from the **Carrier View** tab and send via **Slack DM or email**.

**Include in message:**
- Delivery dates + appointment times per container
- Any contingent containers pending customs clearance (flagged PENDING)
- Site-specific constraints they need to know (lunch window, no-Monday at RIC6, etc.)

**Contacts:**
| Carrier | Contact | Channel |
|---|---|---|
| ATMI | Tyler Domingues — tdomingues@cargomatic.com | Slack + Email |
| ARVY | Tyler Spangler — tspangler@arrivelogistics.com | Slack + Email |
| HUDD RIC6 | Sandji Ruffin — sandji.ruffin@maersk.com | Email |
| HUDD ILM1 | Jerry Nesbit — jerry.nesbit@maersk.com | Email |
| HUDD LAX | Desirae Swain / Ailua Osoimalo — desirae.swain@maersk.com | Email |
""")
            st.caption("The By Site → Daily Notification expander generates tomorrow's delivery list formatted for Slack.")

        elif _sop_step == "Step 9 — Mid-Week Adjustments":
            st.markdown("#### Step 9: Mid-Week Adjustments")
            st.markdown("""
Replan as needed throughout the delivery week when any of these occur:

| Trigger | Action |
|---|---|
| Site stops receiving | Notify affected carriers to hold; reschedule to later in week |
| Container refused at site | Add note (REJECTED), reschedule to next available day; priority for next plan |
| Carrier requests change | Accommodate within site constraints |
| Contingent container clears customs | Slot into plan; update status PENDING → SCHEDULED |

**Mid-week replan steps:**
1. Check updated carrier DBR for which containers didn't deliver today
2. Identify which carriers had deliveries that day
3. Use **By Site → Mid-Week Adjustment** expander to draft hold/reschedule messages
4. Update delivery plan — mark delivered as DELIVERED, rejected as REJECTED
5. Reschedule remaining containers to later days in the week
""")
            st.caption("The Mid-Week Adjustment tool is in the By Site tab — select the site, choose the scenario, and it drafts the carrier message.")

        elif _sop_step == "GVT Status Glossary":
            st.markdown("#### GVT Container Status Meanings")
            st.markdown("""
| Status | Meaning | Action |
|---|---|---|
| **In Yard Full** | Container at dray yard — ready to deliver | ✅ Schedule |
| **Dispatched to Destination** | En route to FC (or refused / returning to yard) | 🔍 Check carrier DBR for delivery outcome |
| **Not Ready** | At port — pending customs clearance or carrier release | ⏳ Add as PENDING if expected to clear |
| **On Water** | Still at sea — not yet at port | 📅 Note ETA; not schedulable this week |
| **Closed** | Delivered and empty returned | ✅ Complete — no action needed |
""")
            st.info("When carrier DBRs show containers as 'at yard' that seem old, always verify against the **Actual FC Delivery Date** on the site tab — DBRs can be stale.")

        elif _sop_step == "File Naming":
            st.markdown("#### File Naming Conventions")
            st.markdown("""
| File | Convention | Example |
|---|---|---|
| Carrier DBR | `[CARRIER] DBR M.DD.YY.xlsx` | `ATMI DBR 7.17.26.xlsx` |
| HUDD site-specific | `HUDD [SITE] DBR M.DD.YY.xlsx` | `HUDD RIC6 DBR 7.17.26.xlsx` |
| Delivery Plan | `[SITE] Delivery Plan M.DD-M.DD.xlsx` | `RIC6 Delivery Plan 7.21-7.25.xlsx` |
| GVT Export | `GVT Data WK##.xlsx` | `GVT Data WK30.xlsx` |
| Inbound Loads | Standard report name | Keep as-is |
""")
            st.caption("App exports automatically follow the Delivery Plan convention.")



# ══════════════════════════════════════════════════════════════════════════════
# TAB 9 — WBR (Dashboard + Generator merged)
# ─────────────────────────────────────────────────────────────────────────────

with tab9:
    _wbr_sum, _wbr_trends, _wbr_build, _wbr_hist = st.tabs([
        "Summary", "Trends", "Build", "History"
    ])

    with _wbr_build:
        st.subheader("WBR Slide Generator")
        st.caption("Upload all three source files to generate both WBR deliverables in one click.")

        from wbr_engine import (
            load_gvt, load_oblt, load_inbound_loads,
            compute_metrics, compute_totals, compute_carrier_scorecard,
            week_bounds, guess_week,
            save_week_to_db, load_weeks_from_db, parse_wbr_pdf,
            load_market_context, format_market_context_block,
            _latest,
        )
        from wbr_pdf import generate_standard_wbr, _fmt
        from wbr_pptx import pdf_to_pptx
        import urllib.parse

        # ── Week identification + due date ─────────────────────────────────────────
        # Anchor to LAST Monday — the current WBR always covers the week that just
        # closed (Sun–Sat ending the prior Saturday). Flips to the next week only
        # when the new submission Monday arrives.
        _wbr_today2   = datetime.now(_EASTERN).date()
        _dsm2         = _wbr_today2.weekday()              # Mon=0, Sun=6
        _last_mon2    = _wbr_today2 - timedelta(days=_dsm2)  # most recent Monday (today if Mon)
        _wbr_sat      = _last_mon2  - timedelta(days=2)    # week-end Saturday (prior Sat)
        _wbr_sun      = _wbr_sat    - timedelta(days=6)    # week-start Sunday
        _wbr_wnum     = int(_wbr_sat.strftime("%V"))
        _next_mon2    = _last_mon2  + timedelta(days=7)    # next submission Monday
        _wbr_sun_str  = _wbr_sun.strftime("%m/%d/%Y")
        _wbr_sat_str  = _wbr_sat.strftime("%m/%d/%Y")
        _wbr_mon_str  = _last_mon2.strftime("%m/%d/%Y")    # submission Monday for current WBR
        _wbr_sun_nice = _wbr_sun.strftime("%b %d")
        _wbr_sat_nice = _wbr_sat.strftime("%b %d, %Y")

        # Due date banner — check DB first; if WBR already generated this week, show submitted state
        _wbr_chk_conn = get_db()
        _wbr_submitted_row = _wbr_chk_conn.execute(
            "SELECT generated_at FROM wbr_results "
            "WHERE year=? AND week_num=? AND generated_at != 'seed' "
            "ORDER BY generated_at DESC LIMIT 1",
            (_wbr_today2.year, _wbr_wnum)
        ).fetchone()
        _wbr_chk_conn.close()

        if _wbr_submitted_row:
            _gen_ts = _wbr_submitted_row[0]
            try:
                _gen_dt = datetime.fromisoformat(_gen_ts).astimezone(_EASTERN)
                _gen_nice = _gen_dt.strftime("%b %d at %I:%M %p ET").lstrip("0")
            except Exception:
                _gen_nice = _gen_ts[:16]
            st.success(f"✅ **WBR W{_wbr_wnum} submitted** — generated {_gen_nice} · Next WBR **W{_wbr_wnum+1}** due **{_next_mon2.strftime('%b %d')}**")
        elif _dsm2 == 0:
            st.error(f"🔴 **WBR W{_wbr_wnum} is DUE TODAY** by 2:00 PM CT ({_ct_to_mst(14)} MST) — submit to `doc+destops-36@fusion.amazon.dev`")
        elif _dsm2 <= 2:
            st.warning(f"🟡 **WBR W{_wbr_wnum}** was due **{_last_mon2.strftime('%b %d')}** — submit ASAP if not yet sent · Next WBR W{_wbr_wnum+1} due {_next_mon2.strftime('%b %d')}")
        else:
            st.info(f"**WBR W{_wbr_wnum}** was due {_last_mon2.strftime('%b %d')} · Next WBR **W{_wbr_wnum+1}** due **{_next_mon2.strftime('%A, %b %d')}** — {(_next_mon2 - _wbr_today2).days} days away")

        st.caption(f"Reporting week: **W{_wbr_wnum}** · {_wbr_sun_str} to {_wbr_sat_str} (Sun–Sat)")

        # ── Context notes ─────────────────────────────────────────────────────────
        # Notes saved per-week to DB; auto-injected into Perjen bridge at generate time.
        # Default week: Thu–Sun → next WBR week; Mon–Wed → current WBR week
        _cn_default_wk = _wbr_wnum + 1 if _dsm2 >= 3 else _wbr_wnum  # Thu=3..Sun=6 → next wk
        _cn_year       = _wbr_today2.year
        with st.expander(f"📝 Context Notes — W{_cn_default_wk} (injected into bridge at generate time)", expanded=False):
            _cn_wk_input = st.number_input(
                "Week # for notes", min_value=1, max_value=53,
                value=_cn_default_wk, key="cn_week_input",
            )
            _cn_load_conn = get_db()
            _cn_saved = _cn_load_conn.execute(
                "SELECT notes FROM wbr_context_notes WHERE year=? AND week_num=?",
                (_cn_year, _cn_wk_input)
            ).fetchone()
            _cn_load_conn.close()
            _cn_existing = _cn_saved["notes"] if _cn_saved and _cn_saved["notes"] else ""
            _cn_text = st.text_area(
                "Operational context, wins, misses, callouts:",
                value=_cn_existing,
                height=100,
                placeholder="e.g. RIC6 shutdown Fri 7/17, 9 containers rerouted to NCKR; ATMI last DBM6 delivered",
                key="cn_notes_text",
            )
            if st.button("Save Context Notes", key="cn_save_btn", type="secondary"):
                _cn_save_conn = get_db()
                _cn_save_conn.execute(
                    "INSERT INTO wbr_context_notes (year, week_num, notes, updated_at) VALUES (?,?,?,?) "
                    "ON CONFLICT(year, week_num) DO UPDATE SET notes=excluded.notes, updated_at=excluded.updated_at",
                    (_cn_year, _cn_wk_input, _cn_text.strip(), datetime.now(_EASTERN).isoformat())
                )
                _cn_save_conn.commit()
                _cn_save_conn.close()
                if S3_ENABLED:
                    data_sync.push_db_to_s3(AWS_KEY, AWS_SECRET, AWS_REGION, S3_BUCKET)
                st.success(f"Notes saved for W{_cn_wk_input}.")

        # ── Slide preview helper (called pre-generate and post-generate) ───────────
        def _wbr_slide_preview(has_gvt, has_oblt, has_il, has_iss=False, curr=None, prior=None):
            n_ready = sum([has_gvt, has_oblt, has_il])  # ISS is optional — doesn't gate generate
            overall_op = 0.12 + 0.88 * (n_ready / 3)

            def _seg(label, done, optional=False):
                bg = "#22c55e" if done else "#1e293b"
                fg = "#000" if done else "#4b5563"
                lbl = f"{label}{'*' if optional else ''}"
                return f'<span style="background:{bg};color:{fg};padding:2px 7px;border-radius:3px;font-size:10px;margin:0 2px;">{lbl}</span>'

            segs = _seg("GVT", has_gvt) + _seg("OBLT", has_oblt) + _seg("IL", has_il) + _seg("ISS", has_iss, optional=True)

            def _box(label, key, need_gvt, need_oblt, need_il, suffix="", sla=False, thresh=85):
                req = (not need_gvt or has_gvt) and (not need_oblt or has_oblt) and (not need_il or has_il)
                if curr and curr.get(key) is not None:
                    v = curr[key]
                    v_str = f"{v}{'%' if sla else suffix}"
                    clr = ("#22c55e" if v >= thresh else "#ef4444") if sla else "#60a5fa"
                    op = 1.0; note = f"✅ W{_wbr_wnum}"
                elif req and prior and prior.get(key) is not None:
                    v = prior[key]; wn = prior.get("week_num","?")
                    v_str = f"{v}{'%' if sla else suffix}"
                    clr = "#f59e0b"; op = 0.65; note = f"↖ W{wn} (prior — upload files)"
                elif prior and prior.get(key) is not None:
                    v = prior[key]; wn = prior.get("week_num","?")
                    v_str = f"{v}{'%' if sla else suffix}"
                    clr = "#374151"; op = 0.22; note = f"W{wn} placeholder"
                else:
                    v_str = "—"; clr = "#1f2937"; op = 0.18; note = "no data yet"
                border = "#1d4ed8" if op >= 1.0 else ("#92400e" if op >= 0.5 else "#1f2937")
                return f"""<div style="background:#111827;border:1px solid {border};border-radius:6px;
                            padding:10px 12px;opacity:{op};transition:opacity 0.5s ease;">
                    <div style="color:#6b7280;font-size:9px;text-transform:uppercase;letter-spacing:.5px;margin-bottom:3px;">{label}</div>
                    <div style="color:{clr};font-size:22px;font-weight:700;line-height:1.1;">{v_str}</div>
                    <div style="color:#374151;font-size:9px;margin-top:4px;">{note}</div>
                </div>"""

            boxes = [
                _box("Containers",   "containers",    True,  False, False),
                _box("AV→OA SLA",   "av_oa_sla_pct", True,  True,  False, sla=True),
                _box("AV→OA Avg",   "av_oa_avg",     True,  True,  False, suffix="d"),
                _box("OA→Del SLA",  "oa_del_sla_pct",False, True,  False, sla=True),
                _box("OA→Del Avg",  "oa_del_avg",    False, True,  False, suffix="d"),
                _box("OTP",         "otp_pct",       False, False, True,  sla=True),
            ]

            status_txt = "✅ Ready — click Generate" if n_ready == 3 else f"⏳ {3-n_ready} file(s) remaining"
            status_clr = "#22c55e" if n_ready == 3 else "#6b7280"
            chart_note  = "✅ Charts ready in PDF" if curr else "📈 Charts rendered in PDF after Generate"
            chart_clr   = "#22c55e" if curr else "#374151"

            return f"""<div style="font-family:system-ui,sans-serif;background:#0d1117;border-radius:10px;
                        padding:16px;border:1px solid #21262d;margin:6px 0;">
                <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px;">
                    <div>
                        <span style="color:#f97316;font-weight:700;font-size:13px;">📊 WBR W{_wbr_wnum} Slide Preview</span>
                        <span style="color:#4b5563;font-size:11px;margin-left:8px;">· {_wbr_sun_nice} – {_wbr_sat_nice}</span>
                    </div>
                    <div style="display:flex;align-items:center;gap:6px;">{segs}
                        <span style="color:{status_clr};font-size:10px;margin-left:4px;">{status_txt}</span>
                    </div>
                </div>
                <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin-bottom:10px;">
                    {"".join(boxes)}
                </div>
                <div style="background:#111827;border-radius:5px;padding:7px 12px;display:flex;
                            justify-content:space-between;align-items:center;opacity:{overall_op:.2f};">
                    <span style="color:#374151;font-size:10px;">AV→OA trend line · OA→Del trend line · Table W{_wbr_wnum-5}–W{_wbr_wnum}</span>
                    <span style="color:{chart_clr};font-size:10px;">{chart_note}</span>
                </div>
            </div>"""

        # ── File uploaders ────────────────────────────────────────────────────────
        wu1, wu2, wu3, wu4 = st.columns(4)
        with wu1:
            with st.expander(f"ℹ️ How to pull GVT — W{_wbr_wnum} ({_wbr_sun_str}–{_wbr_sat_str})"):
                st.markdown(f"""**Global Visibility Tool (GVT) — W{_wbr_wnum} pull**
    - Navigate to GVT → Container Search
    - Filter **Ready Date/Time**: `{_wbr_sun_str}` to `{_wbr_sat_str}` *(W{_wbr_wnum}, Sun–Sat)*
    - Filter: Robotics BU · markets: BOS, ORF, SAV, LAX (all)
    - Export to Excel → save as `GVT Data WK{_wbr_wnum}.xlsx`
    - **Verify:** `Ready Date/Time` and `Container` columns are populated
    - ⚠️ If `Ready Date/Time` is mostly blank, the weekly cut hasn't processed yet — re-pull after EOD {_wbr_sat_str}""")
            wbr_gvt_file = st.file_uploader("GVT Data (.xlsx)", type=["xlsx"], key="wbr_gvt")

        with wu2:
            with st.expander(f"ℹ️ How to pull OBLT — W{_wbr_wnum}"):
                st.markdown(f"""**OBLT (Ocean Bridge Logistics Tracking) — W{_wbr_wnum} pull**
    - Navigate to OBLT → Export
    - Pull containers with **AV date between `{_wbr_sun_str}` and `{_wbr_sat_str}`** (W{_wbr_wnum})
    - Required sheet: `ObltData`
    - Key columns: `tracking_id`, `status` (AV/OA/VD/RD), `status_date`
    - Status key: AV=Available · OA=Outgate/Pickup · VD=Vessel Departed · RD=Received
    - Save as `OBLT WK{_wbr_wnum}.xlsx`""")
            wbr_oblt_file = st.file_uploader("OBLT Data (.xlsx)", type=["xlsx"], key="wbr_oblt")

        with wu3:
            with st.expander(f"ℹ️ How to pull Inbound Loads — W{_wbr_wnum} ({_wbr_mon_str})"):
                st.markdown(f"""**Amazon Robotics Inbound Loads Report — W{_wbr_wnum} pull**
    - Run report **Monday morning {_wbr_mon_str}** before generating the slide
    - Go to: Reporting Portal → Inbound Loads → Amazon Robotics
    - Required sheet: `Inbound Loads`
    - Key columns: `Container Id`, `PO Promised Date`, `Actual Arrival At Final Destination`
    - Expected filename: `Amazon Robotics Inbound Loads Report{_next_mon2.strftime("%d-%b-%Y")} ######.xlsx`
    - ⚠️ Actual Arrival blank for undelivered containers — expected. OTP scored on delivered only.
    - **Supplemental:** DBR Tracker (SharePoint → AGLRobotics) fills gaps for GF/BF deliveries""")
            wbr_il_file = st.file_uploader("Inbound Loads (.xlsx)", type=["xlsx"], key="wbr_il")

        with wu4:
            with st.expander(f"ℹ️ How to pull Import Shipment Status — W{_wbr_wnum} ({_wbr_mon_str})"):
                st.markdown(f"""**Import Shipment Status — W{_wbr_wnum} pull**
    - Go to: Reporting Portal → Import Shipment Status → Amazon Robotics
    - Run as of **Monday morning {_wbr_mon_str}** (same time as Inbound Loads)
    - Download the Excel export — no filters needed, pull all active Robotics containers
    - Key columns used: `ContainerNo`, `EDIContainerAvailable`, `EDIContainerOut`, `EDIDelivered`, `VesselName`, `VoyageNo`, `FinalDestETA`
    - **Optional** — Generate works without this file, but the Enhanced WBR forward look will be richer with it
    - Provides: EDI-sourced tracking dates to fill OBLT gaps + vessel pipeline for next 2 weeks""")
            wbr_iss_file = st.file_uploader("Import Shipment Status (.xlsx)", type=["xlsx"], key="wbr_iss",
                                             help="Optional — enriches Enhanced WBR forward look with vessel/ETA data")

        st.markdown("---")
        wbr_prior_pdf = None  # Prior weeks auto-loaded from DB


        # ── Report date ───────────────────────────────────────────────────────────
        # Auto-default to most recent Monday — WBR always submitted Monday morning
        _wbr_today = datetime.now(_EASTERN).date()
        _last_monday = _wbr_today - timedelta(days=_wbr_today.weekday())  # weekday(): Mon=0
        wbr_report_date = _last_monday

        # ── Slide preview (live, pre-generate) ────────────────────────────────────
        _pconn = get_db()
        _prow  = _pconn.execute(
            "SELECT week_num,containers,av_oa_avg,av_oa_sla_pct,oa_del_avg,oa_del_sla_pct,"
            "empty_term_pct,e2e_avg,otp_pct FROM wbr_results "
            "WHERE year=? ORDER BY week_num DESC LIMIT 1",
            (_wbr_today2.year,)
        ).fetchone()
        _pconn.close()
        _prior_preview = dict(_prow) if _prow else None
        _curr_preview  = st.session_state.get("wbr_preview_metrics")
        st.markdown(
            _wbr_slide_preview(
                has_gvt=wbr_gvt_file is not None,
                has_oblt=wbr_oblt_file is not None,
                has_il=wbr_il_file is not None,
                has_iss=wbr_iss_file is not None,
                curr=_curr_preview,
                prior=_prior_preview,
            ),
            unsafe_allow_html=True,
        )

        # ── Main block ────────────────────────────────────────────────────────────
        if wbr_gvt_file and wbr_oblt_file and wbr_il_file:
            gvt_bytes  = wbr_gvt_file.read()
            oblt_bytes = wbr_oblt_file.read()
            il_bytes   = wbr_il_file.read()
            iss_bytes  = wbr_iss_file.read() if wbr_iss_file else None

            with st.spinner("Parsing files and computing metrics..."):
                try:
                    gvt_df     = load_gvt(gvt_bytes)
                    oblt_df    = load_oblt(oblt_bytes)
                    inbound_df = load_inbound_loads(il_bytes)
                    iss_df     = parse_shipment_status_file(iss_bytes) if iss_bytes else pd.DataFrame()

                    week_num, year = guess_week(gvt_df["ready_date"], wbr_report_date)
                    if week_num is None:
                        st.error("❌ GVT file has no parseable Ready Date/Time values. "
                                 "Re-pull GVT and ensure the Ready Date column is populated before continuing.")
                        st.stop()
                    wk_start, wk_end = week_bounds(week_num, year)
                    st.info(f"Detected: **W{week_num}** ({wk_start} to {wk_end})")

                    # ── Data Preview ──────────────────────────────────────────────
                    with st.expander("📋 Data Preview — inspect raw files before generating", expanded=False):
                        pv1, pv2, pv3, pv4 = st.tabs(["GVT", "OBLT", "Inbound Loads", "Import Shipment Status"])
                        pop_gvt = gvt_df[
                            gvt_df["ready_date"].notna() &
                            (gvt_df["ready_date"] >= wk_start) &
                            (gvt_df["ready_date"] <= wk_end)
                        ].copy()
                        _pop_ctrs = set(pop_gvt["container"])

                        with pv1:
                            st.caption(f"GVT — {len(gvt_df):,} total rows | **{len(pop_gvt)} containers in W{week_num}**")
                            _gvt_cols = [c for c in ["container","ready_date","facility","market","port","scac","status","container_empty","return_to_port","enter_facility"] if c in pop_gvt.columns]
                            st.dataframe(pop_gvt[_gvt_cols].sort_values("ready_date").reset_index(drop=True), use_container_width=True, height=220)
                            _rdates = gvt_df["ready_date"].dropna()
                            if len(_rdates):
                                st.code(f"Ready Date range: {_rdates.min()} to {_rdates.max()}")
                            if "facility" in pop_gvt.columns:
                                st.caption("Containers by facility:")
                                st.dataframe(pop_gvt["facility"].value_counts().reset_index().rename(columns={"index":"facility","facility":"count","count":"containers"}), hide_index=True, use_container_width=True)

                        with pv2:
                            oblt_pop = oblt_df[oblt_df["container"].isin(_pop_ctrs)]
                            st.caption(f"OBLT — {len(oblt_df):,} total rows | **{len(oblt_pop)} events for W{week_num} containers** ({oblt_pop['container'].nunique()} unique)")
                            st.caption("Status distribution (full file):")
                            st.dataframe(oblt_df["status"].value_counts().reset_index().rename(columns={"index":"status","status":"count","count":"events"}), hide_index=True, use_container_width=True)
                            st.caption(f"W{week_num} OBLT events:")
                            _oblt_cols = [c for c in ["container","status","date"] if c in oblt_pop.columns]
                            st.dataframe(oblt_pop[_oblt_cols].sort_values(["container","date"]).reset_index(drop=True), use_container_width=True, height=220)

                        with pv3:
                            il_pop = inbound_df[inbound_df["container"].isin(_pop_ctrs)]
                            st.caption(f"Inbound Loads — {len(inbound_df):,} total rows | **{len(il_pop)} rows matched to W{week_num}**")
                            _il_cols = [c for c in ["container","po_promised","actual_arrival"] if c in il_pop.columns]
                            if _il_cols:
                                st.dataframe(il_pop[_il_cols].reset_index(drop=True), use_container_width=True, height=220)
                            _has_both = (il_pop["po_promised"].notna() & il_pop["actual_arrival"].notna()).sum() if len(il_pop) else 0
                            st.code(f"With PO Promised: {il_pop['po_promised'].notna().sum() if len(il_pop) else 0}   With Actual Arrival: {il_pop['actual_arrival'].notna().sum() if len(il_pop) else 0}   Both: {_has_both}")

                        with pv4:
                            if iss_df.empty:
                                st.info("Upload Import Shipment Status file (4th uploader) to preview EDI data.")
                            else:
                                iss_pop = iss_df[iss_df["container_no"].isin(_pop_ctrs)] if "container_no" in iss_df.columns else pd.DataFrame()
                                st.caption(f"Import Shipment Status — {len(iss_df):,} total rows | **{len(iss_pop)} matched to W{week_num}**")
                                _iss_cols = [c for c in ["container_no","current_status","vessel_name","voyage_no","final_dest_eta","container_available","container_out","delivered","fc"] if c in iss_df.columns]
                                if len(iss_pop):
                                    st.dataframe(iss_pop[_iss_cols].reset_index(drop=True), use_container_width=True, height=220)
                                else:
                                    st.warning("No W{} containers matched in ISS. Verify you uploaded the right file.".format(week_num))
                                    st.dataframe(iss_df[_iss_cols].head(20), use_container_width=True, height=220)

                    # ── Compute metrics ───────────────────────────────────────────
                    curr_metrics = compute_metrics(
                        gvt_df, oblt_df, inbound_df,
                        wk_start, wk_end, wbr_report_date,
                    )
                    carrier_sc = compute_carrier_scorecard(
                        gvt_df, oblt_df, wk_start, wk_end, wbr_report_date
                    )

                    prior_nums = list(range(week_num - 5, week_num))
                    wbr_conn   = get_db()
                    prior_data = load_weeks_from_db(wbr_conn, year, prior_nums)

                    wbr_conn.close()

                    display_nums   = prior_nums + [week_num]
                    display_labels = [f"W{n}" for n in display_nums]
                    display_data   = [prior_data.get(n) for n in prior_nums] + [curr_metrics]
                    totals = compute_totals([d for d in display_data if d])

                    # Store in session state so slide preview lights up immediately
                    st.session_state["wbr_preview_metrics"] = curr_metrics

                    # ── Metrics tiles ─────────────────────────────────────────────
                    st.markdown("#### W{} — Computed Metrics".format(week_num))
                    mc1, mc2, mc3, mc4, mc5 = st.columns(5)
                    mc1.metric("Containers",    curr_metrics.get("containers", "–"))
                    mc2.metric("AV→OA SLA%",    "{}%".format(curr_metrics.get("av_oa_sla_pct", "–")))
                    mc3.metric("OA→Del SLA%",   "{}%".format(curr_metrics.get("oa_del_sla_pct", "–")))
                    mc4.metric("E2E Avg",        curr_metrics.get("e2e_avg", "–"))
                    mc5.metric("OTP%",           "{}%".format(curr_metrics.get("otp_pct", "–")))

                    # ── Data quality check ────────────────────────────────────────
                    with st.expander("🔍 Data Quality & Validation", expanded=False):
                        n_pop = len(_pop_ctrs)
                        n_ctr = curr_metrics.get("containers", 0) or 0
                        if n_ctr == 0:
                            st.error("Zero containers — verify GVT file is the correct week.")
                        elif n_ctr < 20:
                            st.warning(f"Low count ({n_ctr}). Typical weeks are 30–100+.")
                        else:
                            st.success(f"Container count: {n_ctr}")

                        oblt_set = set(oblt_df["container"])
                        oblt_cov = len(_pop_ctrs & oblt_set) / n_pop * 100 if n_pop else 0
                        if oblt_cov < 60:
                            st.warning(f"OBLT coverage low ({oblt_cov:.0f}%). AV→OA and OA→Del may be understated.")
                        else:
                            st.success(f"OBLT coverage: {oblt_cov:.0f}%")

                        oa_set = set(oblt_df[oblt_df["status"] == "OA"]["container"]) & _pop_ctrs
                        rd_set = set(oblt_df[oblt_df["status"] == "RD"]["container"]) & _pop_ctrs
                        in_tr  = len(oa_set - rd_set)
                        if in_tr > 0:
                            st.info(f"{in_tr} container(s) have OA but no RD — in-transit. Elapsed time included in OA→Del avg.")

                        il_m = inbound_df[inbound_df["container"].isin(_pop_ctrs)].dropna(subset=["po_promised","actual_arrival"])
                        otp_m = len(il_m) / n_pop * 100 if n_pop else 0
                        if otp_m < 40:
                            st.warning(f"OTP match rate: only {otp_m:.0f}%. Verify Inbound Loads covers the right containers.")
                        else:
                            st.success(f"OTP match rate: {otp_m:.0f}% ({len(il_m)}/{n_pop})")

                        st.markdown(f"**Week detected:** W{week_num} | {wk_start} (Sun) → {wk_end} (Sat)")
                        st.info("OA→Del excludes A320-RBTCS (ORF, non-operational). BOS-market containers scored.")

                        # ── Additional cross-checks ────────────────────────────────
                        st.markdown("**Cross-file checks**")

                        # Check 1: GVT week vs report_date mismatch
                        _rd = wbr_report_date
                        _rd_week_sat = _rd - timedelta(days=(_rd.weekday() + 2) % 7)  # prior Saturday
                        _expected_sat = wk_end  # week_end from detected GVT week
                        _date_delta = abs((_expected_sat - _rd_week_sat).days)
                        if _date_delta > 7:
                            st.warning(
                                f"⚠️ Report date ({_rd}) and GVT week end ({wk_end}) are {_date_delta} days apart. "
                                "Verify you're uploading the right week's GVT file and the report date is set correctly."
                            )
                        else:
                            st.success(f"Report date vs GVT week: aligned ({wk_end} end, {_rd} submission)")

                        # Check 2: Duplicate container IDs within GVT population
                        _dup_ctrs = pop_gvt[pop_gvt.duplicated("container", keep=False)]
                        if len(_dup_ctrs):
                            _dup_list = _dup_ctrs["container"].unique().tolist()
                            st.warning(
                                f"⚠️ {len(_dup_list)} container ID(s) appear more than once in GVT for W{week_num}. "
                                f"Duplicate(s): {', '.join(_dup_list[:5])}{'...' if len(_dup_list) > 5 else ''}. "
                                "Re-pull GVT — this inflates the container count."
                            )
                        else:
                            st.success(f"No duplicate container IDs in GVT population")

                        # Check 3a: ISS coverage (if provided)
                        if not iss_df.empty and "container_no" in iss_df.columns:
                            iss_pop_v = iss_df[iss_df["container_no"].isin(_pop_ctrs)]
                            iss_cov = len(iss_pop_v) / n_pop * 100 if n_pop else 0
                            if iss_cov < 30:
                                st.warning(f"⚠️ ISS only matches {iss_cov:.0f}% of W{week_num} containers. Verify you uploaded the correct Import Shipment Status file.")
                            else:
                                st.success(f"ISS coverage: {iss_cov:.0f}% ({len(iss_pop_v)}/{n_pop} containers)")
                        else:
                            st.info("ℹ️ Import Shipment Status not uploaded — forward look will show placeholder data. Upload the 4th file to enrich the Enhanced WBR.")

                        # Check 3b: OBLT has any AV timestamps AFTER OA timestamps (data quality)
                        _av_ts = dict(_latest(oblt_df, "AV"))
                        _oa_ts = dict(_latest(oblt_df, "OA"))
                        _bad_ts = [(c, _av_ts[c], _oa_ts[c]) for c in set(_av_ts) & set(_oa_ts)
                                   if _oa_ts[c] < _av_ts[c]]
                        if _bad_ts:
                            st.warning(
                                f"⚠️ {len(_bad_ts)} OBLT row(s) have OA timestamp earlier than AV (bad source data). "
                                "These are excluded from AV→OA averages automatically but counted toward denominator."
                            )
                        else:
                            st.success("OBLT timestamps: no AV/OA inversions detected")

                        st.markdown("**Operational context — added directly to bridge**")
                        _nc1, _nc2 = st.columns(2)
                        with _nc1:
                            st.text_area(
                                "AV→OA context",
                                placeholder="e.g. ATMI ORF: 7 containers picked up day 4 — capacity stretched\nATMI SAV: 10 containers delayed — Tighe receiver constraint\nSite constraint, not carrier execution failure.",
                                key="wbr_op_avoa", height=100,
                            )
                            st.text_area(
                                "OA→Del context",
                                placeholder="e.g. RIC6 not receiving 7/17–7/18 — 9 containers held at ODY\nILM1 resumed receiving 7/19",
                                key="wbr_op_oadel", height=68,
                            )
                        with _nc2:
                            st.text_area(
                                "Path to Green rows (one per line: Action | Impact | Timeline)",
                                placeholder="e.g. ATMI ORF coaching — 1-day miss, capacity adj needed | ORF misses eliminated | WK{wk_next}\nMonitor Tighe capacity with ATMI | SAV pickups resume | TBD",
                                key="wbr_op_ptg", height=100,
                            )
                            st.text_area(
                                "E2E / OTP context",
                                placeholder="e.g. OTP miss driven by ocean delay — not dray-controllable",
                                key="wbr_op_e2e", height=68,
                            )
                        # Legacy key kept for backward compat (not shown; used only if new keys are empty)
                        st.session_state.setdefault("wbr_op_notes", "")

                        st.markdown("---")
                        st.text_area(
                            "🌐 External context / market intel",
                            placeholder="Port congestion, rate spikes, weather events, carrier advisories, vessel delays, ILA activity. Injected into bridge under [External Context]. e.g.:\nPort of LA congestion — avg 3-day vessel delay as of 8/1\nBNSF dwell elevated at Chicago ramp — expect +1-2 days inland transit",
                            key="wbr_ext_ctx", height=90,
                            help="Freeform external/market context injected into the bridge.",
                        )

                    # ── Prior weeks missing ────────────────────────────────────────
                    missing = [f"W{n}" for n in prior_nums if n not in prior_data]
                    if missing:
                        st.warning(
                            f"**{len(missing)} prior week(s) still missing** ({', '.join(missing)}). "
                            "Enter values manually below — or they will be shown as `–` on the slide."
                        )
                        with st.expander(f"✏️ Manual entry — {', '.join(missing)}", expanded=False):
                            st.caption("Fallback: copy values from your last published WBR slide.")
                            manual_inputs = {}
                            for mn in missing:
                                wn = int(mn[1:])
                                st.markdown(f"**{mn}**")
                                mi1,mi2,mi3,mi4,mi5,mi6,mi7,mi8,mi9,mi10 = st.columns(10)
                                manual_inputs[wn] = {
                                    "containers":     mi1.number_input("Vol",    key=f"m_{mn}_vol",   min_value=0, value=0),
                                    "av_oa_avg":      mi2.number_input("AO avg", key=f"m_{mn}_aoavg", min_value=0.0, step=0.1),
                                    "av_oa_sla_pct":  mi3.number_input("AO%",    key=f"m_{mn}_aopct", min_value=0, max_value=100),
                                    "av_oa_p90":      mi4.number_input("AO p90", key=f"m_{mn}_aop90", min_value=0),
                                    "oa_del_avg":     mi5.number_input("OD avg", key=f"m_{mn}_odavg", min_value=0.0, step=0.1),
                                    "oa_del_sla_pct": mi6.number_input("OD%",    key=f"m_{mn}_odpct", min_value=0, max_value=100),
                                    "oa_del_p90":     mi7.number_input("OD p90", key=f"m_{mn}_odp90", min_value=0),
                                    "empty_term_pct": mi8.number_input("ET%",    key=f"m_{mn}_etpct", min_value=0, max_value=100),
                                    "e2e_avg":        mi9.number_input("E2E",    key=f"m_{mn}_e2e",   min_value=0),
                                    "otp_pct":        mi10.number_input("OTP%",  key=f"m_{mn}_otp",   min_value=0, max_value=100),
                                    "week_start": None, "week_end": None,
                                }
                            if st.button("Apply manual values", key="wbr_apply_manual"):
                                for wn, mv in manual_inputs.items():
                                    display_data[prior_nums.index(wn)] = mv if mv["containers"] > 0 else None
                                st.rerun()

                    st.markdown("---")

                    # ── Generate ──────────────────────────────────────────────────
                    wbr_gen_btn = st.button("⚙️ Generate Both WBR Outputs", type="primary", key="wbr_generate")

                    if wbr_gen_btn:
                        with st.spinner("Building WBR outputs..."):
                            # Auto-compute footer time at generation moment
                            _now_phx = datetime.now(_PHX)
                            wbr_time_str = (
                                _now_phx.astimezone(_CENTRAL).strftime("%I:%M %p").lstrip("0") + " CT"
                                "  |  " + _now_phx.strftime("%I:%M %p").lstrip("0") + " MST"
                                "  |  " + _now_phx.astimezone(ZoneInfo("America/Los_Angeles")).strftime("%I:%M %p").lstrip("0") + " PT"
                            )
                            pdf_bytes = generate_standard_wbr(
                                week_labels = display_labels,
                                weeks_data  = display_data,
                                totals      = totals,
                                report_date = wbr_report_date,
                                report_time = wbr_time_str,
                            )
                            wbr_conn2 = get_db()
                            save_week_to_db(wbr_conn2, year, week_num, curr_metrics, datetime.now(_EASTERN).isoformat())
                            # Save per-carrier scorecard so WBR Dashboard can filter by carrier
                            _ts_gen = datetime.now(_EASTERN).isoformat()
                            for _sc_row in (carrier_sc or []):
                                try:
                                    wbr_conn2.execute(
                                        "INSERT OR REPLACE INTO wbr_carrier_results "
                                        "(year,week_num,carrier,volume,av_oa_sla_pct,av_oa_avg,"
                                        "av_oa_misses,oa_del_sla_pct,oa_del_avg,oa_del_misses,generated_at) "
                                        "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                                        (year, week_num,
                                         _sc_row.get("carrier"),
                                         _sc_row.get("volume"),
                                         _sc_row.get("av_oa_sla_pct"),
                                         _sc_row.get("av_oa_avg"),
                                         _sc_row.get("av_oa_misses"),
                                         _sc_row.get("oa_del_sla_pct"),
                                         _sc_row.get("oa_del_avg"),
                                         _sc_row.get("oa_del_misses"),
                                         _ts_gen)
                                    )
                                except Exception:
                                    pass  # non-fatal
                            wbr_conn2.commit(); wbr_conn2.close()
                            if S3_ENABLED:
                                data_sync.push_db_to_s3(AWS_KEY, AWS_SECRET, AWS_REGION, S3_BUCKET)
                            # Persist generated PDF in session state so dashboard survives reruns
                            st.session_state["wbr_last_pdf"]  = pdf_bytes
                            st.session_state["wbr_last_week"] = week_num
                            st.session_state["wbr_last_rd"]   = wbr_report_date.isoformat()
                            # Push to S3 so dashboard persists across sessions / new tabs
                            if S3_ENABLED:
                                data_sync.push_wbr_pdf_to_s3(
                                    pdf_bytes, week_num, wbr_report_date.isoformat(),
                                    AWS_KEY, AWS_SECRET, AWS_REGION, S3_BUCKET
                                )

                        rd = wbr_report_date
                        fname = f"GLS_Robotics_{rd.year}-{rd.month}-{rd.day}.pdf"

                        # ── OUTPUT SECTION ────────────────────────────────────────
                        st.markdown("## Generated Outputs")

                        # ── Metrics explorer (interactive — WoW deltas, trend) ─────
                        with st.expander("📊 Metrics Explorer — WoW deltas & trends", expanded=False):
                            _kpi_defs = [
                                ("Containers",    "containers",     False, ""),
                                ("AV→OA SLA",  "av_oa_sla_pct",  True,  "%"),
                                ("AV→OA Avg",  "av_oa_avg",      False, "d"),
                                ("OA→Del SLA", "oa_del_sla_pct", True,  "%"),
                                ("E2E Avg",        "e2e_avg",        False, "d"),
                                ("OTP",            "otp_pct",        True,  "%"),
                            ]
                            _kpi_cols = st.columns(6)
                            _pw_kpi = prior_data.get(week_num - 1, {}) or {}
                            for _kc, (_klbl, _kkey, _is_sla, _ksufx) in zip(_kpi_cols, _kpi_defs):
                                _kv  = curr_metrics.get(_kkey)
                                _kpv = _pw_kpi.get(_kkey)
                                _kdsp = (
                                    f"{int(round(float(_kv)))}{_ksufx}"
                                    if _kv is not None else "–"
                                )
                                _kdelta = None
                                if _kv is not None and _kpv is not None:
                                    _kdiff = round(float(_kv) - float(_kpv), 1)
                                    _kdelta = (f"+{_kdiff}" if _kdiff >= 0 else str(_kdiff)) + _ksufx
                                _kc.metric(
                                    _klbl, _kdsp, delta=_kdelta,
                                    delta_color="normal" if (not _is_sla or (_kv or 0) >= 95) else "inverse",
                                )

                            import pandas as _pd_dash
                            _wk_idx = display_labels
                            _ch_r1, _ch_r2 = st.columns(3), st.columns(3)
                            for _col, _ct, _ck in [
                                (_ch_r1[0], "AV→OA Avg (d)",   "av_oa_avg"),
                                (_ch_r1[1], "AV→OA SLA %",     "av_oa_sla_pct"),
                                (_ch_r1[2], "OA→Del SLA %",    "oa_del_sla_pct"),
                                (_ch_r2[0], "E2E Transit (d)",      "e2e_avg"),
                                (_ch_r2[1], "OTP %",                "otp_pct"),
                                (_ch_r2[2], "Volume",               "containers"),
                            ]:
                                with _col:
                                    _cdf = _pd_dash.DataFrame({
                                        "Week": _wk_idx,
                                        _ct:   [d.get(_ck) if d else None for d in display_data],
                                    }).dropna(subset=[_ct]).set_index("Week")
                                    st.caption(f"**{_ct}**")
                                    st.line_chart(_cdf, height=130, use_container_width=True)

                        st.markdown("---")

                        out1, out2 = st.columns(2)

                        # ── LEFT: Standard WBR (Mitch) ────────────────────────────
                        with out1:
                            st.markdown("### 📊 Standard WBR — Monday")
                            st.caption(f"Submit to: `doc+destops-36@fusion.amazon.dev` · Subject: `NA Destination Ops WBR_Robotics` · Deadline: **2:00 PM CT ({_ct_to_mst(14)} MST)**")
                            _dl_c1, _dl_c2 = st.columns(2)
                            with _dl_c1:
                                st.download_button(
                                    label=f"⬇️ PDF — {fname}",
                                    data=pdf_bytes,
                                    file_name=fname,
                                    mime="application/pdf",
                                    key="wbr_dl_std",
                                )
                            with _dl_c2:
                                try:
                                    _pptx_fname = f"GLS_Robotics_{wbr_report_date.year}-{wbr_report_date.month}-{wbr_report_date.day}.pptx"
                                    _pptx_bytes = pdf_to_pptx(pdf_bytes, dpi=200)
                                    st.download_button(
                                        label=f"⬇️ PPTX — {_pptx_fname}",
                                        data=_pptx_bytes,
                                        file_name=_pptx_fname,
                                        mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
                                        key="wbr_dl_pptx",
                                    )
                                except Exception as _pptx_err:
                                    st.warning(f"PPTX export unavailable: {_pptx_err}")

                            # ── Bridge — Perjen format ───────────────────────────────
                            def _s(v, sfx=""): return f"{v}{sfx}" if v is not None else "–"
                            _vol    = curr_metrics.get("containers")
                            _av_sla = curr_metrics.get("av_oa_sla_pct")
                            _av_avg = curr_metrics.get("av_oa_avg")
                            _av_p90 = curr_metrics.get("av_oa_p90")
                            _od_sla = curr_metrics.get("oa_del_sla_pct")
                            _od_avg = curr_metrics.get("oa_del_avg")
                            _od_p90 = curr_metrics.get("oa_del_p90")
                            _et     = curr_metrics.get("empty_term_pct")
                            _e2e    = curr_metrics.get("e2e_avg")
                            _otp    = curr_metrics.get("otp_pct")

                            # WoW deltas vs. prior week (W-1)
                            _pw = prior_data.get(week_num - 1, {}) or {}
                            def _wow(curr_val, prior_val, higher_good=True):
                                if curr_val is None or prior_val is None: return ""
                                d = curr_val - prior_val
                                sign = "+" if d >= 0 else ""
                                return f" ({sign}{d} WoW)"
                            _vol_wow  = _wow(_vol,    _pw.get("containers"),    True)
                            _av_wow   = _wow(_av_sla, _pw.get("av_oa_sla_pct"), True)
                            _od_wow   = _wow(_od_sla, _pw.get("oa_del_sla_pct"),True)
                            _e2e_wow  = _wow(_e2e,    _pw.get("e2e_avg"),       False)
                            _otp_wow  = _wow(_otp,    _pw.get("otp_pct"),       True)

                            # Primary carrier for AV→OA (highest miss volume)
                            _dom_carrier = "–"
                            if carrier_sc:
                                _worst = sorted(carrier_sc, key=lambda x: x.get("av_oa_misses") or 0, reverse=True)
                                if _worst and _worst[0].get("av_oa_misses", 0) > 0:
                                    _dom_carrier = _worst[0]["carrier"]

                            # Context notes — dedicated per-section inputs
                            _cn_conn = get_db()
                            _cn_row  = _cn_conn.execute(
                                "SELECT notes FROM wbr_context_notes WHERE year=? AND week_num=?",
                                (year, week_num)
                            ).fetchone()
                            _cn_conn.close()
                            # New per-section fields (preferred); legacy wbr_op_notes as fallback for AV→OA
                            _op_avoa  = (st.session_state.get("wbr_op_avoa")  or "").strip()
                            _op_oadel = (st.session_state.get("wbr_op_oadel") or "").strip()
                            _op_ptg   = (st.session_state.get("wbr_op_ptg")   or "").strip()
                            _op_e2e   = (st.session_state.get("wbr_op_e2e")   or "").strip()
                            _legacy   = (st.session_state.get("wbr_op_notes")  or "").strip()
                            # DB notes go to AV→OA (historical behavior)
                            _db_notes = (_cn_row["notes"] or "").strip() if _cn_row and _cn_row["notes"] else ""
                            _ctx_avoa  = "\n".join(filter(None, [_db_notes, _op_avoa, _legacy if not _op_avoa else ""]))
                            _ctx_oadel = _op_oadel
                            _ctx_ptg   = _op_ptg
                            _ctx_e2e   = _op_e2e
                            # Legacy: keep _ctx_notes pointing to avoa for any remaining old references
                            _ctx_notes = _ctx_avoa

                            # Prior week values for reference lines
                            _p_av  = _s(_pw.get("av_oa_sla_pct"), "%")
                            _p_od  = _s(_pw.get("oa_del_sla_pct"), "%")
                            _p_et  = _s(_pw.get("empty_term_pct"), "%")
                            _p_vol = _s(_pw.get("containers"))

                            # ── Perjen plain-English bridge ─────────────────────────────────
                            def _dir(a, b):
                                if a is None or b is None: return "moved"
                                return "increased" if a > b else ("decreased" if a < b else "remained flat")
                            def _dir_tc(a, b):
                                # Title-case verbs matching Perjen headline style
                                if a is None or b is None: return "Moves to"
                                return "Improves to" if a > b else ("Declines to" if a < b else "Holds at")
                            def _flatcheck(a, b, thresh=2):
                                return a is not None and b is not None and abs(a - b) <= thresh

                            try:
                                _wk_range = f"{wk_start.strftime('%m/%d')}–{wk_end.strftime('%m/%d/%y')}"
                            except Exception:
                                _wk_range = f"WK{week_num}"

                            # Headline: Title Case, carrier callout, matching Perjen style
                            # e.g. "AV→OA Declines to 70% on ATMI Misses; OA→Del Drops to 31%; Empty→Term Holds at 100%"
                            _headline_parts = []
                            if _av_sla is not None:
                                _av_hl = f"AV→OA {_dir_tc(_av_sla, _pw.get('av_oa_sla_pct'))} {_av_sla}%"
                                if _dom_carrier != "–":
                                    _av_hl += f" on {_dom_carrier} Misses"
                                _headline_parts.append(_av_hl)
                            if _od_sla is not None:
                                _headline_parts.append(f"OA→Del {_dir_tc(_od_sla, _pw.get('oa_del_sla_pct'))} {_od_sla}%")
                            if _et is not None:
                                _et_hl_verb = "Holds at" if _et == _pw.get("empty_term_pct") else _dir_tc(_et, _pw.get("empty_term_pct"))
                                _headline_parts.append(f"Empty→Term {_et_hl_verb} {_et}%")
                            if _otp is not None:
                                _headline_parts.append(f"OTP {_dir_tc(_otp, _pw.get('otp_pct'))} {_otp}%")
                            _headline = "; ".join(_headline_parts) if _headline_parts else "Week in Review"

                            # Port mix from GVT (appended to Volume narrative)
                            _port_mix_str = ""
                            try:
                                _pop_gvt = gvt_df[
                                    gvt_df['ready_date'].notna() &
                                    (gvt_df['ready_date'] >= wk_start.date()) &
                                    (gvt_df['ready_date'] <= wk_end.date())
                                ]
                                _ports = _pop_gvt['port'].dropna().astype(str).str.strip().str.upper()
                                _port_counts = _ports.value_counts()
                                if len(_port_counts) > 1:
                                    _pmix = ", ".join(f"{p} ({n})" for p, n in _port_counts.items())
                                    _port_mix_str = f" Port mix: {_pmix}."
                                elif len(_port_counts) == 1:
                                    _port_mix_str = f" All containers at {_port_counts.index[0]}."
                            except Exception:
                                pass

                            # Volume narrative
                            _p_vol_v = _pw.get("containers")
                            if _vol is not None and _p_vol_v and _p_vol_v > 0:
                                _vol_chg = round((_vol - _p_vol_v) / _p_vol_v * 100)
                                _vol_narrative = (f"Container volume {_dir(_vol, _p_vol_v)} {abs(_vol_chg)}% to {_s(_vol)}"
                                                  f" (WK{week_num-1}: {_s(_p_vol_v)}; WK{week_num}: {_s(_vol)}).{_port_mix_str}")
                            else:
                                _vol_narrative = f"Container volume: {_s(_vol)} for WK{week_num}.{_port_mix_str}"

                            # Multi-week AV→OA trend (up to 5 prior + current), shown in parens after P90
                            _trend_str = ""
                            try:
                                _tw = sorted([w for w in prior_data.keys() if w < week_num], reverse=True)[:5]
                                _tv = [prior_data[w].get("av_oa_sla_pct") for w in reversed(_tw)]
                                if len(_tv) >= 2 and all(v is not None for v in _tv):
                                    _chain = "→".join(f"{v}%" for v in _tv)
                                    if _av_sla is not None:
                                        _chain += f"→{_av_sla}%"
                                    _trend_str = f" ({len(_tv) + (1 if _av_sla is not None else 0)}-week trend: {_chain})"
                            except Exception:
                                pass

                            # AV→OA narrative — inline carrier detail matching Perjen:
                            # "Decline driven by ATMI (11/28 = 39%) — HDDR (9/9), ARVY (7/7) all executed at 100%."
                            _av_body = (
                                f"AV→OA SLA {_dir(_av_sla, _pw.get('av_oa_sla_pct'))} to {_s(_av_sla, '%')}"
                                f" (WK{week_num-1}: {_p_av}; WK{week_num}: {_s(_av_sla, '%')})."
                                f" Avg transit {_dir(_av_avg, _pw.get('av_oa_avg'))} to {_s(_av_avg)}d;"
                                f" P90 {_dir(_av_p90, _pw.get('av_oa_p90'))} to {_s(_av_p90)}d.{_trend_str}"
                            )
                            _av_carrier_inline = ""
                            _miss_carriers = []
                            if carrier_sc:
                                _sorted_sc = sorted(carrier_sc, key=lambda x: x.get("av_oa_misses") or 0, reverse=True)
                                _miss_carriers = [c for c in _sorted_sc if (c.get("av_oa_misses") or 0) > 0]
                                _clean_carriers = [c for c in _sorted_sc if (c.get("av_oa_misses") or 0) == 0]
                                _miss_parts = [
                                    f"{c['carrier']} ({c.get('av_oa_misses', 0)}/{c.get('volume', 0)} = {c['av_oa_sla_pct']}%)"
                                    if c.get("av_oa_sla_pct") is not None else
                                    f"{c['carrier']} ({c.get('av_oa_misses', 0)}/{c.get('volume', 0)})"
                                    for c in _miss_carriers
                                ]
                                _clean_parts = [
                                    f"{c['carrier']} ({c.get('volume', 0)}/{c.get('volume', 0)})"
                                    for c in _clean_carriers
                                ]
                                if _miss_parts and _clean_parts:
                                    _av_carrier_inline = (
                                        f"Decline driven by {', '.join(_miss_parts)} — "
                                        f"{', '.join(_clean_parts)} all executed at 100%."
                                    )
                                elif _miss_parts:
                                    _av_carrier_inline = f"Misses: {', '.join(_miss_parts)}."
                                elif _clean_parts:
                                    _av_carrier_inline = "All carriers executed at 100%."
                            # AV→OA context: all lines from the dedicated input field
                            _avoa_ctx_lines = [l for l in _ctx_avoa.splitlines() if l.strip()] if _ctx_avoa else []
                            _av_context_block = "\n".join(_avoa_ctx_lines) if _avoa_ctx_lines else ""

                            # OA→Del narrative
                            _od_body = (
                                f"OA→Del SLA {_dir(_od_sla, _pw.get('oa_del_sla_pct'))} to {_s(_od_sla, '%')}"
                                f" (WK{week_num-1}: {_p_od}; WK{week_num}: {_s(_od_sla, '%')})."
                                f" P90 {_dir(_od_p90, _pw.get('oa_del_p90'))} to {_s(_od_p90)}d [BOS; A320-RBTCS excl.]."
                            )
                            _od_root = (f"\n{_ctx_oadel}" if _ctx_oadel else "")

                            # Empty→Term narrative
                            _et_pw = _pw.get("empty_term_pct")
                            if _et is not None and _et_pw is not None:
                                if _flatcheck(_et, _et_pw):
                                    _et_body = f"Flat at {_s(_et, '%')}. No issues."
                                elif _et >= 95:
                                    _et_body = f"{'Improved' if _et > _et_pw else 'Declined'} to {_s(_et, '%')} (WK{week_num-1}: {_p_et}). Within SLA."
                                else:
                                    _et_body = f"{'Improved' if _et > _et_pw else 'Declined'} to {_s(_et, '%')} (WK{week_num-1}: {_p_et}). Below 95% SLA threshold."
                            else:
                                _et_body = f"{_s(_et, '%')} SLA."

                            # E2E / OTP narrative
                            _e2e_pw = _pw.get("e2e_avg")
                            _otp_pw = _pw.get("otp_pct")
                            _e2e_body = (
                                f"On-Time to Promise {_dir(_otp, _otp_pw)} to {_s(_otp, '%')}"
                                f" (WK{week_num-1}: {_s(_otp_pw, '%')}; WK{week_num}: {_s(_otp, '%')})."
                                f" E2E Transit Avg {_dir(_e2e, _e2e_pw)} to {_s(_e2e)}d (WK{week_num-1}: {_s(_e2e_pw)}d)."
                            )
                            # Append E2E context (dedicated field, or auto-inference as fallback)
                            if _ctx_e2e:
                                _e2e_body += f" {_ctx_e2e}"
                            elif _otp is not None and _otp_pw is not None and _otp < _otp_pw:
                                _e2e_body += (" Miss driven by ocean transit delay — outside dray control."
                                              if (_e2e or 0) > 40 else
                                              " Dray execution contributing to miss — carrier follow-up required.")

                            # Path to Green — typed PTG rows first, then auto-inferred, then placeholder
                            _ptg_rows = []
                            if _ctx_ptg:
                                # User-entered rows: "Action | Impact | Timeline"
                                for _ptg_line in _ctx_ptg.splitlines():
                                    _ptg_line = _ptg_line.strip()
                                    if not _ptg_line: continue
                                    _ptg_parts = [p.strip() for p in _ptg_line.split("|")]
                                    if len(_ptg_parts) >= 3:
                                        _ptg_rows.append((_ptg_parts[0], _ptg_parts[1], _ptg_parts[2]))
                                    elif len(_ptg_parts) == 2:
                                        _ptg_rows.append((_ptg_parts[0], _ptg_parts[1], "TBD"))
                                    else:
                                        _ptg_rows.append((_ptg_line, "–", "TBD"))
                            # Auto-inferred rows for any unaddressed misses
                            if not any("AV" in r[0] or "AV→OA" in r[0] or (_dom_carrier != "–" and _dom_carrier in r[0]) for r in _ptg_rows):
                                if _dom_carrier != "–" and _av_sla is not None and _av_sla < 95:
                                    _ptg_rows.append((f"Carrier performance review with {_dom_carrier}", "Reduce AV→OA SLA misses", "This week"))
                            if not any("OA→Del" in r[0] or "OA-Del" in r[0] for r in _ptg_rows):
                                if _od_sla is not None and _od_sla < 95:
                                    _ptg_rows.append(("Investigate OA→Del delays (BOS market)", "Restore SLA to ≥95%", "Next 2 weeks"))
                            if not any("Empty" in r[0] for r in _ptg_rows):
                                if _et is not None and _et < 95:
                                    _ptg_rows.append(("Empty return follow-up with all carriers", "Clear backlog", "This week"))
                            if not _ptg_rows:
                                _ptg_rows.append(("[Additional action]", "[Expected impact]", "[Timeline]"))
                            _ptg = (
                                "Path to Green:\n\n"
                                "  Action | Expected Impact | Timeline\n"
                                "  " + "-" * 70 + "\n"
                                + "\n".join(f"  {a} | {b} | {c}" for a, b, c in _ptg_rows)
                            )

                            # Market context — load from DB (populated daily by port-intel Lambda)
                            _mkt_ctx = load_market_context(get_db())
                            _mkt_block = format_market_context_block(_mkt_ctx)
                            # Manual external context from UI input
                            _ext_ctx = (st.session_state.get("wbr_ext_ctx") or "").strip()

                            _bridge_std = (
                                f"WK{week_num} ({_wk_range})\n"
                                f"Headline: {_headline}\n\n"
                                + (f"{_mkt_block}\n\n" if _mkt_block else "")
                                + (f"[External Context]\n{_ext_ctx}\n\n" if _ext_ctx else "")
                                + f"[Volume] {_vol_narrative}\n\n"
                                f"[AV→OA] {_av_body}"
                                + (f"\n{_av_carrier_inline}" if _av_carrier_inline else "")
                                + (f"\n{_av_context_block}" if _av_context_block else "")
                                + "\n"
                                + f"\n[OA→Del] {_od_body}"
                                + (_od_root + "\n" if _od_root else "\n")
                                + f"\n[Empty→Term] {_et_body}\n"
                                + f"\n[E2E / OTP] {_e2e_body}\n"
                                + f"\n{_ptg}"
                            )

                            st.text_area(
                                "Bridge — Perjen format (editable before sending)",
                                value=_bridge_std, height=400, key="wbr_bridge_std",
                            )

                            # ── SOP Callout box ───────────────────────────────────
                            st.markdown("---")
                            st.info(
                                "**📧 WBR Submission SOP**\n\n"
                                f"**To:** `doc+destops-36@fusion.amazon.dev`  ·  "
                                f"**Subject:** `NA Destination Ops WBR_Robotics`  ·  "
                                f"**Deadline:** Monday by 2:00 PM CT ({_ct_to_mst(14)} MST)  ·  "
                                f"**Attachment:** `{fname}`"
                            )

                            _email_to   = "doc+destops-36@fusion.amazon.dev"
                            _email_subj = "NA Destination Ops WBR_Robotics"
                            _email_body_txt = (
                                "WBR: NA Destination Ops WBR\n"
                                "Slide: Robotics-Rocus\n"
                                "Owner: Dominique Kennedy\n"
                                "Alias: kennewdo\n"
                                "Frequency: Weekly\n"
                            )

                            # ── Fusion submission check ───────────────────────────────────
                            _fusion_url = "https://fusion.amazon.dev/documentset/DOCUMENTSET%23f42d63bd-a96a-4f29-a356-a66ac18602a9"
                            _fusion_cols = st.columns([2, 1])
                            with _fusion_cols[0]:
                                st.markdown(
                                    f"**WBR Slides are published to Fusion.** "
                                    f"After sending, verify your W{week_num} slide appears in the document set:"
                                )
                                st.link_button("🔗 Open Fusion Document Set", _fusion_url)
                            with _fusion_cols[1]:
                                _submitted_key = f"wbr_submitted_w{week_num}_{year}"
                                _already_marked = st.session_state.get(_submitted_key, False)
                                if _already_marked:
                                    st.success(f"✅ W{week_num} marked as sent")
                                else:
                                    if st.button("✅ Mark as Sent", key=f"wbr_mark_sent_{week_num}"):
                                        st.session_state[_submitted_key] = True
                                        st.rerun()
                            _mailto_url = (
                                f"mailto:{_email_to}"
                                f"?subject={urllib.parse.quote(_email_subj)}"
                                f"&body={urllib.parse.quote(_email_body_txt)}"
                            )

                            _btn_col, _ = st.columns([1, 2])
                            with _btn_col:
                                st.link_button("📧 Open in Email Client", _mailto_url, type="primary")

                            # ── Manual fallback — step-by-step ───────────────────
                            with st.expander("📋 Manual submission instructions (if mailto doesn't work)", expanded=False):
                                st.markdown(f"""**Step-by-step:**

    1. **Download the PDF** — click the ⬇️ Download button above  
    2. **Open a new email** (Outlook / webmail)  
    3. **To:** `doc+destops-36@fusion.amazon.dev`  
    4. **Subject:** `NA Destination Ops WBR_Robotics`  
    5. **Attach** the PDF you just downloaded  
    6. **Paste the email body below** into the message body  
    7. **Send before 2:00 PM CT ({_ct_to_mst(14)} MST) on Monday**
    """)
                                st.caption("Email body (copy all of this):")
                                st.code(_email_body_txt, language="")

                        # ── RIGHT: Enhanced WBR (Robotics team) ──────────────────
                        with out2:
                            st.markdown("### 📦 Enhanced WBR — Wednesday")
                            st.caption("3-page package for Robotics meeting. Carrier scorecard + root cause + forward look.")

                            # Carrier scorecard
                            if carrier_sc:
                                st.markdown("**Carrier Scorecard — AV→OA**")
                                av_rows = []
                                for cs in carrier_sc:
                                    av_rows.append({
                                        "Carrier":  cs["carrier"],
                                        "Vol":      cs["volume"],
                                        "SLA%":     "{}%".format(cs["av_oa_sla_pct"]) if cs["av_oa_sla_pct"] is not None else "–",
                                        "Misses":   cs["av_oa_misses"],
                                        "Avg d":    cs["av_oa_avg"] if cs["av_oa_avg"] is not None else "–",
                                        "P90":      cs["av_oa_p90"] if cs["av_oa_p90"] is not None else "–",
                                    })
                                st.dataframe(pd.DataFrame(av_rows), hide_index=True, use_container_width=True)

                                st.markdown("**Carrier Scorecard — OA→Del** (BOS only)")
                                od_rows = []
                                for cs in carrier_sc:
                                    if cs["oa_del_sla_pct"] is None and cs["oa_del_avg"] is None:
                                        continue
                                    od_rows.append({
                                        "Carrier":  cs["carrier"],
                                        "Vol":      cs["volume"],
                                        "SLA%":     "{}%".format(cs["oa_del_sla_pct"]) if cs["oa_del_sla_pct"] is not None else "–",
                                        "Misses":   cs["oa_del_misses"],
                                        "Avg d":    cs["oa_del_avg"] if cs["oa_del_avg"] is not None else "–",
                                        "P90":      cs["oa_del_p90"] if cs["oa_del_p90"] is not None else "–",
                                    })
                                if od_rows:
                                    st.dataframe(pd.DataFrame(od_rows), hide_index=True, use_container_width=True)

                                # Root cause notes per carrier
                                st.markdown("**Root Cause Notes**")
                                rc_n = min(len(carrier_sc), 4)
                                rc_cols = st.columns(rc_n)
                                for i, cs in enumerate(carrier_sc):
                                    with rc_cols[i % rc_n]:
                                        st.text_input(
                                            "{} ({})".format(cs["carrier"], cs["volume"]),
                                            placeholder="e.g. flip delay, fee clearance",
                                            key="wbr_rc_{}".format(cs["carrier"]),
                                        )
                            else:
                                st.info("No carrier breakdown — verify GVT has Dray SCAC(FL) column.")

                            # Root cause classification
                            with st.expander("Root Cause Classification"):
                                rc1c, rc2c = st.columns(2)
                                with rc1c:
                                    rc_a = st.number_input("(a) Carrier Execution",  min_value=0, key="wbr_rc_a", help="Carrier failed to pick up in SLA window")
                                    rc_b = st.number_input("(b) Process Gap",        min_value=0, key="wbr_rc_b", help="Flip delay, fee clearance, VRID, admin hold")
                                    rc_c = st.number_input("(c) Site Constraint",    min_value=0, key="wbr_rc_c", help="Site shutdown, capacity, stop work")
                                    rc_d = st.number_input("(d) Upstream Ocean",     min_value=0, key="wbr_rc_d", help="Container late to port; PO promise past before dray")
                                    rc_e = st.number_input("(e) Data / Timing",      min_value=0, key="wbr_rc_e", help="In-transit at report cut; resolves next week")
                                with rc2c:
                                    total_rc = rc_a + rc_b + rc_c + rc_d + rc_e
                                    if total_rc > 0:
                                        for label, cnt in [("(a) Carrier", rc_a), ("(b) Process", rc_b), ("(c) Site", rc_c), ("(d) Ocean", rc_d), ("(e) Timing", rc_e)]:
                                            if cnt > 0:
                                                st.markdown(f"`{label}` {cnt} ({round(cnt/total_rc*100)}%)")

                            # Forward look — ISS-driven if available, manual otherwise
                            with st.expander("📦 2-Week Forward Look — Inbound Pipeline", expanded=True):
                                if not iss_df.empty and "final_dest_eta" in iss_df.columns:
                                    _fwd_cut1 = wbr_report_date + timedelta(days=7)
                                    _fwd_cut2 = wbr_report_date + timedelta(days=14)
                                    _iss_fwd  = iss_df.copy()
                                    _iss_fwd["_eta"] = pd.to_datetime(_iss_fwd["final_dest_eta"], errors="coerce").dt.date
                                    _w1 = _iss_fwd[_iss_fwd["_eta"].notna() & (_iss_fwd["_eta"] > wbr_report_date) & (_iss_fwd["_eta"] <= _fwd_cut1)]
                                    _w2 = _iss_fwd[_iss_fwd["_eta"].notna() & (_iss_fwd["_eta"] > _fwd_cut1)        & (_iss_fwd["_eta"] <= _fwd_cut2)]
                                    st.caption(f"ETA within 7d: **{len(_w1)} containers** | within 14d: **{len(_w2)} containers**")
                                    _in_window = _iss_fwd[_iss_fwd["_eta"].notna() & (_iss_fwd["_eta"] > wbr_report_date) & (_iss_fwd["_eta"] <= _fwd_cut2)]
                                    if len(_in_window):
                                        _grp_cols = [c for c in ["vessel_name","voyage_no","_eta"] if c in _in_window.columns]
                                        _fc_agg   = lambda x: ", ".join(x.dropna().astype(str).unique()[:3])
                                        if "fc" in _in_window.columns:
                                            _vg = (_in_window.groupby(_grp_cols)
                                                   .agg(Containers=("container_no","count"), FC=("fc", _fc_agg))
                                                   .reset_index()
                                                   .rename(columns={"vessel_name":"Vessel","voyage_no":"Voyage","_eta":"ETA"})
                                                   .sort_values("ETA"))
                                        else:
                                            _vg = (_in_window.groupby(_grp_cols)
                                                   .agg(Containers=("container_no","count"))
                                                   .reset_index()
                                                   .rename(columns={"vessel_name":"Vessel","voyage_no":"Voyage","_eta":"ETA"})
                                                   .sort_values("ETA"))
                                        st.dataframe(_vg, hide_index=True, use_container_width=True)
                                    else:
                                        st.info("No containers with FinalDestETA in the next 14 days in this ISS file.")
                                else:
                                    st.caption("⬆️ Upload Import Shipment Status (4th uploader) to auto-populate from live vessel ETAs.")
                                    fwd_data = pd.DataFrame([
                                        {"Port": "ORF",     "W+1 Proj": 0, "W+2 Proj": 0, "Carrier(s)": "", "Notes": ""},
                                        {"Port": "BOS",     "W+1 Proj": 0, "W+2 Proj": 0, "Carrier(s)": "", "Notes": ""},
                                        {"Port": "SAV",     "W+1 Proj": 0, "W+2 Proj": 0, "Carrier(s)": "", "Notes": ""},
                                        {"Port": "LAX/LGB", "W+1 Proj": 0, "W+2 Proj": 0, "Carrier(s)": "", "Notes": ""},
                                        {"Port": "OAK",     "W+1 Proj": 0, "W+2 Proj": 0, "Carrier(s)": "", "Notes": ""},
                                    ])
                                    st.data_editor(fwd_data, key="wbr_fwd_look", use_container_width=True, hide_index=True)

                            # Enhanced bridge
                            op_notes_val    = st.session_state.get("wbr_op_notes", "")
                            rc_insight_val  = "\n".join(
                                f"  {cs['carrier']}: {st.session_state.get(f'wbr_rc_{cs["carrier"]}', '') or '—'}"
                                for cs in carrier_sc
                            ) if carrier_sc else ""
                            carrier_summary = "\n".join(
                                "  {}: Vol={}, AV-OA {}%{}".format(
                                    cs["carrier"], cs["volume"],
                                    cs["av_oa_sla_pct"] or "–",
                                    " | OA-Del {}%".format(cs["oa_del_sla_pct"]) if cs["oa_del_sla_pct"] is not None else ""
                                ) for cs in carrier_sc
                            ) if carrier_sc else "  No carrier breakdown available"

                            _bridge_enh = f"""WK{week_num} NA Robotics Destination Dray — Enhanced WBR Bridge
    Reporting Week: {wk_start} to {wk_end}

    PERFORMANCE SUMMARY
      Volume: {_s(_vol)} containers
      AV to OA:        {_s(_av_sla, "%")} SLA | {_s(_av_avg)} avg | P90 {_s(_av_p90)}d
      OA to Delivery:  {_s(_od_sla, "%")} SLA | {_s(_od_avg)} avg | P90 {_s(_od_p90)}d  [BOS market]
      Empty to Term:   {_s(_et, "%")} SLA
      E2E Transit:     {_s(_e2e)} days avg
      On-Time to Promise: {_s(_otp, "%")}

    CARRIER SCORECARD
    {carrier_summary}

    ROOT CAUSE (AV→OA misses)
    {rc_insight_val if rc_insight_val else "  Add root cause insight above"}

    OPERATIONAL CALLOUTS
    {op_notes_val if op_notes_val else "  Add in the Validation section above"}

    PATH TO GREEN
      [Action | Owner | Target Date]
    """
                            st.text_area("Bridge — Enhanced (editable)", value=_bridge_enh, height=320, key="wbr_bridge_enh")

                        st.success(f"W{week_num} metrics saved to DB for future carry-forward.")


                except Exception as e:
                    st.error(f"Error: {e}")

    with _wbr_hist:
            # Saved weeks
            wbr_conn3 = get_db()
            saved = wbr_conn3.execute(
                "SELECT year, week_num, containers, av_oa_sla_pct, oa_del_sla_pct, otp_pct, generated_at FROM wbr_results ORDER BY year DESC, week_num DESC LIMIT 12"
            ).fetchall()
            wbr_conn3.close()
            if saved:
                st.markdown("**Previously saved weeks (carry-forward available):**")
                st.dataframe(pd.DataFrame([dict(r) for r in saved]), hide_index=True, use_container_width=True)

            # Seed historical weeks
            st.markdown("---")
            with st.expander("🗂️ Seed W24–W28 carry-forward values (from WK29 slide)", expanded=False):
                st.caption("Pre-loads W24–W28 from the submitted GLS_Robotics_2026-7-20 slide. Run once.")
                _sc = get_db()
                _seeded_rows = _sc.execute(
                    "SELECT COUNT(*) FROM wbr_results WHERE year=2026 AND week_num IN (24,25,26,27,28)"
                ).fetchone()[0]
                _sc.close()
                if _seeded_rows == 5:
                    st.success("W24–W28 already in DB.")
                else:
                    st.info(f"{_seeded_rows}/5 historical weeks in DB.")
                    if st.button("Seed W24–W28 to DB", key="wbr_seed_hist", type="primary"):
                        _seed = {
                            24: dict(containers=89,  av_oa_avg=2.0, av_oa_sla_pct=68,  av_oa_p90=3, oa_del_avg=0.5, oa_del_sla_pct=100, oa_del_p90=1,  empty_term_pct=100, e2e_avg=44, otp_pct=97,  week_start="2026-06-07", week_end="2026-06-13"),
                            25: dict(containers=101, av_oa_avg=1.1, av_oa_sla_pct=100, av_oa_p90=1, oa_del_avg=2.9, oa_del_sla_pct=71,  oa_del_p90=10, empty_term_pct=100, e2e_avg=45, otp_pct=45,  week_start="2026-06-14", week_end="2026-06-20"),
                            26: dict(containers=83,  av_oa_avg=2.9, av_oa_sla_pct=70,  av_oa_p90=4, oa_del_avg=7.0, oa_del_sla_pct=31,  oa_del_p90=14, empty_term_pct=100, e2e_avg=41, otp_pct=66,  week_start="2026-06-21", week_end="2026-06-27"),
                            27: dict(containers=60,  av_oa_avg=1.7, av_oa_sla_pct=100, av_oa_p90=3, oa_del_avg=4.8, oa_del_sla_pct=28,  oa_del_p90=9,  empty_term_pct=100, e2e_avg=36, otp_pct=48,  week_start="2026-06-28", week_end="2026-07-04"),
                            28: dict(containers=32,  av_oa_avg=0.8, av_oa_sla_pct=100, av_oa_p90=2, oa_del_avg=3.1, oa_del_sla_pct=38,  oa_del_p90=5,  empty_term_pct=100, e2e_avg=55, otp_pct=46,  week_start="2026-07-05", week_end="2026-07-11"),
                        }
                        from wbr_engine import save_week_to_db as _swdb
                        _sconn = get_db()
                        _ts = datetime.now(_EASTERN).isoformat()
                        for _wn, _m in _seed.items():
                            _swdb(_sconn, 2026, _wn, _m, _ts)
                        _sconn.close()
                        if S3_ENABLED:
                            data_sync.push_db_to_s3(AWS_KEY, AWS_SECRET, AWS_REGION, S3_BUCKET)
                        st.success("W24–W28 seeded. Upload files above to generate.")
                        st.rerun()




    with _wbr_sum:
        st.markdown("### WBR Summary")
        # ── Load latest WBR result ────────────────────────────────────────────
        try:
            _conn_s = get_db()
            _res = _conn_s.execute(
                "SELECT week_num, year, week_start, week_end, containers, "
                "av_oa_sla_pct, oa_del_sla_pct, e2e_avg, otp_pct "
                "FROM wbr_results ORDER BY year DESC, week_num DESC LIMIT 1"
            ).fetchone()
            _car_rows = _conn_s.execute(
                "SELECT carrier, volume, av_oa_sla_pct, av_oa_avg, av_oa_misses, "
                "oa_del_sla_pct, oa_del_avg, oa_del_misses "
                "FROM wbr_carrier_results ORDER BY year DESC, week_num DESC, volume DESC"
            ).fetchall()
            _conn_s.close()
        except Exception:
            _res = None
            _car_rows = []
        if _res:
            _wk, _yr, _wstart, _wend, _containers, _av_oa, _oa_del, _e2e, _otp = _res
            st.info(f"**Latest saved week: W{_wk} {_yr}** ({_wstart} – {_wend})")
            _c1, _c2, _c3, _c4, _c5 = st.columns(5)
            _c1.metric("Containers", int(_containers) if _containers else "—")
            _c2.metric("AV→OA SLA", f"{_av_oa:.1f}%" if _av_oa is not None else "—",
                       delta="✓" if _av_oa and _av_oa >= 90 else "⚠")
            _c3.metric("OA→Del SLA", f"{_oa_del:.1f}%" if _oa_del is not None else "—",
                       delta="✓" if _oa_del and _oa_del >= 90 else "⚠")
            _c4.metric("E2E Avg (days)", f"{_e2e:.1f}" if _e2e is not None else "—")
            _c5.metric("On-Time %", f"{_otp:.1f}%" if _otp is not None else "—")
            if _car_rows:
                st.markdown("#### Carrier Scorecards")
                import pandas as _pd_sum
                _df_car = _pd_sum.DataFrame(_car_rows, columns=[
                    "Carrier", "Volume", "AV→OA SLA%", "AV→OA Avg(d)",
                    "AV→OA Misses", "OA→Del SLA%", "OA→Del Avg(d)", "OA→Del Misses"
                ])
                st.dataframe(_df_car, use_container_width=True, hide_index=True)
        else:
            st.info("No saved WBR weeks yet. Use the **Build** tab to generate and save your first week.")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 10 — Inbound Forecast & Carrier Allocation
# ══════════════════════════════════════════════════════════════════════════════

    with _wbr_trends:
        import plotly.graph_objects as go
        from datetime import date as _dt_cls


        # ── Load all DB data ───────────────────────────────────────────────────────
        _d_conn = get_db()
        try:
            _all_wbr = pd.read_sql_query(
                "SELECT * FROM wbr_results ORDER BY year, week_num", _d_conn
            )
            try:
                _car_wbr = pd.read_sql_query(
                    "SELECT * FROM wbr_carrier_results ORDER BY year, week_num, volume DESC",
                    _d_conn
                )
            except Exception:
                _car_wbr = pd.DataFrame()
        finally:
            _d_conn.close()

        if _all_wbr.empty:
            st.info("No WBR data yet. Run the **WBR Generator** tab — this dashboard updates automatically.")
            st.stop()

        # ── Year + availability ────────────────────────────────────────────────────
        _avail_years = sorted(_all_wbr["year"].dropna().unique().tolist(), reverse=True)
        _yr_col, _trend_col, _carr_col = st.columns([1, 3, 2])

        with _yr_col:
            _sel_yr = int(st.selectbox("Year", options=_avail_years, index=0, key="d_yr"))

        _yr_wbr = _all_wbr[_all_wbr["year"] == _sel_yr].copy()
        _yr_wbr["week_num"] = _yr_wbr["week_num"].astype(int)

        # parse dates for each row — fall back to computing from week_num if blank
        def _wk_dates(row):
            try:
                ws = _dt_cls.fromisoformat(str(row["week_start"])[:10])
                we = _dt_cls.fromisoformat(str(row["week_end"])[:10])
                return ws, we
            except Exception:
                from datetime import date as _d, timedelta as _td
                import math as _math
                yr, wn = int(row["year"]), int(row["week_num"])
                try:
                    mon = _d.fromisocalendar(yr, wn, 1)
                    return mon - _td(days=1), mon + _td(days=5)
                except Exception:
                    return None, None

        _avail_wks = sorted(_yr_wbr["week_num"].unique().tolist())
        _max_wk    = max(_avail_wks)

        with _trend_col:
            _trend = st.radio(
                "Trend window", ["4WK", "6WK", "12WK", "All"],
                index=1, horizontal=True, key="d_trend",
                label_visibility="collapsed"
            )

        with _carr_col:
            _carr_opts = (sorted(_car_wbr["carrier"].dropna().unique().tolist())
                          if not _car_wbr.empty else [])
            _sel_carr = st.multiselect(
                "Carrier filter", options=_carr_opts,
                placeholder="All carriers", key="d_carr"
            ) if _carr_opts else []

        # ── Row 2: Ending week anchor + Date range + Week multiselect ─────────────
        _anc_col, _dr_col, _wk_col = st.columns([1, 2, 3])

        with _anc_col:
            # Ending week — trend window counts backward from here.
            # Defaults to most recent available; change to anchor any earlier week.
            _anchor_wk = st.selectbox(
                "Ending week",
                options=list(reversed(_avail_wks)),
                index=0,
                format_func=lambda w: f"W{w}",
                key="d_anchor",
            )

        with _dr_col:
            _all_starts, _all_ends = [], []
            for _, _r in _yr_wbr.iterrows():
                _ws, _we = _wk_dates(_r)
                if _ws: _all_starts.append(_ws)
                if _we: _all_ends.append(_we)
            _min_d = min(_all_starts) if _all_starts else _dt_cls(_sel_yr, 1, 1)
            _max_d = max(_all_ends)   if _all_ends   else _dt_cls.today()
            _date_range = st.date_input(
                "Date range (maps to weeks)",
                value=(_min_d, _max_d),
                min_value=_dt_cls(_sel_yr, 1, 1),
                max_value=_dt_cls(_sel_yr, 12, 31),
                key="d_dates",
            )

        # Resolve trend window backward from anchor week
        _trend_n = {"4WK": 4, "6WK": 6, "12WK": 12, "All": len(_avail_wks)}[_trend]
        _anchor_idx  = _avail_wks.index(_anchor_wk)
        _start_idx   = max(0, _anchor_idx - _trend_n + 1)
        _trend_wks   = _avail_wks[_start_idx : _anchor_idx + 1]

        # Resolve date range → matching weeks
        _date_wks = list(_avail_wks)
        if isinstance(_date_range, (list, tuple)) and len(_date_range) >= 2:
            _dr_s, _dr_e = _date_range[0], _date_range[1]
            _date_wks = []
            for _, _r in _yr_wbr.iterrows():
                _ws, _we = _wk_dates(_r)
                if _ws and _we and _ws <= _dr_e and _we >= _dr_s:
                    _date_wks.append(int(_r["week_num"]))
        elif isinstance(_date_range, _dt_cls):
            _date_wks = []
            for _, _r in _yr_wbr.iterrows():
                _ws, _we = _wk_dates(_r)
                if _ws and _we and _ws <= _date_range <= _we:
                    _date_wks.append(int(_r["week_num"]))

        # Combine: intersection of trend window + date range
        _auto_wks = sorted(set(_trend_wks) & set(_date_wks)) or _trend_wks

        with _wk_col:
            _sel_wks = st.multiselect(
                "Weeks (override above selection)",
                options=_avail_wks,
                default=_auto_wks,
                format_func=lambda w: f"W{w}",
                key="d_wks",
                placeholder="Select weeks...",
            )

        if not _sel_wks:
            st.warning("Select at least one week.")
            st.stop()

        # ── Final filtered data ────────────────────────────────────────────────────
        _wdf = _yr_wbr[_yr_wbr["week_num"].isin(_sel_wks)].copy().sort_values("week_num")
        _wdf["wl"] = "W" + _wdf["week_num"].astype(str)

        if _wdf.empty:
            st.warning("No data for selected filters.")
            st.stop()

        # ── KPI Cards — latest week with WoW delta ─────────────────────────────────
        _lw = _wdf.iloc[-1]
        _pw = _wdf.iloc[-2] if len(_wdf) > 1 else None
        _ws_str = str(_lw.get("week_start") or "")[:10]
        _we_str = str(_lw.get("week_end")   or "")[:10]
        st.markdown(f"#### W{int(_lw['week_num'])}  ·  {_ws_str} → {_we_str}")

        def _vv(row, col):
            v = row[col] if col in row.index else None
            return None if (v is None or (isinstance(v, float) and pd.isna(v))) else v

        def _d(col, invert=False):
            if _pw is None: return None
            c, p = _vv(_lw, col), _vv(_pw, col)
            if c is None or p is None: return None
            diff = round(c - p, 1)
            return -diff if invert else diff

        _kc = st.columns(5)
        _kc[0].metric("Containers",    str(int(_vv(_lw,"containers")))       if _vv(_lw,"containers")    is not None else "—", delta=int(_d("containers"))      if _d("containers")      is not None else None)
        _kc[1].metric("AV→OA SLA%",   f"{int(_vv(_lw,'av_oa_sla_pct'))}%"   if _vv(_lw,"av_oa_sla_pct") is not None else "—", delta=f"{int(_d('av_oa_sla_pct'))}pp" if _d("av_oa_sla_pct") is not None else None)
        _kc[2].metric("OA→Del SLA%",  f"{int(_vv(_lw,'oa_del_sla_pct'))}%"  if _vv(_lw,"oa_del_sla_pct") is not None else "—", delta=f"{int(_d('oa_del_sla_pct'))}pp" if _d("oa_del_sla_pct") is not None else None)
        _kc[3].metric("E2E Avg (days)",str(int(_vv(_lw,"e2e_avg")))          if _vv(_lw,"e2e_avg")       is not None else "—", delta=round(_d("e2e_avg",invert=True),1) if _d("e2e_avg") is not None else None, delta_color="normal")
        _kc[4].metric("OTP%",         f"{int(_vv(_lw,'otp_pct'))}%"          if _vv(_lw,"otp_pct")       is not None else "—", delta=f"{int(_d('otp_pct'))}pp"         if _d("otp_pct")      is not None else None)

        # ── SLA Goals box ─────────────────────────────────────────────────────────
        _sg1, _sg2 = st.columns([1, 3])
        with _sg1:
            st.markdown(
                """<div style="border:1px solid #555;border-radius:6px;padding:10px 14px;
                background:#1a1a1a;font-size:13px;line-height:1.8;">
                <b style="font-size:14px;">SLA Goals</b><hr style="border-color:#444;margin:4px 0 6px 0;">
                &bull; AV to OA &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&le; 3 Days<br>
                &bull; OA to Delivery &le; 3 Days<br>
                &bull; Empty to Term &nbsp;&le; 3 Days<br>
                &bull; E2E / OTP &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&ge; 95%
                </div>""",
                unsafe_allow_html=True
            )

        st.markdown("---")

        # ── Shared Plotly style ────────────────────────────────────────────────────
        _CS = dict(
            plot_bgcolor="rgba(0,0,0,0)",
            paper_bgcolor="rgba(0,0,0,0)",
            font=dict(color="#dddddd", size=11),
            margin=dict(t=36, b=8, l=5, r=5),
            showlegend=False,
        )
        _XAXIS = dict(gridcolor="#252525", linecolor="#444", tickfont=dict(size=10))
        _YAXIS = dict(gridcolor="#252525", linecolor="#444")

        def _lc(col, name, color, suffix="", sla_line=None):
            fig = go.Figure()
            fig.add_trace(go.Scatter(
                x=_wdf["wl"], y=_wdf[col], name=name,
                mode="lines+markers",
                line=dict(color=color, width=2.5), marker=dict(size=7, color=color),
                hovertemplate=f"%{{x}}: <b>%{{y}}{suffix}</b><extra></extra>"
            ))
            if sla_line is not None:
                fig.add_hline(y=sla_line, line_dash="dash", line_color="#E8A838",
                              line_width=1.5,
                              annotation_text=f"Goal: {sla_line}{suffix}",
                              annotation_font_color="#E8A838",
                              annotation_font_size=10)
            return fig

        # ── Performance Trends — 2 rows × 3 charts, exact slide titles + style ────
        # Single line color (#4BACC6 teal), orange goal lines (#E8A838) — matches slide
        _C_LINE = "#4BACC6"   # teal — same as C_DOT on slide
        _C_GOAL = "#E8A838"   # orange — same as C_ACCENT on slide

        st.markdown("#### Performance Trends")
        _r1a, _r1b, _r1c = st.columns(3)

        with _r1a:
            _f = _lc("av_oa_avg", "Leg: AV to OA (avg. days)", _C_LINE, "d", sla_line=3)
            _f.update_layout(title="Leg: AV to OA (avg. days)", yaxis=dict(**_YAXIS), xaxis=_XAXIS, **_CS)
            st.plotly_chart(_f, use_container_width=True)

        with _r1b:
            _f = _lc("oa_del_avg", "Leg: OA to Delivery (avg. days)", _C_LINE, "d", sla_line=3)
            _f.update_layout(title="Leg: OA to Delivery (avg. days)", yaxis=dict(**_YAXIS), xaxis=_XAXIS, **_CS)
            st.plotly_chart(_f, use_container_width=True)

        with _r1c:
            _f = _lc("e2e_avg", "Leg: E2E Transit (avg. days)", _C_LINE, "d")
            _f.update_layout(title="Leg: E2E Transit (avg. days)", yaxis=dict(**_YAXIS), xaxis=_XAXIS, **_CS)
            st.plotly_chart(_f, use_container_width=True)

        # ── Row 2 ─────────────────────────────────────────────────────────────────
        _r2a, _r2b, _r2c = st.columns(3)

        with _r2a:
            _f = _lc("empty_term_pct", "Leg: Empty to Termination (%)", _C_LINE, "%")
            _f.update_layout(title="Leg: Empty to Termination (%)",
                             yaxis=dict(range=[0,110], ticksuffix="%", **_YAXIS),
                             xaxis=_XAXIS, **_CS)
            st.plotly_chart(_f, use_container_width=True)

        with _r2b:
            _f = _lc("otp_pct", "Leg: On-Time to Promise %", _C_LINE, "%", sla_line=95)
            _f.update_layout(title="Leg: On-Time to Promise %",
                             yaxis=dict(range=[0,110], ticksuffix="%", **_YAXIS),
                             xaxis=_XAXIS, **_CS)
            st.plotly_chart(_f, use_container_width=True)

        with _r2c:
            _f = _lc("containers", "Volume (containers)", _C_LINE, "")
            _f.update_layout(title="Volume (containers)", yaxis=dict(**_YAXIS), xaxis=_XAXIS, **_CS)
            st.plotly_chart(_f, use_container_width=True)

        # ── Weekly Performance Table ───────────────────────────────────────────────
        st.markdown("---")
        st.markdown("#### Weekly Performance")

        _METRICS = [
            ("Containers",     "containers",     lambda v: str(int(v)) if v is not None else "—"),
            ("AV→OA SLA%",     "av_oa_sla_pct",  lambda v: f"{int(v)}%" if v is not None else "—"),
            ("AV→OA Avg (d)",  "av_oa_avg",      lambda v: f"{v:.1f}" if v is not None else "—"),
            ("AV→OA P90 (d)",  "av_oa_p90",      lambda v: str(int(v)) if v is not None else "—"),
            ("OA→Del SLA%",    "oa_del_sla_pct", lambda v: f"{int(v)}%" if v is not None else "—"),
            ("OA→Del Avg (d)", "oa_del_avg",     lambda v: f"{v:.1f}" if v is not None else "—"),
            ("OA→Del P90 (d)", "oa_del_p90",     lambda v: str(int(v)) if v is not None else "—"),
            ("Empty→Term%",    "empty_term_pct", lambda v: f"{int(v)}%" if v is not None else "—"),
            ("E2E Avg (d)",    "e2e_avg",        lambda v: str(int(v)) if v is not None else "—"),
            ("OTP%",           "otp_pct",        lambda v: f"{int(v)}%" if v is not None else "—"),
        ]

        _tbl = {"Metric": [m[0] for m in _METRICS]}
        for _, _wr in _wdf.iterrows():
            _ws_short = str(_wr.get("week_start") or "")[:5]
            _col_h = f"W{int(_wr['week_num'])}" + (f"\n{_ws_short}" if _ws_short else "")
            _tbl[_col_h] = [_fmt(_vv(_wr, col)) for _, col, _fmt in _METRICS]

        # ── Totals column — volume sum; all other metrics weighted avg by containers
        def _wtot(col):
            """Weighted average of col across _wdf, weighted by containers."""
            _rows = _wdf[["containers", col]].dropna()
            if _rows.empty: return None
            _wts = _rows["containers"].astype(float)
            if _wts.sum() == 0: return None
            return (_rows[col].astype(float) * _wts).sum() / _wts.sum()

        _tot_vals = []
        for _, col, _fmt in _METRICS:
            if col == "containers":
                _v = _wdf["containers"].dropna().sum()
                _tot_vals.append(str(int(_v)) if _v else "—")
            else:
                _v = _wtot(col)
                _tot_vals.append(_fmt(round(_v, 1) if _v is not None else None))
        _tbl["Totals"] = _tot_vals

        _tbl_df = pd.DataFrame(_tbl)
        st.dataframe(_tbl_df, use_container_width=True, hide_index=True,
                     column_config={"Metric": st.column_config.TextColumn(width="medium"),
                                    "Totals": st.column_config.TextColumn(width="small")})

        # ── Carrier Performance ────────────────────────────────────────────────────
        if not _car_wbr.empty:
            _cdf = _car_wbr[
                (_car_wbr["year"] == _sel_yr) &
                (_car_wbr["week_num"].isin(_sel_wks))
            ].copy()
            if _sel_carr:
                _cdf = _cdf[_cdf["carrier"].isin(_sel_carr)]

            if not _cdf.empty:
                st.markdown("---")
                st.markdown("#### Carrier Performance")
                _cagg = (
                    _cdf.groupby("carrier", as_index=False)
                    .agg(
                        Containers    =("volume",         "sum"),
                        **{"AV→OA SLA%":    ("av_oa_sla_pct",  "mean"),
                           "OA→Del SLA%":   ("oa_del_sla_pct", "mean"),
                           "AV→OA Avg (d)": ("av_oa_avg",      "mean"),
                           "OA→Del Avg (d)":("oa_del_avg",     "mean")}
                    )
                    .round(1)
                    .sort_values("Containers", ascending=False)
                )
                _fc = go.Figure()
                _fc.add_trace(go.Bar(
                    x=_cagg["carrier"], y=_cagg["AV→OA SLA%"],
                    name="AV→OA SLA%", marker_color="#FF6B35"
                ))
                _fc.add_trace(go.Bar(
                    x=_cagg["carrier"], y=_cagg["OA→Del SLA%"],
                    name="OA→Del SLA%", marker_color="#4BACC6"
                ))
                _fc.add_hline(y=85, line_dash="dash", line_color="#555")
                _fc.update_layout(
                    barmode="group", title="Carrier SLA Performance",
                    yaxis=dict(range=[0,110], ticksuffix="%", **_YAXIS),
                    xaxis=_XAXIS, showlegend=True,
                    legend=dict(orientation="h", y=-0.2),
                    **{k:v for k,v in _CS.items() if k != "showlegend"},
                )
                st.plotly_chart(_fc, use_container_width=True)
                st.dataframe(
                    _cagg.rename(columns={"carrier": "Carrier"}),
                    use_container_width=True, hide_index=True
                )
            else:
                st.info("Carrier breakdown available after the next WBR Generator run.")

        # ── Export to Slide ────────────────────────────────────────────────────────
        st.markdown("---")
        st.markdown("#### Export to Slide")
        st.caption("Generates the standard WBR slide from DB data — no source file re-upload needed.")

        _ex1, _ex2 = st.columns([2, 1])
        with _ex1:
            _exp_wk = st.selectbox(
                "Week to export",
                options=sorted(_sel_wks, reverse=True),
                format_func=lambda w: f"W{w}",
                key="d_exp_wk"
            )
        with _ex2:
            _exp_rd = st.date_input("Report date", value=_dt_cls.today(), key="d_exp_rd")

        if st.button("Generate Slide", type="primary", key="d_gen_slide"):
            with st.spinner("Generating slide from database..."):
                try:
                    from wbr_pdf   import generate_standard_wbr as _gwbr
                    from wbr_pptx  import pdf_to_pptx as _g2pptx
                    from wbr_engine import load_weeks_from_db as _lwdb, compute_totals as _ctot

                    _ec   = get_db()
                    _wns  = list(range(max(1, _exp_wk - 5), _exp_wk + 1))
                    _ewd  = _lwdb(_ec, _sel_yr, _wns)
                    _ec.close()

                    # load_weeks_from_db returns a dict keyed by week_num;
                    # generate_standard_wbr and compute_totals expect an ordered list
                    _weeks_list = [_ewd.get(w) for w in _wns]

                    _epdf = _gwbr(
                        week_labels=[ f"W{w}" for w in _wns ],
                        weeks_data=_weeks_list,
                        totals=_ctot(_weeks_list),
                        report_date=_exp_rd,
                        current_week_label=f"W{_exp_wk}",
                    )
                    _efname = f"GLS_Robotics_{_exp_rd.year}-{_exp_rd.month}-{_exp_rd.day}.pdf"

                    import fitz as _ftz
                    _edoc = _ftz.open(stream=_epdf, filetype="pdf")
                    _epng = _edoc[0].get_pixmap(matrix=_ftz.Matrix(2.0, 2.0), alpha=False).tobytes("png")
                    _edoc.close()
                    st.image(_epng, use_container_width=True, caption=f"W{_exp_wk} · {_exp_rd}")

                    _dl1, _dl2 = st.columns(2)
                    with _dl1:
                        st.download_button(
                            label=f"⬇️ PDF — {_efname}",
                            data=_epdf, file_name=_efname,
                            mime="application/pdf", key="d_dl_pdf"
                        )
                    with _dl2:
                        try:
                            _epptx = _g2pptx(_epdf, dpi=200)
                            st.download_button(
                                label=f"⬇️ PPTX — {_efname.replace('.pdf','.pptx')}",
                                data=_epptx,
                                file_name=_efname.replace(".pdf", ".pptx"),
                                mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
                                key="d_dl_pptx"
                            )
                        except Exception as _pe:
                            st.warning(f"PPTX error: {_pe}")
                except Exception as _ee:
                    st.error(f"Export failed: {_ee}")
                    import traceback; st.code(traceback.format_exc())



with tab10:
    _ifc_conn = get_db()
    render_inbound_forecast_tab(_ifc_conn)
    _ifc_conn.close()




# ══════════════════════════════════════════════════════════════════════════════
# TAB — DBR Dashboard
# ══════════════════════════════════════════════════════════════════════════════
with tabDBR:
    st.subheader("DBR Dashboard")
    st.caption("Live view of container lifecycle: port to yard to FC to empty return. Updated each time a carrier submits their daily DBR.")

    _dbr_conn = get_db()
    _ic_all   = pd.read_sql("SELECT * FROM inbound_containers ORDER BY fc_sched DESC", _dbr_conn)
    _cc_all   = pd.read_sql("SELECT * FROM carrier_contacts ORDER BY scac", _dbr_conn)
    _30d_ago  = (datetime.now(_PHX).date() - timedelta(days=30)).isoformat()
    _rcpt_all = pd.read_sql(
        "SELECT carrier, received_date FROM dbr_receipts WHERE received_date >= ?",
        _dbr_conn, params=(_30d_ago,)
    )
    _dbr_conn.close()

    _today_dbr    = datetime.now(_PHX).date()
    _tomorrow_dbr = _today_dbr + timedelta(days=1)

    if not _ic_all.empty:
        _ic_all["status"] = _ic_all["status"].fillna("pending").str.lower().str.strip()

    _kpi_total  = len(_ic_all)
    _kpi_yard   = int((_ic_all["status"] == "at_yard").sum())                              if not _ic_all.empty else 0
    _kpi_del    = int(_ic_all["status"].isin(["delivered", "empty_returned"]).sum())       if not _ic_all.empty else 0
    _kpi_pend   = int((_ic_all["status"] == "pending").sum())                              if not _ic_all.empty else 0
    _kpi_await  = int((_ic_all["fc_actual"].notna() & _ic_all["empty_returned"].isna()).sum()) if not _ic_all.empty else 0
    _kpi_empret = int(_ic_all["empty_returned"].notna().sum())                             if not _ic_all.empty else 0

    _km1, _km2, _km3, _km4, _km5, _km6 = st.columns(6)
    _km1.metric("Total Containers",     _kpi_total)
    _km2.metric("At Yard",              _kpi_yard)
    _km3.metric("Delivered to FC",      _kpi_del)
    _km4.metric("Pending FC Delivery",  _kpi_pend)
    _km5.metric("Awaiting Empty Return", _kpi_await)
    _km6.metric("Empty Returned",       _kpi_empret)

    st.divider()

    _dt2, _dt3, _dt4, _dt5 = st.tabs([
        "Today / Upcoming",
        "Late / At Risk",
        "Empty Returns",
        "By Carrier",
    ])

    # ── 2: TODAY / UPCOMING ──────────────────────────────────────────────────
    with _dt2:
        st.markdown("#### Scheduled FC Deliveries \u2014 Today & Tomorrow")
        _t2_opts   = sorted(_ic_all["carrier"].dropna().unique().tolist()) if not _ic_all.empty else []
        _t2_filter = st.multiselect("Filter by carrier", _t2_opts, key="dt2_carrier")

        def _upcoming(df, d):
            return df[(df["fc_sched"] == d.isoformat()) & ~df["status"].isin(["delivered","empty_returned"])].copy()

        _td_rows = _upcoming(_ic_all, _today_dbr)
        _tm_rows = _upcoming(_ic_all, _tomorrow_dbr)
        if _t2_filter:
            _td_rows = _td_rows[_td_rows["carrier"].isin(_t2_filter)]
            _tm_rows = _tm_rows[_tm_rows["carrier"].isin(_t2_filter)]

        _UP = ["container_id","carrier","fc_dest","yard_actual","fc_sched","status","notes"]

        st.markdown("**Today \u2014 " + _today_dbr.strftime("%A, %B %d") + "** (" + str(len(_td_rows)) + " containers)")
        if _td_rows.empty:
            st.info("No FC deliveries scheduled for today.")
        else:
            _d1 = _td_rows[[c for c in _UP if c in _td_rows.columns]].copy()
            _d1.columns = [c.replace("_"," ").title() for c in _d1.columns]
            st.dataframe(_d1, use_container_width=True, hide_index=True)

        st.markdown("**Tomorrow \u2014 " + _tomorrow_dbr.strftime("%A, %B %d") + "** (" + str(len(_tm_rows)) + " containers)")
        if _tm_rows.empty:
            st.info("No FC deliveries scheduled for tomorrow.")
        else:
            _d2 = _tm_rows[[c for c in _UP if c in _tm_rows.columns]].copy()
            _d2.columns = [c.replace("_"," ").title() for c in _d2.columns]
            st.dataframe(_d2, use_container_width=True, hide_index=True)

    # ── 3: LATE / AT RISK ────────────────────────────────────────────────────
    with _dt3:
        st.markdown("#### Late Containers \u2014 Past Scheduled FC Delivery")
        _t3_opts   = sorted(_ic_all["carrier"].dropna().unique().tolist()) if not _ic_all.empty else []
        _t3_filter = st.multiselect("Filter by carrier", _t3_opts, key="dt3_carrier")

        _late_df = _ic_all[
            _ic_all["fc_sched"].notna()
            & (_ic_all["fc_sched"] < _today_dbr.isoformat())
            & ~_ic_all["status"].isin(["delivered","empty_returned"])
        ].copy()
        if _t3_filter:
            _late_df = _late_df[_late_df["carrier"].isin(_t3_filter)]

        if _late_df.empty:
            st.success("No containers past their scheduled FC delivery date.")
        else:
            _late_df["Days Late"] = _late_df["fc_sched"].apply(
                lambda s: (_today_dbr - date.fromisoformat(str(s)[:10])).days if s else None)
            _ld = _late_df[["container_id","carrier","fc_dest","fc_sched","Days Late","status","notes"]].copy()
            _ld.columns = ["Container","Carrier","FC","FC Sched","Days Late","Status","Notes"]
            _ld = _ld.sort_values("Days Late", ascending=False)
            st.dataframe(_ld, use_container_width=True, hide_index=True)
            st.caption(str(len(_ld)) + " containers | Worst: " + str(int(_ld["Days Late"].max())) + " days late")

    # ── 4: EMPTY RETURNS ─────────────────────────────────────────────────────
    with _dt4:
        st.markdown("#### Delivered to FC \u2014 Awaiting Empty Return")
        _t4_opts   = sorted(_ic_all["carrier"].dropna().unique().tolist()) if not _ic_all.empty else []
        _t4_filter = st.multiselect("Filter by carrier", _t4_opts, key="dt4_carrier")

        _aw_df = _ic_all[_ic_all["fc_actual"].notna() & _ic_all["empty_returned"].isna()].copy()
        if _t4_filter:
            _aw_df = _aw_df[_aw_df["carrier"].isin(_t4_filter)]

        if _aw_df.empty:
            st.success("No containers awaiting empty return.")
        else:
            _aw_df["Days at FC"] = _aw_df["fc_actual"].apply(
                lambda s: (_today_dbr - date.fromisoformat(str(s)[:10])).days if s else None)
            _ad = _aw_df[["container_id","carrier","fc_dest","fc_actual","Days at FC","notes"]].copy()
            _ad.columns = ["Container","Carrier","FC","FC Actual Del","Days at FC","Notes"]
            _ad = _ad.sort_values("Days at FC", ascending=False)
            st.dataframe(_ad, use_container_width=True, hide_index=True)
            st.caption(str(len(_ad)) + " containers | Longest dwell: " + str(int(_ad["Days at FC"].max())) + " days")

    # ── 5: BY CARRIER ────────────────────────────────────────────────────────
    with _dt5:
        _car_opts5 = sorted(_ic_all["carrier"].dropna().unique().tolist()) if not _ic_all.empty else []
        _sel_c     = st.selectbox("Select carrier", _car_opts5, key="dt5_carrier_sel")

        if _sel_c:
            _cdf = _ic_all[_ic_all["carrier"] == _sel_c].copy()

            _ck1, _ck2, _ck3, _ck4, _ck5 = st.columns(5)
            _ck1.metric("Total",          len(_cdf))
            _ck2.metric("At Yard",        int((_cdf["status"] == "at_yard").sum()))
            _ck3.metric("Delivered",      int(_cdf["status"].isin(["delivered","empty_returned"]).sum()))
            _ck4.metric("Pending FC",     int((_cdf["status"] == "pending").sum()))
            _ck5.metric("Awaiting Empty", int((_cdf["fc_actual"].notna() & _cdf["empty_returned"].isna()).sum()))

            _ddf = _cdf[_cdf["fc_actual"].notna() & _cdf["fc_sched"].notna()].copy()
            if not _ddf.empty:
                _ddf["_ot"] = _ddf["fc_actual"] <= _ddf["fc_sched"]
                st.metric("On-Time Delivery %", str(int(_ddf["_ot"].mean() * 100)) + "%")

            _ls5 = _rcpt_all[_rcpt_all["carrier"] == _sel_c]["received_date"].max() if not _rcpt_all.empty else None
            st.caption("Last DBR received: **" + (_ls5 or "No submissions on record") + "**")

            st.divider()
            st.markdown("#### All Containers \u2014 " + _sel_c)

            _cf1, _cf2, _cf3 = st.columns(3)
            _fc_f = _cf1.multiselect("FC",     sorted(_cdf["fc_dest"].dropna().unique().tolist()), key="dt5_fc_filt")
            _st_f = _cf2.multiselect("Status", sorted(_cdf["status"].dropna().unique().tolist()),  key="dt5_st_filt")
            _dr_m = _cf3.selectbox("Date range",
                                    ["All time","Last 30 days","Last 90 days","Custom range"],
                                    key="dt5_dr_mode")

            _cv = _cdf.copy()
            if _fc_f:
                _cv = _cv[_cv["fc_dest"].isin(_fc_f)]
            if _st_f:
                _cv = _cv[_cv["status"].isin(_st_f)]
            if _dr_m == "Last 30 days":
                _cv = _cv[_cv["fc_sched"] >= (_today_dbr - timedelta(30)).isoformat()]
            elif _dr_m == "Last 90 days":
                _cv = _cv[_cv["fc_sched"] >= (_today_dbr - timedelta(90)).isoformat()]
            elif _dr_m == "Custom range":
                _r1, _r2 = st.columns(2)
                _rfrom = _r1.date_input("From", value=_today_dbr - timedelta(90), key="dt5_dr_from")
                _rto   = _r2.date_input("To",   value=_today_dbr,                key="dt5_dr_to")
                _cv = _cv[(_cv["fc_sched"] >= _rfrom.isoformat()) & (_cv["fc_sched"] <= _rto.isoformat())]

            _VC = ["container_id","fc_dest","port_region","yard_sched","yard_actual",
                   "fc_sched","fc_actual","empty_returned","status","notes"]
            _cvd = _cv[[c for c in _VC if c in _cv.columns]].copy()
            _cvd.columns = [c.replace("_"," ").title() for c in _cvd.columns]
            st.caption("Showing " + str(len(_cvd)) + " of " + str(len(_cdf)) + " containers")
            st.dataframe(_cvd, use_container_width=True, hide_index=True)

            st.download_button(
                "Download " + _sel_c + " data (.csv)",
                data=_cvd.to_csv(index=False).encode(),
                file_name=_sel_c + "_containers_" + str(_today_dbr) + ".csv",
                mime="text/csv",
                key="dt5_export",
            )

            st.divider()
            st.markdown("#### Carrier Contact")
            st.caption("Save the email address here to enable reminder sending from Overview.")
            _ccm = _cc_all[_cc_all["scac"] == _sel_c]
            if not _ccm.empty:
                _cce = _ccm[["scac","display_name","contact_name","email","reminder_enabled"]].copy()
                _cced = st.data_editor(
                    _cce, use_container_width=True, hide_index=True,
                    column_config={
                        "scac":             st.column_config.TextColumn("SCAC", disabled=True),
                        "display_name":     st.column_config.TextColumn("Display Name"),
                        "contact_name":     st.column_config.TextColumn("Contact Name"),
                        "email":            st.column_config.TextColumn("Email"),
                        "reminder_enabled": st.column_config.CheckboxColumn("Reminders On"),
                    },
                    key="dt5_cc_editor",
                )
                if st.button("Save Contact Info", key="dt5_cc_save"):
                    _sv = get_db()
                    for _, _r5 in _cced.iterrows():
                        _sv.execute(
                            "UPDATE carrier_contacts SET display_name=?,contact_name=?,email=?,"
                            "reminder_enabled=?,updated_at=datetime('now') WHERE scac=?",
                            (_r5["display_name"], _r5["contact_name"],
                             _r5["email"], int(_r5["reminder_enabled"]), _r5["scac"])
                        )
                    _sv.commit(); _sv.close()
                    if S3_ENABLED:
                        data_sync.push_db_to_s3(AWS_KEY, AWS_SECRET, AWS_REGION, S3_BUCKET)
                    st.success("Saved.")
                    st.rerun()

                # ── Portal Password ───────────────────────────────────────────────
                st.divider()
                st.markdown("#### Portal Password")
                st.caption(
                    "Set this carrier\'s unique login password for the vendor submission portal. "
                    "The password is stored as a one-way hash — it cannot be recovered, only reset."
                )
                with st.form("dt5_pw_form_" + _sel_c):
                    _pw1 = st.text_input("New password", type="password", key="dt5_pw1")
                    _pw2 = st.text_input("Confirm password", type="password", key="dt5_pw2")
                    _pw_save = st.form_submit_button("Set Password", type="primary")
                if _pw_save:
                    if not _pw1:
                        st.warning("Password cannot be empty.")
                    elif _pw1 != _pw2:
                        st.error("Passwords do not match.")
                    else:
                        import hashlib
                        _hashed = hashlib.sha256(_pw1.encode()).hexdigest()
                        _pw_conn = get_db()
                        _pw_conn.execute(
                            "UPDATE carrier_contacts SET portal_password=?, updated_at=datetime(\'now\') WHERE scac=?",
                            (_hashed, _sel_c)
                        )
                        _pw_conn.commit(); _pw_conn.close()
                        if S3_ENABLED:
                            data_sync.push_db_to_s3(AWS_KEY, AWS_SECRET, AWS_REGION, S3_BUCKET)
                        st.success("Portal password set for " + _sel_c + ".")
            else:
                st.info("No contact record for " + _sel_c + ". Will be created on next DBR submission.")

