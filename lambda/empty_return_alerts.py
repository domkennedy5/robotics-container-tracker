"""
lambda/empty_return_alerts.py

Scheduled daily (EventBridge, 8 AM EST) — checks the DBR's Empty Returns
sheet and emails alerts for containers overdue or due within 2 days.
Deduplicates: won't re-alert on the same container same day.

Environment variables:
  S3_BUCKET         = robotics-container-tracker
  DBR_S3_KEY        = dbr/latest.xlsx
  DB_S3_KEY         = db/tracker.db
  SES_FROM_EMAIL    = containers-noreply@yourdomain.com
  ALERT_RECIPIENTS  = email1@amazon.com,email2@amazon.com
  AWS_REGION        = us-east-1
"""
import os
import io
import json
import sqlite3
import logging
from datetime import datetime, date, timezone, timedelta

import boto3
import openpyxl

logger = logging.getLogger()
logger.setLevel(logging.INFO)

S3_BUCKET        = os.environ["S3_BUCKET"]
DBR_S3_KEY       = os.environ.get("DBR_S3_KEY", "dbr/latest.xlsx")
DB_S3_KEY        = os.environ.get("DB_S3_KEY", "db/tracker.db")
SES_FROM_EMAIL   = os.environ["SES_FROM_EMAIL"]
ALERT_RECIPIENTS = [r.strip() for r in os.environ.get("ALERT_RECIPIENTS", "").split(",") if r.strip()]
REGION           = os.environ.get("AWS_REGION", "us-east-1")

s3  = boto3.client("s3")
ses = boto3.client("ses", region_name=REGION)


def load_empty_returns() -> list:
    """Pull DBR from S3 and extract Empty Returns rows."""
    obj  = s3.get_object(Bucket=S3_BUCKET, Key=DBR_S3_KEY)
    data = obj["Body"].read()
    wb   = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
    if "Empty Returns" not in wb.sheetnames:
        logger.warning("No 'Empty Returns' sheet in DBR")
        return []
    ws   = wb["Empty Returns"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    result  = []
    for row in rows[1:]:
        rec = dict(zip(headers, row))
        if not rec.get("Container #"):
            continue
        result.append(rec)
    wb.close()
    return result


def already_alerted(conn, alert_type: str, container_id: str) -> bool:
    today = date.today().isoformat()
    row = conn.execute(
        """SELECT id FROM alerts_log
           WHERE alert_type=? AND container_id=? AND sent_at >= ?""",
        (alert_type, container_id, today)
    ).fetchone()
    return row is not None


def load_db() -> tuple:
    local = "/tmp/tracker.db"
    try:
        obj = s3.get_object(Bucket=S3_BUCKET, Key=DB_S3_KEY)
        with open(local, "wb") as f:
            f.write(obj["Body"].read())
    except Exception:
        pass
    conn = sqlite3.connect(local)
    conn.execute("""CREATE TABLE IF NOT EXISTS alerts_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        alert_type TEXT, container_id TEXT, sent_at TEXT, recipient TEXT)""")
    conn.commit()
    return local, conn


def save_db(path: str):
    with open(path, "rb") as f:
        s3.put_object(Bucket=S3_BUCKET, Key=DB_S3_KEY, Body=f.read())


def send_alert(subject: str, rows: list):
    if not ALERT_RECIPIENTS or not rows:
        return
    header = "Container #    | Terminal | Due Date   | Days\n"
    header += "-" * 55 + "\n"
    lines  = "\n".join(
        f"{r['container']:<16}| {r['terminal'] or 'N/A':<9}| {r['due_date']:<11}| {r['days']}"
        for r in rows
    )
    body = f"The following containers require attention:\n\n{header}{lines}\n\n— AGL Container Tracker"
    try:
        ses.send_email(
            Source=SES_FROM_EMAIL,
            Destination={"ToAddresses": ALERT_RECIPIENTS},
            Message={
                "Subject": {"Data": subject},
                "Body":    {"Text": {"Data": body}},
            },
        )
        logger.info(f"Alert sent: {subject} ({len(rows)} containers)")
    except Exception as e:
        logger.error(f"SES send failed: {e}")


def handler(event, context):
    today     = date.today()
    threshold = today + timedelta(days=2)

    er_rows = load_empty_returns()
    if not er_rows:
        logger.info("No empty return rows found")
        return

    db_path, conn = load_db()
    overdue_list, due_soon_list = [], []

    for rec in er_rows:
        status = str(rec.get("Status", "")).upper()
        if status == "TERMINATED":
            continue

        due_raw = rec.get("Empty Return Due Date")
        if not due_raw:
            continue

        try:
            if isinstance(due_raw, (datetime, date)):
                due = due_raw.date() if hasattr(due_raw, "date") else due_raw
            else:
                due = datetime.strptime(str(due_raw)[:10], "%Y-%m-%d").date()
        except Exception:
            continue

        container = str(rec.get("Container #", "")).strip()
        terminal  = str(rec.get("Terminal", "")).strip()
        days_val  = (due - today).days
        row_data  = {"container": container, "terminal": terminal,
                     "due_date": due.isoformat(), "days": days_val}

        if days_val < 0:
            atype = "empty_return_overdue"
            if not already_alerted(conn, atype, container):
                overdue_list.append(row_data)
                conn.execute(
                    "INSERT INTO alerts_log (alert_type, container_id, sent_at) VALUES (?,?,?)",
                    (atype, container, datetime.now(timezone.utc).isoformat())
                )
        elif days_val <= 2:
            atype = "empty_return_due_soon"
            if not already_alerted(conn, atype, container):
                due_soon_list.append(row_data)
                conn.execute(
                    "INSERT INTO alerts_log (alert_type, container_id, sent_at) VALUES (?,?,?)",
                    (atype, container, datetime.now(timezone.utc).isoformat())
                )

    conn.commit()

    if overdue_list:
        send_alert(
            f"[ACTION REQUIRED] {len(overdue_list)} Overdue Empty Return(s) — {today}",
            overdue_list
        )
    if due_soon_list:
        send_alert(
            f"[REMINDER] {len(due_soon_list)} Empty Return(s) Due Within 2 Days — {today}",
            due_soon_list
        )

    conn.close()
    save_db(db_path)

    logger.info(f"Done — {len(overdue_list)} overdue, {len(due_soon_list)} due soon")
    return {
        "statusCode": 200,
        "body": json.dumps({"overdue": len(overdue_list), "due_soon": len(due_soon_list)})
    }
