"""
lambda/email_ingest.py

Triggered by S3 ObjectCreated on prefix: inbound/
Flow: SES receives carrier email → saves raw .eml to S3 →
      this Lambda parses it, extracts Excel/CSV attachments,
      inserts container rows into tracker.db (via S3), replies to sender.

Environment variables (set in Lambda console):
  S3_BUCKET       = robotics-container-tracker
  DB_S3_KEY       = db/tracker.db
  SES_FROM_EMAIL  = containers-noreply@yourdomain.com
  AWS_REGION      = us-east-1  (set automatically by Lambda runtime)
"""
import os
import io
import json
import email
import sqlite3
import logging
import tempfile
from datetime import datetime, timezone
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

import boto3

# Lambda runtime provides boto3 — openpyxl/pandas must be in a Lambda Layer
import sys
sys.path.insert(0, "/opt/python")  # Lambda layer path

logger = logging.getLogger()
logger.setLevel(logging.INFO)

S3_BUCKET      = os.environ["S3_BUCKET"]
DB_S3_KEY      = os.environ.get("DB_S3_KEY", "db/tracker.db")
SES_FROM_EMAIL = os.environ["SES_FROM_EMAIL"]
REGION         = os.environ.get("AWS_REGION", "us-east-1")

s3  = boto3.client("s3")
ses = boto3.client("ses", region_name=REGION)


# ── utils (duplicated from utils.py for Lambda self-containment) ──────────────

def normalize_container(cid: str) -> str:
    import re
    return re.sub(r"[-\s]", "", str(cid).strip().upper())


def parse_carrier_file_bytes(file_bytes: bytes, filename: str) -> list:
    """Returns list of dicts with keys: container_id, terminal, status, notes."""
    rows_all = []
    try:
        import openpyxl
        import re
        if filename.lower().endswith(".csv"):
            import csv
            reader = csv.DictReader(io.StringIO(file_bytes.decode("utf-8", errors="replace")))
            for row in reader:
                cont_key = next((k for k in row if "container" in k.lower()), None)
                if cont_key and row[cont_key].strip():
                    rows_all.append(_extract(row, cont_key))
        else:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
            for sname in wb.sheetnames:
                ws = wb[sname]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                header_idx = next(
                    (i for i, r in enumerate(rows[:6])
                     if any(c and "container" in str(c).lower() for c in r)),
                    None
                )
                if header_idx is None:
                    continue
                headers = [str(h).strip() if h else f"col_{i}"
                           for i, h in enumerate(rows[header_idx])]
                cont_col = next((h for h in headers if "container" in h.lower()), None)
                if not cont_col:
                    continue
                for row in rows[header_idx + 1:]:
                    if not any(row):
                        continue
                    rec = dict(zip(headers, row))
                    val = rec.get(cont_col)
                    if val and str(val).strip():
                        rows_all.append(_extract(rec, cont_col))
            wb.close()
    except Exception as e:
        logger.error(f"parse error for {filename}: {e}")
    return rows_all


def _extract(rec: dict, cont_col: str) -> dict:
    def find(*keys):
        for k in rec:
            if any(kw in str(k).lower() for kw in keys):
                v = rec[k]
                if v and str(v).strip() not in ("None", "nan", ""):
                    return str(v).strip()
        return None
    return {
        "container_id": normalize_container(str(rec[cont_col])),
        "terminal":     find("terminal", "port"),
        "status":       find("status", "state"),
        "notes":        find("note", "comment", "reason"),
    }


# ── DB helpers ─────────────────────────────────────────────────────────────────

def load_db() -> tuple:
    """Pull tracker.db from S3 to /tmp, return (path, conn)."""
    local = "/tmp/tracker.db"
    try:
        obj = s3.get_object(Bucket=S3_BUCKET, Key=DB_S3_KEY)
        with open(local, "wb") as f:
            f.write(obj["Body"].read())
    except s3.exceptions.NoSuchKey:
        pass  # fresh DB
    conn = sqlite3.connect(local)
    conn.execute("""CREATE TABLE IF NOT EXISTS carrier_submissions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        submitted_at TEXT, carrier_name TEXT, container_id TEXT,
        terminal TEXT, status TEXT, notes TEXT, source_file TEXT, source TEXT)""")
    conn.commit()
    return local, conn


def save_db(local_path: str):
    """Push local tracker.db back to S3."""
    with open(local_path, "rb") as f:
        s3.put_object(Bucket=S3_BUCKET, Key=DB_S3_KEY, Body=f.read())


# ── email helpers ──────────────────────────────────────────────────────────────

def send_reply(to_addr: str, subject: str, containers: list, count: int):
    container_list = "\n".join(f"  • {c}" for c in containers[:50])
    if len(containers) > 50:
        container_list += f"\n  … and {len(containers) - 50} more"
    body = (
        f"Hi,\n\n"
        f"We received your submission and logged {count} container(s).\n\n"
        f"Containers processed:\n{container_list}\n\n"
        f"Submitted at: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}\n\n"
        f"If anything looks wrong, reply to this email.\n\n"
        f"— Amazon Global Logistics Container Tracker"
    )
    try:
        ses.send_email(
            Source=SES_FROM_EMAIL,
            Destination={"ToAddresses": [to_addr]},
            Message={
                "Subject": {"Data": f"Re: {subject} — Received ✓"},
                "Body":    {"Text": {"Data": body}},
            },
        )
        logger.info(f"Reply sent to {to_addr}")
    except Exception as e:
        logger.error(f"Failed to send reply to {to_addr}: {e}")


# ── handler ────────────────────────────────────────────────────────────────────

def handler(event, context):
    record    = event["Records"][0]["s3"]
    bucket    = record["bucket"]["name"]
    key       = record["object"]["key"]

    logger.info(f"Processing s3://{bucket}/{key}")

    # fetch raw email from S3
    obj      = s3.get_object(Bucket=bucket, Key=key)
    raw_mail = obj["Body"].read()
    msg      = email.message_from_bytes(raw_mail)

    sender  = email.utils.parseaddr(msg.get("From", ""))[1]
    subject = msg.get("Subject", "(no subject)")

    # extract carrier name from sender or subject
    carrier_name = sender.split("@")[0] if sender else "Unknown"

    # find and parse attachments
    all_containers = []
    for part in msg.walk():
        ct   = part.get_content_type()
        disp = str(part.get("Content-Disposition", ""))
        fn   = part.get_filename() or ""

        if "attachment" not in disp and ct not in (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
            "text/csv",
        ):
            continue

        payload = part.get_payload(decode=True)
        if not payload:
            continue

        rows = parse_carrier_file_bytes(payload, fn or "attachment.xlsx")
        all_containers.extend(rows)
        logger.info(f"Parsed {len(rows)} containers from attachment: {fn}")

    if not all_containers:
        logger.warning(f"No containers found in email from {sender}")
        # still reply so carrier knows we got it
        send_reply(sender, subject, [], 0)
        return {"statusCode": 200, "body": "no containers found"}

    # write to DB
    db_path, conn = load_db()
    now = datetime.now(timezone.utc).isoformat()
    for row in all_containers:
        conn.execute(
            """INSERT INTO carrier_submissions
               (submitted_at, carrier_name, container_id, terminal,
                status, notes, source_file, source)
               VALUES (?,?,?,?,?,?,?,?)""",
            (now, carrier_name, row["container_id"],
             row.get("terminal"), row.get("status"),
             row.get("notes"), fn, "email")
        )
    conn.commit()
    conn.close()
    save_db(db_path)

    # log to S3 ingest log
    log_entry = {
        "timestamp": now, "sender": sender, "subject": subject,
        "s3_key": key, "containers_logged": len(all_containers),
    }
    log_key = f"ingest-log/{datetime.now(timezone.utc).strftime('%Y-%m-%d')}.jsonl"
    try:
        existing = b""
        try:
            existing = s3.get_object(Bucket=S3_BUCKET, Key=log_key)["Body"].read()
        except Exception:
            pass
        s3.put_object(
            Bucket=S3_BUCKET, Key=log_key,
            Body=existing + json.dumps(log_entry).encode() + b"\n"
        )
    except Exception as e:
        logger.warning(f"Could not write ingest log: {e}")

    send_reply(sender, subject, [r["container_id"] for r in all_containers], len(all_containers))

    return {"statusCode": 200, "body": f"logged {len(all_containers)} containers"}
