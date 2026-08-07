from __future__ import annotations
"""
utils.py — shared logic used by both app.py and Lambda functions.
Keeps container parsing, matching, and file parsing in one place.
"""
import re
import io
import openpyxl
import pandas as pd


# ── container ID helpers ───────────────────────────────────────────────────────

def normalize_container(cid: str) -> str:
    """Strip dashes and spaces so 'TCNU389902-4' and 'TCNU3899024' are the same."""
    return re.sub(r"[-\s]", "", str(cid).strip().upper())


def containers_match(query_norm: str, dbr_norm: str) -> bool:
    """Prefix-tolerant match — handles IDs submitted without the trailing check digit.
    'GCXU542076' matches 'GCXU5420760' because lengths differ by exactly 1.
    """
    if query_norm == dbr_norm:
        return True
    if abs(len(query_norm) - len(dbr_norm)) == 1:
        short, long_ = (
            (query_norm, dbr_norm) if len(query_norm) < len(dbr_norm)
            else (dbr_norm, query_norm)
        )
        return long_.startswith(short)
    return False


def parse_container_list(text: str) -> list:
    """Accept newline, comma, semicolon, or space-separated container IDs."""
    ids = re.split(r"[\n,;\s]+", text.strip())
    return [i.strip() for i in ids if i.strip()]


# ── carrier file parser ────────────────────────────────────────────────────────

def parse_carrier_file(file_bytes: bytes, filename: str) -> pd.DataFrame:
    """Parse any carrier submission Excel or CSV.

    Handles multiple formats:
    - AGL standard template (header on row 2, row 1 is a note)
    - Legacy ARVY / Carrier DBR formats (header on row 1 or 2)
    - CSV files

    Searches every sheet for ANY row containing 'Container' to find the header,
    then extracts container IDs and associated fields.
    Returns DataFrame with columns: container_id, sheet_source, terminal, status, notes.
    """
    rows_all = []

    if filename.lower().endswith(".csv"):
        try:
            df = pd.read_csv(io.BytesIO(file_bytes), dtype=str)
            cont_col = next((c for c in df.columns if "container" in c.lower()), None)
            if cont_col:
                for _, row in df.iterrows():
                    if pd.notna(row[cont_col]) and str(row[cont_col]).strip():
                        rows_all.append(_extract_row_fields(row.to_dict(), "Sheet1", cont_col))
        except Exception:
            pass
    else:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
        for sname in wb.sheetnames:
            ws = wb[sname]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # Find header row — scan first 8 rows for one containing 'Container'
            header_idx = next(
                (i for i, row in enumerate(rows[:8])
                 if any(c and "container" in str(c).lower() for c in row)),
                None
            )
            if header_idx is None:
                continue

            headers = [
                str(h).strip() if h and str(h).strip() else f"_col_{i}"
                for i, h in enumerate(rows[header_idx])
            ]
            cont_col = next((h for h in headers if "container" in h.lower()), None)
            if not cont_col:
                continue

            for row in rows[header_idx + 1:]:
                if not any(c for c in row if c is not None):
                    continue
                rec = dict(zip(headers, row))
                val = rec.get(cont_col)
                if not val or not str(val).strip() or str(val).strip() in ("None", "nan"):
                    continue
                rows_all.append(_extract_row_fields(rec, sname, cont_col))
        wb.close()

    if not rows_all:
        return pd.DataFrame()
    return pd.DataFrame(rows_all)


def _extract_row_fields(rec: dict, sheet: str, cont_col: str) -> dict:
    """Pull standardized fields from a raw carrier row dict."""
    def _find(keys):
        for k in rec:
            if any(kw in str(k).lower() for kw in keys):
                v = rec[k]
                if v and str(v).strip() not in ("None", "nan", ""):
                    return str(v).strip()
        return None

    return {
        "container_id":  normalize_container(str(rec[cont_col])),
        "sheet_source":  sheet,
        "terminal":      _find(["terminal", "port"]),
        "status":        _find(["status", "state"]),
        "notes":         _find(["note", "comment", "reason", "bridge"]),
    }


# ── Standard template full parser ─────────────────────────────────────────────

# Maps each sheet name to the columns we care about, with fallback aliases
TEMPLATE_SHEET_MAP = {
    "Static Delivery Plan": {
        "sheet_type": "delivery",
        "fields": {
            "container_id":    ["Container #", "Container"],
            "port":            ["Port"],
            "terminal":        ["Terminal", "Terminal "],
            "fc_building":     ["FC/Building", "FC/Building "],
            "flexi_id":        ["Flexi ID", "Flexi ID*"],
            "outgate_date":    ["Outgate Date"],
            "delivery_date":   ["Delivery Date/Time", "Delivery date/time"],
            "status":          ["Status"],
            "within_sla":      ["Within SLA?"],
            "sla_notes":       ["If No, Why?", "If no, why?"],
        },
    },
    # HUDD (Maersk) uses "Delivery Plan" instead of "Static Delivery Plan"
    "Delivery Plan": {
        "sheet_type": "delivery",
        "fields": {
            "container_id":    ["Container #", "Container"],
            "port":            ["Port"],
            "terminal":        ["Terminal", "Terminal "],
            "fc_building":     ["FC/Building", "FC/Building "],
            "flexi_id":        ["Flexi ID", "Flexi ID*"],
            "outgate_date":    ["Outgate Date"],
            "delivery_date":   ["Delivery Date/Time", "Delivery date/time"],
            "status":          ["Status"],
            "within_sla":      ["Within SLA?"],
            "sla_notes":       ["If No, Why?", "If no, why?"],
        },
    },
    # TGHE (Tighe) puts container data in the USBOS sheet
    "USBOS": {
        "sheet_type": "delivery",
        "fields": {
            "container_id":    ["Container #", "Container"],
            "port":            ["Port"],
            "terminal":        ["First Deliver to ", "First Deliver to", "2nd Del To ", "2nd Del To"],
            "outgate_date":    ["P/U from port Date", "Port Arrival Date"],
            "delivery_date":   ["Act Del Date", "Sched Del Date"],
            "notes":           ["Notes"],
        },
    },
    "Empty Returns": {
        "sheet_type": "empty_return",
        "fields": {
            "container_id":        ["Container #", "Container"],
            "terminal":            ["Terminal"],
            "empty_return_due":    ["Empty Return Due Date"],
            "appointment_date":    ["Appointment Date"],
            "status":              ["Status"],
            "sla_notes":           ["If Outside Window - Reason"],
        },
    },
    "ILM1": {
        "sheet_type": "delivery_ilm1",
        "fields": {
            "container_id":    ["Container #", "Container"],
            "port":            ["Arriving Port"],
            "outgate_date":    ["Arrive Date"],
            "delivery_date":   ["Act Del Date"],
            "status":          ["Container returned to Port"],
            "notes":           ["Notes"],
        },
    },
    "RIC6": {
        "sheet_type": "delivery_ric6",
        "fields": {
            "container_id":    ["Container #", "Container"],
            "port":            ["Arriving Port"],
            "outgate_date":    ["Arrive Date"],
            "delivery_date":   ["Act Del Date"],
            "status":          ["Container returned to Port"],
            "notes":           ["Notes"],
        },
    },
    "DBM6": {
        "sheet_type": "delivery_dbm6",
        "fields": {
            "container_id":    ["Container #", "Container"],
            "port":            ["Port", "Arriving Port"],
            "outgate_date":    ["Arrive Date", "Port Arrival Date"],
            "delivery_date":   ["Act Del Date"],
            "status":          ["Container returned to Port"],
            "notes":           ["Notes"],
        },
    },
    "Storage ODY": {
        "sheet_type": "ody",
        "fields": {
            "container_id":    ["Container #", "Container"],
            "terminal":        ["Yard", "Yard "],
            "outgate_date":    ["Pre-Pull Date"],
            "notes":           ["Days Stored", "Notes", "Bridge"],
        },
    },
    "Demurrage": {
        "sheet_type": "demurrage",
        "fields": {
            "container_id":    ["Container #", "Container"],
            "terminal":        ["Terminal", "Terminal "],
            "accessorial_type":["Type of Hold"],
            "notes":           ["Days in Hold", "Bridge", "Notes"],
        },
    },
    "Accessorials": {
        "sheet_type": "accessorial",
        "fields": {
            "container_id":    ["Container #", "Container"],
            "accessorial_type":["Accessorial Type", "Accessorial"],
            "outgate_date":    ["Date(s) Incurred"],
            "notes":           ["Reason", "Reason for Charge", "Notes"],
        },
    },
    "Accessorials Tracker": {
        "sheet_type": "accessorial",
        "fields": {
            "container_id":    ["Container #", "Container"],
            "accessorial_type":["Accessorial"],
            "outgate_date":    ["Date(s) Incurred"],
            "notes":           ["Reason for Charge", "Notes"],
        },
    },
}


def _resolve_field(headers: list, aliases: list) -> str | None:
    """Return the first header that matches any alias (case-insensitive strip)."""
    h_lower = {h.lower().strip(): h for h in headers}
    for alias in aliases:
        match = h_lower.get(alias.lower().strip())
        if match:
            return match
    return None


def parse_carrier_template(file_bytes: bytes, filename: str) -> dict:
    """Parse a carrier submission file (standard template OR legacy format).

    Returns a dict keyed by sheet_type, each value a list of row dicts.
    Row dicts use standardized keys from TEMPLATE_SHEET_MAP fields.
    """
    result = {}  # sheet_type -> list of row dicts

    if filename.lower().endswith(".csv"):
        # CSV — treat as Static Delivery Plan
        import pandas as _pd
        try:
            df = _pd.read_csv(io.BytesIO(file_bytes), dtype=str)
            rows = _parse_sheet_df(df, TEMPLATE_SHEET_MAP["Static Delivery Plan"]["fields"],
                                   "Static Delivery Plan")
            if rows:
                result["delivery"] = rows
        except Exception:
            pass
        return result

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
    for sname in wb.sheetnames:
        cfg = TEMPLATE_SHEET_MAP.get(sname)
        if not cfg:
            continue

        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Find header row (scan up to row 8)
        header_idx = next(
            (i for i, row in enumerate(rows[:8])
             if any(c and "container" in str(c).lower() for c in row)),
            None
        )
        if header_idx is None:
            continue

        headers = [str(h).strip() if h and str(h).strip() not in ("", "None") else f"_col_{i}"
                   for i, h in enumerate(rows[header_idx])]

        cont_aliases = cfg["fields"].get("container_id", ["Container #"])
        cont_col = _resolve_field(headers, cont_aliases)
        if not cont_col:
            continue

        parsed_rows = []
        for row in rows[header_idx + 1:]:
            if not any(c for c in row if c is not None):
                continue
            rec = dict(zip(headers, row))
            raw_id = rec.get(cont_col)
            if not raw_id or str(raw_id).strip() in ("", "None", "nan"):
                continue

            out = {"_raw_container": str(raw_id).strip(),
                   "container_id": normalize_container(str(raw_id))}
            for field_key, aliases in cfg["fields"].items():
                if field_key == "container_id":
                    continue
                col = _resolve_field(headers, aliases)
                if col:
                    val = rec.get(col)
                    out[field_key] = str(val).strip() if val and str(val).strip() not in ("None", "nan") else None
                else:
                    out[field_key] = None

            parsed_rows.append(out)

        if parsed_rows:
            sheet_type = cfg["sheet_type"]
            result.setdefault(sheet_type, []).extend(parsed_rows)

    wb.close()
    return result


def _parse_sheet_df(df, fields: dict, sname: str) -> list:
    """Helper: parse a pandas DataFrame using field aliases."""
    import pandas as _pd
    df.columns = [str(c).strip() for c in df.columns]
    cont_aliases = fields.get("container_id", ["Container #"])
    cont_col = _resolve_field(list(df.columns), cont_aliases)
    if not cont_col:
        return []
    rows = []
    for _, row in df.iterrows():
        raw_id = row.get(cont_col)
        if not raw_id or str(raw_id).strip() in ("", "None", "nan"):
            continue
        out = {"container_id": normalize_container(str(raw_id)),
               "_raw_container": str(raw_id).strip()}
        for field_key, aliases in fields.items():
            if field_key == "container_id":
                continue
            col = _resolve_field(list(df.columns), aliases)
            if col:
                val = row.get(col)
                out[field_key] = str(val).strip() if _pd.notna(val) else None
            else:
                out[field_key] = None
        rows.append(out)
    return rows
